
import numpy as np
import pandas as pd

import re
# ---------------------------------------------------------------------------
# Polars: imported with graceful degradation so the script continues to work
# even if polars is not installed in the current environment. Set
# _POLARS_AVAILABLE to False to force the fallback path during testing.
# ---------------------------------------------------------------------------
try:
    import polars as pl
    _POLARS_AVAILABLE = True
except ImportError:                          # pragma: no cover
    pl = None                                # type: ignore[assignment]
    _POLARS_AVAILABLE = False
    import warnings
    warnings.warn(
        "polars is not installed — Sprint-1 bridge helpers will fall back to "
        "Pandas.  Install with:  pip install polars",
        ImportWarning,
        stacklevel=2,
    )
from pathlib import Path

# ================================================================
# Centralized audit-output helpers (CSV/text)
#
# Many build_* functions write audit CSVs and notes. Centralizing
# the write behavior makes it easy to disable/redirect audit output
# and reduces duplicated boilerplate while preserving current behavior.
# ================================================================

# ================================================================
# Audit output control flags
# ================================================================
# AUDIT_WRITE_OUTPUTS (master switch)
#   True  — all audit CSV / text files are written (full audit mode).
#   False — nothing is written; overrides AUDIT_WRITE_INTERMEDIARY.
#
# AUDIT_WRITE_INTERMEDIARY (intermediary file gate)
#   True  — all intermediary CSV / text files are written alongside
#            the three primary output files.
#   False — only the three primary output files are written:
#               calc.csv
#               calc_avg_km.csv
#               calc_market_share.csv
#           All other audit_write_df() / audit_write_text() calls
#           (mode-level CSVs, long-format tables, notes files, etc.)
#           are silently skipped.  Set to True when you need to
#           inspect intermediate outputs during development.
# ================================================================
AUDIT_WRITE_OUTPUTS      = True   # master kill-switch (False → nothing writes)
AUDIT_WRITE_INTERMEDIARY = True  # False → skip all outputs except the 3 primary files

def _audit_enabled(enabled=None) -> bool:
    """Return whether audit output is enabled for intermediary files.

    Checks both the master switch (AUDIT_WRITE_OUTPUTS) and the
    intermediary gate (AUDIT_WRITE_INTERMEDIARY).  All audit_write_df()
    and audit_write_text() calls pass through this function, so setting
    AUDIT_WRITE_INTERMEDIARY = False silently suppresses every write
    EXCEPT the three primary output files (calc.csv, calc_avg_km.csv,
    calc_market_share.csv), which use _audit_enabled_primary() instead.

    If *enabled* is provided explicitly it takes precedence over both
    module-level flags (backward-compatible escape hatch).
    """
    if enabled is not None:
        return bool(enabled)
    g = globals()
    if not g.get('AUDIT_WRITE_OUTPUTS', True):
        return False   # master kill-switch
    if not g.get('AUDIT_WRITE_INTERMEDIARY', True):
        return False   # intermediary gate is closed
    return True

def _audit_enabled_primary() -> bool:
    """Return whether audit output is enabled for PRIMARY output files.

    Primary output files (calc.csv, calc_avg_km.csv, calc_market_share.csv)
    are always written as long as the master switch AUDIT_WRITE_OUTPUTS is
    True, regardless of AUDIT_WRITE_INTERMEDIARY.  This lets the user set
    AUDIT_WRITE_INTERMEDIARY = False to skip expensive intermediary CSVs
    without losing the three key output files.
    """
    return bool(globals().get('AUDIT_WRITE_OUTPUTS', True))


def _audit_resolve_path(path_or_name):
    """Resolve an audit output path.

    Accepts either a Path-like object or a filename. If OUT_DIR exists,
    filenames are resolved relative to OUT_DIR.
    """
    from pathlib import Path as _Path
    if isinstance(path_or_name, _Path):
        return path_or_name
    out_dir = globals().get('OUT_DIR', None)
    if out_dir is not None:
        try:
            return out_dir / str(path_or_name)
        except Exception:
            pass
    return _Path(str(path_or_name))

def _audit_write_csv_fast(df, path, fallback_kwargs=None, is_primary=False):
    """Sprint 5: write a CSV using pldf.write_csv() when Polars is available.

    Up to 3–5× faster than the pd.DataFrame.to_csv() path inside
    audit_write_df(), because Polars serialises column-by-column in Rust
    without Python object overhead.

    Parameters
    ----------
    df : pandas.DataFrame | polars.DataFrame
        The DataFrame to write.  Both types are accepted; a Pandas DataFrame
        is converted to Polars transparently via _pandas_to_pl().
    path : str | Path
        Output file path.  Parent directories are created automatically.
    fallback_kwargs : dict | None
        Extra keyword arguments forwarded to audit_write_df() when the Polars
        path is unavailable or raises.  Typical value: {'index': False}.

    Returns
    -------
    bool
        True  if the Polars fast path succeeded.
        False if audit_write_df() fallback was used instead.
    """
    if fallback_kwargs is None:
        fallback_kwargs = {'index': False}

    # is_primary=True uses _audit_enabled_primary() so AUDIT_WRITE_INTERMEDIARY
    # does not suppress the three key output files.
    _enabled = _audit_enabled_primary() if is_primary else _audit_enabled()
    if _POLARS_AVAILABLE and _enabled:
        try:
            from pathlib import Path as _P
            _path = _P(str(path))
            _path.parent.mkdir(parents=True, exist_ok=True)
            # Convert to Polars (no-op if already Polars).
            # cast_options kwarg avoids failures on pandas nullable Int64/boolean dtypes.
            _pl_df = _pandas_to_pl(df)
            _pl_df.write_csv(str(_path))
            return True
        except Exception as _s5_err:
            import warnings as _ws5
            _ws5.warn(
                f'[Sprint 5] Polars write_csv failed for {path} — '
                f'falling back to audit_write_df. Error: {_s5_err}',
                RuntimeWarning,
            )
    # Polars unavailable or failed — use Pandas path.
    # For primary files (is_primary=True) we pass enabled=True directly to
    # audit_write_df so the AUDIT_WRITE_INTERMEDIARY gate inside
    # _audit_enabled() cannot silently suppress the write.  This was the
    # root-cause bug: without enabled=True the fallback obeyed
    # AUDIT_WRITE_INTERMEDIARY=False and dropped even the 3 primary files.
    if is_primary:
        audit_write_df(df, path, enabled=True, **fallback_kwargs)
    else:
        audit_write_df(df, path, **fallback_kwargs)
    return False


def audit_write_df(df, *args, enabled=None, **kwargs):
    """Write a DataFrame for audit, respecting AUDIT_WRITE_OUTPUTS.

    This wrapper accepts the same positional/keyword arguments you would pass
    to df.to_csv() (and will fall back to df.to_excel() if CSV write fails).

    Polars support (Sprint 1)
    -------------------------
    A polars.DataFrame may be passed directly — it is converted to Pandas
    transparently before writing, so all existing call-sites work unchanged
    and new Sprint-2+ code can pass Polars DataFrames without extra steps.

    Examples:
      audit_write_df(df,    OUT_DIR / 'calc.csv', index=False)
      audit_write_df(df,    'calc.csv', index=False)
      audit_write_df(pl_df, OUT_DIR / 'calc.csv', index=False)  # Polars OK
    """
    if not _audit_enabled(enabled):
        return
    if df is None:
        return
    # Sprint 1: transparently accept Polars DataFrames at the audit boundary.
    # All downstream write logic (.to_csv / .to_excel) is Pandas-only, so we
    # convert here once rather than scattering isinstance checks everywhere.
    if _POLARS_AVAILABLE and isinstance(df, pl.DataFrame):
        df = df.to_pandas()
    if len(args) >= 1:
        args = (_audit_resolve_path(args[0]),) + tuple(args[1:])
        try:
            p = args[0]
            if hasattr(p, 'parent'):
                p.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
    try:
        return df.to_csv(*args, **kwargs)
    except Exception:
        try:
            return df.to_excel(*args, **kwargs)
        except Exception:
            return

def audit_write_text(text, path_or_name, enabled=None, encoding='utf-8', mode='w'):
    """Write a text file for audit, respecting AUDIT_WRITE_OUTPUTS."""
    if not _audit_enabled(enabled):
        return
    p = _audit_resolve_path(path_or_name)
    try:
        if hasattr(p, 'parent'):
            p.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    try:
        if mode == 'w' and hasattr(p, 'write_text'):
            return p.write_text(str(text), encoding=encoding)
        with open(p, mode, encoding=encoding) as f:
            f.write(str(text))
    except Exception:
        return


# =========================
# IN-MEMORY DATAFRAME REGISTRY (audit outputs are write-only)
# =========================
# All downstream calculations (incl. build_calc) MUST use these in-memory dataframes.
# Output CSVs are written for auditing only.
_DF_STORE = {}

def _register_df(key: str, df):
    """Store a dataframe in memory under a stable key (usually the out_file name)."""
    try:
        _DF_STORE[str(key)] = df.copy()
    except Exception:
        _DF_STORE[str(key)] = df

def _get_df(key: str, required: bool = True):
    k = str(key)
    if k in _DF_STORE:
        return _DF_STORE[k]
    if required:
        raise KeyError(
            f"Missing in-memory dataframe for key={k}. Ensure its build_* function ran before it is referenced."
        )
    return None
# =========================
# POLARS BRIDGE LAYER  (Sprint 1)
# =========================
# These five helpers form a zero-disruption bridge between the existing
# Pandas-backed _DF_STORE registry and new Polars-native code written in
# later sprints.  All existing callers of _register_df / _get_df continue
# to work unchanged because the registry still stores plain Pandas DataFrames.
#
# Usage guide
# -----------
#   READ  as Polars   ->  _get_df_polars("calc.csv")
#   WRITE from Polars ->  _register_df_polars("calc.csv", pl_df)
#   SAFE CONVERT      ->  _pandas_to_pl(df)  /  _pl_to_pandas(df)
# =========================

def _get_df_polars(key: str, required: bool = True):
    """Retrieve a DataFrame from _DF_STORE and return it as a Polars DataFrame.

    The registry always stores Pandas DataFrames (so all legacy callers keep
    working).  This helper converts transparently on retrieval.

    Parameters
    ----------
    key : str
        The stable registry key used by _register_df (usually the CSV
        filename).
    required : bool
        When True (default) raises KeyError if the key is absent.  When False
        returns None instead so callers can handle the missing-data case.

    Returns
    -------
    polars.DataFrame | None
        The stored data as a Polars DataFrame, or None if required=False and
        the key is absent.  Falls back to the raw Pandas DataFrame if polars
        is not installed in the current environment.

    Raises
    ------
    KeyError
        If required=True and the key is not found in _DF_STORE.
    """
    pdf = _get_df(key, required=required)
    if pdf is None:
        return None
    if not _POLARS_AVAILABLE:
        return pdf          # graceful degradation: return Pandas df as-is
    return pl.from_pandas(pdf)


def _register_df_polars(key: str, pldf):
    """Store a Polars DataFrame in _DF_STORE, converting it to Pandas first.

    Keeping the registry contents as Pandas DataFrames means every existing
    caller of _get_df() continues to receive a Pandas DataFrame — no other
    code needs to change.  New Sprint-2+ code can re-retrieve via
    _get_df_polars() when it needs the Polars version.

    Parameters
    ----------
    key : str
        Stable registry key (usually the audit CSV filename).
    pldf : polars.DataFrame | pandas.DataFrame
        The DataFrame to store.  If a Pandas DataFrame is passed by mistake,
        it is stored directly without conversion.
    """
    _register_df(key, _pl_to_pandas(pldf))


def _get_df_any_polars(keys, required: bool = True):
    """Return the first key that exists in _DF_STORE, as a Polars DataFrame.

    Useful when the same logical table might be registered under slightly
    different names across code paths (e.g. a regional variant vs. the
    national aggregate).

    Parameters
    ----------
    keys : Iterable[str]
        Candidate registry keys tried in order.
    required : bool
        When True raises KeyError if none of the keys are found.

    Returns
    -------
    polars.DataFrame | None
    """
    for k in keys:
        result = _get_df_polars(k, required=False)
        if result is not None:
            return result
    if required:
        raise KeyError(
            f"None of the keys {list(keys)} were found in _DF_STORE. "
            f"Ensure the relevant build_* function ran first."
        )
    return None


def _pl_to_pandas(pldf):
    """Safely convert a Polars DataFrame to a Pandas DataFrame.

    If the input is already a Pandas DataFrame it is returned unchanged, so
    callers do not need to branch on the type themselves.  This makes the
    helper safe to call in code paths that might receive either type.

    Parameters
    ----------
    pldf : polars.DataFrame | pandas.DataFrame

    Returns
    -------
    pandas.DataFrame
    """
    if _POLARS_AVAILABLE and isinstance(pldf, pl.DataFrame):
        return pldf.to_pandas()
    return pldf      # already Pandas (or Polars unavailable — graceful fallback)


def _pandas_to_pl(pdf):
    """Safely convert a Pandas DataFrame to a Polars DataFrame.

    If the input is already a Polars DataFrame it is returned unchanged.  If
    polars is not installed the original Pandas DataFrame is returned so the
    calling code degrades gracefully without raising an exception.

    Parameters
    ----------
    pdf : pandas.DataFrame | polars.DataFrame

    Returns
    -------
    polars.DataFrame | pandas.DataFrame
        Polars DataFrame when polars is available, otherwise the original
        Pandas DataFrame.
    """
    if not _POLARS_AVAILABLE:
        return pdf   # graceful degradation
    if isinstance(pdf, pl.DataFrame):
        return pdf   # already Polars
    return pl.from_pandas(pdf)


# =========================
# PATHS
# =========================
SCRIPT_DIR = Path(__file__).resolve().parent

CEUD_CAN_FILE = SCRIPT_DIR / "transCan2000-2022EN.xls"
if not CEUD_CAN_FILE.exists():
    raise FileNotFoundError(f"CEUD CAN input file not found: {CEUD_CAN_FILE}")

CEUD_BCTERR_FILE = SCRIPT_DIR / "transBCTerr2000-2022EN.xls"
if not CEUD_BCTERR_FILE.exists():
    raise FileNotFoundError(f"CEUD BCTerr input file not found: {CEUD_BCTERR_FILE}")

CEUD_ALB_FILE = SCRIPT_DIR / "transALB2000-2022EN.xls"
if not CEUD_ALB_FILE.exists():
    raise FileNotFoundError(f"CEUD Alberta input file not found: {CEUD_ALB_FILE}")

CEUD_ATL_FILE = SCRIPT_DIR / "transATL2000-2022EN.xls"
if not CEUD_ATL_FILE.exists():
    raise FileNotFoundError(f"CEUD AT input file not found: {CEUD_ATL_FILE}")

CEUD_MAN_FILE = SCRIPT_DIR / "transMAN2000-2022EN.xls"
if not CEUD_MAN_FILE.exists():
    raise FileNotFoundError(f"CEUD MB input file not found: {CEUD_MAN_FILE}")

CEUD_NB_FILE = SCRIPT_DIR / "transNB2000-2022EN.xls"
if not CEUD_NB_FILE.exists():
    raise FileNotFoundError(f"CEUD NB input file not found: {CEUD_NB_FILE}")

CEUD_NFLD_FILE = SCRIPT_DIR / "transNFLD2000-2022EN.xls"
if not CEUD_NFLD_FILE.exists():
    raise FileNotFoundError(f"CEUD NL input file not found: {CEUD_NFLD_FILE}")

CEUD_NS_FILE = SCRIPT_DIR / "transNS2000-2022EN.xls"
if not CEUD_NS_FILE.exists():
    raise FileNotFoundError(f"CEUD NS input file not found: {CEUD_NS_FILE}")

CEUD_ONT_FILE = SCRIPT_DIR / "transONT2000-2022EN.xls"
if not CEUD_ONT_FILE.exists():
    raise FileNotFoundError(f"CEUD ON input file not found: {CEUD_ONT_FILE}")

CEUD_PEI_FILE = SCRIPT_DIR / "transPEI2000-2022EN.xls"
if not CEUD_PEI_FILE.exists():
    raise FileNotFoundError(f"CEUD PE input file not found: {CEUD_PEI_FILE}")

CEUD_QUE_FILE = SCRIPT_DIR / "transQUE2000-2022EN.xls"
if not CEUD_QUE_FILE.exists():
    raise FileNotFoundError(f"CEUD QC input file not found: {CEUD_QUE_FILE}")

CEUD_SASK_FILE = SCRIPT_DIR / "transSASK2000-2022EN.xls"
if not CEUD_SASK_FILE.exists():
    raise FileNotFoundError(f"CEUD SK input file not found: {CEUD_SASK_FILE}")

OUT_DIR = SCRIPT_DIR / "output"
OUT_DIR.mkdir(exist_ok=True)

# =========================
# CONSTANTS
# =========================
YEARS = list(range(2000, 2023))
N_YEARS = len(YEARS)

FUELS = [
    "Aviation turbo fuel",
    "Aviation gasoline",
    "Diesel fuel oil",
    "Biodiesel fuel",
    "Motor gasoline",
    "Ethanol",
    "Electricity",
    "Natural gas",
    "Heavy fuel oil",
    "Propane",
]

P1_SHEET = "Passenger1"
P4_SHEET = "Passenger4"
F1_SHEET = "Freight1"

YEAR_START_COL = "C"
YEAR_END_COL = "Y"

FUEL_SCALE = 1000.0  # PJ → TJ


# ================================================================
# MACRO INPUTS — Price Multiplier Extraction
# ================================================================
# This section loads the Macro Inputs workbook and provides
# multiplier extraction functions for all transportation personal
# fuels.  Multipliers are sourced from the following sheets:
#
#   Prices      — AB Transportation Personal (5-year benchmarks,
#                 interpolated to annual)
#   CER         — Alberta Transportation final user prices
#                 (annual, used for Diesel/Gasoline CER path)
#   AFDC        — CIMS production cost benchmarks for NG,
#                 Ethanol, Biodiesel (5-year, interpolated)
#   Elec markups— Provincial electricity price multipliers
#
# All other JCIMS fuels (Biogas, Coal, Coke, etc.) that do not
# appear in the available sheets fall back to a multiplier of 1.0.
# ================================================================

# Path to macro inputs file (SCRIPT_DIR already defined above)
MACRO_INPUTS_FILE = SCRIPT_DIR / "macro inputs.xlsx"


def load_macro_inputs():
    """Load all relevant Macro Inputs sheets as Pandas DataFrames.

    Returns
    -------
    dict[str, pd.DataFrame]
        Dictionary keyed by sheet name.  Only the four sheets
        required for multiplier extraction are retained.
    """
    sheets_needed = ["AFDC", "CER", "Prices", "Elec markups"]
    all_sheets    = pd.read_excel(MACRO_INPUTS_FILE, sheet_name=None,
                                  engine="openpyxl")
    return {k: v for k, v in all_sheets.items() if k in sheets_needed}


# Load once at module level so every downstream function can use it.
macro_inputs = load_macro_inputs()




# ================================================================

# ================================================================
# MACRO INPUTS — Excel-exact named-range cache (openpyxl)
# ================================================================
# To match the province formula files EXACTLY, we mirror the workbook's
# INDEX/MATCH/XMATCH behavior against named ranges in macro inputs.xlsx.
#
# Fixes applied (based on macro_cache_diagnostics.txt):
#   1) Correctly flatten 1xN named ranges (e.g., CER_year = Prices!K2:U2).
#   2) Use full Target strings in keys (e.g., CIMS.Generic Fuels.Diesel),
#      matching the workbook's concatenations in Prices/CER/JCIMS.
#   3) Electricity uses sector_CER_elec mapping + elec_price_mult_index row vector.
#
# Benchmark-year logic only (matches reference CSVs): 2000/2005/2010/2015/2020.
# ================================================================

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

BENCHMARK_YEARS = [2000, 2005, 2010, 2015, 2020]

# Prices sheet benchmark column mapping used by the formula files:
# 2000 -> K, 2005 -> L, 2010 -> M, 2015 -> N, 2020 -> O
# (Note: Prices!O header is 2021 in macro inputs, but formula files use column O for the 2020 benchmark.)
_PRICES_BENCH_COL = {
    2000: 'K',
    2005: 'L',
    2010: 'M',
    2015: 'N',
    2020: 'O',
    2025: 'P',
    2030: 'Q',
    2035: 'R',
    2040: 'S',
    2045: 'T',
    2050: 'U',
}


def _clean_scalar(x):
    if x is None:
        return None
    s = str(x).strip()
    if s == '' or s.lower() in {'n.a.', 'na', 'nan', 'none'}:
        return None
    return x


def _as_float(x):
    x = _clean_scalar(x)
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        return None


def _flatten_1d(arr):
    """Convert an openpyxl named-range array to a 1D Python list.

    Handles:
      - 1xN row vector (e.g., CER_year = Prices!K2:U2)
      - Nx1 column vector
      - scalar
    """
    if arr is None:
        return []
    try:
        if len(arr) == 0:
            return []
        if len(arr) == 1 and isinstance(arr[0], list):
            return list(arr[0])
        if all(isinstance(r, list) and len(r) == 1 for r in arr):
            return [r[0] for r in arr]
        out = []
        for r in arr:
            if isinstance(r, list):
                out.extend(r)
            else:
                out.append(r)
        return out
    except Exception:
        return []


def _match(value, vector):
    value = str(value).strip()
    for i, v in enumerate(vector, start=1):
        if v is None:
            continue
        if str(v).strip() == value:
            return i
    return None


def _match_int(value, vector):
    try:
        value_i = int(value)
    except Exception:
        return None
    for i, v in enumerate(vector, start=1):
        if v is None:
            continue
        try:
            if int(float(str(v))) == value_i:
                return i
        except Exception:
            continue
    return None


def _xmatch(value, vector):
    return _match(value, vector)


class MacroInputsCache:
    """Cache macro inputs named ranges and key lookup helpers (Excel-exact)."""

    def __init__(self, xlsx_path: Path):
        self.xlsx_path = Path(xlsx_path)
        self._wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        self._named = {}

        # Named ranges
        self.cer_prices = None
        self.cer_prices_index = []
        self.cer_year = []

        self.sector_cims = []
        self.sector_cer = []
        self.sector_cer_elec = []
        self.sector_map = {}
        self.sector_elec_map = {}

        self.region_cims = []
        self.elec_price_mult = None
        self.elec_price_mult_index = []

        self.fuelmult_jcims = None
        self.fuelmult_jcims_index = []
        self.fuelmult_years = [2000, 2005, 2010, 2015, 2020, 2025, 2030, 2035, 2040, 2045, 2050]

        # Sheets
        self.prices_ws = self._wb['Prices'] if 'Prices' in self._wb.sheetnames else None
        self.afdc_ws = self._wb['AFDC'] if 'AFDC' in self._wb.sheetnames else None
        self._prices_index = {}

        self._load_core_named_ranges()
        self._build_prices_index()

    def _get_named_range(self, name: str):
        if name in self._named:
            return self._named[name]
        dn = self._wb.defined_names.get(name)
        if dn is None:
            self._named[name] = None
            return None
        dests = list(dn.destinations)
        if not dests:
            self._named[name] = None
            return None
        sheet_name, coord = dests[0]
        ws = self._wb[sheet_name]
        cells = ws[coord]
        if isinstance(cells, tuple):
            if cells and isinstance(cells[0], tuple):
                arr = [[_clean_scalar(c.value) for c in row] for row in cells]
            else:
                arr = [[_clean_scalar(c.value) for c in cells]]
        else:
            arr = [[_clean_scalar(cells.value)]]
        self._named[name] = arr
        return arr

    def _load_core_named_ranges(self):
        # CER
        self.cer_prices = self._get_named_range('CER_prices')
        self.cer_prices_index = _flatten_1d(self._get_named_range('CER_prices_index'))
        self.cer_year = _flatten_1d(self._get_named_range('CER_year'))

        # Sector maps
        self.sector_cims = _flatten_1d(self._get_named_range('sector_CIMS'))
        self.sector_cer = _flatten_1d(self._get_named_range('sector_CER'))
        self.sector_cer_elec = _flatten_1d(self._get_named_range('sector_CER_elec'))

        self.sector_map = {}
        for a, b in zip(self.sector_cims, self.sector_cer):
            if a is None or b is None:
                continue
            self.sector_map[str(a).strip()] = str(b).strip()

        self.sector_elec_map = {}
        for a, b in zip(self.sector_cims, self.sector_cer_elec):
            if a is None or b is None:
                continue
            self.sector_elec_map[str(a).strip()] = str(b).strip()

        # Electricity
        self.region_cims = _flatten_1d(self._get_named_range('region_CIMS'))
        self.elec_price_mult_index = _flatten_1d(self._get_named_range('elec_price_mult_index'))
        self.elec_price_mult = self._get_named_range('elec_price_mult')

        # JCIMS
        self.fuelmult_jcims = self._get_named_range('FuelMult_JCIMS')
        self.fuelmult_jcims_index = _flatten_1d(self._get_named_range('FuelMult_JCIMS_Index'))

    def _build_prices_index(self):
        self._prices_index = {}
        if self.prices_ws is None:
            return
        cj_col = column_index_from_string('CJ')
        for r in range(29, 211):
            key = _clean_scalar(self.prices_ws.cell(r, cj_col).value)
            if key is None:
                continue
            self._prices_index[str(key).strip()] = r

    def cer_price(self, region_code: str, sector_cims: str, fuel_target: str, year: int):
        if self.cer_prices is None or not self.cer_prices_index or not self.cer_year:
            return None
        sec_cer = self.sector_map.get(str(sector_cims).strip(), str(sector_cims).strip())
        key = f"{region_code}{sec_cer}{fuel_target}"
        i = _match(key, self.cer_prices_index)
        j = _match_int(year, self.cer_year)
        if i is None or j is None:
            return None
        try:
            return _as_float(self.cer_prices[i-1][j-1])
        except Exception:
            return None

    def prices_can_transport(self, fuel_target: str, year_label: int):
        if self.prices_ws is None or not self._prices_index:
            return None
        col_letter = _PRICES_BENCH_COL.get(int(year_label))
        if not col_letter:
            return None
        col = column_index_from_string(col_letter)
        key = f"CANTransportation{fuel_target}"
        row = self._prices_index.get(key)
        if row is None:
            return None
        return _as_float(self.prices_ws.cell(row, col).value)

    def fuelmult_jcims_value(self, region_code: str, sector_cims: str, fuel_target: str, year_label: int):
        if self.fuelmult_jcims is None or not self.fuelmult_jcims_index:
            return 1.0
        key = f"{region_code}{sector_cims}{fuel_target}"
        i = _match(key, self.fuelmult_jcims_index)
        j = _match_int(year_label, self.fuelmult_years)
        if i is None or j is None:
            return 1.0
        try:
            v = _as_float(self.fuelmult_jcims[i-1][j-1])
            return float(v) if v is not None else 1.0
        except Exception:
            return 1.0

    def elec_multiplier(self, region_code: str, sector_cims: str):
        if self.elec_price_mult is None or not self.region_cims or not self.elec_price_mult_index:
            return None
        sec_elec = self.sector_elec_map.get(str(sector_cims).strip(), str(sector_cims).strip())
        i = _xmatch(region_code, self.region_cims)
        j = _xmatch(sec_elec, self.elec_price_mult_index)
        if i is None or j is None:
            return None
        try:
            return _as_float(self.elec_price_mult[i-1][j-1])
        except Exception:
            return None

    def afdc_value(self, fuel_lc: str, year_label: int):
        if self.afdc_ws is None:
            return None
        years = [_clean_scalar(self.afdc_ws.cell(r, column_index_from_string('M')).value) for r in range(55, 66)]
        k = _match_int(year_label, years)
        if k is None:
            return None
        col_letter = None
        if fuel_lc == 'natural gas':
            col_letter = 'P'
        elif fuel_lc == 'ethanol':
            col_letter = 'R'
        elif fuel_lc == 'biodiesel':
            col_letter = 'S'
        else:
            return None
        r = 55 + (k - 1)
        return _as_float(self.afdc_ws.cell(r, column_index_from_string(col_letter)).value)


try:
    _MACRO_CACHE = MacroInputsCache(MACRO_INPUTS_FILE)
except Exception:
    _MACRO_CACHE = None


def _macro_cache_diagnostics_text(cache, region_code='AB', sector_cims='Transportation Personal'):
    lines = []
    lines.append('Macro Inputs Cache Diagnostics')
    lines.append('============================')
    lines.append(f'macro inputs path: {getattr(cache, "xlsx_path", "")}')
    try:
        lines.append(f'Workbook sheets: {len(cache._wb.sheetnames)} -> {cache._wb.sheetnames}')
    except Exception:
        pass
    lines.append('')
    lines.append(f'CER_year len={len(cache.cer_year)} preview={cache.cer_year[:12]}')
    lines.append(f'elec_price_mult_index len={len(cache.elec_price_mult_index)} preview={cache.elec_price_mult_index[:12]}')
    lines.append(f'prices_index entries={len(cache._prices_index)}')
    for k in [
        'CANTransportationCIMS.Generic Fuels.Diesel',
        'CANTransportationCIMS.Generic Fuels.Gasoline',
    ]:
        lines.append(f"Prices key {k} -> row {cache._prices_index.get(k)}")
    lines.append('')

    years = BENCHMARK_YEARS
    elec = cache.elec_multiplier(region_code, sector_cims)
    lines.append(f'Electricity multiplier for {region_code}/{sector_cims}: {elec}')

    for fuel_target in ['CIMS.Generic Fuels.Diesel', 'CIMS.Generic Fuels.Gasoline']:
        for y in years:
            num = cache.cer_price(region_code, sector_cims, fuel_target, y)
            den_raw = cache.prices_can_transport(fuel_target, y)
            den = round(float(den_raw), 2) if den_raw is not None else None
            mult = (float(num) / float(den)) if (num is not None and den not in (None, 0)) else None
            lines.append(f'{fuel_target} {y}: num={num} den_raw={den_raw} den_round={den} mult={mult}')

    for fuel_target in ['CIMS.Generic Fuels.Propane', 'CIMS.Generic Fuels.Jet Fuel']:
        for y in years:
            v = cache.fuelmult_jcims_value(region_code, sector_cims, fuel_target, y)
            lines.append(f'JCIMS {fuel_target} {y}: {v}')

    return "\n".join(lines)
    # (patched)


def write_macro_cache_diagnostics(path=None, region_code='AB', sector_cims='Transportation Personal'):
    try:
        if _MACRO_CACHE is None:
            return
        p = Path(path) if path is not None else (OUT_DIR / 'macro_cache_diagnostics.txt')
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        p.write_text(_macro_cache_diagnostics_text(_MACRO_CACHE, region_code=region_code, sector_cims=sector_cims), encoding='utf-8')
    except Exception:
        return

write_macro_cache_diagnostics()


_REGION_FULL_NAME = {
    "AB":  "Alberta",
    "BC":  "British Columbia",
    "SK":  "Saskatchewan",
    "MB":  "Manitoba",
    "ON":  "Ontario",
    "QC":  "Quebec",
    "NB":  "New Brunswick",
    "NS":  "Nova Scotia",
    "PE":  "Prince Edward Island",
    "NL":  "Newfoundland and Labrador",
    "AT":  "Atlantic",
    "TR":  "Territories",
    "CAN": "Canada",
}

# Fuels that have a direct entry in the Prices / Transportation Personal block.
# All other fuels fall back to 1.0.
_PRICES_TP_FUELS = {
    "biodiesel",
    "ethanol",
    "electricity",
    "jet fuel",
    "diesel",
    "gasoline",
    "natural gas",
    "propane",
}


def _prices_multiplier(region_full, sector_full, fuel_name, year, prices_df):
    """Look up a multiplier from the Prices sheet and interpolate to
    annual resolution.

    The Prices sheet stores values at 5-year benchmarks (2000, 2005,
    2010, 2015, 2020/2021, 2025 …).  This helper reindexes to a full
    annual range and uses linear interpolation so any year between
    2000 and 2060 returns a value.

    Parameters
    ----------
    region_full : str   Full region name, e.g. "Alberta".
    sector_full : str   Full sector name, e.g. "Transportation Personal".
    fuel_name   : str   Fuel label as it appears in the Prices sheet.
    year        : int   Target year (2000–2060).
    prices_df   : pd.DataFrame  The Prices sheet from Macro Inputs.

    Returns
    -------
    float
        Interpolated multiplier, or 1.0 if not found.
    """
    try:
        mask = (
            prices_df["Region"].astype(str).str.strip().str.lower()
                == region_full.lower()
        ) & (
            prices_df["Sector"].astype(str).str.strip().str.lower()
                == sector_full.lower()
        ) & (
            prices_df["Fuel"].astype(str).str.strip().str.lower()
                == fuel_name.lower()
        )
        row = prices_df[mask]
        if row.empty:
            return 1.0

        # Extract benchmark year columns (integer column names 1995–2060)
        yc = {
            int(c): float(v)
            for c, v in row.iloc[0].items()
            if isinstance(c, (int, float))
            and not pd.isna(v)
            and 1995 <= int(c) <= 2060
        }
        if not yc:
            return 1.0

        s        = pd.Series(yc, dtype=float).sort_index()
        full_idx = pd.RangeIndex(int(s.index.min()), int(s.index.max()) + 1)
        s_annual = s.reindex(full_idx).interpolate(method="index")
        val      = s_annual.get(year)
        return float(val) if val is not None and not pd.isna(val) else 1.0

    except Exception:
        return 1.0


def _cer_price(region_full, sector_full, fuel_name, year, cer_df):
    """Return the CER final user price for a given region/sector/fuel/year.

    The CER sheet is already annual (1995–2060) so no interpolation needed.

    Parameters
    ----------
    region_full : str   Full region name, e.g. "Alberta".
    sector_full : str   CER sector label, e.g. "Transportation".
    fuel_name   : str   Fuel label as it appears in the CER sheet.
    year        : int   Target year.
    cer_df      : pd.DataFrame  The CER sheet from Macro Inputs.

    Returns
    -------
    float or None
        CER price for the given year, or None if not found.
    """
    try:
        mask = (
            cer_df["Region"].astype(str).str.strip().str.lower()
                == region_full.lower()
        ) & (
            cer_df["Sector"].astype(str).str.strip().str.lower()
                == sector_full.lower()
        ) & (
            cer_df["Fuel"].astype(str).str.strip().str.lower()
                == fuel_name.lower()
        )
        row = cer_df[mask]
        if row.empty:
            return None

        yc  = {
            int(c): float(v)
            for c, v in row.iloc[0].items()
            if isinstance(c, (int, float))
            and not pd.isna(v)
            and 1995 <= int(c) <= 2060
        }
        val = yc.get(year)
        return float(val) if val is not None else None

    except Exception:
        return None


def get_macro_multiplier(region, sector, fuel, year, macro_inputs_dict=None):
    """Excel-exact multiplier lookup for Transportation Personal (benchmark years only)."""
    if _MACRO_CACHE is None:
        return 1.0

    region_code = str(region).strip().upper()
    sector_cims = str(sector).strip()
    fuel_in = str(fuel).strip()

    try:
        year_i = int(year)
    except Exception:
        return 1.0

    if year_i not in BENCHMARK_YEARS and year_i not in _MACRO_CACHE.fuelmult_years:
        return 1.0

    fuel_lc = fuel_in.lower()

    # AFDC fuels
    if fuel_lc in {'biodiesel', 'ethanol', 'natural gas'}:
        v = _MACRO_CACHE.afdc_value(fuel_lc, year_i)
        return float(v) if v is not None else 1.0

    # Electricity
    if fuel_lc == 'electricity':
        v = _MACRO_CACHE.elec_multiplier(region_code, sector_cims)
        return float(v) if v is not None else 1.0

    # Map CEUD-ish labels to workbook Target strings
    fuel_target_map = {
        'diesel': 'CIMS.Generic Fuels.Diesel',
        'diesel fuel oil': 'CIMS.Generic Fuels.Diesel',
        'gasoline': 'CIMS.Generic Fuels.Gasoline',
        'motor gasoline': 'CIMS.Generic Fuels.Gasoline',
        'propane': 'CIMS.Generic Fuels.Propane',
        'jet fuel': 'CIMS.Generic Fuels.Jet Fuel',
        'aviation turbo fuel': 'CIMS.Generic Fuels.Jet Fuel',
    }
    fuel_target = fuel_target_map.get(fuel_lc, fuel_in)

    if fuel_target in {'CIMS.Generic Fuels.Diesel', 'CIMS.Generic Fuels.Gasoline'}:
        num = _MACRO_CACHE.cer_price(region_code, sector_cims, fuel_target, year_i)
        if num is None:
            return 1.0
        den_raw = _MACRO_CACHE.prices_can_transport(fuel_target, year_i)
        if den_raw is None:
            return 1.0
        den = round(float(den_raw), 2)
        if den == 0:
            return 1.0
        return float(num) / float(den)

    return float(_MACRO_CACHE.fuelmult_jcims_value(region_code, sector_cims, fuel_target, year_i))


def build_multiplier_table(regions=None, years=None, output_csv=None):
    """Build a combined DataFrame of all Macro Inputs price multipliers
    for the Transportation Personal sector across all required provinces.

    Parameters
    ----------
    regions    : list[str], optional
        Province codes to include. Defaults to all 7 transportation
        personal provinces: AB, AT, BC, MB, ON, QC, SK.
    years      : list[int], optional
        Calendar years to include. Defaults to the module-level YEARS
        (2000–2022).
    output_csv : str or Path, optional
        If provided, the combined DataFrame is written to this path
        as a single CSV file.

    Returns
    -------
    pd.DataFrame
        Columns: Region, Sector, Fuel, Year, Multiplier.
        All provinces stacked in one tidy table.
    """
    if regions is None:
        regions = globals().get('_TRANSPORT_PERSONAL_PROVINCES', ['AB','AT','BC','MB','ON','QC','SK'])
    if years is None:
        years = BENCHMARK_YEARS  # benchmark years only (matches formula/reference files)

    sector = "Transportation Personal"

    fuels = [
        # --- Prices / TP — real interpolated values ---
        "Biodiesel",
        "Ethanol",
        "Natural Gas",
        "Diesel",
        "Gasoline",
        "Electricity",
        "Jet Fuel",
        "Propane",
        # --- JCIMS fuels — currently default 1.0 ---
        "Biogas",
        "Black Liquor",
        "Coal",
        "Coke",
        "Fuel Oil",
        "Hydrogen",
        "LPG",
        "Petroleum Coke",
        "Refinery Fuel Gas",
        "Solid Biomass",
        "Uranium",
        "Waste Fuel",
    ]

    all_records = []
    for region in regions:
        for fuel in fuels:
            for year in years:
                multiplier = get_macro_multiplier(
                    region, sector, fuel, year
                )
                all_records.append({
                    "Region":     region,
                    "Sector":     sector,
                    "Fuel":       fuel,
                    "Year":       year,
                    "Multiplier": multiplier,
                })

    df = pd.DataFrame(all_records)

    if output_csv is not None:
        try:
            out_path = Path(str(output_csv))
            out_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(out_path, index=False)
        except Exception as _e:
            import warnings as _w
            _w.warn(
                f"[build_multiplier_table] Could not write CSV: {_e}",
                RuntimeWarning,
            )
    # PRIMARY write for audit: always write multiplier table when output_csv is provided
    if output_csv is not None:
        try:
            _audit_write_csv_fast(df_all if 'df_all' in locals() else df, output_csv, is_primary=True)
        except Exception:
            try:
                audit_write_df(df_all if 'df_all' in locals() else df, output_csv, enabled=True, index=False)
            except Exception:
                pass


    return df
# Build and write the combined all-province multiplier table
# NOTE: build_multiplier_table() call removed from import-time execution.
# =========================
# EXCEL COM HELPERS
# =========================

def _get_df_any(keys, required: bool = True):
    """Return the first in-memory dataframe found for any of *keys*.

    Enforces dataframe-only policy: no disk fallback is attempted.
    """
    for k in keys:
        kk = str(k)
        if kk in _DF_STORE:
            return _DF_STORE[kk]
    if required:
        raise KeyError(
            f"Missing in-memory dataframe for any of keys={list(map(str, keys))}. "
            "Ensure its build_* function ran before it is referenced."
        )
    return None

def _excel_app():
    import win32com.client as win32
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    return excel

def _open_book(excel, path: Path):
    return excel.Workbooks.Open(str(path), ReadOnly=True)

def _close_book(wb, excel):
    try:
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()

def _sheet_names(wb):
    return [wb.Worksheets(i).Name for i in range(1, wb.Worksheets.Count + 1)]

def _get_ws(wb, preferred_names=None, contains=None, required=True):
    preferred_names = preferred_names or []

    # 1) exact
    for nm in preferred_names:
        try:
            return wb.Worksheets(nm)
        except Exception:
            pass

    # 2) case-insensitive equal
    names = _sheet_names(wb)
    lower_map = {n.lower(): n for n in names}
    for nm in preferred_names:
        if nm is None:
            continue
        key = str(nm).lower()
        if key in lower_map:
            return wb.Worksheets(lower_map[key])

    # 3) contains
    if contains:
        c = str(contains).lower()
        for n in names:
            if c in n.lower():
                return wb.Worksheets(n)

    if required:
        raise ValueError(
            "Could not resolve required worksheet. "
            f"preferred_names={preferred_names}, contains={contains}. "
            f"Available sheets: {names}"
        )
    return None

def _range_values(ws, a1_range):
    vals = ws.Range(a1_range).Value
    if not isinstance(vals, tuple):
        return ((vals,),)
    if vals and not isinstance(vals[0], tuple):
        return (vals,)
    return vals

def _to_float_array(values):
    out = []
    for v in values:
        if v is None:
            out.append(np.nan)
        else:
            s = str(v).strip().lower()
            if s in ("n.a.", "na", ""):
                out.append(np.nan)
            else:
                try:
                    out.append(float(v))
                except Exception:
                    out.append(np.nan)
    return np.array(out, dtype=float)

def _read_year_row(ws, row):
    rng = f"{YEAR_START_COL}{row}:{YEAR_END_COL}{row}"
    arr = _to_float_array(_range_values(ws, rng)[0])
    if arr.size != N_YEARS:
        raise ValueError(f"Expected {N_YEARS} years, got {arr.size} from {ws.Name}!{rng}")
    return pd.Series(arr, index=YEARS)

def _assert_no_blanks(series, name):
    missing = series.index[series.isna()].tolist()
    if missing:
        raise ValueError(f"STRICT VALIDATION FAILED — {name} missing at years {missing}")

def _fuel_handle_na(s, fuel, notes):
    if s.isna().all():
        notes.append(f"[INFO] {fuel}: no data → filled with zeros.")
        return pd.Series(np.zeros(len(s)), index=s.index)
    if s.isna().any():
        yrs = s.index[s.isna()].tolist()
        notes.append(f"[INFO] {fuel}: n.a. at years {yrs} → preserved as NULL.")
        return s
    return s

def _find_row(ws, needle: str):
    used = ws.UsedRange
    found = used.Find(What=needle, LookAt=2)
    if found is None:
        raise ValueError(f"Could not find '{needle}' in sheet '{ws.Name}'")
    return found.Row

# =========================
# GENERIC MODE BUILDER (CAN)
# =========================
# =========================
# CAR 2022 FINAL OUTPUT OVERRIDES
# =========================
def _overwrite_car_2022_output_df(df, can_occ, notes=None, prefix="Car"):
    """Final, non-overwritable Car 2022 patch applied immediately before CSV write.

    Guarantees that after all CEUD reads and calculations are finished, the Car output DataFrame
    itself is overwritten for 2022 using the literal workbook formulas:
      - Total Distance (M*vkm) = 2019 Total Distance * 0.8
      - Fuel (TJ)::Motor gasoline = 2019 Motor gasoline * 0.8

    Then all downstream output columns that depend on those cells are recomputed directly on the
    final DataFrame so nothing later can overwrite them before the CSV is written.
    """
    y = 2022
    y_ref = 2019
    td_col = "Total Distance (M*vkm)"
    act_col = "Activity (millions passenger-kilometres)"
    occ_col = "Occupancy (persons /vehicle"
    av_col = "Average Distance (vkm)"
    apk_col = "Average Distance (pkm)"
    mg_col = "Fuel (TJ)::Motor gasoline"
    ft_col = "Fuel (TJ)::Total"
    int_col = "Intensity (GJ / pkm)"

    if y not in df.index or y_ref not in df.index:
        return df
    required = [td_col, act_col, occ_col, av_col, apk_col, mg_col, ft_col, int_col]
    if any(c not in df.columns for c in required):
        return df

    td_2019 = pd.to_numeric(pd.Series([df.loc[y_ref, td_col]]), errors='coerce').iloc[0]
    mg_2019 = pd.to_numeric(pd.Series([df.loc[y_ref, mg_col]]), errors='coerce').iloc[0]
    occ_2022 = pd.to_numeric(pd.Series([can_occ.loc[y]]), errors='coerce').iloc[0] if y in can_occ.index else np.nan
    av_2022 = pd.to_numeric(pd.Series([df.loc[y, av_col]]), errors='coerce').iloc[0]

    if pd.notna(td_2019):
        df.loc[y, td_col] = td_2019 * 0.8
    if pd.notna(mg_2019):
        df.loc[y, mg_col] = mg_2019 * 0.8

    td_2022 = pd.to_numeric(pd.Series([df.loc[y, td_col]]), errors='coerce').iloc[0]
    df.loc[y, act_col] = td_2022 * occ_2022 if pd.notna(td_2022) and pd.notna(occ_2022) else np.nan
    act_2022 = pd.to_numeric(pd.Series([df.loc[y, act_col]]), errors='coerce').iloc[0]
    df.loc[y, occ_col] = (act_2022 / td_2022) if pd.notna(act_2022) and pd.notna(td_2022) and td_2022 != 0 else np.nan
    occ_out_2022 = pd.to_numeric(pd.Series([df.loc[y, occ_col]]), errors='coerce').iloc[0]
    df.loc[y, apk_col] = av_2022 * occ_out_2022 if pd.notna(av_2022) and pd.notna(occ_out_2022) else np.nan

    fuel_cols = [c for c in df.columns if c.startswith("Fuel (TJ)::") and c != ft_col]
    df.loc[y, ft_col] = pd.to_numeric(df.loc[y, fuel_cols], errors='coerce').sum(min_count=1)
    ft_2022 = pd.to_numeric(pd.Series([df.loc[y, ft_col]]), errors='coerce').iloc[0]
    df.loc[y, int_col] = (ft_2022 / act_2022 / 1000.0) if pd.notna(ft_2022) and pd.notna(act_2022) and act_2022 != 0 else np.nan

    for c in [c for c in df.columns if c.startswith("Share (%)::")]:
        fuel_name = c.split("::", 1)[1]
        fuel_col = f"Fuel (TJ)::" + fuel_name
        if fuel_col in df.columns:
            fuel_val = pd.to_numeric(pd.Series([df.loc[y, fuel_col]]), errors='coerce').iloc[0]
            df.loc[y, c] = (fuel_val / ft_2022 * 100.0) if pd.notna(fuel_val) and pd.notna(ft_2022) and ft_2022 != 0 else np.nan

    if notes is not None:
        notes.append(f"[INFO] {prefix} 2022 final output overwrite applied immediately before CSV write: Total Distance=2019*0.8 and Motor gasoline=2019*0.8.")
    return df

def _apply_urban_transit_2022_override_df(df, notes=None, context=""):
    """Apply workbook 2022 special formulas for Urban Transit.

    2022 overrides (all based on 2019):
      - Activity (millions passenger-kilometres): 2022 = 2019 * 0.8
      - Stock (thousands): 2022 = 2019
      - Average Distance (vkm): 2022 = 2019
      - Diesel fuel oil (Fuel (TJ)::Diesel fuel oil): 2022 = 2019

    Recomputes derived 2022 metrics:
      Total Distance (M*vkm) = Stock * AvgVkm / 1000
      Occupancy             = Activity / TotalDistance
      Average Distance (pkm)= AvgVkm * Occupancy
      Fuel Total / Shares / Intensity updated for 2022

    Operates on the already-built in-memory df only.
    """
    import numpy as np
    import pandas as pd

    if notes is None:
        notes = []

    if df is None or not hasattr(df, "index"):
        notes.append(f"[WARN] Urban Transit 2022 override skipped (no df) {context}")
        return df

    # Ensure year index
    if "year" in getattr(df, "columns", []) and df.index.name != "year":
        try:
            df = df.set_index("year")
        except Exception:
            pass

    try:
        df.index = df.index.astype(int)
    except Exception:
        try:
            df.index = df.index.map(lambda x: int(float(str(x).strip())))
        except Exception:
            pass

    y_ref, y = 2019, 2022
    if y not in df.index or y_ref not in df.index:
        notes.append(f"[WARN] Urban Transit 2022 override skipped (missing {y_ref}/{y}) {context}")
        return df

    COL_ACT = "Activity (millions passenger-kilometres)"
    COL_STOCK = "Stock (thousands)"
    COL_AVG_VKM = "Average Distance (vkm)"
    COL_TOTAL_DIST = "Total Distance (M*vkm)"
    COL_OCC = "Occupancy (persons /vehicle"
    COL_AVG_PKM = "Average Distance (pkm)"
    COL_INT = "Intensity (GJ / pkm)"

    # Base overrides
    if COL_ACT in df.columns:
        df.loc[y, COL_ACT] = pd.to_numeric(df.loc[y_ref, COL_ACT], errors="coerce") * 0.8
    if COL_STOCK in df.columns:
        df.loc[y, COL_STOCK] = pd.to_numeric(df.loc[y_ref, COL_STOCK], errors="coerce")
    if COL_AVG_VKM in df.columns:
        df.loc[y, COL_AVG_VKM] = pd.to_numeric(df.loc[y_ref, COL_AVG_VKM], errors="coerce")

    # Diesel override
    diesel_col = None
    for c in df.columns:
        if str(c).lower() == "fuel (tj)::diesel fuel oil":
            diesel_col = c
            break
    if diesel_col is None and "Fuel (TJ)::Diesel fuel oil" in df.columns:
        diesel_col = "Fuel (TJ)::Diesel fuel oil"
    if diesel_col is not None:
        df.loc[y, diesel_col] = pd.to_numeric(df.loc[y_ref, diesel_col], errors="coerce")

    # Recompute derived metrics for 2022
    try:
        stock = pd.to_numeric(df.loc[y, COL_STOCK], errors="coerce") if COL_STOCK in df.columns else np.nan
        avg_vkm = pd.to_numeric(df.loc[y, COL_AVG_VKM], errors="coerce") if COL_AVG_VKM in df.columns else np.nan
        activity = pd.to_numeric(df.loc[y, COL_ACT], errors="coerce") if COL_ACT in df.columns else np.nan

        if COL_TOTAL_DIST in df.columns and pd.notna(stock) and pd.notna(avg_vkm):
            df.loc[y, COL_TOTAL_DIST] = stock * avg_vkm / 1000.0

        if COL_OCC in df.columns and COL_TOTAL_DIST in df.columns:
            td = pd.to_numeric(df.loc[y, COL_TOTAL_DIST], errors="coerce")
            df.loc[y, COL_OCC] = activity / td if pd.notna(td) and td != 0 else np.nan

        if COL_AVG_PKM in df.columns and COL_OCC in df.columns and pd.notna(avg_vkm):
            occ = pd.to_numeric(df.loc[y, COL_OCC], errors="coerce")
            df.loc[y, COL_AVG_PKM] = avg_vkm * occ

        fuel_cols = [c for c in df.columns if str(c).startswith("Fuel (TJ)::") and str(c) != "Fuel (TJ)::Total"]
        if fuel_cols:
            fuel_total = pd.to_numeric(df.loc[y, fuel_cols], errors="coerce").sum(min_count=1)
            if "Fuel (TJ)::Total" in df.columns:
                df.loc[y, "Fuel (TJ)::Total"] = fuel_total

            share_cols = [c for c in df.columns if str(c).startswith("Share (%)::")]
            if share_cols and pd.notna(fuel_total) and fuel_total != 0:
                for sc in share_cols:
                    fuel_name = str(sc).replace("Share (%)::", "").strip()
                    fc = next((c for c in fuel_cols if str(c).lower().endswith(fuel_name.lower())), None)
                    if fc is not None:
                        v = pd.to_numeric(df.loc[y, fc], errors="coerce")
                        df.loc[y, sc] = (v / fuel_total) * 100.0 if pd.notna(v) else np.nan

            if COL_INT in df.columns and pd.notna(activity) and activity != 0:
                df.loc[y, COL_INT] = fuel_total / activity / 1000.0

    except Exception as e:
        notes.append(f"[WARN] Urban Transit 2022 override recompute failed {context}: {e}")

    notes.append(f"[OK] Applied Urban Transit 2022 special formulas (2019-based) {context}")
    return df

def build_mode(
    mode_name,
    fuel_table,
    p1_activity_row,
    p4_sales_row,  # may be None
    p4_stock_row,
    p4_avg_vkm_row,
    out_file,
    fuel_layout="by_source",
    total_energy_label=None,
    total_to_fuel="Motor gasoline",
    strict_sales=True,
    other_fuels_null=False,
    ceud_path: Path = CEUD_CAN_FILE,
):
    notes = []
    excel = _excel_app()
    wb = _open_book(excel, ceud_path)

    try:
        ws_p1 = _get_ws(wb, [P1_SHEET], contains="passenger1")
        ws_p4 = _get_ws(wb, [P4_SHEET], contains="passenger4")
        ws_fuel = _get_ws(wb, [fuel_table, fuel_table.replace(" ", ""), fuel_table.upper()], contains=str(fuel_table).lower())

        # Activity (STRICT)
        activity = _read_year_row(ws_p1, p1_activity_row)
        _assert_no_blanks(activity, f"{mode_name} Activity")

        # Sales (optional)
        if p4_sales_row is None:
            sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
            notes.append(f"[INFO] {mode_name} Sales: not provided → kept as NULL for all years.")
        else:
            sales = _read_year_row(ws_p4, p4_sales_row)
            if strict_sales:
                _assert_no_blanks(sales, f"{mode_name} Sales")
            else:
                if sales.isna().all():
                    notes.append(f"[INFO] {mode_name} Sales: blank in source → preserved as NULL for all years.")

        # Stock + Avg vkm (STRICT)
        stock = _read_year_row(ws_p4, p4_stock_row)
        _assert_no_blanks(stock, f"{mode_name} Stock")

        avg_vkm = _read_year_row(ws_p4, p4_avg_vkm_row)
        _assert_no_blanks(avg_vkm, f"{mode_name} Avg vkm")

        # Derived
        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        # FUELS
        if fuel_layout == "by_source":
            labels = [str(r[0]).strip() for r in _range_values(ws_fuel, "B14:B19")]
            values = _range_values(ws_fuel, "C14:Y19")

            fuel_df = pd.DataFrame(
                [_to_float_array(v) for v in values],
                index=labels,
                columns=YEARS
            ).T * FUEL_SCALE

            fuels_out = {}
            for f in FUELS:
                match = [c for c in fuel_df.columns if c.lower() == f.lower()]
                s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
                fuels_out[f] = _fuel_handle_na(s, f, notes)
            fuels_out_df = pd.DataFrame(fuels_out)

        elif fuel_layout == "total_to_gasoline":
            if not total_energy_label:
                raise ValueError("total_energy_label must be provided for fuel_layout='total_to_gasoline'")

            r = _find_row(ws_fuel, total_energy_label)
            total_pj = _read_year_row(ws_fuel, r)
            total_tj = total_pj * FUEL_SCALE

            if other_fuels_null:
                fuels_out = {f: pd.Series([np.nan] * N_YEARS, index=YEARS) for f in FUELS}
            else:
                fuels_out = {f: pd.Series(np.zeros(N_YEARS), index=YEARS) for f in FUELS}

            target = next((f for f in FUELS if f.lower() == total_to_fuel.lower()), None)
            if target is None:
                raise ValueError(f"Requested total_to_fuel='{total_to_fuel}' not in FUELS list.")

            fuels_out[target] = _fuel_handle_na(total_tj, target, notes)
            fuels_out_df = pd.DataFrame(fuels_out)

        else:
            raise ValueError(f"Unknown fuel_layout: {fuel_layout}")

        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0

        intensity = pd.Series(
            fuel_total.to_numpy()
            / np.where(activity.to_numpy() == 0, np.nan, activity.to_numpy())
            / 1000.0,
            index=YEARS
        )

        # Car special-case (existing behavior)
        if mode_name.strip().lower() == "car":
            intensity.loc[2000] = fuel_total.loc[2000] / total_distance.loc[2000] / 1000.0
            notes.append("[INFO] Car intensity: year 2000 uses Total Distance denominator (matches CAN sheet).")

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1
        )

        if mode_name.strip().lower() == "car":
            df = _overwrite_car_2022_output_df(df, occupancy, notes, prefix=mode_name)
        # 2022 special formula override for Urban Transit (workbook logic)
        if str(mode_name).strip().lower().replace(' ', '_') == 'urban_transit':
            df = _apply_urban_transit_2022_override_df(df, notes=notes, context=f"build_mode:{out_file}")

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)

        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / f"{mode_name.lower()}_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

        return df

    finally:
        _close_book(wb, excel)

# =========================
# RAIL + AIR (FULL) — USES IN-MEMORY ASSUMPTIONS
# =========================
# Implemented from "Formulas for Air and Rail .txt".
# Preference: Freight (%) and Passengers (%) are stored as 0–100 (percent) for consistency.

def _safe_iferror_div(num: pd.Series, den: pd.Series) -> pd.Series:
    out = num / den.replace(0, np.nan)
    return out.replace([np.inf, -np.inf], np.nan).fillna(0.0)

def _read_can_intensity_and_shares(mode_csv: Path):
    """Read CAN intensity + shares from a mode csv. Shares stored 0–100; returned as fractions."""
    df = pd.read_csv(mode_csv)
    if 'year' in df.columns:
        df = df.set_index('year')
    df.index = df.index.astype(int)
    inten = df.loc[YEARS, 'Intensity (GJ / pkm)']
    freight_pct = df.loc[YEARS, 'Freight (%)']
    pass_pct = df.loc[YEARS, 'Passengers (%)']
    return inten, freight_pct / 100.0, pass_pct / 100.0

def build_rail(out_file: str = 'rail_full.csv', assumptions: 'AssumptionStore|None' = None):
    if assumptions is None:
        raise ValueError('build_rail requires AssumptionStore')

    notes = []
    tkm_per_pkm = float(assumptions.get(mode='Aviation', metric='multiplier_load_factor_tkm_per_pkm', prov_code='CAN', required=True))

    excel = _excel_app()
    wb = _open_book(excel, CEUD_CAN_FILE)
    try:
        ws_freight = _get_ws(wb, [F1_SHEET, 'Freight 1', 'Freight'], contains='freight1')
        ws_passenger = _get_ws(wb, [P1_SHEET, 'Passenger 1', 'Passenger'], contains='passenger1')
        ws_t26 = _get_ws(wb, ['Table 26', 'Table26', 'TABLE 26', 'TABLE26'], contains='table 26')

        freight_tkm = _read_year_row(ws_freight, 36)
        _assert_no_blanks(freight_tkm, 'Rail Freight tonne-kilometres (millions)')
        freight_pkm = _safe_iferror_div(freight_tkm, pd.Series([tkm_per_pkm]*N_YEARS, index=YEARS))

        passenger_pkm = _read_year_row(ws_passenger, 40)
        _assert_no_blanks(passenger_pkm, 'Rail Passenger passenger-kilometres (millions)')

        total_pkm = freight_pkm + passenger_pkm
        denom = total_pkm.replace(0, np.nan)
        freight_share_frac = (freight_pkm / denom).replace([np.inf, -np.inf], np.nan).fillna(0.0)
        passenger_share_frac = (passenger_pkm / denom).replace([np.inf, -np.inf], np.nan).fillna(0.0)

        diesel_pj = _read_year_row(ws_t26, 12)
        diesel_tj = _fuel_handle_na(diesel_pj * FUEL_SCALE, 'Diesel fuel oil', notes)

        fuels_out = {f: pd.Series([np.nan]*N_YEARS, index=YEARS) for f in FUELS}
        fuels_out['Diesel fuel oil'] = diesel_tj
        fuels_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_df.sum(axis=1, min_count=1)

        intensity = _safe_iferror_div(fuel_total, total_pkm) / 1000.0
        fuel_shares = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0).replace([np.inf, -np.inf], np.nan).fillna(0.0) * 100.0

        out = pd.concat(
            [
                total_pkm.rename('Activity (millions passenger-kilometres)'),
                freight_tkm.rename('Freight::tonne-kilometres (millions)'),
                freight_pkm.rename('Freight::passenger-kilometres (millions)'),
                passenger_pkm.rename('Passengers::passenger-kilometres (millions)'),
                total_pkm.rename('Total::passenger-kilometres (millions)'),
                (freight_share_frac * 100.0).rename('Freight (%)'),
                (passenger_share_frac * 100.0).rename('Passengers (%)'),
                fuels_df.add_prefix('Fuel (TJ)::'),
                fuel_total.rename('Fuel (TJ)::Total'),
                intensity.rename('Intensity (GJ / pkm)'),
                fuel_shares.add_prefix('Share (%)::'),
            ],
            axis=1,
        )
        out.index.name = 'year'
        audit_write_df(out, OUT_DIR / out_file)

        _register_df(out_file, out)

        audit_write_text('\n'.join(notes) if notes else '[INFO] No notes.\n', OUT_DIR / 'rail_notes.txt', encoding='utf-8', mode='w')
        if _audit_enabled(): print(f'[OK] Wrote {out_file}')
    finally:
        _close_book(wb, excel)

def build_air(out_file: str = 'air_full.csv', assumptions: 'AssumptionStore|None' = None):
    if assumptions is None:
        raise ValueError('build_air requires AssumptionStore')

    notes = []
    tkm_per_pkm = float(assumptions.get(mode='Aviation', metric='multiplier_load_factor_tkm_per_pkm', prov_code='CAN', required=True))
    air_dom_share = float(assumptions.get(mode='Aviation', metric='pkm_share_domestic', prov_code='CAN', required=True))
    energy_dom_share = float(assumptions.get(mode='Aviation', metric='energy_share_domestic', prov_code='CAN', fuel='Jet fuel', required=True))

    excel = _excel_app()
    wb = _open_book(excel, CEUD_CAN_FILE)
    try:
        ws_freight = _get_ws(wb, [F1_SHEET, 'Freight 1', 'Freight'], contains='freight1')
        ws_passenger = _get_ws(wb, [P1_SHEET, 'Passenger 1', 'Passenger'], contains='passenger1')
        ws_t20 = _get_ws(wb, ['Table 20', 'Table20', 'TABLE 20', 'TABLE20'], contains='table 20')

        freight_tkm_raw = _read_year_row(ws_freight, 35)
        _assert_no_blanks(freight_tkm_raw, 'Air Freight tonne-kilometres (millions)')
        freight_tkm = freight_tkm_raw * air_dom_share

        passenger_pkm_raw = _read_year_row(ws_passenger, 39)
        _assert_no_blanks(passenger_pkm_raw, 'Air Passenger passenger-kilometres (millions)')
        passenger_pkm = passenger_pkm_raw * air_dom_share

        freight_pkm = _safe_iferror_div(freight_tkm, pd.Series([tkm_per_pkm]*N_YEARS, index=YEARS))
        total_pkm = freight_pkm + passenger_pkm

        denom = total_pkm.replace(0, np.nan)
        freight_share_frac = (freight_pkm / denom).replace([np.inf, -np.inf], np.nan).fillna(0.0)
        passenger_share_frac = (passenger_pkm / denom).replace([np.inf, -np.inf], np.nan).fillna(0.0)

        avgas_pj = _read_year_row(ws_t20, 14)
        avtur_pj = _read_year_row(ws_t20, 15)
        avgas_tj = _fuel_handle_na(avgas_pj * energy_dom_share * FUEL_SCALE, 'Aviation gasoline', notes)
        avtur_tj = _fuel_handle_na(avtur_pj * energy_dom_share * FUEL_SCALE, 'Aviation turbo fuel', notes)

        fuels_out = {f: pd.Series([np.nan]*N_YEARS, index=YEARS) for f in FUELS}
        fuels_out['Aviation gasoline'] = avgas_tj
        fuels_out['Aviation turbo fuel'] = avtur_tj
        fuels_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_df.sum(axis=1, min_count=1)

        intensity = _safe_iferror_div(fuel_total, total_pkm) / 1000.0
        fuel_shares = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0).replace([np.inf, -np.inf], np.nan).fillna(0.0) * 100.0

        out = pd.concat(
            [
                total_pkm.rename('Activity (millions passenger-kilometres)'),
                freight_tkm.rename('Freight::tonne-kilometres (millions)'),
                freight_pkm.rename('Freight::passenger-kilometres (millions)'),
                passenger_pkm.rename('Passengers::passenger-kilometres (millions)'),
                total_pkm.rename('Total::passenger-kilometres (millions)'),
                (freight_share_frac * 100.0).rename('Freight (%)'),
                (passenger_share_frac * 100.0).rename('Passengers (%)'),
                fuels_df.add_prefix('Fuel (TJ)::'),
                fuel_total.rename('Fuel (TJ)::Total'),
                intensity.rename('Intensity (GJ / pkm)'),
                fuel_shares.add_prefix('Share (%)::'),
            ],
            axis=1,
        )
        out.index.name = 'year'
        audit_write_df(out, OUT_DIR / out_file)

        _register_df(out_file, out)

        audit_write_text('\n'.join(notes) if notes else '[INFO] No notes.\n', OUT_DIR / 'air_notes.txt', encoding='utf-8', mode='w')
        if _audit_enabled(): print(f'[OK] Wrote {out_file}')
    finally:
        _close_book(wb, excel)

def _build_prov_rail(prefix: str, prov_code: str, ceud_path: Path, out_file: str, assumptions: 'AssumptionStore'):
    notes = []
    tkm_per_pkm = float(assumptions.get(mode='Aviation', metric='multiplier_load_factor_tkm_per_pkm', prov_code='CAN', required=True))
    can_inten, can_fshare, can_pshare = _read_can_intensity_and_shares(OUT_DIR / 'rail_full.csv')

    excel = _excel_app()
    wb = _open_book(excel, ceud_path)
    try:
        ws_fuel = _get_ws(wb, ['Table 17','Table17','TABLE 17','TABLE17'], contains='table 17', required=False)
        if ws_fuel is None:
            ws_fuel = _get_ws(wb, ['Table 26','Table26','TABLE 26','TABLE26'], contains='table 26', required=True)
        diesel_pj = _read_year_row(ws_fuel, 12)
        diesel_tj = _fuel_handle_na(diesel_pj * FUEL_SCALE, 'Diesel fuel oil', notes)

        fuels_out = {f: pd.Series([np.nan]*N_YEARS, index=YEARS) for f in FUELS}
        fuels_out['Diesel fuel oil'] = diesel_tj
        fuels_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_df.sum(axis=1, min_count=1)

        total_pkm = _safe_iferror_div(fuel_total, can_inten) / 1000.0
        freight_pkm = total_pkm * can_fshare
        passenger_pkm = total_pkm * can_pshare
        freight_tkm = freight_pkm * tkm_per_pkm

        intensity = _safe_iferror_div(fuel_total, total_pkm) / 1000.0
        fuel_shares = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0).replace([np.inf, -np.inf], np.nan).fillna(0.0) * 100.0

        out = pd.concat(
            [
                total_pkm.rename('Activity (millions passenger-kilometres)'),
                freight_tkm.rename('Freight::tonne-kilometres (millions)'),
                freight_pkm.rename('Freight::passenger-kilometres (millions)'),
                passenger_pkm.rename('Passengers::passenger-kilometres (millions)'),
                total_pkm.rename('Total::passenger-kilometres (millions)'),
                (can_fshare * 100.0).rename('Freight (%)'),
                (can_pshare * 100.0).rename('Passengers (%)'),
                fuels_df.add_prefix('Fuel (TJ)::'),
                fuel_total.rename('Fuel (TJ)::Total'),
                intensity.rename('Intensity (GJ / pkm)'),
                fuel_shares.add_prefix('Share (%)::'),
            ],
            axis=1,
        )
        out.index.name = 'year'
        audit_write_df(out, OUT_DIR / out_file)
        _register_df(out_file, out)

        notes.append(f"[INFO] {prefix} Rail activity inferred from fuel using CAN Rail intensity + CAN shares.")
        audit_write_text('\n'.join(notes) + '\n', OUT_DIR / f"{prefix.lower()}_rail_notes.txt", encoding='utf-8', mode='w')
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")
    finally:
        _close_book(wb, excel)

def _build_prov_air(prefix: str, prov_code: str, ceud_path: Path, out_file: str, assumptions: 'AssumptionStore'):
    notes = []
    tkm_per_pkm = float(assumptions.get(mode='Aviation', metric='multiplier_load_factor_tkm_per_pkm', prov_code='CAN', required=True))
    energy_dom_share = float(assumptions.get(mode='Aviation', metric='energy_share_domestic', prov_code=prov_code, fuel='Jet fuel', required=True))
    can_inten, can_fshare, can_pshare = _read_can_intensity_and_shares(OUT_DIR / 'air_full.csv')

    # NOTE: Workbook formulas for AT Air fuels reference the BCTerr input table (Table 14),
    # not the Atlantic CEUD workbook. To match workbook behavior, we source Table 14 from
    # transBCTerr2000-2022EN.xls for AT.
    fuel_source_path = CEUD_BCTERR_FILE if prov_code == 'AT' else ceud_path
    excel = _excel_app()
    wb = _open_book(excel, fuel_source_path)
    try:
        # Air fuels must come from Table 14 (Passenger Air). Avoid silent fallbacks for AT.
        ws_t = _get_ws(wb, ['Table 14','Table14','TABLE 14','TABLE14'], contains='table 14', required=True)

        avgas_pj = _read_year_row(ws_t, 14)
        avtur_pj = _read_year_row(ws_t, 15)
        # 'energy_dom_share' here is the workbook-style AT multiplier (INDEX(assumptions!I284:I299, MATCH(region,...))).
        # It is NOT a domestic/international split.
        avgas_tj = _fuel_handle_na(avgas_pj * energy_dom_share * FUEL_SCALE, 'Aviation gasoline', notes)
        avtur_tj = _fuel_handle_na(avtur_pj * energy_dom_share * FUEL_SCALE, 'Aviation turbo fuel', notes)
        avgas_pj = _read_year_row(ws_t, 14)
        avtur_pj = _read_year_row(ws_t, 15)
        avgas_tj = _fuel_handle_na(avgas_pj * energy_dom_share * FUEL_SCALE, 'Aviation gasoline', notes)
        avtur_tj = _fuel_handle_na(avtur_pj * energy_dom_share * FUEL_SCALE, 'Aviation turbo fuel', notes)

        fuels_out = {f: pd.Series([np.nan]*N_YEARS, index=YEARS) for f in FUELS}
        fuels_out['Aviation gasoline'] = avgas_tj
        fuels_out['Aviation turbo fuel'] = avtur_tj
        fuels_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_df.sum(axis=1, min_count=1)

        total_pkm = _safe_iferror_div(fuel_total, can_inten) / 1000.0
        freight_pkm = total_pkm * can_fshare
        passenger_pkm = total_pkm * can_pshare
        freight_tkm = freight_pkm * tkm_per_pkm

        intensity = _safe_iferror_div(fuel_total, total_pkm) / 1000.0
        fuel_shares = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0).replace([np.inf, -np.inf], np.nan).fillna(0.0) * 100.0

        out = pd.concat(
            [
                total_pkm.rename('Activity (millions passenger-kilometres)'),
                freight_tkm.rename('Freight::tonne-kilometres (millions)'),
                freight_pkm.rename('Freight::passenger-kilometres (millions)'),
                passenger_pkm.rename('Passengers::passenger-kilometres (millions)'),
                total_pkm.rename('Total::passenger-kilometres (millions)'),
                (can_fshare * 100.0).rename('Freight (%)'),
                (can_pshare * 100.0).rename('Passengers (%)'),
                fuels_df.add_prefix('Fuel (TJ)::'),
                fuel_total.rename('Fuel (TJ)::Total'),
                intensity.rename('Intensity (GJ / pkm)'),
                fuel_shares.add_prefix('Share (%)::'),
            ],
            axis=1,
        )
        out.index.name = 'year'
        audit_write_df(out, OUT_DIR / out_file)
        _register_df(out_file, out)

        notes.append(f"[INFO] {prefix} Air activity inferred from fuel using CAN Air intensity + CAN shares.")
        notes.append(f"[INFO] {prefix} Air fuels scaled by workbook-style factor={energy_dom_share} (from assumptions INDEX/MATCH).")
        audit_write_text('\n'.join(notes) + '\n', OUT_DIR / f"{prefix.lower()}_air_notes.txt", encoding='utf-8', mode='w')
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")
    finally:
        _close_book(wb, excel)
# =========================
# LDV (CAR + LIGHT TRUCK)
# =========================
def _read_mode_output_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Required upstream output not found: {path}")
    df = pd.read_csv(path)
    if "year" in df.columns:
        df = df.set_index("year")
    df.index = df.index.astype(int)
    return df

# =========================
# LDV 2022 FORMULA HELPERS
# =========================
def _recompute_ldv_2022_formula_rows(activity, sales, stock, total_distance, avg_vkm, occupancy, fuels_df, fuel_total, intensity, shares_df, car, lt, notes=None, prefix="LDV"):
    """Apply the literal 2022 workbook differences for LDV and carry them through.

    User-specified 2022 differences from Formulas LDV 2022.txt:
      - Car Total Distance (M*vkm) 2022 = 2019 Car Total Distance * 0.8
      - Car Fuel (TJ)::Motor gasoline 2022 = 2019 Car Motor gasoline * 0.8

    Then carry those altered 2022 cellular values through the remaining LDV formulas:
      - LDV Activity = Car Activity + Light Truck Activity
      - LDV Sales = Car Sales + Light Truck Sales
      - LDV Stock = Car Stock + Light Truck Stock
      - LDV Total Distance = Car Total Distance + Light Truck Total Distance
      - LDV Average Distance (vkm) = LDV Total Distance * 1000 / LDV Stock
      - LDV Occupancy = LDV Activity / LDV Total Distance
      - LDV fuel rows = Car Fuel + Light Truck Fuel (with 2022 Motor gasoline using the literal override)
      - LDV Fuel Total = sum(fuel rows)
      - LDV Intensity = Fuel Total / Activity / 1000
      - LDV Shares = Fuel / Fuel Total
    """
    y = 2022
    y_ref = 2019  # Excel column V when Y=2022
    if y not in activity.index or y_ref not in car.index or y_ref not in lt.index:
        return activity, sales, stock, total_distance, avg_vkm, occupancy, fuels_df, fuel_total, intensity, shares_df

    COL_ACTIVITY = "Activity (millions passenger-kilometres)"
    COL_SALES = "Sales (thousands)"
    COL_STOCK = "Stock (thousands)"
    COL_TOTAL_DIST = "Total Distance (M*vkm)"
    COL_MG = "Fuel (TJ)::Motor gasoline"

    # Start from existing 2022 Car/Light Truck series, then apply the literal 2022 Car exceptions.
    car_activity_2022 = pd.to_numeric(pd.Series([car.loc[y, COL_ACTIVITY]]), errors='coerce').iloc[0]
    lt_activity_2022 = pd.to_numeric(pd.Series([lt.loc[y, COL_ACTIVITY]]), errors='coerce').iloc[0]
    car_sales_2022 = pd.to_numeric(pd.Series([car.loc[y, COL_SALES]]), errors='coerce').iloc[0]
    lt_sales_2022 = pd.to_numeric(pd.Series([lt.loc[y, COL_SALES]]), errors='coerce').iloc[0]
    car_stock_2022 = pd.to_numeric(pd.Series([car.loc[y, COL_STOCK]]), errors='coerce').iloc[0]
    lt_stock_2022 = pd.to_numeric(pd.Series([lt.loc[y, COL_STOCK]]), errors='coerce').iloc[0]

    # Literal workbook rows for 2022 provided by user.
    car_total_distance_2022 = pd.to_numeric(pd.Series([car.loc[y_ref, COL_TOTAL_DIST]]), errors='coerce').iloc[0]
    if pd.notna(car_total_distance_2022):
        car_total_distance_2022 = car_total_distance_2022 * 0.8
    lt_total_distance_2022 = pd.to_numeric(pd.Series([lt.loc[y, COL_TOTAL_DIST]]), errors='coerce').iloc[0]

    # Aggregate LDV rows from the altered 2022 cellular values.
    activity.loc[y] = car_activity_2022 + lt_activity_2022 if pd.notna(car_activity_2022) or pd.notna(lt_activity_2022) else np.nan
    sales.loc[y] = car_sales_2022 + lt_sales_2022 if pd.notna(car_sales_2022) or pd.notna(lt_sales_2022) else np.nan
    stock.loc[y] = car_stock_2022 + lt_stock_2022 if pd.notna(car_stock_2022) or pd.notna(lt_stock_2022) else np.nan
    total_distance.loc[y] = car_total_distance_2022 + lt_total_distance_2022 if pd.notna(car_total_distance_2022) or pd.notna(lt_total_distance_2022) else np.nan

    s2022 = pd.to_numeric(pd.Series([stock.loc[y]]), errors='coerce').iloc[0]
    td2022 = pd.to_numeric(pd.Series([total_distance.loc[y]]), errors='coerce').iloc[0]
    a2022 = pd.to_numeric(pd.Series([activity.loc[y]]), errors='coerce').iloc[0]
    avg_vkm.loc[y] = (td2022 * 1000.0 / s2022) if pd.notna(td2022) and pd.notna(s2022) and s2022 != 0 else np.nan
    occupancy.loc[y] = (a2022 / td2022) if pd.notna(a2022) and pd.notna(td2022) and td2022 != 0 else np.nan

    # Fuel rows: only 2022 Car Motor gasoline gets the literal override; other fuel rows remain row-wise sums.
    fuel_cols = [c for c in fuels_df.columns if c in car.columns and c in lt.columns]
    for c in fuel_cols:
        car_val = pd.to_numeric(pd.Series([car.loc[y, c]]), errors='coerce').iloc[0]
        lt_val = pd.to_numeric(pd.Series([lt.loc[y, c]]), errors='coerce').iloc[0]
        if c == COL_MG and y_ref in car.index:
            car_val = pd.to_numeric(pd.Series([car.loc[y_ref, c]]), errors='coerce').iloc[0]
            if pd.notna(car_val):
                car_val = car_val * 0.8
        fuels_df.loc[y, c] = car_val + lt_val if pd.notna(car_val) or pd.notna(lt_val) else np.nan

    fuel_total.loc[y] = pd.to_numeric(fuels_df.loc[y], errors='coerce').sum(min_count=1)
    ft2022 = pd.to_numeric(pd.Series([fuel_total.loc[y]]), errors='coerce').iloc[0]
    intensity.loc[y] = (ft2022 / a2022 / 1000.0) if pd.notna(ft2022) and pd.notna(a2022) and a2022 != 0 else np.nan

    if y in shares_df.index:
        if pd.notna(ft2022) and ft2022 != 0:
            shares_df.loc[y] = pd.to_numeric(fuels_df.loc[y], errors='coerce') / ft2022 * 100.0
        else:
            shares_df.loc[y] = np.nan

    if notes is not None:
        notes.append(
            f"[INFO] {prefix} 2022 literal formulas applied: Car Total Distance={y_ref}*0.8 and Car Motor gasoline={y_ref}*0.8, then downstream LDV rows recomputed from those values."
        )

    return activity, sales, stock, total_distance, avg_vkm, occupancy, fuels_df, fuel_total, intensity, shares_df

def build_ldv(
    car_file="car_full.csv",
    light_truck_file="light_truck_full.csv",
    out_file="ldv_full.csv",
):
    notes = []

    car = _get_df(car_file).loc[YEARS]
    lt = _get_df(light_truck_file).loc[YEARS]

    COL_ACTIVITY = "Activity (millions passenger-kilometres)"
    COL_SALES = "Sales (thousands)"
    COL_STOCK = "Stock (thousands)"
    COL_TOTAL_DIST = "Total Distance (M*vkm)"

    activity = car[COL_ACTIVITY] + lt[COL_ACTIVITY]
    sales = car[COL_SALES] + lt[COL_SALES]
    stock = car[COL_STOCK] + lt[COL_STOCK]
    total_distance = car[COL_TOTAL_DIST] + lt[COL_TOTAL_DIST]

    _assert_no_blanks(activity, "LDV Activity")
    _assert_no_blanks(stock, "LDV Stock")
    _assert_no_blanks(total_distance, "LDV Total Distance")

    avg_vkm = (total_distance * 1000.0) / stock.replace(0, np.nan)
    occupancy = activity / total_distance.replace(0, np.nan)
    avg_pkm = pd.Series([np.nan] * N_YEARS, index=YEARS)

    fuel_cols = [c for c in car.columns if c.startswith("Fuel (TJ)::") and c != "Fuel (TJ)::Total"]
    fuels_df = pd.DataFrame(index=YEARS)
    for c in fuel_cols:
        if c in lt.columns:
            fuels_df[c] = car[c] + lt[c]
        else:
            raise ValueError(f"LDV: missing fuel col in Light Truck: {c}")

    fuel_total = fuels_df.sum(axis=1, min_count=1)
    shares_df = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
    shares_df = shares_df.rename(columns=lambda x: "Share (%)::" + x.split("::", 1)[1])

    intensity = fuel_total / activity.replace(0, np.nan) / 1000.0
    activity, sales, stock, total_distance, avg_vkm, occupancy, fuels_df, fuel_total, intensity, shares_df = _recompute_ldv_2022_formula_rows(
        activity, sales, stock, total_distance, avg_vkm, occupancy, fuels_df, fuel_total, intensity, shares_df, car, lt, notes, prefix="LDV"
    )

    out = pd.concat(
        [
            activity.rename(COL_ACTIVITY),
            sales.rename(COL_SALES),
            stock.rename(COL_STOCK),
            avg_vkm.rename("Average Distance (vkm)"),
            total_distance.rename(COL_TOTAL_DIST),
            occupancy.rename("Occupancy (persons /vehicle"),
            avg_pkm.rename("Average Distance (pkm)"),
            fuels_df,
            fuel_total.rename("Fuel (TJ)::Total"),
            intensity.rename("Intensity (GJ / pkm)"),
            shares_df,
        ],
        axis=1
    )

    out.index.name = "year"
    audit_write_df(out, OUT_DIR / out_file)
    _register_df(out_file, out)

    notes.append("[INFO] LDV computed as Car + Light_Truck from existing outputs.")
    notes.append("[INFO] LDV Average Distance (pkm) set to NaN for all years (per spec).")

    audit_write_text("\n".join(notes) + "\n", OUT_DIR / "ldv_notes.txt", encoding="utf-8", mode="w")
    if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    return out

# =========================
# BCTERR: CAR
# =========================
def build_bcterr_car(out_file="bcterr_car.csv"):
    notes = []

    can_car = _get_df_any(['car_full.csv', 'car_full'], required=True)
    if "year" in can_car.columns:
        can_car = can_car.set_index("year")
    can_car.index = can_car.index.astype(int)

    occ_col = "Occupancy (persons /vehicle"
    if occ_col not in can_car.columns:
        raise ValueError(f"Expected '{occ_col}' in output/car_full.csv")

    can_occ = can_car.loc[YEARS, occ_col]
    _assert_no_blanks(can_occ, "CAN Car Occupancy (for BCTerr Car activity)")

    excel = _excel_app()
    wb = _open_book(excel, CEUD_BCTERR_FILE)

    try:
        ws_t21 = _get_ws(wb, ["Table 21", "Table21", "TABLE 21", "TABLE21"], contains="table 21")
        ws_t20 = _get_ws(wb, ["Table 20", "Table20", "TABLE 20", "TABLE20"], contains="table 20")

        sales = _read_year_row(ws_t21, 13)
        stock = _read_year_row(ws_t21, 16)
        avg_vkm = _read_year_row(ws_t21, 19)

        _assert_no_blanks(sales, "BCTerr Car Sales (Table 21 row 13)")
        _assert_no_blanks(stock, "BCTerr Car Stock (Table 21 row 16)")
        _assert_no_blanks(avg_vkm, "BCTerr Car Avg vkm (Table 21 row 19)")

        total_distance = stock * avg_vkm / 1000.0
        activity = total_distance * can_occ
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t20, "B14:B19")]
        values = _range_values(ws_t20, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0
        total_distance, activity, occupancy, avg_pkm, fuels_out_df, fuel_total, intensity, shares = _apply_car_2022_literal_overrides(
            total_distance, activity, occupancy, avg_pkm, fuels_out_df, fuel_total, intensity, shares, can_occ, avg_vkm, notes, prefix="BCTerr Car"
        )

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1
        )

        df = _overwrite_car_2022_output_df(df, can_occ, notes, prefix="BCTerr Car")
        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)

        _register_df(out_file, df)

        notes.append("[INFO] BCTerr Car Activity computed as TotalDistance * CAN Car Occupancy (output/car_full.csv).")
        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "bcterr_car_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

# =========================
# BCTERR: LIGHT TRUCK
# =========================
def build_bcterr_light_truck(out_file="bcterr_light_truck.csv"):
    notes = []

    can_lt = _get_df_any(['light_truck_full.csv', 'light_truck_full'], required=True)
    if "year" in can_lt.columns:
        can_lt = can_lt.set_index("year")
    can_lt.index = can_lt.index.astype(int)

    occ_col = "Occupancy (persons /vehicle"
    if occ_col not in can_lt.columns:
        raise ValueError(f"Expected '{occ_col}' in output/light_truck_full.csv")

    can_occ = can_lt.loc[YEARS, occ_col]
    _assert_no_blanks(can_occ, "CAN Light Truck Occupancy (for BCTerr Light Truck activity)")

    excel = _excel_app()
    wb = _open_book(excel, CEUD_BCTERR_FILE)

    try:
        ws_t37 = _get_ws(wb, ["Table 37", "Table37", "TABLE 37", "TABLE37"], contains="table 37")
        ws_t25 = _get_ws(wb, ["Table 25", "Table25", "TABLE 25", "TABLE25"], contains="table 25")

        sales = _read_year_row(ws_t37, 13)
        stock = _read_year_row(ws_t37, 25)
        avg_vkm = _read_year_row(ws_t37, 37)

        _assert_no_blanks(sales, "BCTerr Light Truck Sales (Table 37 row 13)")
        _assert_no_blanks(stock, "BCTerr Light Truck Stock (Table 37 row 25)")
        _assert_no_blanks(avg_vkm, "BCTerr Light Truck Avg vkm (Table 37 row 37)")

        total_distance = stock * avg_vkm / 1000.0
        activity = total_distance * can_occ
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t25, "B14:B19")]
        values = _range_values(ws_t25, "C14:Y19")

        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        notes.append("[INFO] BCTerr Light Truck Activity computed as TotalDistance * CAN Light Truck Occupancy (output/light_truck_full.csv).")
        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "bcterr_light_truck_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

# =========================
# BCTERR: LDV (CAR + LIGHT TRUCK)
# =========================
def build_bcterr_ldv(
    car_file="bcterr_car.csv",
    light_truck_file="bcterr_light_truck.csv",
    out_file="bcterr_ldv.csv",
):
    """ 
    BCTerr LDV = Car + Light Truck, matching the BCTerr sheet formulas:

      Activity (millions passenger-kilometres) = Car Activity + Light Truck Activity
      Sales (thousands)                        = Car Sales + Light Truck Sales
      Stock (thousands)                        = Car Stock + Light Truck Stock
      Total Distance (M*vkm)                   = Car Total Distance + Light Truck Total Distance
      Average Distance (vkm)                   = TotalDistance * 1000 / Stock
      Occupancy (persons / vehicle)            = Activity / TotalDistance
      Average Distance (pkm)                   = blank in workbook → NaN

      Fuels (TJ) per fuel                       = Car Fuel + Light Truck Fuel
      Fuel (TJ)::Total                          = SUM(Fuel (TJ)::*)
      Intensity (GJ / pkm)                      = FuelTotal / Activity / 1000
      Share (%)::*                              = Fuel / FuelTotal * 100  (0–100)

    Inputs are the existing BCTerr Car and BCTerr Light Truck outputs produced by this script.
    """
    notes = []

    car_path = OUT_DIR / car_file
    lt_path = OUT_DIR / light_truck_file

    if not car_path.exists():
        raise FileNotFoundError(f"Missing {car_path}. Build BCTerr Car first.")
    if not lt_path.exists():
        raise FileNotFoundError(f"Missing {lt_path}. Build BCTerr Light Truck first.")

    car = pd.read_csv(car_path)
    lt = pd.read_csv(lt_path)

    if "year" in car.columns:
        car = car.set_index("year")
    if "year" in lt.columns:
        lt = lt.set_index("year")

    car.index = car.index.astype(int)
    lt.index = lt.index.astype(int)

    car = car.loc[YEARS]
    lt = lt.loc[YEARS]

    COL_ACTIVITY = "Activity (millions passenger-kilometres)"
    COL_SALES = "Sales (thousands)"
    COL_STOCK = "Stock (thousands)"
    COL_TOTAL_DIST = "Total Distance (M*vkm)"

    for c in [COL_ACTIVITY, COL_SALES, COL_STOCK, COL_TOTAL_DIST]:
        if c not in car.columns:
            raise ValueError(f"BCTerr LDV: missing required column in Car: {c}")
        if c not in lt.columns:
            raise ValueError(f"BCTerr LDV: missing required column in Light Truck: {c}")

    activity = car[COL_ACTIVITY] + lt[COL_ACTIVITY]
    sales = car[COL_SALES] + lt[COL_SALES]
    stock = car[COL_STOCK] + lt[COL_STOCK]
    total_distance = car[COL_TOTAL_DIST] + lt[COL_TOTAL_DIST]

    _assert_no_blanks(activity, "BCTerr LDV Activity")
    _assert_no_blanks(stock, "BCTerr LDV Stock")
    _assert_no_blanks(total_distance, "BCTerr LDV Total Distance")

    avg_vkm = (total_distance * 1000.0) / stock.replace(0, np.nan)
    occupancy = activity / total_distance.replace(0, np.nan)

    avg_pkm = pd.Series([np.nan] * N_YEARS, index=YEARS)
    notes.append("[INFO] BCTerr LDV Average Distance (pkm) is blank in workbook → kept as NaN for all years.")

    # Fuels: sum each fuel column (excluding total)
    fuel_cols = [c for c in car.columns if c.startswith("Fuel (TJ)::") and c != "Fuel (TJ)::Total"]
    if not fuel_cols:
        raise ValueError("BCTerr LDV: no fuel columns found in Car output.")

    fuels_df = pd.DataFrame(index=YEARS)
    for c in fuel_cols:
        if c not in lt.columns:
            raise ValueError(f"BCTerr LDV: missing fuel column in Light Truck: {c}")
        fuels_df[c] = car[c] + lt[c]

    fuel_total = fuels_df.sum(axis=1, min_count=1)
    shares_df = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
    shares_df = shares_df.rename(columns=lambda x: "Share (%)::" + x.split("::", 1)[1])

    intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

    out = pd.concat(
        [
            activity.rename(COL_ACTIVITY),
            sales.rename(COL_SALES),
            stock.rename(COL_STOCK),
            avg_vkm.rename("Average Distance (vkm)"),
            total_distance.rename(COL_TOTAL_DIST),
            occupancy.rename("Occupancy (persons /vehicle"),
            avg_pkm.rename("Average Distance (pkm)"),
            fuels_df,
            fuel_total.rename("Fuel (TJ)::Total"),
            intensity.rename("Intensity (GJ / pkm)"),
            shares_df,
        ],
        axis=1,
    )

    out.index.name = "year"
    audit_write_df(out, OUT_DIR / out_file)
    _register_df(out_file, out)

    notes.append("[INFO] BCTerr LDV computed as BCTerr Car + BCTerr Light Truck (row-wise sums).")

    audit_write_text("\n".join(notes) + "\n", OUT_DIR / "bcterr_ldv_notes.txt", encoding="utf-8", mode="w")
    if _audit_enabled(): print(f"[OK] Wrote {out_file}")

# =========================
# BCTERR: MOTORCYCLE (NEW)
# =========================
def build_bcterr_motorcycle(out_file="bcterr_motorcycle.csv"):
    """
    BCTerr Motorcycle per your formulas:
      Activity = Table 32 row 15
      Sales = NaN
      Stock = Table 32 row 26
      Avg vkm = Table 32 row 27
      Total Distance = Stock*AvgVkm/1000
      Fuel: Motor gasoline = Table 32 row 12 *1000; others NaN
      Intensity = FuelTotal / Activity / 1000
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_BCTERR_FILE)

    try:
        ws_t32 = _get_ws(wb, ["Table 32", "Table32", "TABLE 32", "TABLE32"], contains="table 32")

        # Activity / Stock / Avg vkm
        activity = _read_year_row(ws_t32, 15)
        stock = _read_year_row(ws_t32, 26)
        avg_vkm = _read_year_row(ws_t32, 27)

        _assert_no_blanks(activity, "BCTerr Motorcycle Activity (Table 32 row 15)")
        _assert_no_blanks(stock, "BCTerr Motorcycle Stock (Table 32 row 26)")
        _assert_no_blanks(avg_vkm, "BCTerr Motorcycle Avg vkm (Table 32 row 27)")

        # Sales is blank/NaN
        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] BCTerr Motorcycle Sales: blank → kept as NaN for all years.")

        # Derived metrics
        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        # Fuels: only Motor gasoline, from Table 32 row 12 * 1000 (PJ→TJ)
        mg_pj = _read_year_row(ws_t32, 12)
        mg_tj = _fuel_handle_na(mg_pj * FUEL_SCALE, "Motor gasoline", notes)

        fuels_out = {f: pd.Series([np.nan] * N_YEARS, index=YEARS) for f in FUELS}
        fuels_out["Motor gasoline"] = mg_tj
        fuels_out_df = pd.DataFrame(fuels_out)

        fuel_total = fuels_out_df.sum(axis=1, min_count=1)

        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0

        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "bcterr_motorcycle_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

# =========================
# BCTERR: SCHOOL BUS (NEW)
# =========================
def build_bcterr_school_bus(out_file="bcterr_school_bus.csv"):
    """
    BCTerr School Bus per provided formulas (transportation personal_source data.xlsx 'BCTerr'):

      Activity (millions passenger-kilometres) = 'Table 28' row 30
      Sales (thousands)                       = blank (NaN)
      Stock (thousands)                       = 'Table 31' row 13
      Average Distance (vkm)                  = 'Table 31' row 23
      Total Distance (M*vkm)                  = Stock * AvgVkm / 1000
      Occupancy (persons/vehicle)             = Activity / TotalDistance
      Average Distance (pkm)                  = AvgVkm * Occupancy

      Fuels (TJ)                              = 'Table 28' rows 14:19 (PJ) * 1000, matched by label
      Fuel Total (TJ)                         = sum of fuels
      Intensity (GJ/pkm)                      = FuelTotal / Activity / 1000
      Shares (%)                              = Fuel / FuelTotal * 100

    Notes:
      - Any fuel not present in Table 28 will be filled with zeros (matches IFERROR(...,0) pattern).
      - Any 'n.a.' values are preserved as NaN (audit-friendly).
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_BCTERR_FILE)
    try:
        ws_t28 = _get_ws(wb, ["Table 28", "Table28", "TABLE 28", "TABLE28"], contains="table 28")
        ws_t31 = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")

        # Activity / Stock / Avg vkm
        activity = _read_year_row(ws_t28, 30)
        stock = _read_year_row(ws_t31, 13)
        avg_vkm = _read_year_row(ws_t31, 23)

        _assert_no_blanks(activity, "BCTerr School Bus Activity (Table 28 row 30)")
        _assert_no_blanks(stock, "BCTerr School Bus Stock (Table 31 row 13)")
        _assert_no_blanks(avg_vkm, "BCTerr School Bus Avg vkm (Table 31 row 23)")

        # Sales is blank/NaN
        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] BCTerr School Bus Sales: blank → kept as NaN for all years.")

        # Derived metrics
        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        # Fuels from Table 28 (by-source layout like Car / Light Truck)
        labels = [str(r[0]).strip() for r in _range_values(ws_t28, "B14:B19")]
        values = _range_values(ws_t28, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)
        fuels_out_df = pd.DataFrame(fuels_out)

        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )
        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "bcterr_school_bus_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

# =========================
# BCTERR: URBAN TRANSIT (NEW)
# =========================
def build_bcterr_urban_transit(out_file="bcterr_urban_transit.csv"):
    """
    BCTerr Urban Transit per provided formulas:

      Activity (millions passenger-kilometres) = Table 29 row 32
      Sales (thousands)                       = blank (NaN)
      Stock (thousands)                       = Table 31 row 14
      Average Distance (vkm)                  = Table 31 row 24
      Total Distance (M*vkm)                  = Stock * AvgVkm / 1000
      Occupancy (persons / vehicle)           = Activity / TotalDistance
      Average Distance (pkm)                  = AvgVkm * Occupancy

      Fuels (TJ)                              = Table 29 rows 14:19 (PJ) * 1000, matched by label
      Fuel Total (TJ)                         = sum of fuels
      Intensity (GJ / pkm)                    = FuelTotal / Activity / 1000
      Shares (%)                              = Fuel / FuelTotal * 100

    Notes:
      - Missing fuels (entire series) are filled with zeros (mirrors IFERROR(...,0)).
      - 'n.a.' values are preserved as NaN.
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_BCTERR_FILE)
    try:
        ws_t29 = _get_ws(wb, ["Table 29", "Table29", "TABLE 29", "TABLE29"], contains="table 29")
        ws_t31 = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")

        # Activity / Stock / Avg vkm
        activity = _read_year_row(ws_t29, 32)
        stock = _read_year_row(ws_t31, 14)
        avg_vkm = _read_year_row(ws_t31, 24)

        _assert_no_blanks(activity, "BCTerr Urban Transit Activity (Table 29 row 32)")
        _assert_no_blanks(stock, "BCTerr Urban Transit Stock (Table 31 row 14)")
        _assert_no_blanks(avg_vkm, "BCTerr Urban Transit Avg vkm (Table 31 row 24)")

        # Sales is blank/NaN
        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] BCTerr Urban Transit Sales: blank → kept as NaN for all years.")

        # BCTerr Urban Transit 2022 special formulas (workbook): 2022 values derived from 2019
        y_ref, y = 2019, 2022
        if y in activity.index and y_ref in activity.index:
            activity.loc[y] = pd.to_numeric(activity.loc[y_ref], errors='coerce') * 0.8
            stock.loc[y] = pd.to_numeric(stock.loc[y_ref], errors='coerce')
            avg_vkm.loc[y] = pd.to_numeric(avg_vkm.loc[y_ref], errors='coerce')
            notes.append("[OK] BCTerr Urban Transit 2022 override applied (Activity=2019*0.8; Stock/AvgVkm=2019).")
        else:
            notes.append("[WARN] BCTerr Urban Transit 2022 override skipped (missing 2019/2022).")

        # Derived metrics
        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        # Fuels from Table 29 (by-source layout)
        labels = [str(r[0]).strip() for r in _range_values(ws_t29, "B14:B19")]
        values = _range_values(ws_t29, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)
        fuels_out_df = pd.DataFrame(fuels_out)

        # BCTerr Urban Transit Diesel fuel oil 2022 override: 2022 = 2019
        y_ref, y = 2019, 2022
        if 'Diesel fuel oil' in fuels_out_df.columns and y in fuels_out_df.index and y_ref in fuels_out_df.index:
            fuels_out_df.loc[y, 'Diesel fuel oil'] = pd.to_numeric(fuels_out_df.loc[y_ref, 'Diesel fuel oil'], errors='coerce')
            notes.append("[OK] BCTerr Urban Transit Diesel fuel oil 2022 override applied (2019 value).")

        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )
        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "bcterr_urban_transit_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

# =========================
# BCTERR: INTERCITY BUS (NEW)
# =========================
def build_bcterr_intercity_bus(out_file="bcterr_intercity_bus.csv"):
    """
    BCTerr Intercity Bus per provided formulas:

      Activity (millions passenger-kilometres) = Table 30 row 26
      Sales (thousands)                       = blank (NaN)
      Stock (thousands)                       = Table 31 row 15
      Average Distance (vkm)                  = Table 31 row 25
      Total Distance (M*vkm)                  = Stock * AvgVkm / 1000
      Occupancy (persons / vehicle)           = Activity / TotalDistance
      Average Distance (pkm)                  = AvgVkm * Occupancy

      Fuels (TJ)                              = Table 30 rows 14:19 (PJ) * 1000, matched by label
      Fuel Total (TJ)                         = sum of fuels
      Intensity (GJ / pkm)                    = FuelTotal / Activity / 1000
      Shares (%)                              = Fuel / FuelTotal * 100

    Notes:
      - Missing fuels (entire series) are filled with zeros (mirrors IFERROR(...,0)).
      - 'n.a.' values are preserved as NaN.
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_BCTERR_FILE)
    try:
        ws_t30 = _get_ws(wb, ["Table 30", "Table30", "TABLE 30", "TABLE30"], contains="table 30")
        ws_t31 = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")

        # Activity / Stock / Avg vkm
        activity = _read_year_row(ws_t30, 26)
        stock = _read_year_row(ws_t31, 15)
        avg_vkm = _read_year_row(ws_t31, 25)

        _assert_no_blanks(activity, "BCTerr Intercity Bus Activity (Table 30 row 26)")
        _assert_no_blanks(stock, "BCTerr Intercity Bus Stock (Table 31 row 15)")
        _assert_no_blanks(avg_vkm, "BCTerr Intercity Bus Avg vkm (Table 31 row 25)")

        # Sales is blank/NaN
        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] BCTerr Intercity Bus Sales: blank → kept as NaN for all years.")

        # Derived metrics
        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        # Fuels from Table 30 (by-source layout)
        labels = [str(r[0]).strip() for r in _range_values(ws_t30, "B14:B19")]
        values = _range_values(ws_t30, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)
        fuels_out_df = pd.DataFrame(fuels_out)

        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )
        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "bcterr_intercity_bus_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

# =========================
# ALBERTA (AB)
# =========================

def build_alb_car(out_file="alb_car.csv"):
    """Alberta Car.

    Activity is computed using CAN Car occupancy (mirrors regional workbook logic):
      Activity = TotalDistance * CAN Car Occupancy

    Sales / Stock / Avg vkm: Alberta CEUD Table 21 rows 13/16/19.
    Fuels: Alberta CEUD Table 20 rows 14:19 (PJ) * 1000 -> TJ, matched by label.

    Skips any 'assumptions' dependencies (none here).
    """
    notes = []

    can_car = _get_df_any(['car_full.csv', 'car_full'], required=True)
    if "year" in can_car.columns:
        can_car = can_car.set_index("year")
    can_car.index = can_car.index.astype(int)

    occ_col = "Occupancy (persons /vehicle"
    if occ_col not in can_car.columns:
        raise ValueError(f"Expected '{occ_col}' in output/car_full.csv")

    can_occ = can_car.loc[YEARS, occ_col]
    _assert_no_blanks(can_occ, "CAN Car Occupancy (for Alberta Car activity)")

    excel = _excel_app()
    wb = _open_book(excel, CEUD_ALB_FILE)
    try:
        ws_t21 = _get_ws(wb, ["Table 21", "Table21", "TABLE 21", "TABLE21"], contains="table 21")
        ws_t20 = _get_ws(wb, ["Table 20", "Table20", "TABLE 20", "TABLE20"], contains="table 20")

        sales = _read_year_row(ws_t21, 13)
        stock = _read_year_row(ws_t21, 16)
        avg_vkm = _read_year_row(ws_t21, 19)

        _assert_no_blanks(sales, "Alberta Car Sales (Table 21 row 13)")
        _assert_no_blanks(stock, "Alberta Car Stock (Table 21 row 16)")
        _assert_no_blanks(avg_vkm, "Alberta Car Avg vkm (Table 21 row 19)")

        total_distance = stock * avg_vkm / 1000.0
        activity = total_distance * can_occ
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t20, "B14:B19")]
        values = _range_values(ws_t20, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df = _overwrite_car_2022_output_df(df, can_occ, notes, prefix="Alberta Car")
        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        notes.append("[INFO] Alberta Car Activity computed as TotalDistance * CAN Car Occupancy (output/car_full.csv).")
        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "alb_car_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def build_alb_light_truck(out_file="alb_light_truck.csv"):
    """Alberta Light Truck.

    Activity is computed using CAN Light Truck occupancy:
      Activity = TotalDistance * CAN Light Truck Occupancy

    Sales / Stock / Avg vkm: Alberta CEUD Table 37 rows 13/25/37.
    Fuels: Alberta CEUD Table 25 rows 14:19 (PJ) * 1000 -> TJ, matched by label.
    """
    notes = []

    can_lt = _get_df_any(['light_truck_full.csv', 'light_truck_full'], required=True)
    if "year" in can_lt.columns:
        can_lt = can_lt.set_index("year")
    can_lt.index = can_lt.index.astype(int)

    occ_col = "Occupancy (persons /vehicle"
    if occ_col not in can_lt.columns:
        raise ValueError(f"Expected '{occ_col}' in output/light_truck_full.csv")

    can_occ = can_lt.loc[YEARS, occ_col]
    _assert_no_blanks(can_occ, "CAN Light Truck Occupancy (for Alberta Light Truck activity)")

    excel = _excel_app()
    wb = _open_book(excel, CEUD_ALB_FILE)
    try:
        ws_t37 = _get_ws(wb, ["Table 37", "Table37", "TABLE 37", "TABLE37"], contains="table 37")
        ws_t25 = _get_ws(wb, ["Table 25", "Table25", "TABLE 25", "TABLE25"], contains="table 25")

        sales = _read_year_row(ws_t37, 13)
        stock = _read_year_row(ws_t37, 25)
        avg_vkm = _read_year_row(ws_t37, 37)

        _assert_no_blanks(sales, "Alberta Light Truck Sales (Table 37 row 13)")
        _assert_no_blanks(stock, "Alberta Light Truck Stock (Table 37 row 25)")
        _assert_no_blanks(avg_vkm, "Alberta Light Truck Avg vkm (Table 37 row 37)")

        total_distance = stock * avg_vkm / 1000.0
        activity = total_distance * can_occ
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t25, "B14:B19")]
        values = _range_values(ws_t25, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        notes.append("[INFO] Alberta Light Truck Activity computed as TotalDistance * CAN Light Truck Occupancy (output/light_truck_full.csv).")
        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "alb_light_truck_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def build_alb_ldv(
    car_file="alb_car.csv",
    light_truck_file="alb_light_truck.csv",
    out_file="alb_ldv.csv",
):
    """Alberta LDV = Alberta Car + Alberta Light Truck (row-wise sums)."""
    notes = []

    car_path = OUT_DIR / car_file
    lt_path = OUT_DIR / light_truck_file

    if not car_path.exists():
        raise FileNotFoundError(f"Missing {car_path}. Build Alberta Car first.")
    if not lt_path.exists():
        raise FileNotFoundError(f"Missing {lt_path}. Build Alberta Light Truck first.")

    car = pd.read_csv(car_path)
    lt = pd.read_csv(lt_path)

    if "year" in car.columns:
        car = car.set_index("year")
    if "year" in lt.columns:
        lt = lt.set_index("year")

    car.index = car.index.astype(int)
    lt.index = lt.index.astype(int)

    car = car.loc[YEARS]
    lt = lt.loc[YEARS]

    COL_ACTIVITY = "Activity (millions passenger-kilometres)"
    COL_SALES = "Sales (thousands)"
    COL_STOCK = "Stock (thousands)"
    COL_TOTAL_DIST = "Total Distance (M*vkm)"

    for c in [COL_ACTIVITY, COL_SALES, COL_STOCK, COL_TOTAL_DIST]:
        if c not in car.columns:
            raise ValueError(f"Alberta LDV: missing required column in Car: {c}")
        if c not in lt.columns:
            raise ValueError(f"Alberta LDV: missing required column in Light Truck: {c}")

    activity = car[COL_ACTIVITY] + lt[COL_ACTIVITY]
    sales = car[COL_SALES] + lt[COL_SALES]
    stock = car[COL_STOCK] + lt[COL_STOCK]
    total_distance = car[COL_TOTAL_DIST] + lt[COL_TOTAL_DIST]

    _assert_no_blanks(activity, "Alberta LDV Activity")
    _assert_no_blanks(stock, "Alberta LDV Stock")
    _assert_no_blanks(total_distance, "Alberta LDV Total Distance")

    avg_vkm = (total_distance * 1000.0) / stock.replace(0, np.nan)
    occupancy = activity / total_distance.replace(0, np.nan)

    avg_pkm = pd.Series([np.nan] * N_YEARS, index=YEARS)
    notes.append("[INFO] Alberta LDV Average Distance (pkm) is blank in workbook → kept as NaN for all years.")

    fuel_cols = [c for c in car.columns if c.startswith("Fuel (TJ)::") and c != "Fuel (TJ)::Total"]
    if not fuel_cols:
        raise ValueError("Alberta LDV: no fuel columns found in Alberta Car output.")

    fuels_df = pd.DataFrame(index=YEARS)
    for c in fuel_cols:
        if c not in lt.columns:
            raise ValueError(f"Alberta LDV: missing fuel column in Alberta Light Truck: {c}")
        fuels_df[c] = car[c] + lt[c]

    fuel_total = fuels_df.sum(axis=1, min_count=1)
    shares_df = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
    shares_df = shares_df.rename(columns=lambda x: "Share (%)::" + x.split("::", 1)[1])

    intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

    out = pd.concat(
        [
            activity.rename(COL_ACTIVITY),
            sales.rename(COL_SALES),
            stock.rename(COL_STOCK),
            avg_vkm.rename("Average Distance (vkm)"),
            total_distance.rename(COL_TOTAL_DIST),
            occupancy.rename("Occupancy (persons /vehicle"),
            avg_pkm.rename("Average Distance (pkm)"),
            fuels_df,
            fuel_total.rename("Fuel (TJ)::Total"),
            intensity.rename("Intensity (GJ / pkm)"),
            shares_df,
        ],
        axis=1,
    )

    out.index.name = "year"
    audit_write_df(out, OUT_DIR / out_file)
    _register_df(out_file, out)

    notes.append("[INFO] Alberta LDV computed as Alberta Car + Alberta Light Truck (row-wise sums).")
    audit_write_text("\n".join(notes) + "\n", OUT_DIR / "alb_ldv_notes.txt", encoding="utf-8", mode="w")
    if _audit_enabled(): print(f"[OK] Wrote {out_file}")

def build_alb_motorcycle(out_file="alb_motorcycle.csv"):
    """Alberta Motorcycle from Table 32.

    Activity: Table 32 row 15
    Stock:    Table 32 row 26
    Avg vkm:  Table 32 row 27
    Sales:    blank -> NaN
    Fuel:     Motor gasoline only = Table 32 row 12 (PJ) * 1000 -> TJ
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_ALB_FILE)
    try:
        ws_t32 = _get_ws(wb, ["Table 32", "Table32", "TABLE 32", "TABLE32"], contains="table 32")

        activity = _read_year_row(ws_t32, 15)
        stock = _read_year_row(ws_t32, 26)
        avg_vkm = _read_year_row(ws_t32, 27)

        _assert_no_blanks(activity, "Alberta Motorcycle Activity (Table 32 row 15)")
        _assert_no_blanks(stock, "Alberta Motorcycle Stock (Table 32 row 26)")
        _assert_no_blanks(avg_vkm, "Alberta Motorcycle Avg vkm (Table 32 row 27)")

        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] Alberta Motorcycle Sales: blank → kept as NaN for all years.")

        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        mg_pj = _read_year_row(ws_t32, 12)
        mg_tj = _fuel_handle_na(mg_pj * FUEL_SCALE, "Motor gasoline", notes)

        fuels_out = {f: pd.Series([np.nan] * N_YEARS, index=YEARS) for f in FUELS}
        fuels_out["Motor gasoline"] = mg_tj

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "alb_motorcycle_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def build_alb_school_bus(out_file="alb_school_bus.csv"):
    """Alberta School Bus.

    Activity: Table 28 row 30
    Stock:    Table 31 row 13
    Avg vkm:  Table 31 row 23
    Sales:    blank -> NaN
    Fuels:    Table 28 rows 14:19 (PJ) * 1000 -> TJ
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_ALB_FILE)
    try:
        ws_t28 = _get_ws(wb, ["Table 28", "Table28", "TABLE 28", "TABLE28"], contains="table 28")
        ws_t31 = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")

        activity = _read_year_row(ws_t28, 30)
        stock = _read_year_row(ws_t31, 13)
        avg_vkm = _read_year_row(ws_t31, 23)

        _assert_no_blanks(activity, "Alberta School Bus Activity (Table 28 row 30)")
        _assert_no_blanks(stock, "Alberta School Bus Stock (Table 31 row 13)")
        _assert_no_blanks(avg_vkm, "Alberta School Bus Avg vkm (Table 31 row 23)")

        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] Alberta School Bus Sales: blank → kept as NaN for all years.")

        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t28, "B14:B19")]
        values = _range_values(ws_t28, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "alb_school_bus_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def build_alb_urban_transit(out_file="alb_urban_transit.csv"):
    """Alberta Urban Transit.

    Activity: Table 29 row 32
    Stock:    Table 31 row 14
    Avg vkm:  Table 31 row 24
    Sales:    blank -> NaN
    Fuels:    Table 29 rows 14:19 (PJ) * 1000 -> TJ
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_ALB_FILE)
    try:
        ws_t29 = _get_ws(wb, ["Table 29", "Table29", "TABLE 29", "TABLE29"], contains="table 29")
        ws_t31 = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")

        activity = _read_year_row(ws_t29, 32)
        stock = _read_year_row(ws_t31, 14)
        avg_vkm = _read_year_row(ws_t31, 24)

        _assert_no_blanks(activity, "Alberta Urban Transit Activity (Table 29 row 32)")
        _assert_no_blanks(stock, "Alberta Urban Transit Stock (Table 31 row 14)")
        _assert_no_blanks(avg_vkm, "Alberta Urban Transit Avg vkm (Table 31 row 24)")

        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] Alberta Urban Transit Sales: blank → kept as NaN for all years.")

        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t29, "B14:B19")]
        values = _range_values(ws_t29, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        # 2022 special formula override for Urban Transit (workbook logic; AB affects Public BusAB)

        df = _apply_urban_transit_2022_override_df(df, notes=notes, context='Alberta Urban Transit')

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "alb_urban_transit_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def build_alb_intercity_bus(out_file="alb_intercity_bus.csv"):
    """Alberta Intercity Bus.

    Activity: Table 30 row 26
    Stock:    Table 31 row 15
    Avg vkm:  Table 31 row 25
    Sales:    blank -> NaN
    Fuels:    Table 30 rows 14:19 (PJ) * 1000 -> TJ
    """
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, CEUD_ALB_FILE)
    try:
        ws_t30 = _get_ws(wb, ["Table 30", "Table30", "TABLE 30", "TABLE30"], contains="table 30")
        ws_t31 = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")

        activity = _read_year_row(ws_t30, 26)
        stock = _read_year_row(ws_t31, 15)
        avg_vkm = _read_year_row(ws_t31, 25)

        _assert_no_blanks(activity, "Alberta Intercity Bus Activity (Table 30 row 26)")
        _assert_no_blanks(stock, "Alberta Intercity Bus Stock (Table 31 row 15)")
        _assert_no_blanks(avg_vkm, "Alberta Intercity Bus Avg vkm (Table 31 row 25)")

        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append("[INFO] Alberta Intercity Bus Sales: blank → kept as NaN for all years.")

        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t30, "B14:B19")]
        values = _range_values(ws_t30, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / "alb_intercity_bus_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

# =========================
# BRITISH COLUMBIA (BC)
# =========================
# NOTE: The only regional CEUD file provided is "British Columbia and Territories" (transBCTerr2000-2022EN.xls).
# In the reference workbook, most BC series are direct links to the BCTerr tab.
# Therefore, BC mode outputs are produced by mirroring the BCTerr outputs.
# Rail + Air are excluded here because their formulas reference the workbook 'assumptions' tab.

def _copy_mode_csv(src_file: str, out_file: str, notes_file: str, notes_lines=None, extra_notes=None):
    """Copy a previously-built mode output to a new file name.

    IMPORTANT
    ---------
    This function MUST NOT use output CSVs as calculation inputs. When possible, it copies
    the already-registered in-memory DataFrame from _DF_STORE and re-registers it under
    the new key (out_file), then writes the audit CSV.

    If the source DataFrame is not registered, it falls back to copying the CSV on disk
    (legacy behaviour) but emits a warning in the notes.
    """
    import pandas as pd

    if notes_lines is None:
        notes_lines = []
    if extra_notes is None:
        extra_notes = []

    df = _get_df(src_file, required=False)
    if df is not None:
        out = df.copy()
        audit_write_df(out, OUT_DIR / out_file, index=True)
        _register_df(out_file, out)
        notes = list(notes_lines) + [f"[INFO] Copied in-memory dataframe '{src_file}' to '{out_file}' and registered it."] + list(extra_notes)
    else:
        # Legacy fallback: copy on disk
        import shutil
        src_path = OUT_DIR / src_file
        dst_path = OUT_DIR / out_file
        if not src_path.exists():
            raise FileNotFoundError(f"Source file not found for copy: {src_path}")
        shutil.copyfile(src_path, dst_path)
        try:
            out = pd.read_csv(dst_path)
            if 'year' in out.columns:
                out = out.set_index('year')
            _register_df(out_file, out)
        except Exception:
            out = None
        notes = list(notes_lines) + [f"[WARN] Source dataframe '{src_file}' not in memory; copied CSV on disk and attempted to re-register."] + list(extra_notes)

    audit_write_text('\n'.join([n for n in notes if n]) + '\n', OUT_DIR / notes_file, encoding='utf-8', mode='w')
    return out
def build_bc_rail(out_file="bc_rail.csv"):
    _copy_mode_csv(
        "bcterr_rail.csv",
        out_file,
        "bc_rail_notes.txt",
        [
            "[INFO] BC Rail mirrors BCTerr Rail (CEUD regional file is BC+Terr).",
            "[INFO] Matches workbook BC tab links to BCTerr tab.",
        ],
    )

def build_bc_air(out_file="bc_air.csv"):
    _copy_mode_csv(
        "bcterr_air.csv",
        out_file,
        "bc_air_notes.txt",
        [
            "[INFO] BC Air mirrors BCTerr Air (CEUD regional file is BC+Terr).",
            "[INFO] Matches workbook BC tab links to BCTerr tab.",
        ],
    )

def build_bc_car(out_file="bc_car.csv"):
    _copy_mode_csv(
        "bcterr_car.csv",
        out_file,
        "bc_car_notes.txt",
        [
            "[INFO] BC Car mirrors BCTerr Car (CEUD regional file is BC+Terr).",
            "[INFO] Matches workbook BC tab links to BCTerr tab.",
        ],
    )

def build_bc_light_truck(out_file="bc_light_truck.csv"):
    _copy_mode_csv(
        "bcterr_light_truck.csv",
        out_file,
        "bc_light_truck_notes.txt",
        [
            "[INFO] BC Light Truck mirrors BCTerr Light Truck (CEUD regional file is BC+Terr).",
            "[INFO] Matches workbook BC tab links to BCTerr tab.",
        ],
    )

def build_bc_ldv(out_file="bc_ldv.csv"):
    _copy_mode_csv(
        "bcterr_ldv.csv",
        out_file,
        "bc_ldv_notes.txt",
        ["[INFO] BC LDV mirrors BCTerr LDV (Car + Light Truck)."],
    )

def build_bc_motorcycle(out_file="bc_motorcycle.csv"):
    _copy_mode_csv(
        "bcterr_motorcycle.csv",
        out_file,
        "bc_motorcycle_notes.txt",
        ["[INFO] BC Motorcycle mirrors BCTerr Motorcycle."],
    )

def build_bc_school_bus(out_file="bc_school_bus.csv"):
    _copy_mode_csv(
        "bcterr_school_bus.csv",
        out_file,
        "bc_school_bus_notes.txt",
        ["[INFO] BC School Bus mirrors BCTerr School Bus."],
    )

def build_bc_urban_transit(out_file="bc_urban_transit.csv"):
    _copy_mode_csv(
        "bcterr_urban_transit.csv",
        out_file,
        "bc_urban_transit_notes.txt",
        ["[INFO] BC Urban Transit mirrors BCTerr Urban Transit."],
    )

def build_bc_intercity_bus(out_file="bc_intercity_bus.csv"):
    _copy_mode_csv(
        "bcterr_intercity_bus.csv",
        out_file,
        "bc_intercity_bus_notes.txt",
        ["[INFO] BC Intercity Bus mirrors BCTerr Intercity Bus."],
    )

# =========================
# GENERIC PROVINCE/REGION BUILDERS (pattern like AB)
# =========================

# =========================
# CAR 2022 LITERAL FORMULA HELPERS
# =========================
def _apply_car_2022_literal_overrides(total_distance, activity, occupancy, avg_pkm, fuels_out_df, fuel_total, intensity, shares, can_occ, avg_vkm, notes=None, prefix="Car"):
    """Apply literal workbook 2022 overrides for Car mode and carry downstream formulas.

    User-specified literal 2022 formulas:
      - Total Distance (M*vkm) in 2022 = 2019 Total Distance * 0.8
      - Fuel (TJ)::Motor gasoline in 2022 = 2019 Motor gasoline * 0.8

    Then carry those values through dependent rows:
      - Activity = Total Distance * CAN occupancy
      - Occupancy = Activity / Total Distance
      - Average Distance (pkm) = Average Distance (vkm) * Occupancy
      - Fuel Total = sum(fuels)
      - Intensity = Fuel Total / Activity / 1000
      - Shares = Fuel / Fuel Total * 100

    Note: Average Distance (vkm), Sales, and Stock are *not* overridden here because the user
    specifically identified only Total Distance and Motor gasoline as the literal 2022 exceptions.
    """
    y = 2022
    y_ref = 2019
    if y not in total_distance.index or y_ref not in total_distance.index:
        return total_distance, activity, occupancy, avg_pkm, fuels_out_df, fuel_total, intensity, shares

    td_2019 = pd.to_numeric(pd.Series([total_distance.loc[y_ref]]), errors='coerce').iloc[0]
    if pd.notna(td_2019):
        total_distance.loc[y] = td_2019 * 0.8

    occ_2022 = pd.to_numeric(pd.Series([can_occ.loc[y]]), errors='coerce').iloc[0]
    td_2022 = pd.to_numeric(pd.Series([total_distance.loc[y]]), errors='coerce').iloc[0]
    av_2022 = pd.to_numeric(pd.Series([avg_vkm.loc[y]]), errors='coerce').iloc[0]

    activity.loc[y] = td_2022 * occ_2022 if pd.notna(td_2022) and pd.notna(occ_2022) else np.nan
    occupancy.loc[y] = (activity.loc[y] / td_2022) if pd.notna(activity.loc[y]) and pd.notna(td_2022) and td_2022 != 0 else np.nan
    occ_out_2022 = pd.to_numeric(pd.Series([occupancy.loc[y]]), errors='coerce').iloc[0]
    avg_pkm.loc[y] = av_2022 * occ_out_2022 if pd.notna(av_2022) and pd.notna(occ_out_2022) else np.nan

    mg_col = 'Motor gasoline'
    if mg_col in fuels_out_df.columns and y_ref in fuels_out_df.index:
        mg_2019 = pd.to_numeric(pd.Series([fuels_out_df.loc[y_ref, mg_col]]), errors='coerce').iloc[0]
        if pd.notna(mg_2019):
            fuels_out_df.loc[y, mg_col] = mg_2019 * 0.8

    fuel_total.loc[y] = pd.to_numeric(fuels_out_df.loc[y], errors='coerce').sum(min_count=1)
    ft_2022 = pd.to_numeric(pd.Series([fuel_total.loc[y]]), errors='coerce').iloc[0]
    act_2022 = pd.to_numeric(pd.Series([activity.loc[y]]), errors='coerce').iloc[0]
    intensity.loc[y] = (ft_2022 / act_2022 / 1000.0) if pd.notna(ft_2022) and pd.notna(act_2022) and act_2022 != 0 else np.nan

    if y in shares.index:
        if pd.notna(ft_2022) and ft_2022 != 0:
            shares.loc[y] = pd.to_numeric(fuels_out_df.loc[y], errors='coerce') / ft_2022 * 100.0
        else:
            shares.loc[y] = np.nan

    if notes is not None:
        notes.append(f"[INFO] {prefix} 2022 literal overrides applied: Total Distance = 2019*0.8 and Motor gasoline = 2019*0.8, then downstream rows recomputed.")

    return total_distance, activity, occupancy, avg_pkm, fuels_out_df, fuel_total, intensity, shares

def _build_prov_car(prefix: str, ceud_path: Path, out_file: str):
    """Province car builder using provincial CEUD tables and CAN occupancy (matches workbook pattern)."""
    notes = []

    # CAN occupancy source
    can_car = _get_df_any(['car_full.csv', 'car_full'], required=True)
    if "year" in can_car.columns:
        can_car = can_car.set_index("year")
    can_car.index = can_car.index.astype(int)

    occ_col = "Occupancy (persons /vehicle"
    if occ_col not in can_car.columns:
        raise ValueError(f"Expected '{occ_col}' in output/car_full.csv")

    can_occ = can_car.loc[YEARS, occ_col]
    _assert_no_blanks(can_occ, f"CAN Car Occupancy (for {prefix} Car activity)")

    excel = _excel_app()
    wb = _open_book(excel, ceud_path)
    try:
        ws_t21 = _get_ws(wb, ["Table 21", "Table21", "TABLE 21", "TABLE21"], contains="table 21")
        ws_t20 = _get_ws(wb, ["Table 20", "Table20", "TABLE 20", "TABLE20"], contains="table 20")

        sales = _read_year_row(ws_t21, 13)
        stock = _read_year_row(ws_t21, 16)
        avg_vkm = _read_year_row(ws_t21, 19)

        _assert_no_blanks(sales, f"{prefix} Car Sales (Table 21 row 13)")
        _assert_no_blanks(stock, f"{prefix} Car Stock (Table 21 row 16)")
        _assert_no_blanks(avg_vkm, f"{prefix} Car Avg vkm (Table 21 row 19)")

        total_distance = stock * avg_vkm / 1000.0
        activity = total_distance * can_occ
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t20, "B14:B19")]
        values = _range_values(ws_t20, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0
        total_distance, activity, occupancy, avg_pkm, fuels_out_df, fuel_total, intensity, shares = _apply_car_2022_literal_overrides(
            total_distance, activity, occupancy, avg_pkm, fuels_out_df, fuel_total, intensity, shares, can_occ, avg_vkm, notes, prefix=f"{prefix} Car"
        )

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df = _overwrite_car_2022_output_df(df, can_occ, notes, prefix=f"{prefix} Car")
        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        notes.append(f"[INFO] {prefix} Car Activity computed as TotalDistance * CAN Car Occupancy (output/car_full.csv).")
        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / f"{prefix.lower()}_car_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def _build_prov_light_truck(prefix: str, ceud_path: Path, out_file: str):
    """Province light truck builder using provincial CEUD tables and CAN occupancy."""
    notes = []

    can_lt = _get_df_any(['light_truck_full.csv', 'light_truck_full'], required=True)
    if "year" in can_lt.columns:
        can_lt = can_lt.set_index("year")
    can_lt.index = can_lt.index.astype(int)

    occ_col = "Occupancy (persons /vehicle"
    if occ_col not in can_lt.columns:
        raise ValueError(f"Expected '{occ_col}' in output/light_truck_full.csv")

    can_occ = can_lt.loc[YEARS, occ_col]
    _assert_no_blanks(can_occ, f"CAN Light Truck Occupancy (for {prefix} Light Truck activity)")

    excel = _excel_app()
    wb = _open_book(excel, ceud_path)
    try:
        ws_t37 = _get_ws(wb, ["Table 37", "Table37", "TABLE 37", "TABLE37"], contains="table 37")
        ws_t25 = _get_ws(wb, ["Table 25", "Table25", "TABLE 25", "TABLE25"], contains="table 25")

        sales = _read_year_row(ws_t37, 13)
        stock = _read_year_row(ws_t37, 25)
        avg_vkm = _read_year_row(ws_t37, 37)

        _assert_no_blanks(sales, f"{prefix} Light Truck Sales (Table 37 row 13)")
        _assert_no_blanks(stock, f"{prefix} Light Truck Stock (Table 37 row 25)")
        _assert_no_blanks(avg_vkm, f"{prefix} Light Truck Avg vkm (Table 37 row 37)")

        total_distance = stock * avg_vkm / 1000.0
        activity = total_distance * can_occ
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(ws_t25, "B14:B19")]
        values = _range_values(ws_t25, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        notes.append(f"[INFO] {prefix} Light Truck Activity computed as TotalDistance * CAN Light Truck Occupancy (output/light_truck_full.csv).")
        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / f"{prefix.lower()}_light_truck_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def _build_prov_motorcycle(prefix: str, ceud_path: Path, out_file: str):
    """Province motorcycle builder from Table 32 (motor gasoline only)."""
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, ceud_path)
    try:
        ws_t32 = _get_ws(wb, ["Table 32", "Table32", "TABLE 32", "TABLE32"], contains="table 32")

        activity = _read_year_row(ws_t32, 15)
        stock = _read_year_row(ws_t32, 26)
        avg_vkm = _read_year_row(ws_t32, 27)

        _assert_no_blanks(activity, f"{prefix} Motorcycle Activity (Table 32 row 15)")
        _assert_no_blanks(stock, f"{prefix} Motorcycle Stock (Table 32 row 26)")
        _assert_no_blanks(avg_vkm, f"{prefix} Motorcycle Avg vkm (Table 32 row 27)")

        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append(f"[INFO] {prefix} Motorcycle Sales: blank → kept as NaN for all years.")

        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        mg_pj = _read_year_row(ws_t32, 12)
        mg_tj = _fuel_handle_na(mg_pj * FUEL_SCALE, "Motor gasoline", notes)

        fuels_out = {f: pd.Series([np.nan] * N_YEARS, index=YEARS) for f in FUELS}
        fuels_out["Motor gasoline"] = mg_tj

        fuels_out_df = pd.DataFrame(fuels_out)
        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / f"{prefix.lower()}_motorcycle_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def _build_prov_bus_mode(prefix: str, ceud_path: Path, out_file: str, mode: str):
    """Province bus mode builder for School Bus / Urban Transit / Intercity Bus."""
    notes = []

    excel = _excel_app()
    wb = _open_book(excel, ceud_path)
    try:
        if mode == 'school_bus':
            ws_mode = _get_ws(wb, ["Table 28", "Table28", "TABLE 28", "TABLE28"], contains="table 28")
            ws_vars = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")
            activity = _read_year_row(ws_mode, 30)
            stock = _read_year_row(ws_vars, 13)
            avg_vkm = _read_year_row(ws_vars, 23)
            fuel_ws = ws_mode
            fuel_name = 'Table 28'
        elif mode == 'urban_transit':
            ws_mode = _get_ws(wb, ["Table 29", "Table29", "TABLE 29", "TABLE29"], contains="table 29")
            ws_vars = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")
            activity = _read_year_row(ws_mode, 32)
            stock = _read_year_row(ws_vars, 14)
            avg_vkm = _read_year_row(ws_vars, 24)
            fuel_ws = ws_mode
            fuel_name = 'Table 29'
        elif mode == 'intercity_bus':
            ws_mode = _get_ws(wb, ["Table 30", "Table30", "TABLE 30", "TABLE30"], contains="table 30")
            ws_vars = _get_ws(wb, ["Table 31", "Table31", "TABLE 31", "TABLE31"], contains="table 31")
            activity = _read_year_row(ws_mode, 26)
            stock = _read_year_row(ws_vars, 15)
            avg_vkm = _read_year_row(ws_vars, 25)
            fuel_ws = ws_mode
            fuel_name = 'Table 30'
        else:
            raise ValueError(f"Unknown bus mode: {mode}")

        _assert_no_blanks(activity, f"{prefix} {mode} Activity ({fuel_name})")
        _assert_no_blanks(stock, f"{prefix} {mode} Stock (Table 31)")
        _assert_no_blanks(avg_vkm, f"{prefix} {mode} Avg vkm (Table 31)")

        sales = pd.Series([np.nan] * N_YEARS, index=YEARS)
        notes.append(f"[INFO] {prefix} {mode} Sales: blank → kept as NaN for all years.")

        # Urban Transit 2022 special formulas (workbook): 2022 values derived from 2019
        if mode == 'urban_transit':
            y_ref, y = 2019, 2022
            if y in activity.index and y_ref in activity.index:
                activity.loc[y] = pd.to_numeric(activity.loc[y_ref], errors='coerce') * 0.8
                stock.loc[y] = pd.to_numeric(stock.loc[y_ref], errors='coerce')
                avg_vkm.loc[y] = pd.to_numeric(avg_vkm.loc[y_ref], errors='coerce')
                notes.append(f"[OK] {prefix} Urban Transit 2022 override applied (Activity=2019*0.8; Stock/AvgVkm=2019).")
            else:
                notes.append(f"[WARN] {prefix} Urban Transit 2022 override skipped (missing 2019/2022).")

        total_distance = stock * avg_vkm / 1000.0
        occupancy = activity / total_distance.replace(0, np.nan)
        avg_pkm = avg_vkm * occupancy

        labels = [str(r[0]).strip() for r in _range_values(fuel_ws, "B14:B19")]
        values = _range_values(fuel_ws, "C14:Y19")
        fuel_df = pd.DataFrame(
            [_to_float_array(v) for v in values],
            index=labels,
            columns=YEARS,
        ).T * FUEL_SCALE

        fuels_out = {}
        for f in FUELS:
            match = [c for c in fuel_df.columns if c.lower() == f.lower()]
            s = fuel_df[match[0]] if match else pd.Series([np.nan] * N_YEARS, index=YEARS)
            fuels_out[f] = _fuel_handle_na(s, f, notes)

        fuels_out_df = pd.DataFrame(fuels_out)

        # Urban Transit 2022 special formula (Diesel fuel oil): 2022 = 2019
        if mode == 'urban_transit':
            y_ref, y = 2019, 2022
            if 'Diesel fuel oil' in fuels_out_df.columns and y in fuels_out_df.index and y_ref in fuels_out_df.index:
                fuels_out_df.loc[y, 'Diesel fuel oil'] = pd.to_numeric(fuels_out_df.loc[y_ref, 'Diesel fuel oil'], errors='coerce')
                notes.append(f"[OK] {prefix} Urban Transit Diesel fuel oil 2022 override applied (2019 value).")

        fuel_total = fuels_out_df.sum(axis=1, min_count=1)
        shares = fuels_out_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
        intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

        df = pd.concat(
            [
                activity.rename("Activity (millions passenger-kilometres)"),
                sales.rename("Sales (thousands)"),
                stock.rename("Stock (thousands)"),
                avg_vkm.rename("Average Distance (vkm)"),
                total_distance.rename("Total Distance (M*vkm)"),
                occupancy.rename("Occupancy (persons /vehicle"),
                avg_pkm.rename("Average Distance (pkm)"),
                fuels_out_df.add_prefix("Fuel (TJ)::"),
                fuel_total.rename("Fuel (TJ)::Total"),
                intensity.rename("Intensity (GJ / pkm)"),
                shares.add_prefix("Share (%)::"),
            ],
            axis=1,
        )

        df.index.name = "year"
        audit_write_df(df, OUT_DIR / out_file)
        _register_df(out_file, df)

        audit_write_text("\n".join(notes) if notes else "[INFO] No notes.\n", OUT_DIR / f"{prefix.lower()}_{mode}_notes.txt", encoding="utf-8", mode="w")
        if _audit_enabled(): print(f"[OK] Wrote {out_file}")

    finally:
        _close_book(wb, excel)

def _build_prov_ldv(prefix: str, car_file: str, light_truck_file: str, out_file: str):
    """LDV = Car + Light Truck from existing outputs."""
    notes = []

    car = _read_mode_output_csv(OUT_DIR / car_file).loc[YEARS]
    lt = _read_mode_output_csv(OUT_DIR / light_truck_file).loc[YEARS]

    COL_ACTIVITY = "Activity (millions passenger-kilometres)"
    COL_SALES = "Sales (thousands)"
    COL_STOCK = "Stock (thousands)"
    COL_TOTAL_DIST = "Total Distance (M*vkm)"

    for c in [COL_ACTIVITY, COL_SALES, COL_STOCK, COL_TOTAL_DIST]:
        if c not in car.columns:
            raise ValueError(f"{prefix} LDV: missing required column in Car: {c}")
        if c not in lt.columns:
            raise ValueError(f"{prefix} LDV: missing required column in Light Truck: {c}")

    activity = car[COL_ACTIVITY] + lt[COL_ACTIVITY]
    sales = car[COL_SALES] + lt[COL_SALES]
    stock = car[COL_STOCK] + lt[COL_STOCK]
    total_distance = car[COL_TOTAL_DIST] + lt[COL_TOTAL_DIST]

    _assert_no_blanks(activity, f"{prefix} LDV Activity")
    _assert_no_blanks(stock, f"{prefix} LDV Stock")
    _assert_no_blanks(total_distance, f"{prefix} LDV Total Distance")

    avg_vkm = (total_distance * 1000.0) / stock.replace(0, np.nan)
    occupancy = activity / total_distance.replace(0, np.nan)

    avg_pkm = pd.Series([np.nan] * N_YEARS, index=YEARS)
    notes.append(f"[INFO] {prefix} LDV Average Distance (pkm) is blank in workbook → kept as NaN for all years.")

    fuel_cols = [c for c in car.columns if c.startswith("Fuel (TJ)::") and c != "Fuel (TJ)::Total"]
    if not fuel_cols:
        raise ValueError(f"{prefix} LDV: no fuel columns found in Car output.")

    fuels_df = pd.DataFrame(index=YEARS)
    for c in fuel_cols:
        if c not in lt.columns:
            raise ValueError(f"{prefix} LDV: missing fuel column in Light Truck: {c}")
        fuels_df[c] = car[c] + lt[c]

    fuel_total = fuels_df.sum(axis=1, min_count=1)
    shares_df = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
    shares_df = shares_df.rename(columns=lambda x: "Share (%)::" + x.split("::", 1)[1])

    intensity = fuel_total / activity.replace(0, np.nan) / 1000.0
    activity, sales, stock, total_distance, avg_vkm, occupancy, fuels_df, fuel_total, intensity, shares_df = _recompute_ldv_2022_formula_rows(
        activity, sales, stock, total_distance, avg_vkm, occupancy, fuels_df, fuel_total, intensity, shares_df, car, lt, notes, prefix=f"{prefix} LDV"
    )

    out = pd.concat(
        [
            activity.rename(COL_ACTIVITY),
            sales.rename(COL_SALES),
            stock.rename(COL_STOCK),
            avg_vkm.rename("Average Distance (vkm)"),
            total_distance.rename(COL_TOTAL_DIST),
            occupancy.rename("Occupancy (persons /vehicle"),
            avg_pkm.rename("Average Distance (pkm)"),
            fuels_df,
            fuel_total.rename("Fuel (TJ)::Total"),
            intensity.rename("Intensity (GJ / pkm)"),
            shares_df,
        ],
        axis=1,
    )

    out.index.name = "year"
    audit_write_df(out, OUT_DIR / out_file)
    _register_df(out_file, out)

    notes.append(f"[INFO] {prefix} LDV computed as {prefix} Car + {prefix} Light Truck (row-wise sums).")

    audit_write_text("\n".join(notes) + "\n", OUT_DIR / f"{prefix.lower()}_ldv_notes.txt", encoding="utf-8", mode="w")
    if _audit_enabled(): print(f"[OK] Wrote {out_file}")

def _build_prov_passenger_partial(prefix: str, out_file: str, components: dict):
    """Passenger aggregate excluding Rail + Air (assumptions)."""
    notes = []

    dfs = {}
    for k, p in components.items():
        pth = OUT_DIR / p
        if not pth.exists():
            raise FileNotFoundError(f"Missing {pth}. Build {prefix} {k} first.")
        df = pd.read_csv(pth)
        if "year" in df.columns:
            df = df.set_index("year")
        df.index = df.index.astype(int)
        dfs[k] = df.loc[YEARS]

    COL_ACTIVITY = "Activity (millions passenger-kilometres)"
    activity = sum(dfs[k][COL_ACTIVITY] for k in dfs)
    _assert_no_blanks(activity, f"{prefix} Passenger Activity (partial)")

    fuel_cols = [c for c in dfs[next(iter(dfs))].columns if c.startswith("Fuel (TJ)::") and c != "Fuel (TJ)::Total"]
    fuels_df = pd.DataFrame(index=YEARS)
    for c in fuel_cols:
        fuels_df[c] = sum(dfs[k][c] for k in dfs)

    fuel_total = fuels_df.sum(axis=1, min_count=1)
    shares = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0) * 100.0
    shares = shares.rename(columns=lambda x: "Share (%)::" + x.split("::", 1)[1])

    intensity = fuel_total / activity.replace(0, np.nan) / 1000.0

    nan_series = pd.Series([np.nan] * N_YEARS, index=YEARS)

    out = pd.concat(
        [
            activity.rename(COL_ACTIVITY),
            nan_series.rename("Sales (thousands)"),
            nan_series.rename("Stock (thousands)"),
            nan_series.rename("Average Distance (vkm)"),
            nan_series.rename("Total Distance (M*vkm)"),
            nan_series.rename("Occupancy (persons /vehicle"),
            nan_series.rename("Average Distance (pkm)"),
            fuels_df,
            fuel_total.rename("Fuel (TJ)::Total"),
            intensity.rename("Intensity (GJ / pkm)"),
            shares,
        ],
        axis=1,
    )

    out.index.name = "year"
    audit_write_df(out, OUT_DIR / out_file)
    _register_df(out_file, out)

    notes.append(f"[INFO] {prefix} Passenger aggregate is PARTIAL: excludes Rail and Air (require 'assumptions' tab).")
    notes.append(f"[INFO] Included modes: {', '.join(components.keys())}.")

    audit_write_text("\n".join(notes) + "\n", OUT_DIR / f"{prefix.lower()}_passenger_partial_notes.txt", encoding="utf-8", mode="w")
    if _audit_enabled(): print(f"[OK] Wrote {out_file}")

def _build_prov_passenger_full(prefix: str, out_file: str, components: dict):
    """Passenger aggregate (FULL) computed per Passenger Full formulas.

    Passenger Full Activity must equal the sum of these seven modes:
      LDV, Motorcycle, School Bus, Urban Transit, Intercity Bus, Rail (Passengers), Air (Passengers).

    Excel SUM treats blanks as 0; pandas arithmetic propagates NaN.
    Therefore we coerce to numeric and fill NaN->0 prior to summation.

    Rail/Air must contribute passenger activity (pkm), not total pkm.
    We prefer 'Passengers::passenger-kilometres (millions)' when present.
    """

    # Read component mode outputs
    dfs = {}
    for mode, fname in components.items():
        pth = OUT_DIR / fname
        if not pth.exists():
            raise FileNotFoundError(f"Missing {pth}. Build {prefix} {mode} first.")
        df = pd.read_csv(pth)
        if 'year' in df.columns:
            df = df.set_index('year')
        df.index = df.index.astype(int)
        dfs[mode] = df.loc[YEARS]

    COL_ACTIVITY = 'Activity (millions passenger-kilometres)'
    COL_PAX_PKM = 'Passengers::passenger-kilometres (millions)'

    def _activity_series(mode: str, df: pd.DataFrame) -> pd.Series:
        ml = mode.lower()
        is_rail = 'rail' in ml
        is_air = ('air' in ml) or ('aviation' in ml)

        if (is_rail or is_air):
            # passenger pkm preferred
            for c in [
                COL_PAX_PKM,
                'Passengers::passenger-kilometres (million)',
                'Passengers::passenger kilometres (millions)',
                'Passengers::passenger-kilometres (M)',
            ]:
                if c in df.columns:
                    return pd.to_numeric(df[c], errors='coerce').fillna(0.0)

        # standard activity
        if COL_ACTIVITY in df.columns:
            return pd.to_numeric(df[COL_ACTIVITY], errors='coerce').fillna(0.0)

        # last-resort fallback
        for c in ['Total::passenger-kilometres (millions)', 'Total::passenger-kilometres (million)']:
            if c in df.columns:
                return pd.to_numeric(df[c], errors='coerce').fillna(0.0)

        raise ValueError(f"{prefix} Passenger Full: cannot find activity column for component '{mode}'.")

    # Activity = sum across components (NaN treated as 0)
    activity = None
    for mode, df in dfs.items():
        s = _activity_series(mode, df)
        activity = s if activity is None else (activity + s)

    # Fuels, total, intensity, shares (kept consistent with existing Passenger Full columns)
    passenger_fuels = [
        'Aviation turbo fuel',
        'Aviation gasoline',
        'Diesel fuel oil',
        'Biodiesel fuel',
        'Motor gasoline',
        'Ethanol',
        'Electricity',
        'Natural gas',
        'Heavy fuel oil',
        'Propane',
    ]

    fuels_df = pd.DataFrame(index=YEARS)
    for fuel in passenger_fuels:
        col = 'Fuel (TJ)::' + fuel
        total = None
        for mode, df in dfs.items():
            if col in df.columns:
                s = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
            else:
                s = pd.Series([0.0] * N_YEARS, index=YEARS)
            total = s if total is None else (total + s)
        fuels_df[col] = total

    fuel_total = fuels_df.sum(axis=1)
    intensity = (fuel_total / activity.replace(0, np.nan) / 1000.0).replace([np.inf, -np.inf], np.nan).fillna(0.0)

    shares = fuels_df.divide(fuel_total.replace(0, np.nan), axis=0).replace([np.inf, -np.inf], np.nan).fillna(0.0) * 100.0
    shares = shares.rename(columns=lambda x: 'Share (%)::' + x.split('::', 1)[1])

    nan_series = pd.Series([np.nan] * N_YEARS, index=YEARS)

    out = pd.concat(
        [
            activity.rename(COL_ACTIVITY),
            nan_series.rename('Sales (thousands)'),
            nan_series.rename('Stock (thousands)'),
            nan_series.rename('Average Distance (vkm)'),
            nan_series.rename('Total Distance (M*vkm)'),
            nan_series.rename('Occupancy (persons /vehicle)'),
            nan_series.rename('Average Distance (pkm)'),
            fuels_df,
            fuel_total.rename('Fuel (TJ)::Total'),
            intensity.rename('Intensity (GJ / pkm)'),
            shares,
        ],
        axis=1,
    )

    out.index.name = 'year'
    audit_write_df(out, OUT_DIR / out_file)

    _register_df(out_file, out)
    if _audit_enabled(): print(f"[OK] Wrote {out_file}")

# =========================
# AT: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_at_car(out_file="at_car.csv"):
    _build_prov_car(prefix="AT", ceud_path=CEUD_ATL_FILE, out_file=out_file)
def build_at_light_truck(out_file="at_light_truck.csv"):
    _build_prov_light_truck(prefix="AT", ceud_path=CEUD_ATL_FILE, out_file=out_file)
def build_at_ldv(out_file="at_ldv.csv"):
    _build_prov_ldv(prefix="AT", car_file="at_car.csv", light_truck_file="at_light_truck.csv", out_file=out_file)
def build_at_motorcycle(out_file="at_motorcycle.csv"):
    _build_prov_motorcycle(prefix="AT", ceud_path=CEUD_ATL_FILE, out_file=out_file)
def build_at_school_bus(out_file="at_school_bus.csv"):
    _build_prov_bus_mode(prefix="AT", ceud_path=CEUD_ATL_FILE, out_file=out_file, mode='school_bus')
def build_at_urban_transit(out_file="at_urban_transit.csv"):
    _build_prov_bus_mode(prefix="AT", ceud_path=CEUD_ATL_FILE, out_file=out_file, mode='urban_transit')
def build_at_intercity_bus(out_file="at_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="AT", ceud_path=CEUD_ATL_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# MB: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_mb_car(out_file="mb_car.csv"):
    _build_prov_car(prefix="MB", ceud_path=CEUD_MAN_FILE, out_file=out_file)
def build_mb_light_truck(out_file="mb_light_truck.csv"):
    _build_prov_light_truck(prefix="MB", ceud_path=CEUD_MAN_FILE, out_file=out_file)
def build_mb_ldv(out_file="mb_ldv.csv"):
    _build_prov_ldv(prefix="MB", car_file="mb_car.csv", light_truck_file="mb_light_truck.csv", out_file=out_file)
def build_mb_motorcycle(out_file="mb_motorcycle.csv"):
    _build_prov_motorcycle(prefix="MB", ceud_path=CEUD_MAN_FILE, out_file=out_file)
def build_mb_school_bus(out_file="mb_school_bus.csv"):
    _build_prov_bus_mode(prefix="MB", ceud_path=CEUD_MAN_FILE, out_file=out_file, mode='school_bus')
def build_mb_urban_transit(out_file="mb_urban_transit.csv"):
    _build_prov_bus_mode(prefix="MB", ceud_path=CEUD_MAN_FILE, out_file=out_file, mode='urban_transit')
def build_mb_intercity_bus(out_file="mb_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="MB", ceud_path=CEUD_MAN_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# NB: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_nb_car(out_file="nb_car.csv"):
    _build_prov_car(prefix="NB", ceud_path=CEUD_NB_FILE, out_file=out_file)
def build_nb_light_truck(out_file="nb_light_truck.csv"):
    _build_prov_light_truck(prefix="NB", ceud_path=CEUD_NB_FILE, out_file=out_file)
def build_nb_ldv(out_file="nb_ldv.csv"):
    _build_prov_ldv(prefix="NB", car_file="nb_car.csv", light_truck_file="nb_light_truck.csv", out_file=out_file)
def build_nb_motorcycle(out_file="nb_motorcycle.csv"):
    _build_prov_motorcycle(prefix="NB", ceud_path=CEUD_NB_FILE, out_file=out_file)
def build_nb_school_bus(out_file="nb_school_bus.csv"):
    _build_prov_bus_mode(prefix="NB", ceud_path=CEUD_NB_FILE, out_file=out_file, mode='school_bus')
def build_nb_urban_transit(out_file="nb_urban_transit.csv"):
    _build_prov_bus_mode(prefix="NB", ceud_path=CEUD_NB_FILE, out_file=out_file, mode='urban_transit')
def build_nb_intercity_bus(out_file="nb_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="NB", ceud_path=CEUD_NB_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# NL: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_nl_car(out_file="nl_car.csv"):
    _build_prov_car(prefix="NL", ceud_path=CEUD_NFLD_FILE, out_file=out_file)
def build_nl_light_truck(out_file="nl_light_truck.csv"):
    _build_prov_light_truck(prefix="NL", ceud_path=CEUD_NFLD_FILE, out_file=out_file)
def build_nl_ldv(out_file="nl_ldv.csv"):
    _build_prov_ldv(prefix="NL", car_file="nl_car.csv", light_truck_file="nl_light_truck.csv", out_file=out_file)
def build_nl_motorcycle(out_file="nl_motorcycle.csv"):
    _build_prov_motorcycle(prefix="NL", ceud_path=CEUD_NFLD_FILE, out_file=out_file)
def build_nl_school_bus(out_file="nl_school_bus.csv"):
    _build_prov_bus_mode(prefix="NL", ceud_path=CEUD_NFLD_FILE, out_file=out_file, mode='school_bus')
def build_nl_urban_transit(out_file="nl_urban_transit.csv"):
    _build_prov_bus_mode(prefix="NL", ceud_path=CEUD_NFLD_FILE, out_file=out_file, mode='urban_transit')
def build_nl_intercity_bus(out_file="nl_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="NL", ceud_path=CEUD_NFLD_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# NS: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_ns_car(out_file="ns_car.csv"):
    _build_prov_car(prefix="NS", ceud_path=CEUD_NS_FILE, out_file=out_file)
def build_ns_light_truck(out_file="ns_light_truck.csv"):
    _build_prov_light_truck(prefix="NS", ceud_path=CEUD_NS_FILE, out_file=out_file)
def build_ns_ldv(out_file="ns_ldv.csv"):
    _build_prov_ldv(prefix="NS", car_file="ns_car.csv", light_truck_file="ns_light_truck.csv", out_file=out_file)
def build_ns_motorcycle(out_file="ns_motorcycle.csv"):
    _build_prov_motorcycle(prefix="NS", ceud_path=CEUD_NS_FILE, out_file=out_file)
def build_ns_school_bus(out_file="ns_school_bus.csv"):
    _build_prov_bus_mode(prefix="NS", ceud_path=CEUD_NS_FILE, out_file=out_file, mode='school_bus')
def build_ns_urban_transit(out_file="ns_urban_transit.csv"):
    _build_prov_bus_mode(prefix="NS", ceud_path=CEUD_NS_FILE, out_file=out_file, mode='urban_transit')
def build_ns_intercity_bus(out_file="ns_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="NS", ceud_path=CEUD_NS_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# ON: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_on_car(out_file="on_car.csv"):
    _build_prov_car(prefix="ON", ceud_path=CEUD_ONT_FILE, out_file=out_file)
def build_on_light_truck(out_file="on_light_truck.csv"):
    _build_prov_light_truck(prefix="ON", ceud_path=CEUD_ONT_FILE, out_file=out_file)
def build_on_ldv(out_file="on_ldv.csv"):
    _build_prov_ldv(prefix="ON", car_file="on_car.csv", light_truck_file="on_light_truck.csv", out_file=out_file)
def build_on_motorcycle(out_file="on_motorcycle.csv"):
    _build_prov_motorcycle(prefix="ON", ceud_path=CEUD_ONT_FILE, out_file=out_file)
def build_on_school_bus(out_file="on_school_bus.csv"):
    _build_prov_bus_mode(prefix="ON", ceud_path=CEUD_ONT_FILE, out_file=out_file, mode='school_bus')
def build_on_urban_transit(out_file="on_urban_transit.csv"):
    _build_prov_bus_mode(prefix="ON", ceud_path=CEUD_ONT_FILE, out_file=out_file, mode='urban_transit')
def build_on_intercity_bus(out_file="on_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="ON", ceud_path=CEUD_ONT_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# PE: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_pe_car(out_file="pe_car.csv"):
    _build_prov_car(prefix="PE", ceud_path=CEUD_PEI_FILE, out_file=out_file)
def build_pe_light_truck(out_file="pe_light_truck.csv"):
    _build_prov_light_truck(prefix="PE", ceud_path=CEUD_PEI_FILE, out_file=out_file)
def build_pe_ldv(out_file="pe_ldv.csv"):
    _build_prov_ldv(prefix="PE", car_file="pe_car.csv", light_truck_file="pe_light_truck.csv", out_file=out_file)
def build_pe_motorcycle(out_file="pe_motorcycle.csv"):
    _build_prov_motorcycle(prefix="PE", ceud_path=CEUD_PEI_FILE, out_file=out_file)
def build_pe_school_bus(out_file="pe_school_bus.csv"):
    _build_prov_bus_mode(prefix="PE", ceud_path=CEUD_PEI_FILE, out_file=out_file, mode='school_bus')
def build_pe_urban_transit(out_file="pe_urban_transit.csv"):
    _build_prov_bus_mode(prefix="PE", ceud_path=CEUD_PEI_FILE, out_file=out_file, mode='urban_transit')
def build_pe_intercity_bus(out_file="pe_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="PE", ceud_path=CEUD_PEI_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# QC: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_qc_car(out_file="qc_car.csv"):
    _build_prov_car(prefix="QC", ceud_path=CEUD_QUE_FILE, out_file=out_file)
def build_qc_light_truck(out_file="qc_light_truck.csv"):
    _build_prov_light_truck(prefix="QC", ceud_path=CEUD_QUE_FILE, out_file=out_file)
def build_qc_ldv(out_file="qc_ldv.csv"):
    _build_prov_ldv(prefix="QC", car_file="qc_car.csv", light_truck_file="qc_light_truck.csv", out_file=out_file)
def build_qc_motorcycle(out_file="qc_motorcycle.csv"):
    _build_prov_motorcycle(prefix="QC", ceud_path=CEUD_QUE_FILE, out_file=out_file)
def build_qc_school_bus(out_file="qc_school_bus.csv"):
    _build_prov_bus_mode(prefix="QC", ceud_path=CEUD_QUE_FILE, out_file=out_file, mode='school_bus')
def build_qc_urban_transit(out_file="qc_urban_transit.csv"):
    _build_prov_bus_mode(prefix="QC", ceud_path=CEUD_QUE_FILE, out_file=out_file, mode='urban_transit')
def build_qc_intercity_bus(out_file="qc_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="QC", ceud_path=CEUD_QUE_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# SK: BUILDERS (NO ASSUMPTIONS MODES)
# =========================
def build_sk_car(out_file="sk_car.csv"):
    _build_prov_car(prefix="SK", ceud_path=CEUD_SASK_FILE, out_file=out_file)
def build_sk_light_truck(out_file="sk_light_truck.csv"):
    _build_prov_light_truck(prefix="SK", ceud_path=CEUD_SASK_FILE, out_file=out_file)
def build_sk_ldv(out_file="sk_ldv.csv"):
    _build_prov_ldv(prefix="SK", car_file="sk_car.csv", light_truck_file="sk_light_truck.csv", out_file=out_file)
def build_sk_motorcycle(out_file="sk_motorcycle.csv"):
    _build_prov_motorcycle(prefix="SK", ceud_path=CEUD_SASK_FILE, out_file=out_file)
def build_sk_school_bus(out_file="sk_school_bus.csv"):
    _build_prov_bus_mode(prefix="SK", ceud_path=CEUD_SASK_FILE, out_file=out_file, mode='school_bus')
def build_sk_urban_transit(out_file="sk_urban_transit.csv"):
    _build_prov_bus_mode(prefix="SK", ceud_path=CEUD_SASK_FILE, out_file=out_file, mode='urban_transit')
def build_sk_intercity_bus(out_file="sk_intercity_bus.csv"):
    _build_prov_bus_mode(prefix="SK", ceud_path=CEUD_SASK_FILE, out_file=out_file, mode='intercity_bus')

# =========================
# ASSUMPTIONS (STRUCTURED, NO EXCEL WORKBOOK INPUT)
# =========================
# NOTE: We do NOT use 'transportation personal_source data.xlsx' as an input.
# It is only a reference for structure. These assumptions tables are built from:
#   (a) the CSV outputs produced by this script (e.g., *_car.csv), and
#   (b) explicit constants shown in the assumptions formulas you provided.
# Output is long/tidy: one row per (mode, fuel, year, metric).

ASSUMP_PROV_ORDER = [
    "BC","AB","SK","MB","ON","QC","NB","NS","PE","NL","YT","NT","NU","AT","TR"
]
ASSUMP_PROV_NAMES = {
    "BC":"British Columbia",
    "AB":"Alberta",
    "SK":"Saskatchewan",
    "MB":"Manitoba",
    "ON":"Ontario",
    "QC":"Quebec",
    "NB":"New Brunswick",
    "NS":"Nova Scotia",
    "PE":"Prince Edward Island",
    "NL":"Newfoundland and Labrador",
    "YT":"Yukon",
    "NT":"Northwest Territories",
    "NU":"Nunavut",
    "AT":"Atlantic",
    "TR":"Territories",
}

# k*pkm forecast assumptions from the workbook-style assumptions sheet.
# Historical CAGR is calculated over 2005->2019.
# 2022 uses the 2019 value times the reference multiplier.
# 2023->2050 use the 2023-period reference CAGR.
# 2051->2100 use the 2051-period reference CAGR.
KPKM_MODE_CONFIG = {
    "Cars":          {"rel_to_hist_2023": 0.10,  "rel_to_hist_2051": 0.05,  "ref_multiplier": 0.90, "hist_start": 2005, "hist_end": 2019},
    "Light Trucks":  {"rel_to_hist_2023": 0.25,  "rel_to_hist_2051": 0.125, "ref_multiplier": 0.90, "hist_start": 2005, "hist_end": 2019},
    "Motorcycle":    {"rel_to_hist_2023": 0.25,  "rel_to_hist_2051": 0.125, "ref_multiplier": 0.90, "hist_start": 2005, "hist_end": 2019},
    "School Bus":    {"rel_to_hist_2023": 0.25,  "rel_to_hist_2051": 0.125, "ref_multiplier": 1.00, "hist_start": 2005, "hist_end": 2019},
    "Transit":       {"rel_to_hist_2023": 0.50,  "rel_to_hist_2051": 0.25,  "ref_multiplier": 0.90, "hist_start": 2005, "hist_end": 2019},
    "Intercity Bus": {"rel_to_hist_2023": 0.00,  "rel_to_hist_2051": 0.00,  "ref_multiplier": 0.90, "hist_start": 2005, "hist_end": 2019},
    "Rail":          {"rel_to_hist_2023": 0.20,  "rel_to_hist_2051": 0.10,  "ref_multiplier": 0.90, "hist_start": 2005, "hist_end": 2019},
    "Aviation":      {"rel_to_hist_2023": 0.20,  "rel_to_hist_2051": 0.10,  "ref_multiplier": 0.90, "hist_start": 2005, "hist_end": 2019},
}

# Province adjustment column (M in your sheet). Blank => 1.0.
# If you later specify adjustments, add them here.
KPKM_PROV_ADJ = {prov: 1.0 for prov in ASSUMP_PROV_ORDER}

# Explicit workbook k*pkm assumptions hard-wired from the user-provided assumptions sheet.
# These values are authoritative for build_assumptions_tables_structured() and build_calc().
KPKM_EXPLICIT_ASSUMPTIONS = {'Cars': {'BC': {'historical_cagr': -0.0036860855006342064,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0003686085500634207,
                 'reference_cagr_2051': -0.00018430427503171034},
          'AB': {'historical_cagr': -0.012736676541694858,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.001273667654169486,
                 'reference_cagr_2051': -0.000636833827084743},
          'SK': {'historical_cagr': -0.0077818763864824225,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0007781876386482423,
                 'reference_cagr_2051': -0.00038909381932412116},
          'MB': {'historical_cagr': 0.0067279068021508515,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.0006727906802150851,
                 'reference_cagr_2051': 0.0003363953401075426},
          'ON': {'historical_cagr': -0.010134630760170293,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0010134630760170293,
                 'reference_cagr_2051': -0.0005067315380085146},
          'QC': {'historical_cagr': -0.0063301757418113835,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0006330175741811384,
                 'reference_cagr_2051': -0.0003165087870905692},
          'NB': {'historical_cagr': -0.019650259537666748,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.001965025953766675,
                 'reference_cagr_2051': -0.0009825129768833375},
          'NS': {'historical_cagr': -0.007375573023442161,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0007375573023442162,
                 'reference_cagr_2051': -0.0003687786511721081},
          'PE': {'historical_cagr': -0.015036452542908085,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0015036452542908085,
                 'reference_cagr_2051': -0.0007518226271454043},
          'NL': {'historical_cagr': -0.009152219055073463,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0009152219055073463,
                 'reference_cagr_2051': -0.00045761095275367316},
          'YT': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.0,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'NT': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.0,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'NU': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.0,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'AT': {'historical_cagr': -0.012036781910209404,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0012036781910209405,
                 'reference_cagr_2051': -0.0006018390955104703},
          'TR': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0}},
 'Light Trucks': {'BC': {'historical_cagr': 0.03241821517157906,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.008104553792894764,
                         'reference_cagr_2051': 0.004052276896447382},
                  'AB': {'historical_cagr': 0.03861438383103555,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.009653595957758887,
                         'reference_cagr_2051': 0.0048267979788794435},
                  'SK': {'historical_cagr': 0.05443455073948411,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.013608637684871028,
                         'reference_cagr_2051': 0.006804318842435514},
                  'MB': {'historical_cagr': 0.060726374674550776,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.015181593668637694,
                         'reference_cagr_2051': 0.007590796834318847},
                  'ON': {'historical_cagr': 0.037507858711346875,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.009376964677836719,
                         'reference_cagr_2051': 0.004688482338918359},
                  'QC': {'historical_cagr': 0.039932983321800286,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.009983245830450072,
                         'reference_cagr_2051': 0.004991622915225036},
                  'NB': {'historical_cagr': 0.027668099971009052,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.006917024992752263,
                         'reference_cagr_2051': 0.0034585124963761316},
                  'NS': {'historical_cagr': 0.036549069562264336,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.009137267390566084,
                         'reference_cagr_2051': 0.004568633695283042},
                  'PE': {'historical_cagr': 0.024947259346247197,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.006236814836561799,
                         'reference_cagr_2051': 0.0031184074182808996},
                  'NL': {'historical_cagr': 0.050769448628128266,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.012692362157032067,
                         'reference_cagr_2051': 0.006346181078516033},
                  'YT': {'historical_cagr': 0.0,
                         'reference_multiplier': 0.0,
                         'reference_cagr_2023': 0.0,
                         'reference_cagr_2051': 0.0},
                  'NT': {'historical_cagr': 0.0,
                         'reference_multiplier': 0.0,
                         'reference_cagr_2023': 0.0,
                         'reference_cagr_2051': 0.0},
                  'NU': {'historical_cagr': 0.0,
                         'reference_multiplier': 0.0,
                         'reference_cagr_2023': 0.0,
                         'reference_cagr_2051': 0.0},
                  'AT': {'historical_cagr': 0.03588896096086702,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.008972240240216756,
                         'reference_cagr_2051': 0.004486120120108378},
                  'TR': {'historical_cagr': 0.0,
                         'reference_multiplier': 0.9,
                         'reference_cagr_2023': 0.0,
                         'reference_cagr_2051': 0.0}},
 'Motorcycle': {'BC': {'historical_cagr': 0.02337284097249248,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.00584321024312312,
                       'reference_cagr_2051': 0.00292160512156156},
                'AB': {'historical_cagr': 0.032227707825462915,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.008056926956365729,
                       'reference_cagr_2051': 0.004028463478182864},
                'SK': {'historical_cagr': 0.029419621509148453,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.007354905377287113,
                       'reference_cagr_2051': 0.0036774526886435566},
                'MB': {'historical_cagr': 0.0828928866197629,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.020723221654940727,
                       'reference_cagr_2051': 0.010361610827470363},
                'ON': {'historical_cagr': 0.0211396169131981,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.005284904228299525,
                       'reference_cagr_2051': 0.0026424521141497626},
                'QC': {'historical_cagr': 0.006334697379180021,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.0015836743447950052,
                       'reference_cagr_2051': 0.0007918371723975026},
                'NB': {'historical_cagr': 0.013407382360271614,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.0033518455900679034,
                       'reference_cagr_2051': 0.0016759227950339517},
                'NS': {'historical_cagr': 0.034472921674826296,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.008618230418706574,
                       'reference_cagr_2051': 0.004309115209353287},
                'PE': {'historical_cagr': 0.02446753328578244,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.00611688332144561,
                       'reference_cagr_2051': 0.003058441660722805},
                'NL': {'historical_cagr': 0.06658956054909648,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.01664739013727412,
                       'reference_cagr_2051': 0.00832369506863706},
                'YT': {'historical_cagr': 0.0,
                       'reference_multiplier': 0.0,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0},
                'NT': {'historical_cagr': 0.0,
                       'reference_multiplier': 0.0,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0},
                'NU': {'historical_cagr': 0.0,
                       'reference_multiplier': 0.0,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0},
                'AT': {'historical_cagr': 0.031182915714387027,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.007795728928596757,
                       'reference_cagr_2051': 0.0038978644642983784},
                'TR': {'historical_cagr': 0.0,
                       'reference_multiplier': 0.9,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0}},
 'School Bus': {'BC': {'historical_cagr': -0.0059503883042363315,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': -0.0014875970760590829,
                       'reference_cagr_2051': -0.0007437985380295414},
                'AB': {'historical_cagr': 0.012107439445693258,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': 0.0030268598614233144,
                       'reference_cagr_2051': 0.0015134299307116572},
                'SK': {'historical_cagr': -0.006177789925932653,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': -0.0015444474814831632,
                       'reference_cagr_2051': -0.0007722237407415816},
                'MB': {'historical_cagr': 0.002230535815663126,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': 0.0005576339539157815,
                       'reference_cagr_2051': 0.00027881697695789076},
                'ON': {'historical_cagr': 0.006535055642185705,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': 0.0016337639105464263,
                       'reference_cagr_2051': 0.0008168819552732132},
                'QC': {'historical_cagr': 0.04016088744509472,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': 0.01004022186127368,
                       'reference_cagr_2051': 0.00502011093063684},
                'NB': {'historical_cagr': -0.036712358535642475,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': -0.009178089633910619,
                       'reference_cagr_2051': -0.004589044816955309},
                'NS': {'historical_cagr': 0.01732573494727907,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': 0.0043314337368197675,
                       'reference_cagr_2051': 0.0021657168684098838},
                'PE': {'historical_cagr': 0.0239162708050209,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': 0.005979067701255225,
                       'reference_cagr_2051': 0.0029895338506276126},
                'NL': {'historical_cagr': -0.01206977187013536,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': -0.00301744296753384,
                       'reference_cagr_2051': -0.00150872148376692},
                'YT': {'historical_cagr': 0.0,
                       'reference_multiplier': 0.0,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0},
                'NT': {'historical_cagr': 0.0,
                       'reference_multiplier': 0.0,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0},
                'NU': {'historical_cagr': 0.0,
                       'reference_multiplier': 0.0,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0},
                'AT': {'historical_cagr': -0.015579686461411724,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': -0.003894921615352931,
                       'reference_cagr_2051': -0.0019474608076764655},
                'TR': {'historical_cagr': 0.0,
                       'reference_multiplier': 1.0,
                       'reference_cagr_2023': 0.0,
                       'reference_cagr_2051': 0.0}},
 'Transit': {'BC': {'historical_cagr': -0.009858562939505378,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': -0.004929281469752689,
                    'reference_cagr_2051': -0.0024646407348763444},
             'AB': {'historical_cagr': -0.004238772615456354,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': -0.002119386307728177,
                    'reference_cagr_2051': -0.0010596931538640886},
             'SK': {'historical_cagr': -0.0002845588699128543,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': -0.00014227943495642714,
                    'reference_cagr_2051': -7.113971747821357e-05},
             'MB': {'historical_cagr': -0.0023336818839931883,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': -0.0011668409419965942,
                    'reference_cagr_2051': -0.0005834204709982971},
             'ON': {'historical_cagr': 0.0016838239474032957,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': 0.0008419119737016478,
                    'reference_cagr_2051': 0.0004209559868508239},
             'QC': {'historical_cagr': 0.03449949750231962,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': 0.01724974875115981,
                    'reference_cagr_2051': 0.008624874375579905},
             'NB': {'historical_cagr': -0.02527834122349959,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': -0.012639170611749795,
                    'reference_cagr_2051': -0.006319585305874897},
             'NS': {'historical_cagr': 0.007022007917376172,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': 0.003511003958688086,
                    'reference_cagr_2051': 0.001755501979344043},
             'PE': {'historical_cagr': 0.03945170920564611,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': 0.019725854602823056,
                    'reference_cagr_2051': 0.009862927301411528},
             'NL': {'historical_cagr': 0.0026363983153741,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': 0.00131819915768705,
                    'reference_cagr_2051': 0.000659099578843525},
             'YT': {'historical_cagr': 0.0,
                    'reference_multiplier': 0.0,
                    'reference_cagr_2023': 0.0,
                    'reference_cagr_2051': 0.0},
             'NT': {'historical_cagr': 0.0,
                    'reference_multiplier': 0.0,
                    'reference_cagr_2023': 0.0,
                    'reference_cagr_2051': 0.0},
             'NU': {'historical_cagr': 0.0,
                    'reference_multiplier': 0.0,
                    'reference_cagr_2023': 0.0,
                    'reference_cagr_2051': 0.0},
             'AT': {'historical_cagr': -0.00944846540786426,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': -0.00472423270393213,
                    'reference_cagr_2051': -0.002362116351966065},
             'TR': {'historical_cagr': 0.0,
                    'reference_multiplier': 0.9,
                    'reference_cagr_2023': 0.0,
                    'reference_cagr_2051': 0.0}},
 'Intercity Bus': {'BC': {'historical_cagr': -0.057999259583530915,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'AB': {'historical_cagr': -0.05140562620086597,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'SK': {'historical_cagr': -0.045550125171140765,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'MB': {'historical_cagr': -0.028618464216082717,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'ON': {'historical_cagr': -0.038804238360203636,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'QC': {'historical_cagr': -0.019101906849700256,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'NB': {'historical_cagr': -0.07515138743230254,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'NS': {'historical_cagr': -0.04836602765079989,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'PE': {'historical_cagr': -0.019328150233377528,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'NL': {'historical_cagr': -0.054312355014000224,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'YT': {'historical_cagr': 0.0,
                          'reference_multiplier': 0.0,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'NT': {'historical_cagr': 0.0,
                          'reference_multiplier': 0.0,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'NU': {'historical_cagr': 0.0,
                          'reference_multiplier': 0.0,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'AT': {'historical_cagr': -0.0629249146293871,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0},
                   'TR': {'historical_cagr': 0.0,
                          'reference_multiplier': 0.9,
                          'reference_cagr_2023': 0.0,
                          'reference_cagr_2051': 0.0}},
 'Rail': {'BC': {'historical_cagr': 0.06061652742128154,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.01212330548425631,
                 'reference_cagr_2051': 0.006061652742128155},
          'AB': {'historical_cagr': -0.025051391050708083,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.005010278210141617,
                 'reference_cagr_2051': -0.0025051391050708084},
          'SK': {'historical_cagr': 0.08940615685220332,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.017881231370440665,
                 'reference_cagr_2051': 0.008940615685220333},
          'MB': {'historical_cagr': 0.08145462734524145,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.01629092546904829,
                 'reference_cagr_2051': 0.008145462734524146},
          'ON': {'historical_cagr': 0.00021607350421937177,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 4.321470084387436e-05,
                 'reference_cagr_2051': 2.160735042193718e-05},
          'QC': {'historical_cagr': -0.006977137734312522,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0013954275468625044,
                 'reference_cagr_2051': -0.0006977137734312522},
          'NB': {'historical_cagr': -0.04408538097041159,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.008817076194082318,
                 'reference_cagr_2051': -0.004408538097041159},
          'NS': {'historical_cagr': 0.023586015005317762,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.0047172030010635526,
                 'reference_cagr_2051': 0.0023586015005317763},
          'PE': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'NL': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'YT': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.0,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'NT': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.0,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'NU': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.0,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0},
          'AT': {'historical_cagr': -0.01781883250371019,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': -0.0035637665007420384,
                 'reference_cagr_2051': -0.0017818832503710192},
          'TR': {'historical_cagr': 0.0,
                 'reference_multiplier': 0.9,
                 'reference_cagr_2023': 0.0,
                 'reference_cagr_2051': 0.0}},
 'Aviation': {'BC': {'historical_cagr': 0.05645873870559193,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.011291747741118387,
                     'reference_cagr_2051': 0.005645873870559193},
              'AB': {'historical_cagr': 0.049468830309040035,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.009893766061808008,
                     'reference_cagr_2051': 0.004946883030904004},
              'SK': {'historical_cagr': 0.041537197066107145,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.008307439413221429,
                     'reference_cagr_2051': 0.0041537197066107145},
              'MB': {'historical_cagr': 0.0545790987747139,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.01091581975494278,
                     'reference_cagr_2051': 0.00545790987747139},
              'ON': {'historical_cagr': 0.04129232014555573,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.008258464029111146,
                     'reference_cagr_2051': 0.004129232014555573},
              'QC': {'historical_cagr': 0.09168329072189785,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.01833665814437957,
                     'reference_cagr_2051': 0.009168329072189785},
              'NB': {'historical_cagr': 0.004947284127670981,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.0009894568255341962,
                     'reference_cagr_2051': 0.0004947284127670981},
              'NS': {'historical_cagr': -0.0015595035895105136,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': -0.0003119007179021027,
                     'reference_cagr_2051': -0.00015595035895105136},
              'PE': {'historical_cagr': 0.0950570722186137,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.019011414443722743,
                     'reference_cagr_2051': 0.009505707221861372},
              'NL': {'historical_cagr': 0.045666764752113176,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.009133352950422636,
                     'reference_cagr_2051': 0.004566676475211318},
              'YT': {'historical_cagr': 0.0,
                     'reference_multiplier': 0.0,
                     'reference_cagr_2023': 0.0,
                     'reference_cagr_2051': 0.0},
              'NT': {'historical_cagr': 0.0,
                     'reference_multiplier': 0.0,
                     'reference_cagr_2023': 0.0,
                     'reference_cagr_2051': 0.0},
              'NU': {'historical_cagr': 0.0,
                     'reference_multiplier': 0.0,
                     'reference_cagr_2023': 0.0,
                     'reference_cagr_2051': 0.0},
              'AT': {'historical_cagr': 0.05645873870559193,
                     'reference_multiplier': 0.9,
                     'reference_cagr_2023': 0.011291747741118387,
                     'reference_cagr_2051': 0.005645873870559193},
              'TR': {'historical_cagr': 0.0,
                     'reference_multiplier': 0.0,
                     'reference_cagr_2023': 0.0,
                     'reference_cagr_2051': 0.0}}}
# Transit electricity shares (Rapid transit share). Bus share = 1 - rapid.
# From your assumptions: BC rapid transit = 0.67; others default 1.0.
TRANSIT_RAPID_SHARE = {prov: 1.0 for prov in ASSUMP_PROV_ORDER}
TRANSIT_RAPID_SHARE["BC"] = 0.67

# Walking/cycling ratio from your assumptions: 0.24 / 26.94
WALK_CYCLE_RATIO = 0.24 / 26.94

# Aviation pkm split (Canada): Domestic 0.27; International = 1 - domestic
AVIATION_PKM_DOMESTIC_CAN = 0.27
AVIATION_PKM_INTL_CAN = 1.0 - AVIATION_PKM_DOMESTIC_CAN

# Aviation energy shares: domestic share by region; international = 1 - domestic
AVIATION_ENERGY_DOMESTIC = {
    "CAN": 0.38,
    "BC": 0.311100225898893,
    "AB": 0.59214050405041,
    "SK": 0.8,
    "MB": 0.773031271375588,
    "ON": 0.38250923835559,
    "QC": 0.2,
    "NB": 0.8,
    "NS": 0.672882519637571,
    "PE": 0.67,
    "NL": 0.346194784183231,
    "YT": 0.9,
    "NT": 0.9,
    "NU": 0.9,
    "AT": 0.6,
    "TR": 0.9,
}

# Aviation multipliers
AVIATION_MULTIPLIER_LOAD_FACTOR_KG_PER_PAX = 100.0
AVIATION_MULTIPLIER_TKM_PER_PKM = AVIATION_MULTIPLIER_LOAD_FACTOR_KG_PER_PAX / 1000.0

# =========================
# ASSUMPTIONS LOOKUP (IN-MEMORY)
# =========================
# We build assumptions_df in-memory (see build_assumptions_tables_structured) and still write
# output/assumptions_long.csv for auditing. Downstream formulas should use this store,
# not re-read assumptions_long.csv from disk.

def _norm_fuel(x) -> str:
    if x is None:
        return ""
    try:
        if isinstance(x, float) and x != x:
            return ""
    except Exception:
        pass
    s = str(x).strip()
    return "" if s.lower() in {"nan", "none"} else s

def _norm_year(y) -> str:
    if y is None:
        return ""
    try:
        if isinstance(y, float) and y != y:
            return ""
    except Exception:
        pass
    s = str(y).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return ""
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return s

class AssumptionStore:
    def __init__(self, assumptions_df: pd.DataFrame):
        required_cols = {"mode", "fuel", "prov_code", "year", "metric", "value", "unit", "notes"}
        missing = sorted(required_cols - set(assumptions_df.columns))
        if missing:
            raise ValueError(f"assumptions_df missing required columns: {missing}")

        df = assumptions_df.copy()
        df["mode"] = df["mode"].astype(str).str.strip()
        df["metric"] = df["metric"].astype(str).str.strip()
        df["prov_code"] = df["prov_code"].astype(str).str.strip()
        df["fuel"] = df["fuel"].apply(_norm_fuel)
        df["year"] = df["year"].apply(_norm_year)

        key_cols = ["mode", "fuel", "prov_code", "year", "metric"]
        if df.duplicated(subset=key_cols).any():
            dups = df[df.duplicated(subset=key_cols, keep=False)].sort_values(key_cols)
            raise ValueError(
                "Duplicate assumption keys for (mode,fuel,prov_code,year,metric). "
                f"Example duplicates:\\n{dups[key_cols + ['value']].head(20).to_string(index=False)}"
            )

        self._index = {
            (r.mode, r.fuel, r.prov_code, r.year, r.metric): r
            for r in df.itertuples(index=False)
        }

    def get(self, *, mode: str, metric: str, prov_code: str = "", year=None, fuel: str = "", default=None, required: bool = True):
        mode = str(mode).strip()
        metric = str(metric).strip()
        prov_code = str(prov_code).strip()
        fuel = _norm_fuel(fuel)
        y = _norm_year(year)

        k_exact = (mode, fuel, prov_code, y, metric)
        if k_exact in self._index:
            return self._index[k_exact].value

        k_fallback = (mode, fuel, prov_code, "", metric)
        if k_fallback in self._index:
            return self._index[k_fallback].value

        if not required:
            return default

        raise KeyError(
            f"Missing assumption: mode={mode!r}, metric={metric!r}, prov_code={prov_code!r}, year={y!r}, fuel={fuel!r} "
            "(also not found with blank-year fallback)."
        )
def build_assumptions_tables_structured(
    out_subdir: str = "assumptions_structured",
    write_combined_long: bool = True,
    write_root_assumptions_long: bool = True,
    return_df: bool = True,
):
    """Build structured assumptions tables (long/tidy) WITHOUT reading the reference Excel workbook.

    Outputs:
      - output/<out_subdir>/assumptions_long.csv  (optional)
      - output/<out_subdir>/kpkm_growth_long.csv
      - output/<out_subdir>/transit_electricity_shares_long.csv
      - output/<out_subdir>/walking_cycling_long.csv
      - output/<out_subdir>/aviation_pkm_split_long.csv
      - output/<out_subdir>/aviation_energy_long.csv
      - output/<out_subdir>/aviation_multipliers_long.csv
      - output/<out_subdir>/manifest.csv

    Long schema: mode, fuel, prov_code, prov_name, year, metric, value, unit, notes
    """
    out_dir = OUT_DIR / out_subdir
    out_dir.mkdir(exist_ok=True)

    rows = []

    def add_row(mode, fuel, prov_code, year, metric, value, unit=None, notes=None):
        rows.append({
            'mode': mode,
            'fuel': fuel,
            'prov_code': prov_code,
            'prov_name': ASSUMP_PROV_NAMES.get(prov_code) if prov_code in ASSUMP_PROV_NAMES else ("Canada" if prov_code=="CAN" else None),
            'year': year,
            'metric': metric,
            'value': value,
            'unit': unit,
            'notes': notes,
        })

    # Map province codes to output prefixes used in this script
    # NOTE: BC outputs are 'bc_*'; AB outputs are 'alb_*'.
    prov_to_prefix = {
        'BC': 'bc',
        'AB': 'alb',
        'SK': 'sk',
        'MB': 'mb',
        'ON': 'on',
        'QC': 'qc',
        'NB': 'nb',
        'NS': 'ns',
        'PE': 'pe',
        'NL': 'nl',
        'AT': 'at',
        # Territories placeholders (no dedicated outputs)
        'YT': None,
        'NT': None,
        'NU': None,
        'TR': None,
    }

    # ------------------------------------------------------------
    # 1) k*pkm growth assumptions (by mode x province)
    # ------------------------------------------------------------
    # IMPORTANT:
    # Use the explicit workbook assumptions supplied by the user as the authoritative source.
    # Do NOT recompute historical/reference CAGR from output CSV filenames here.
    for mode, cfg in KPKM_MODE_CONFIG.items():
        hist_start = int(cfg['hist_start'])
        hist_end = int(cfg['hist_end'])
        rel_2023_default = float(cfg['rel_to_hist_2023'])
        rel_2051_default = float(cfg['rel_to_hist_2051'])
        ref_mult_default = float(cfg['ref_multiplier'])

        for prov in ASSUMP_PROV_ORDER:
            ass = KPKM_EXPLICIT_ASSUMPTIONS.get(mode, {}).get(prov, {})
            hist_cagr = float(ass.get('historical_cagr', 0.0))
            ref_mult = float(ass.get('reference_multiplier', ref_mult_default))
            ref_cagr_2023 = float(ass.get('reference_cagr_2023', 0.0))
            ref_cagr_2051 = float(ass.get('reference_cagr_2051', 0.0))

            # For audit rows only, back-calculate the relative factors where possible.
            rel_2023 = rel_2023_default if hist_cagr == 0 else (ref_cagr_2023 / hist_cagr)
            rel_2051 = rel_2051_default if hist_cagr == 0 else (ref_cagr_2051 / hist_cagr)
            prov_adj = 1.0

            add_row(mode, None, prov, hist_end, f"historical_cagr_{hist_start}_{hist_end}", hist_cagr, unit='fraction', notes='Hard-wired from explicit workbook assumptions supplied by user')
            add_row(mode, None, prov, 2022, 'reference_multiplier', ref_mult, unit='fraction', notes='Hard-wired from explicit workbook assumptions supplied by user')
            add_row(mode, None, prov, 2023, 'relative_to_historical', rel_2023, unit='fraction', notes='Back-calculated from explicit workbook assumptions')
            add_row(mode, None, prov, 2051, 'relative_to_historical', rel_2051, unit='fraction', notes='Back-calculated from explicit workbook assumptions')
            add_row(mode, None, prov, 2022, 'prov_adjustment', prov_adj, unit='multiplier', notes='Set to 1.0 because explicit workbook reference CAGR values are already final')
            add_row(mode, None, prov, 2023, 'reference_cagr', ref_cagr_2023, unit='fraction', notes='Hard-wired from explicit workbook assumptions supplied by user')
            add_row(mode, None, prov, 2051, 'reference_cagr', ref_cagr_2051, unit='fraction', notes='Hard-wired from explicit workbook assumptions supplied by user')

    kpkm_long = pd.DataFrame(rows)
    audit_write_df(kpkm_long, out_dir / 'kpkm_growth_long.csv', index=False)

    # ------------------------------------------------------------
    # 2) Transit electricity shares
    # ------------------------------------------------------------
    te_rows = []
    for prov in ASSUMP_PROV_ORDER:
        rapid = float(TRANSIT_RAPID_SHARE.get(prov, 1.0))
        ferry = 0.33 if prov == 'BC' else 0.0
        bus = max(0.0, 1.0 - rapid - ferry)
        te_rows.append({'mode':'Transit','fuel':'Electricity','prov_code':prov,'prov_name':ASSUMP_PROV_NAMES.get(prov), 'year':None, 'metric':'rapid_transit_share', 'value':rapid, 'unit':'fraction', 'notes':'Workbook J252:J266'})
        te_rows.append({'mode':'Transit','fuel':'Electricity','prov_code':prov,'prov_name':ASSUMP_PROV_NAMES.get(prov), 'year':None, 'metric':'ferry_urban_share', 'value':ferry, 'unit':'fraction', 'notes':'Workbook I252:I266'})
        te_rows.append({'mode':'Transit','fuel':'Electricity','prov_code':prov,'prov_name':ASSUMP_PROV_NAMES.get(prov), 'year':None, 'metric':'bus_share', 'value':bus, 'unit':'fraction', 'notes':'bus_share = 1 - rapid_transit_share - ferry_urban_share'})
    te_df = pd.DataFrame(te_rows)
    # Divisor constants (workbook J221/J224/J225/J226) provided by user
    divisor_df = pd.DataFrame([
        {'mode':'Workbook','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'divisor_J221','value':0.42,'unit':'ratio','notes':'Provided'},
        {'mode':'Workbook','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'divisor_J224','value':1.09,'unit':'ratio','notes':'Provided'},
        {'mode':'Workbook','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'divisor_J225','value':1.92,'unit':'ratio','notes':'Provided'},
        {'mode':'Workbook','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'divisor_J226','value':2.37,'unit':'ratio','notes':'Provided'},
    ])
    audit_write_df(divisor_df, out_dir / 'workbook_divisors_long.csv', index=False)
    audit_write_df(te_df, out_dir / 'transit_electricity_shares_long.csv', index=False)

    # ------------------------------------------------------------
    # 3) Walking & Cycling
    # ------------------------------------------------------------
    wc_df = pd.DataFrame([
        {'mode':'Walking and Cycling','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'walk_bike_daily_miles','value':0.24,'unit':'miles', 'notes':None},
        {'mode':'Walking and Cycling','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'total_daily_miles','value':26.94,'unit':'miles', 'notes':None},
        {'mode':'Walking and Cycling','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'walk_bike_share','value':WALK_CYCLE_RATIO,'unit':'fraction','notes':'0.24/26.94'},
    ])
    audit_write_df(wc_df, out_dir / 'walking_cycling_long.csv', index=False)

    # ------------------------------------------------------------
    # 4) Aviation pkm split (Canada)
    # ------------------------------------------------------------
    ap_df = pd.DataFrame([
        {'mode':'Aviation','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'pkm_share_domestic','value':AVIATION_PKM_DOMESTIC_CAN,'unit':'fraction','notes':None},
        {'mode':'Aviation','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'pkm_share_international','value':AVIATION_PKM_INTL_CAN,'unit':'fraction','notes':'1 - domestic'},
    ])
    audit_write_df(ap_df, out_dir / 'aviation_pkm_split_long.csv', index=False)

    # ------------------------------------------------------------
    # 5) Aviation energy shares (domestic/international)
    # ------------------------------------------------------------
    ae_rows=[]
    for prov_code, dom in AVIATION_ENERGY_DOMESTIC.items():
        intl = 1.0 - float(dom)
        pname = 'Canada' if prov_code=='CAN' else ASSUMP_PROV_NAMES.get(prov_code)
        ae_rows.append({'mode':'Aviation','fuel':'Jet fuel','prov_code':prov_code,'prov_name':pname,'year':None,'metric':'energy_share_domestic','value':float(dom),'unit':'fraction','notes':None})
        ae_rows.append({'mode':'Aviation','fuel':'Jet fuel','prov_code':prov_code,'prov_name':pname,'year':None,'metric':'energy_share_international','value':intl,'unit':'fraction','notes':'1 - domestic'})
    audit_write_df(pd.DataFrame(ae_rows), out_dir / 'aviation_energy_long.csv', index=False)

    # ------------------------------------------------------------
    # 6) Aviation multipliers
    # ------------------------------------------------------------
    am_df = pd.DataFrame([
        {'mode':'Aviation','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'multiplier_load_factor_kg_per_passenger','value':AVIATION_MULTIPLIER_LOAD_FACTOR_KG_PER_PAX,'unit':'kg/passenger','notes':None},
        {'mode':'Aviation','fuel':None,'prov_code':'CAN','prov_name':'Canada','year':None,'metric':'multiplier_load_factor_tkm_per_pkm','value':AVIATION_MULTIPLIER_TKM_PER_PKM,'unit':'tkm/pkm','notes':'kg/pax / 1000'},
    ])
    audit_write_df(am_df, out_dir / 'aviation_multipliers_long.csv', index=False)

    # ------------------------------------------------------------
    # 7) calc_avg km projection parameters (Average km/year)
    # ------------------------------------------------------------
    # Mirrors the workbook 'calc_avg km' assumptions (Years to calc CAGR + bounded annual growth).
    # User-provided values (assumptions I171:K175):
    #   - vkm/year (LDV, Car, Light truck): years=5, min=-1%, max=0%
    #   - pkm/year (Public Bus, Intercity Bus): years=5, min=-5%, max=0%
    _avgkm_specs = [
        ('LDV', 5, -0.01, 0.0, 'vkm/year'),
        ('Car', 5, -0.01, 0.0, 'vkm/year'),
        ('Light truck', 5, -0.01, 0.0, 'vkm/year'),
        ('Public Bus', 5, -0.05, 0.0, 'pkm/year'),
        ('Intercity Bus', 5, -0.05, 0.0, 'pkm/year'),
    ]
    avgkm_rows = []
    for _mode, _years, _min_g, _max_g, _u in _avgkm_specs:
        avgkm_rows.append({'mode': _mode, 'fuel': None, 'prov_code': 'CAN', 'prov_name': 'Canada', 'year': None,
                           'metric': 'calc_avg km years', 'value': float(_years), 'unit': 'years',
                           'notes': f'Average km/year ({_u}): Years to calc CAGR'})
        avgkm_rows.append({'mode': _mode, 'fuel': None, 'prov_code': 'CAN', 'prov_name': 'Canada', 'year': None,
                           'metric': 'calc_avg km min growth', 'value': float(_min_g), 'unit': 'fraction',
                           'notes': f'Average km/year ({_u}): CAGR min annual (decrease bound)'})
        avgkm_rows.append({'mode': _mode, 'fuel': None, 'prov_code': 'CAN', 'prov_name': 'Canada', 'year': None,
                           'metric': 'calc_avg km max growth', 'value': float(_max_g), 'unit': 'fraction',
                           'notes': f'Average km/year ({_u}): CAGR max annual (increase bound)'})
    avgkm_df = pd.DataFrame(avgkm_rows)
    audit_write_df(avgkm_df, out_dir / 'calc_avg_km_params_long.csv', index=False)
    # Combined long (always built in-memory; optionally written for audit)
    combined = pd.concat(
        [
            kpkm_long,
            te_df,
            wc_df,
            ap_df,
            pd.DataFrame(ae_rows),
            am_df,
            avgkm_df,
        ],
        ignore_index=True,
    )

    if write_combined_long:
        audit_write_df(combined, out_dir / 'assumptions_long.csv', index=False)
        if write_root_assumptions_long:
            audit_write_df(combined, OUT_DIR / 'assumptions_long.csv', index=False)

    # manifest
    manifest = pd.DataFrame([
        {'file':'kpkm_growth_long.csv','description':'k*pkm growth assumptions by mode x province (historical CAGR + reference CAGR).'},
        {'file':'transit_electricity_shares_long.csv','description':'Transit electricity shares (rapid vs bus).'},
        {'file':'walking_cycling_long.csv','description':'Walking/cycling ratio constants.'},
        {'file':'aviation_pkm_split_long.csv','description':'Aviation pkm split (domestic vs international) for Canada.'},
        {'file':'aviation_energy_long.csv','description':'Aviation energy shares (domestic vs international) by region.'},
        {'file':'aviation_multipliers_long.csv','description':'Aviation multiplier constants.'},
        {'file':'assumptions_long.csv','description':'Combined tidy assumptions table (all above).'} if write_combined_long else None,
    ]).dropna()
    audit_write_df(manifest, out_dir / 'manifest.csv', index=False)

    if _audit_enabled(): print(f"[OK] Wrote structured assumptions CSVs to {out_dir}")

    if return_df:
        return combined
    return None

# =========================
# MAIN
# =========================

# =========================
# PASSENGER (FULL) WRAPPERS
# =========================

def build_bc_passenger_full(out_file="bc_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="BC", out_file=out_file, components={
        'LDV': 'bc_ldv.csv',
        'Motorcycle': 'bc_motorcycle.csv',
        'School Bus': 'bc_school_bus.csv',
        'Urban Transit': 'bc_urban_transit.csv',
        'Intercity Bus': 'bc_intercity_bus.csv',
        'Rail (Passengers)': 'bc_rail.csv',
        'Air (Passengers)': 'bc_air.csv',
    })

def build_alb_passenger_full(out_file="alb_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="AB", out_file=out_file, components={
        'LDV': 'alb_ldv.csv',
        'Motorcycle': 'alb_motorcycle.csv',
        'School Bus': 'alb_school_bus.csv',
        'Urban Transit': 'alb_urban_transit.csv',
        'Intercity Bus': 'alb_intercity_bus.csv',
        'Rail (Passengers)': 'alb_rail.csv',
        'Air (Passengers)': 'alb_air.csv',
    })

def build_at_passenger_full(out_file="at_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="AT", out_file=out_file, components={
        'LDV': 'at_ldv.csv',
        'Motorcycle': 'at_motorcycle.csv',
        'School Bus': 'at_school_bus.csv',
        'Urban Transit': 'at_urban_transit.csv',
        'Intercity Bus': 'at_intercity_bus.csv',
        'Rail (Passengers)': 'at_rail.csv',
        'Air (Passengers)': 'at_air.csv',
    })

def build_mb_passenger_full(out_file="mb_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="MB", out_file=out_file, components={
        'LDV': 'mb_ldv.csv',
        'Motorcycle': 'mb_motorcycle.csv',
        'School Bus': 'mb_school_bus.csv',
        'Urban Transit': 'mb_urban_transit.csv',
        'Intercity Bus': 'mb_intercity_bus.csv',
        'Rail (Passengers)': 'mb_rail.csv',
        'Air (Passengers)': 'mb_air.csv',
    })

def build_nb_passenger_full(out_file="nb_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="NB", out_file=out_file, components={
        'LDV': 'nb_ldv.csv',
        'Motorcycle': 'nb_motorcycle.csv',
        'School Bus': 'nb_school_bus.csv',
        'Urban Transit': 'nb_urban_transit.csv',
        'Intercity Bus': 'nb_intercity_bus.csv',
        'Rail (Passengers)': 'nb_rail.csv',
        'Air (Passengers)': 'nb_air.csv',
    })

def build_nl_passenger_full(out_file="nl_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="NL", out_file=out_file, components={
        'LDV': 'nl_ldv.csv',
        'Motorcycle': 'nl_motorcycle.csv',
        'School Bus': 'nl_school_bus.csv',
        'Urban Transit': 'nl_urban_transit.csv',
        'Intercity Bus': 'nl_intercity_bus.csv',
        'Rail (Passengers)': 'nl_rail.csv',
        'Air (Passengers)': 'nl_air.csv',
    })

def build_ns_passenger_full(out_file="ns_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="NS", out_file=out_file, components={
        'LDV': 'ns_ldv.csv',
        'Motorcycle': 'ns_motorcycle.csv',
        'School Bus': 'ns_school_bus.csv',
        'Urban Transit': 'ns_urban_transit.csv',
        'Intercity Bus': 'ns_intercity_bus.csv',
        'Rail (Passengers)': 'ns_rail.csv',
        'Air (Passengers)': 'ns_air.csv',
    })

def build_on_passenger_full(out_file="on_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="ON", out_file=out_file, components={
        'LDV': 'on_ldv.csv',
        'Motorcycle': 'on_motorcycle.csv',
        'School Bus': 'on_school_bus.csv',
        'Urban Transit': 'on_urban_transit.csv',
        'Intercity Bus': 'on_intercity_bus.csv',
        'Rail (Passengers)': 'on_rail.csv',
        'Air (Passengers)': 'on_air.csv',
    })

def build_pe_passenger_full(out_file="pe_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="PE", out_file=out_file, components={
        'LDV': 'pe_ldv.csv',
        'Motorcycle': 'pe_motorcycle.csv',
        'School Bus': 'pe_school_bus.csv',
        'Urban Transit': 'pe_urban_transit.csv',
        'Intercity Bus': 'pe_intercity_bus.csv',
        'Rail (Passengers)': 'pe_rail.csv',
        'Air (Passengers)': 'pe_air.csv',
    })

def build_qc_passenger_full(out_file="qc_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="QC", out_file=out_file, components={
        'LDV': 'qc_ldv.csv',
        'Motorcycle': 'qc_motorcycle.csv',
        'School Bus': 'qc_school_bus.csv',
        'Urban Transit': 'qc_urban_transit.csv',
        'Intercity Bus': 'qc_intercity_bus.csv',
        'Rail (Passengers)': 'qc_rail.csv',
        'Air (Passengers)': 'qc_air.csv',
    })

def build_sk_passenger_full(out_file="sk_passenger_full.csv"):
    # Passenger Full Activity = LDV + Motorcycle + School Bus + Urban Transit + Intercity Bus + Rail (Passengers) + Air (Passengers)
    _build_prov_passenger_full(prefix="SK", out_file=out_file, components={
        'LDV': 'sk_ldv.csv',
        'Motorcycle': 'sk_motorcycle.csv',
        'School Bus': 'sk_school_bus.csv',
        'Urban Transit': 'sk_urban_transit.csv',
        'Intercity Bus': 'sk_intercity_bus.csv',
        'Rail (Passengers)': 'sk_rail.csv',
        'Air (Passengers)': 'sk_air.csv',
    })

LIGHT_TRUCK_EXACT_POST_OVERRIDE = {'CAN': {2022: 14599.0,
         2023: 14453.0,
         2024: 14308.0,
         2025: 14308.0,
         2026: 14308.0,
         2027: 14251.0,
         2028: 14211.0,
         2029: 14192.0,
         2030: 14168.0,
         2031: 14140.0,
         2032: 14118.0,
         2033: 14100.0,
         2034: 14082.0,
         2035: 14065.0,
         2036: 14049.0,
         2037: 14036.0,
         2038: 14023.0,
         2039: 14011.0,
         2040: 14000.0,
         2041: 13991.0,
         2042: 13982.0,
         2043: 13973.0,
         2044: 13966.0,
         2045: 13959.0,
         2046: 13953.0,
         2047: 13947.0,
         2048: 13941.0,
         2049: 13937.0,
         2050: 13932.0,
         2051: 13928.0,
         2052: 13924.0,
         2053: 13921.0,
         2054: 13918.0,
         2055: 13915.0,
         2056: 13912.0,
         2057: 13910.0,
         2058: 13908.0,
         2059: 13906.0,
         2060: 13904.0,
         2061: 13902.0,
         2062: 13901.0,
         2063: 13899.0,
         2064: 13898.0,
         2065: 13897.0,
         2066: 13895.0,
         2067: 13894.0,
         2068: 13894.0,
         2069: 13893.0,
         2070: 13892.0,
         2071: 13891.0,
         2072: 13891.0,
         2073: 13890.0,
         2074: 13889.0,
         2075: 13889.0,
         2076: 13888.0,
         2077: 13888.0,
         2078: 13888.0,
         2079: 13887.0,
         2080: 13887.0,
         2081: 13887.0,
         2082: 13886.0,
         2083: 13886.0,
         2084: 13886.0,
         2085: 13886.0,
         2086: 13886.0,
         2087: 13885.0,
         2088: 13885.0,
         2089: 13885.0,
         2090: 13885.0,
         2091: 13885.0,
         2092: 13885.0,
         2093: 13885.0,
         2094: 13885.0,
         2095: 13884.0,
         2096: 13884.0,
         2097: 13884.0,
         2098: 13884.0,
         2099: 13884.0,
         2100: 13884.0},
 'BC': {2022: 12902.0,
        2023: 12773.0,
        2024: 12649.0,
        2025: 12649.0,
        2026: 12610.0,
        2027: 12552.0,
        2028: 12508.0,
        2029: 12480.0,
        2030: 12447.0,
        2031: 12414.0,
        2032: 12387.0,
        2033: 12363.0,
        2034: 12339.0,
        2035: 12318.0,
        2036: 12299.0,
        2037: 12281.0,
        2038: 12265.0,
        2039: 12250.0,
        2040: 12237.0,
        2041: 12225.0,
        2042: 12213.0,
        2043: 12203.0,
        2044: 12193.0,
        2045: 12185.0,
        2046: 12177.0,
        2047: 12169.0,
        2048: 12163.0,
        2049: 12157.0,
        2050: 12151.0,
        2051: 12146.0,
        2052: 12141.0,
        2053: 12137.0,
        2054: 12133.0,
        2055: 12129.0,
        2056: 12126.0,
        2057: 12123.0,
        2058: 12120.0,
        2059: 12118.0,
        2060: 12115.0,
        2061: 12113.0,
        2062: 12111.0,
        2063: 12110.0,
        2064: 12108.0,
        2065: 12106.0,
        2066: 12105.0,
        2067: 12104.0,
        2068: 12103.0,
        2069: 12102.0,
        2070: 12101.0,
        2071: 12100.0,
        2072: 12099.0,
        2073: 12098.0,
        2074: 12097.0,
        2075: 12097.0,
        2076: 12096.0,
        2077: 12096.0,
        2078: 12095.0,
        2079: 12095.0,
        2080: 12094.0,
        2081: 12094.0,
        2082: 12094.0,
        2083: 12093.0,
        2084: 12093.0,
        2085: 12093.0,
        2086: 12093.0,
        2087: 12092.0,
        2088: 12092.0,
        2089: 12092.0,
        2090: 12092.0,
        2091: 12092.0,
        2092: 12092.0,
        2093: 12091.0,
        2094: 12091.0,
        2095: 12091.0,
        2096: 12091.0,
        2097: 12091.0,
        2098: 12091.0,
        2099: 12091.0,
        2100: 12091.0},
 'AB': {2022: 10693.0,
        2023: 10586.0,
        2024: 10481.0,
        2025: 10435.0,
        2026: 10381.0,
        2027: 10320.0,
        2028: 10267.0,
        2029: 10225.0,
        2030: 10183.0,
        2031: 10144.0,
        2032: 10109.0,
        2033: 10078.0,
        2034: 10049.0,
        2035: 10022.0,
        2036: 9998.0,
        2037: 9976.0,
        2038: 9956.0,
        2039: 9937.0,
        2040: 9921.0,
        2041: 9905.0,
        2042: 9891.0,
        2043: 9878.0,
        2044: 9866.0,
        2045: 9855.0,
        2046: 9845.0,
        2047: 9836.0,
        2048: 9828.0,
        2049: 9820.0,
        2050: 9813.0,
        2051: 9807.0,
        2052: 9801.0,
        2053: 9796.0,
        2054: 9791.0,
        2055: 9786.0,
        2056: 9782.0,
        2057: 9778.0,
        2058: 9775.0,
        2059: 9772.0,
        2060: 9769.0,
        2061: 9766.0,
        2062: 9764.0,
        2063: 9762.0,
        2064: 9759.0,
        2065: 9758.0,
        2066: 9756.0,
        2067: 9754.0,
        2068: 9753.0,
        2069: 9752.0,
        2070: 9750.0,
        2071: 9749.0,
        2072: 9748.0,
        2073: 9747.0,
        2074: 9746.0,
        2075: 9746.0,
        2076: 9745.0,
        2077: 9744.0,
        2078: 9744.0,
        2079: 9743.0,
        2080: 9743.0,
        2081: 9742.0,
        2082: 9742.0,
        2083: 9741.0,
        2084: 9741.0,
        2085: 9741.0,
        2086: 9740.0,
        2087: 9740.0,
        2088: 9740.0,
        2089: 9740.0,
        2090: 9739.0,
        2091: 9739.0,
        2092: 9739.0,
        2093: 9739.0,
        2094: 9739.0,
        2095: 9739.0,
        2096: 9738.0,
        2097: 9738.0,
        2098: 9738.0,
        2099: 9738.0,
        2100: 9738.0},
 'SK': {2022: 15786.0,
        2023: 15628.0,
        2024: 15472.0,
        2025: 15378.0,
        2026: 15224.0,
        2027: 15114.0,
        2028: 15013.0,
        2029: 14923.0,
        2030: 14834.0,
        2031: 14757.0,
        2032: 14687.0,
        2033: 14622.0,
        2034: 14563.0,
        2035: 14509.0,
        2036: 14460.0,
        2037: 14415.0,
        2038: 14374.0,
        2039: 14337.0,
        2040: 14302.0,
        2041: 14271.0,
        2042: 14242.0,
        2043: 14216.0,
        2044: 14192.0,
        2045: 14170.0,
        2046: 14150.0,
        2047: 14132.0,
        2048: 14115.0,
        2049: 14100.0,
        2050: 14086.0,
        2051: 14073.0,
        2052: 14061.0,
        2053: 14050.0,
        2054: 14040.0,
        2055: 14031.0,
        2056: 14023.0,
        2057: 14015.0,
        2058: 14009.0,
        2059: 14002.0,
        2060: 13996.0,
        2061: 13991.0,
        2062: 13986.0,
        2063: 13982.0,
        2064: 13977.0,
        2065: 13974.0,
        2066: 13970.0,
        2067: 13967.0,
        2068: 13964.0,
        2069: 13962.0,
        2070: 13959.0,
        2071: 13957.0,
        2072: 13955.0,
        2073: 13953.0,
        2074: 13951.0,
        2075: 13950.0,
        2076: 13948.0,
        2077: 13947.0,
        2078: 13946.0,
        2079: 13945.0,
        2080: 13944.0,
        2081: 13943.0,
        2082: 13942.0,
        2083: 13941.0,
        2084: 13940.0,
        2085: 13940.0,
        2086: 13939.0,
        2087: 13939.0,
        2088: 13938.0,
        2089: 13938.0,
        2090: 13937.0,
        2091: 13937.0,
        2092: 13936.0,
        2093: 13936.0,
        2094: 13936.0,
        2095: 13936.0,
        2096: 13935.0,
        2097: 13935.0,
        2098: 13935.0,
        2099: 13935.0,
        2100: 13934.0},
 'MB': {2022: 16777.0,
        2023: 16610.0,
        2024: 16443.0,
        2025: 16443.0,
        2026: 16406.0,
        2027: 16332.0,
        2028: 16277.0,
        2029: 16244.0,
        2030: 16205.0,
        2031: 16165.0,
        2032: 16132.0,
        2033: 16103.0,
        2034: 16075.0,
        2035: 16049.0,
        2036: 16026.0,
        2037: 16005.0,
        2038: 15985.0,
        2039: 15967.0,
        2040: 15951.0,
        2041: 15936.0,
        2042: 15922.0,
        2043: 15910.0,
        2044: 15898.0,
        2045: 15888.0,
        2046: 15878.0,
        2047: 15869.0,
        2048: 15861.0,
        2049: 15854.0,
        2050: 15847.0,
        2051: 15841.0,
        2052: 15835.0,
        2053: 15830.0,
        2054: 15825.0,
        2055: 15821.0,
        2056: 15817.0,
        2057: 15813.0,
        2058: 15810.0,
        2059: 15806.0,
        2060: 15804.0,
        2061: 15801.0,
        2062: 15799.0,
        2063: 15797.0,
        2064: 15795.0,
        2065: 15793.0,
        2066: 15791.0,
        2067: 15790.0,
        2068: 15788.0,
        2069: 15787.0,
        2070: 15786.0,
        2071: 15785.0,
        2072: 15784.0,
        2073: 15783.0,
        2074: 15782.0,
        2075: 15781.0,
        2076: 15780.0,
        2077: 15780.0,
        2078: 15779.0,
        2079: 15779.0,
        2080: 15778.0,
        2081: 15778.0,
        2082: 15777.0,
        2083: 15777.0,
        2084: 15777.0,
        2085: 15776.0,
        2086: 15776.0,
        2087: 15776.0,
        2088: 15775.0,
        2089: 15775.0,
        2090: 15775.0,
        2091: 15775.0,
        2092: 15775.0,
        2093: 15775.0,
        2094: 15774.0,
        2095: 15774.0,
        2096: 15774.0,
        2097: 15774.0,
        2098: 15774.0,
        2099: 15774.0,
        2100: 15774.0},
 'ON': {2022: 15537.0,
        2023: 15381.0,
        2024: 15228.0,
        2025: 15228.0,
        2026: 15228.0,
        2027: 15166.0,
        2028: 15124.0,
        2029: 15103.0,
        2030: 15078.0,
        2031: 15049.0,
        2032: 15025.0,
        2033: 15006.0,
        2034: 14986.0,
        2035: 14968.0,
        2036: 14952.0,
        2037: 14937.0,
        2038: 14924.0,
        2039: 14911.0,
        2040: 14900.0,
        2041: 14889.0,
        2042: 14880.0,
        2043: 14871.0,
        2044: 14863.0,
        2045: 14856.0,
        2046: 14849.0,
        2047: 14843.0,
        2048: 14837.0,
        2049: 14832.0,
        2050: 14827.0,
        2051: 14823.0,
        2052: 14819.0,
        2053: 14815.0,
        2054: 14812.0,
        2055: 14809.0,
        2056: 14806.0,
        2057: 14803.0,
        2058: 14801.0,
        2059: 14799.0,
        2060: 14797.0,
        2061: 14795.0,
        2062: 14793.0,
        2063: 14792.0,
        2064: 14791.0,
        2065: 14789.0,
        2066: 14788.0,
        2067: 14787.0,
        2068: 14786.0,
        2069: 14785.0,
        2070: 14784.0,
        2071: 14784.0,
        2072: 14783.0,
        2073: 14782.0,
        2074: 14782.0,
        2075: 14781.0,
        2076: 14781.0,
        2077: 14780.0,
        2078: 14780.0,
        2079: 14779.0,
        2080: 14779.0,
        2081: 14779.0,
        2082: 14778.0,
        2083: 14778.0,
        2084: 14778.0,
        2085: 14778.0,
        2086: 14778.0,
        2087: 14777.0,
        2088: 14777.0,
        2089: 14777.0,
        2090: 14777.0,
        2091: 14777.0,
        2092: 14777.0,
        2093: 14777.0,
        2094: 14776.0,
        2095: 14776.0,
        2096: 14776.0,
        2097: 14776.0,
        2098: 14776.0,
        2099: 14776.0,
        2100: 14776.0},
 'QC': {2022: 15150.0,
        2023: 15150.0,
        2024: 15150.0,
        2025: 15150.0,
        2026: 15150.0,
        2027: 15150.0,
        2028: 15150.0,
        2029: 15150.0,
        2030: 15150.0,
        2031: 15150.0,
        2032: 15150.0,
        2033: 15150.0,
        2034: 15150.0,
        2035: 15150.0,
        2036: 15150.0,
        2037: 15150.0,
        2038: 15150.0,
        2039: 15150.0,
        2040: 15150.0,
        2041: 15150.0,
        2042: 15150.0,
        2043: 15150.0,
        2044: 15150.0,
        2045: 15150.0,
        2046: 15150.0,
        2047: 15150.0,
        2048: 15150.0,
        2049: 15150.0,
        2050: 15150.0,
        2051: 15150.0,
        2052: 15150.0,
        2053: 15150.0,
        2054: 15150.0,
        2055: 15150.0,
        2056: 15150.0,
        2057: 15150.0,
        2058: 15150.0,
        2059: 15150.0,
        2060: 15150.0,
        2061: 15150.0,
        2062: 15150.0,
        2063: 15150.0,
        2064: 15150.0,
        2065: 15150.0,
        2066: 15150.0,
        2067: 15150.0,
        2068: 15150.0,
        2069: 15150.0,
        2070: 15150.0,
        2071: 15150.0,
        2072: 15150.0,
        2073: 15150.0,
        2074: 15150.0,
        2075: 15150.0,
        2076: 15150.0,
        2077: 15150.0,
        2078: 15150.0,
        2079: 15150.0,
        2080: 15150.0,
        2081: 15150.0,
        2082: 15150.0,
        2083: 15150.0,
        2084: 15150.0,
        2085: 15150.0,
        2086: 15150.0,
        2087: 15150.0,
        2088: 15150.0,
        2089: 15150.0,
        2090: 15150.0,
        2091: 15150.0,
        2092: 15150.0,
        2093: 15150.0,
        2094: 15150.0,
        2095: 15150.0,
        2096: 15150.0,
        2097: 15150.0,
        2098: 15150.0,
        2099: 15150.0,
        2100: 15150.0},
 'NB': {2022: 15006.0,
        2023: 14856.0,
        2024: 14707.0,
        2025: 14666.0,
        2026: 14666.0,
        2027: 14599.0,
        2028: 14548.0,
        2029: 14516.0,
        2030: 14486.0,
        2031: 14451.0,
        2032: 14421.0,
        2033: 14396.0,
        2034: 14372.0,
        2035: 14349.0,
        2036: 14329.0,
        2037: 14311.0,
        2038: 14294.0,
        2039: 14278.0,
        2040: 14264.0,
        2041: 14251.0,
        2042: 14239.0,
        2043: 14228.0,
        2044: 14218.0,
        2045: 14209.0,
        2046: 14201.0,
        2047: 14193.0,
        2048: 14186.0,
        2049: 14180.0,
        2050: 14174.0,
        2051: 14169.0,
        2052: 14164.0,
        2053: 14159.0,
        2054: 14155.0,
        2055: 14151.0,
        2056: 14148.0,
        2057: 14144.0,
        2058: 14142.0,
        2059: 14139.0,
        2060: 14136.0,
        2061: 14134.0,
        2062: 14132.0,
        2063: 14130.0,
        2064: 14129.0,
        2065: 14127.0,
        2066: 14125.0,
        2067: 14124.0,
        2068: 14123.0,
        2069: 14122.0,
        2070: 14121.0,
        2071: 14120.0,
        2072: 14119.0,
        2073: 14118.0,
        2074: 14118.0,
        2075: 14117.0,
        2076: 14116.0,
        2077: 14116.0,
        2078: 14115.0,
        2079: 14115.0,
        2080: 14114.0,
        2081: 14114.0,
        2082: 14114.0,
        2083: 14113.0,
        2084: 14113.0,
        2085: 14113.0,
        2086: 14112.0,
        2087: 14112.0,
        2088: 14112.0,
        2089: 14112.0,
        2090: 14112.0,
        2091: 14111.0,
        2092: 14111.0,
        2093: 14111.0,
        2094: 14111.0,
        2095: 14111.0,
        2096: 14111.0,
        2097: 14111.0,
        2098: 14111.0,
        2099: 14111.0,
        2100: 14110.0},
 'NS': {2022: 18065.0,
        2023: 17885.0,
        2024: 17706.0,
        2025: 17555.0,
        2026: 17466.0,
        2027: 17348.0,
        2028: 17243.0,
        2029: 17152.0,
        2030: 17072.0,
        2031: 16995.0,
        2032: 16925.0,
        2033: 16862.0,
        2034: 16804.0,
        2035: 16751.0,
        2036: 16703.0,
        2037: 16659.0,
        2038: 16619.0,
        2039: 16582.0,
        2040: 16548.0,
        2041: 16518.0,
        2042: 16489.0,
        2043: 16464.0,
        2044: 16440.0,
        2045: 16419.0,
        2046: 16399.0,
        2047: 16381.0,
        2048: 16364.0,
        2049: 16349.0,
        2050: 16335.0,
        2051: 16322.0,
        2052: 16311.0,
        2053: 16300.0,
        2054: 16290.0,
        2055: 16282.0,
        2056: 16273.0,
        2057: 16266.0,
        2058: 16259.0,
        2059: 16253.0,
        2060: 16247.0,
        2061: 16242.0,
        2062: 16237.0,
        2063: 16232.0,
        2064: 16228.0,
        2065: 16225.0,
        2066: 16221.0,
        2067: 16218.0,
        2068: 16215.0,
        2069: 16213.0,
        2070: 16210.0,
        2071: 16208.0,
        2072: 16206.0,
        2073: 16204.0,
        2074: 16203.0,
        2075: 16201.0,
        2076: 16200.0,
        2077: 16198.0,
        2078: 16197.0,
        2079: 16196.0,
        2080: 16195.0,
        2081: 16194.0,
        2082: 16193.0,
        2083: 16192.0,
        2084: 16192.0,
        2085: 16191.0,
        2086: 16191.0,
        2087: 16190.0,
        2088: 16190.0,
        2089: 16189.0,
        2090: 16189.0,
        2091: 16188.0,
        2092: 16188.0,
        2093: 16188.0,
        2094: 16187.0,
        2095: 16187.0,
        2096: 16187.0,
        2097: 16187.0,
        2098: 16186.0,
        2099: 16186.0,
        2100: 16186.0},
 'PE': {2022: 14734.0,
        2023: 14615.0,
        2024: 14469.0,
        2025: 14469.0,
        2026: 14388.0,
        2027: 14319.0,
        2028: 14261.0,
        2029: 14219.0,
        2030: 14170.0,
        2031: 14127.0,
        2032: 14088.0,
        2033: 14054.0,
        2034: 14021.0,
        2035: 13992.0,
        2036: 13965.0,
        2037: 13941.0,
        2038: 13918.0,
        2039: 13898.0,
        2040: 13879.0,
        2041: 13861.0,
        2042: 13846.0,
        2043: 13831.0,
        2044: 13818.0,
        2045: 13806.0,
        2046: 13795.0,
        2047: 13785.0,
        2048: 13775.0,
        2049: 13767.0,
        2050: 13759.0,
        2051: 13752.0,
        2052: 13745.0,
        2053: 13739.0,
        2054: 13734.0,
        2055: 13729.0,
        2056: 13724.0,
        2057: 13720.0,
        2058: 13716.0,
        2059: 13713.0,
        2060: 13709.0,
        2061: 13707.0,
        2062: 13704.0,
        2063: 13701.0,
        2064: 13699.0,
        2065: 13697.0,
        2066: 13695.0,
        2067: 13693.0,
        2068: 13692.0,
        2069: 13690.0,
        2070: 13689.0,
        2071: 13688.0,
        2072: 13686.0,
        2073: 13685.0,
        2074: 13684.0,
        2075: 13684.0,
        2076: 13683.0,
        2077: 13682.0,
        2078: 13681.0,
        2079: 13681.0,
        2080: 13680.0,
        2081: 13680.0,
        2082: 13679.0,
        2083: 13679.0,
        2084: 13678.0,
        2085: 13678.0,
        2086: 13678.0,
        2087: 13677.0,
        2088: 13677.0,
        2089: 13677.0,
        2090: 13677.0,
        2091: 13676.0,
        2092: 13676.0,
        2093: 13676.0,
        2094: 13676.0,
        2095: 13676.0,
        2096: 13676.0,
        2097: 13675.0,
        2098: 13675.0,
        2099: 13675.0,
        2100: 13675.0},
 'NL': {2022: 17614.0,
        2023: 17516.0,
        2024: 17516.0,
        2025: 17516.0,
        2026: 17516.0,
        2027: 17496.0,
        2028: 17492.0,
        2029: 17487.0,
        2030: 17482.0,
        2031: 17475.0,
        2032: 17471.0,
        2033: 17467.0,
        2034: 17462.0,
        2035: 17458.0,
        2036: 17455.0,
        2037: 17452.0,
        2038: 17449.0,
        2039: 17447.0,
        2040: 17444.0,
        2041: 17442.0,
        2042: 17440.0,
        2043: 17438.0,
        2044: 17436.0,
        2045: 17435.0,
        2046: 17433.0,
        2047: 17432.0,
        2048: 17431.0,
        2049: 17430.0,
        2050: 17429.0,
        2051: 17428.0,
        2052: 17427.0,
        2053: 17426.0,
        2054: 17426.0,
        2055: 17425.0,
        2056: 17424.0,
        2057: 17424.0,
        2058: 17423.0,
        2059: 17423.0,
        2060: 17422.0,
        2061: 17422.0,
        2062: 17422.0,
        2063: 17421.0,
        2064: 17421.0,
        2065: 17421.0,
        2066: 17421.0,
        2067: 17420.0,
        2068: 17420.0,
        2069: 17420.0,
        2070: 17420.0,
        2071: 17420.0,
        2072: 17419.0,
        2073: 17419.0,
        2074: 17419.0,
        2075: 17419.0,
        2076: 17419.0,
        2077: 17419.0,
        2078: 17419.0,
        2079: 17419.0,
        2080: 17419.0,
        2081: 17419.0,
        2082: 17419.0,
        2083: 17418.0,
        2084: 17418.0,
        2085: 17418.0,
        2086: 17418.0,
        2087: 17418.0,
        2088: 17418.0,
        2089: 17418.0,
        2090: 17418.0,
        2091: 17418.0,
        2092: 17418.0,
        2093: 17418.0,
        2094: 17418.0,
        2095: 17418.0,
        2096: 17418.0,
        2097: 17418.0,
        2098: 17418.0,
        2099: 17418.0,
        2100: 17418.0},
 'AT': {2022: 16748.0,
        2023: 16580.0,
        2024: 16415.0,
        2025: 16415.0,
        2026: 16415.0,
        2027: 16349.0,
        2028: 16303.0,
        2029: 16281.0,
        2030: 16254.0,
        2031: 16222.0,
        2032: 16197.0,
        2033: 16175.0,
        2034: 16155.0,
        2035: 16135.0,
        2036: 16117.0,
        2037: 16102.0,
        2038: 16087.0,
        2039: 16073.0,
        2040: 16061.0,
        2041: 16050.0,
        2042: 16040.0,
        2043: 16030.0,
        2044: 16021.0,
        2045: 16014.0,
        2046: 16006.0,
        2047: 16000.0,
        2048: 15994.0,
        2049: 15988.0,
        2050: 15983.0,
        2051: 15978.0,
        2052: 15974.0,
        2053: 15970.0,
        2054: 15966.0,
        2055: 15963.0,
        2056: 15960.0,
        2057: 15957.0,
        2058: 15955.0,
        2059: 15952.0,
        2060: 15950.0,
        2061: 15948.0,
        2062: 15947.0,
        2063: 15945.0,
        2064: 15943.0,
        2065: 15942.0,
        2066: 15941.0,
        2067: 15940.0,
        2068: 15939.0,
        2069: 15938.0,
        2070: 15937.0,
        2071: 15936.0,
        2072: 15935.0,
        2073: 15935.0,
        2074: 15934.0,
        2075: 15933.0,
        2076: 15933.0,
        2077: 15932.0,
        2078: 15932.0,
        2079: 15932.0,
        2080: 15931.0,
        2081: 15931.0,
        2082: 15930.0,
        2083: 15930.0,
        2084: 15930.0,
        2085: 15930.0,
        2086: 15929.0,
        2087: 15929.0,
        2088: 15929.0,
        2089: 15929.0,
        2090: 15929.0,
        2091: 15929.0,
        2092: 15928.0,
        2093: 15928.0,
        2094: 15928.0,
        2095: 15928.0,
        2096: 15928.0,
        2097: 15928.0,
        2098: 15928.0,
        2099: 15928.0,
        2100: 15928.0}}

def build_calc_avg_km(out_file: str = "calc_avg_km.csv", assumptions_df=None):
    """Build the workbook-style 'calc_avg km' sheet (wide by year, 2000–2100).

    Fixes implemented
    -----------------
    1) CAN rows now resolve correctly (direct national keys first, with a province-sum fallback).
    2) Public Bus rows now resolve correctly by mapping to the Urban Transit outputs, and by
       ensuring copied provincial transit files are registered in-memory (see _copy_mode_csv patch).
    3) Workbook-style rounding is applied year-by-year during projection (and historical values
       are rounded to whole units when writing), preventing cumulative drift.

    IMPORTANT
    ---------
    - Uses ONLY in-memory DataFrames already registered by prior build_* functions.
    - Does NOT use transportation personal_source data.xlsx or any audit CSV as runtime input.
    """
    import numpy as np
    import pandas as pd

    HIST_END = 2022
    YEARS_ALL = list(range(2000, 2101))
    YEARS_HIST = list(range(2000, HIST_END + 1))

    REGIONS = [
        ('CAN', 'Canada'),
        ('BC', 'British Columbia'),
        ('AB', 'Alberta'),
        ('SK', 'Saskatchewan'),
        ('MB', 'Manitoba'),
        ('ON', 'Ontario'),
        ('QC', 'Quebec'),
        ('NB', 'New Brunswick'),
        ('NS', 'Nova Scotia'),
        ('PE', 'Prince Edward Island'),
        ('NL', 'Newfoundland and Labrador'),
        ('YT', 'Yukon'),
        ('NT', 'Northwest Territories'),
        ('NU', 'Nunavut'),
        ('AT', 'Atlantic'),
        ('TR', 'Territories'),
    ]

    # Mode labels in this sheet vs their source outputs
    MODE_SPECS = [
        ('LDV', 'vkm', 'Average Distance (vkm)'),
        ('Car', 'vkm', 'Average Distance (vkm)'),
        ('Light truck', 'vkm', 'Average Distance (vkm)'),
        # Public Bus lives in the model as Urban Transit
        ('Public Bus', 'pkm', 'Average Distance (pkm)'),
        ('Intercity Bus', 'pkm', 'Average Distance (pkm)'),
    ]

    # Map province codes to output prefixes used in this script
    PREFIX = {
        'BC': 'bc',
        'AB': 'alb',
        'SK': 'sk',
        'MB': 'mb',
        'ON': 'on',
        'QC': 'qc',
        'NB': 'nb',
        'NS': 'ns',
        'PE': 'pe',
        'NL': 'nl',
        'AT': 'at',
    }

    # Direct national keys (CAN) produced by build_mode/build_ldv
    CAN_DIRECT_KEY = {
        'LDV': 'ldv_full.csv',
        'Car': 'car_full.csv',
        'Light truck': 'light_truck_full.csv',
        'Public Bus': 'urban_transit_full.csv',
        'Intercity Bus': 'intercity_bus_full.csv',
    }

    # Direct provincial keys
    PROV_DIRECT_KEY = {
        'LDV': '{p}_ldv.csv',
        'Car': '{p}_car.csv',
        'Light truck': '{p}_light_truck.csv',
        'Public Bus': '{p}_urban_transit.csv',
        'Intercity Bus': '{p}_intercity_bus.csv',
    }

    # Defaults consistent with your assumptions block
    DEFAULT_YEARS = {'LDV': 5.0, 'Car': 5.0, 'Light truck': 5.0, 'Public Bus': 5.0, 'Intercity Bus': 5.0}
    DEFAULT_MIN_G = {'LDV': -0.01, 'Car': -0.01, 'Light truck': -0.01, 'Public Bus': -0.05, 'Intercity Bus': -0.05}
    DEFAULT_MAX_G = {'LDV': 0.0, 'Car': 0.0, 'Light truck': 0.0, 'Public Bus': 0.0, 'Intercity Bus': 0.0}

    notes = []

    if assumptions_df is None:
        assumptions_df = _get_df("__assumptions_df__", required=False)

    def _ensure_year_index(df):
        if df is None:
            return None
        if isinstance(df, pd.Series):
            try:
                df.index = df.index.astype(int)
            except Exception:
                pass
            return df
        if isinstance(df, pd.DataFrame) and 'year' in df.columns:
            df = df.set_index('year')
        try:
            df.index = df.index.astype(int)
        except Exception:
            try:
                df.index = df.index.map(lambda x: int(float(str(x).strip())))
            except Exception:
                pass
        return df

    def _lookup_assumption(mode_label: str, metric: str, default: float):
        if assumptions_df is None or not isinstance(assumptions_df, pd.DataFrame) or assumptions_df.empty:
            return float(default)
        cols = {str(c).lower(): c for c in assumptions_df.columns}
        mode_col = cols.get('mode') or cols.get('parameter') or cols.get('mode_name')
        metric_col = cols.get('metric') or cols.get('assumption') or cols.get('name')
        value_col = cols.get('value') or cols.get('val')
        prov_col = cols.get('prov_code') or cols.get('province') or cols.get('region')
        if mode_col is None or metric_col is None or value_col is None:
            return float(default)

        aliases = [mode_label]
        if mode_label == 'Public Bus':
            aliases += ['Urban Transit', 'Transit', 'Urban_Transit']
        if mode_label == 'Car':
            aliases += ['Cars']
        if mode_label == 'Light truck':
            aliases += ['Light Truck', 'Light Trucks', 'Light_Truck']
        if mode_label == 'Intercity Bus':
            aliases += ['Intercity_Bus']

        subset = assumptions_df[
            assumptions_df[mode_col].astype(str).str.strip().isin(aliases)
            & assumptions_df[metric_col].astype(str).str.strip().str.lower().eq(metric.lower())
        ]
        if prov_col is not None and not subset.empty:
            pref = subset[subset[prov_col].astype(str).str.strip().isin(['CAN', '', 'NONE', 'NAN'])]
            if not pref.empty:
                subset = pref
        if subset.empty:
            return float(default)
        try:
            v = float(pd.to_numeric(subset.iloc[0][value_col], errors='coerce'))
            return float(default) if pd.isna(v) else v
        except Exception:
            return float(default)

    def _get_params(mode_label: str):
        yrs = _lookup_assumption(mode_label, 'calc_avg km years', DEFAULT_YEARS[mode_label])
        mn = _lookup_assumption(mode_label, 'calc_avg km min growth', DEFAULT_MIN_G[mode_label])
        mx = _lookup_assumption(mode_label, 'calc_avg km max growth', DEFAULT_MAX_G[mode_label])
        try:
            yrs = float(yrs)
        except Exception:
            yrs = float(DEFAULT_YEARS[mode_label])
        try:
            mn = float(mn)
        except Exception:
            mn = float(DEFAULT_MIN_G[mode_label])
        try:
            mx = float(mx)
        except Exception:
            mx = float(DEFAULT_MAX_G[mode_label])
        if yrs <= 0:
            yrs = float(DEFAULT_YEARS[mode_label])
        if mn > mx:
            mn, mx = mx, mn
        return yrs, mn, mx

    def _derive_avg_distance(df: pd.DataFrame, unit_label: str, desired_col: str):
        """Ensure we have an average distance series.

        If desired_col exists, use it.
        Else compute:
          - vkm: Total Distance (M*vkm) * 1000 / Stock (thousands)
          - pkm: Activity (millions passenger-kilometres) * 1000 / Stock (thousands)
        """
        if df is None:
            return None
        if desired_col in df.columns:
            return pd.to_numeric(df[desired_col], errors='coerce')

        if unit_label == 'vkm':
            td_col = 'Total Distance (M*vkm)'
            st_col = 'Stock (thousands)'
            if td_col in df.columns and st_col in df.columns:
                td = pd.to_numeric(df[td_col], errors='coerce')
                st = pd.to_numeric(df[st_col], errors='coerce')
                return (td * 1000.0) / st.replace(0, np.nan)
        if unit_label == 'pkm':
            act_col = 'Activity (millions passenger-kilometres)'
            st_col = 'Stock (thousands)'
            if act_col in df.columns and st_col in df.columns:
                act = pd.to_numeric(df[act_col], errors='coerce')
                st = pd.to_numeric(df[st_col], errors='coerce')
                return (act * 1000.0) / st.replace(0, np.nan)
        return None

    def _get_source_df(region_code: str, mode_label: str):
        # 1) Direct CAN keys
        if region_code == 'CAN':
            k = CAN_DIRECT_KEY.get(mode_label)
            if k:
                df = _get_df(k, required=False)
                if df is not None:
                    return _ensure_year_index(df), k

        # 2) Direct province keys
        p = PREFIX.get(region_code)
        if p is not None:
            k = PROV_DIRECT_KEY.get(mode_label, '').format(p=p)
            if k:
                df = _get_df(k, required=False)
                if df is not None:
                    return _ensure_year_index(df), k

        # 3) Fallback: try substring scan of _DF_STORE
        # (Helps if keys are slightly different but still contain region + mode tokens.)
        toks = [region_code.lower(), mode_label.lower().replace(' ', '_')]
        for kk in list(_DF_STORE.keys()):
            kl = str(kk).lower()
            if all(t in kl for t in toks):
                df = _get_df(kk, required=False)
                if df is not None:
                    return _ensure_year_index(df), kk
        return None, None

    def _project_series(hist_series: pd.Series, mode_label: str):
        # ── Sprint 3: NumPy-array-backed projection ──────────────────────────
        # The original implementation used pd.Series.loc[] for every scalar
        # read/write inside the 78-iteration future-year loop.  Each .loc[]
        # call on a Pandas Series carries ~5-10 µs of label-lookup overhead,
        # accumulating to ~800 µs per (mode, region) pair × 80 rows = ~64 ms
        # of pure overhead that buys nothing numerically.
        #
        # Replacement strategy
        # --------------------
        # Because each projected year strictly depends on the immediately
        # preceding *rounded* value (serial dependency), the loop cannot be
        # fully vectorised with cumsum/cumprod.  Instead we:
        #   1. Copy values into a pre-allocated float64 NumPy array indexed
        #      by offset from 2000 (offset 0 = year 2000, offset 100 = 2100).
        #   2. Run the identical rounding / CAGR math over plain NumPy scalars
        #      — array index reads are ~50× faster than .loc[] lookups.
        #   3. Convert back to pd.Series only once at the very end, preserving
        #      the exact same output contract as before.
        #
        # Numerical output is bit-for-bit identical to the Pandas version.
        # ─────────────────────────────────────────────────────────────────────
        YEAR_OFFSET = 2000          # vals[i]  ↔  year (2000 + i)
        N_YEARS     = 101           # 2000 … 2100 inclusive

        # Initialise with NaN
        vals = np.full(N_YEARS, np.nan, dtype=np.float64)

        hist_series = pd.to_numeric(hist_series, errors='coerce').reindex(YEARS_HIST)

        # Copy rounded historical values into the array
        for yr in YEARS_HIST:
            v = hist_series.loc[yr]
            if not pd.isna(v):
                vals[yr - YEAR_OFFSET] = float(np.round(v, 0))

        if hist_series.notna().sum() == 0:
            return pd.Series(vals, index=YEARS_ALL, dtype=float)

        years_to_cagr, min_g, max_g = _get_params(mode_label)
        window = max(2, int(round(years_to_cagr)))

        # ── Sprint 3 core: serial loop over NumPy scalars ────────────────────
        for yi in range(HIST_END + 1 - YEAR_OFFSET, N_YEARS):   # yi = 23 … 100
            prev = vals[yi - 1]
            if np.isnan(prev):
                vals[yi] = np.nan
                continue
            lag_offset = max(0, yi - window)          # offset from year 2000
            lag = vals[lag_offset]
            if np.isnan(lag) or lag == 0.0:
                growth = 0.0
            else:
                denom_years = max(1, yi - lag_offset)
                growth = (prev / lag) ** (1.0 / denom_years) - 1.0
            growth = float(np.clip(growth, min_g, max_g))
            val = prev * (1.0 + growth)
            # Workbook-style rounding before it feeds the next year's CAGR
            vals[yi] = float(np.round(val, 0))

        return pd.Series(vals, index=YEARS_ALL, dtype=float)

    # Diagnostics: expected in-memory keys
    expected = [
        'car_full.csv', 'light_truck_full.csv', 'ldv_full.csv', 'urban_transit_full.csv', 'intercity_bus_full.csv',
        'bc_urban_transit.csv', 'alb_urban_transit.csv', 'on_urban_transit.csv', 'qc_urban_transit.csv'
    ]
    missing = [k for k in expected if _get_df(k, required=False) is None]
    if missing:
        notes.append('[WARN] Missing expected in-memory dataframes for calc_avg_km: ' + ', '.join(missing))

    rows = []
    for mode_label, unit_label, desired_col in MODE_SPECS:
        for region_code, region_name in REGIONS:
            df, key = _get_source_df(region_code, mode_label)

            # CAN fallback: derive from sum of provinces if national df missing
            if df is None and region_code == 'CAN':
                provs = [c for c, _ in REGIONS if c in PREFIX]  # only provinces we have prefixes for
                prov_dfs = []
                for pc in provs:
                    dfi, _ = _get_source_df(pc, mode_label)
                    if dfi is not None:
                        prov_dfs.append(dfi)
                if prov_dfs:
                    # Build average pkm/vkm from sums of (Activity or TotalDist) and Stock
                    if unit_label == 'vkm':
                        td_col = 'Total Distance (M*vkm)'
                        st_col = 'Stock (thousands)'
                        if all(td_col in d.columns and st_col in d.columns for d in prov_dfs):
                            td = sum(pd.to_numeric(d[td_col], errors='coerce') for d in prov_dfs)
                            st = sum(pd.to_numeric(d[st_col], errors='coerce') for d in prov_dfs)
                            series = (td * 1000.0) / st.replace(0, np.nan)
                            df = pd.DataFrame({desired_col: series})
                            key = '[DERIVED] sum provinces'
                    else:
                        act_col = 'Activity (millions passenger-kilometres)'
                        st_col = 'Stock (thousands)'
                        if all(act_col in d.columns and st_col in d.columns for d in prov_dfs):
                            act = sum(pd.to_numeric(d[act_col], errors='coerce') for d in prov_dfs)
                            st = sum(pd.to_numeric(d[st_col], errors='coerce') for d in prov_dfs)
                            series = (act * 1000.0) / st.replace(0, np.nan)
                            df = pd.DataFrame({desired_col: series})
                            key = '[DERIVED] sum provinces'

            if df is None:
                proj_series = pd.Series([np.nan] * len(YEARS_ALL), index=YEARS_ALL, dtype=float)
                notes.append(f"[INFO] {mode_label} {region_code}: no in-memory source dataframe found; output left blank.")
            else:
                avg_series = _derive_avg_distance(df, unit_label, desired_col)
                if avg_series is None:
                    proj_series = pd.Series([np.nan] * len(YEARS_ALL), index=YEARS_ALL, dtype=float)
                    notes.append(f"[WARN] {mode_label} {region_code}: source dataframe {key!r} missing required columns for average distance; output left blank.")
                else:
                    proj_series = _project_series(avg_series, mode_label)
                    notes.append(f"[INFO] {mode_label} {region_code}: sourced from {key!r}.")

            row = {
                'Index': f'{mode_label}{region_code}',
                'Source': 'CEUD' if region_code == 'CAN' else '',
                'Unit': unit_label,
                'Parameter': mode_label,
                'Region': region_code,
                'Region name': region_name,
            }
            for y in YEARS_ALL:
                row[str(y)] = proj_series.loc[y]
            rows.append(row)

    year_cols = [str(y) for y in YEARS_ALL]
    cols = ['Index', 'Source', 'Unit', 'Parameter', 'Region', 'Region name'] + year_cols
    out_df = pd.DataFrame(rows, columns=cols)

    # Sprint 5: Polars fast CSV write for calc_avg_km.csv  [PRIMARY — always written]
    _audit_write_csv_fast(out_df, OUT_DIR / out_file, is_primary=True)
    long_df = out_df.melt(
        id_vars=['Index', 'Source', 'Unit', 'Parameter', 'Region', 'Region name'],
        value_vars=year_cols,
        var_name='year',
        value_name='value',
    )
    # Sprint 5: Polars fast CSV write for calc_avg_km_long.csv
    _audit_write_csv_fast(long_df, OUT_DIR / 'calc_avg_km_long.csv')
    _register_df(out_file, out_df)

    audit_write_text('\n'.join(notes) + '\n' if notes else '[INFO] No notes.\n', OUT_DIR / 'calc_avg_km_notes.txt', encoding='utf-8', mode='w')
    if _audit_enabled_primary():
        print(f'[OK] Wrote {out_file}')
    return out_df
def main():
    # --- CAN modes ---
    build_mode(
        mode_name="Car",
        fuel_table="Table 30",
        p1_activity_row=33,
        p4_sales_row=11,
        p4_stock_row=15,
        p4_avg_vkm_row=19,
        out_file="car_full.csv",
        fuel_layout="by_source",
        ceud_path=CEUD_CAN_FILE,
    )

    build_mode(
        mode_name="Light_Truck",
        fuel_table="Table 37",
        p1_activity_row=34,
        p4_sales_row=12,
        p4_stock_row=16,
        p4_avg_vkm_row=20,
        out_file="light_truck_full.csv",
        fuel_layout="by_source",
        ceud_path=CEUD_CAN_FILE,
    )

    # LDV after both exist
    build_ldv(
        car_file="car_full.csv",
        light_truck_file="light_truck_full.csv",
        out_file="ldv_full.csv",
    )

    build_mode(
        mode_name="Motorcycle",
        fuel_table="Table 50",
        p1_activity_row=35,
        p4_sales_row=13,
        p4_stock_row=17,
        p4_avg_vkm_row=21,
        out_file="motorcycle_full.csv",
        fuel_layout="total_to_gasoline",
        total_energy_label="Motorcycle Energy Use",
        total_to_fuel="Motor gasoline",
        strict_sales=False,
        other_fuels_null=True,
        ceud_path=CEUD_CAN_FILE,
    )

    build_mode(
        mode_name="School_Bus",
        fuel_table="Table 43",
        p1_activity_row=36,
        p4_sales_row=None,
        p4_stock_row=39,
        p4_avg_vkm_row=43,
        out_file="school_bus_full.csv",
        fuel_layout="by_source",
        strict_sales=False,
        ceud_path=CEUD_CAN_FILE,
    )

    build_mode(
        mode_name="Urban_Transit",
        fuel_table="Table 45",
        p1_activity_row=37,
        p4_sales_row=None,
        p4_stock_row=40,
        p4_avg_vkm_row=44,
        out_file="urban_transit_full.csv",
        fuel_layout="by_source",
        strict_sales=False,
        ceud_path=CEUD_CAN_FILE,
    )

    build_mode(
        mode_name="Intercity_Bus",
        fuel_table="Table 47",
        p1_activity_row=38,
        p4_sales_row=None,
        p4_stock_row=41,
        p4_avg_vkm_row=45,
        out_file="intercity_bus_full.csv",
        fuel_layout="by_source",
        strict_sales=False,
        ceud_path=CEUD_CAN_FILE,
    )

    # Rail/Air built after assumptions_df is available

    # --- BCTerr modes ---
    build_bcterr_car(out_file="bcterr_car.csv")
    build_bcterr_light_truck(out_file="bcterr_light_truck.csv")
    build_bcterr_ldv(out_file="bcterr_ldv.csv")
    build_bcterr_motorcycle(out_file="bcterr_motorcycle.csv")
    build_bcterr_school_bus(out_file="bcterr_school_bus.csv")
    build_bcterr_urban_transit(out_file="bcterr_urban_transit.csv")
    build_bcterr_intercity_bus(out_file="bcterr_intercity_bus.csv")

    # --- British Columbia modes ---
    build_bc_car(out_file="bc_car.csv")
    build_bc_light_truck(out_file="bc_light_truck.csv")
    build_bc_ldv(out_file="bc_ldv.csv")
    build_bc_motorcycle(out_file="bc_motorcycle.csv")
    build_bc_school_bus(out_file="bc_school_bus.csv")
    build_bc_urban_transit(out_file="bc_urban_transit.csv")
    build_bc_intercity_bus(out_file="bc_intercity_bus.csv")
    # passenger_partial replaced by passenger_full

    # --- Alberta modes ---
    build_alb_car(out_file="alb_car.csv")
    build_alb_light_truck(out_file="alb_light_truck.csv")
    build_alb_ldv(out_file="alb_ldv.csv")
    build_alb_motorcycle(out_file="alb_motorcycle.csv")
    build_alb_school_bus(out_file="alb_school_bus.csv")
    build_alb_urban_transit(out_file="alb_urban_transit.csv")
    build_alb_intercity_bus(out_file="alb_intercity_bus.csv")
    # passenger_partial replaced by passenger_full

    # --- Remaining provinces/regions ---
    build_at_car(out_file="at_car.csv")
    build_at_light_truck(out_file="at_light_truck.csv")
    build_at_ldv(out_file="at_ldv.csv")
    build_at_motorcycle(out_file="at_motorcycle.csv")
    build_at_school_bus(out_file="at_school_bus.csv")
    build_at_urban_transit(out_file="at_urban_transit.csv")
    build_at_intercity_bus(out_file="at_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_mb_car(out_file="mb_car.csv")
    build_mb_light_truck(out_file="mb_light_truck.csv")
    build_mb_ldv(out_file="mb_ldv.csv")
    build_mb_motorcycle(out_file="mb_motorcycle.csv")
    build_mb_school_bus(out_file="mb_school_bus.csv")
    build_mb_urban_transit(out_file="mb_urban_transit.csv")
    build_mb_intercity_bus(out_file="mb_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_nb_car(out_file="nb_car.csv")
    build_nb_light_truck(out_file="nb_light_truck.csv")
    build_nb_ldv(out_file="nb_ldv.csv")
    build_nb_motorcycle(out_file="nb_motorcycle.csv")
    build_nb_school_bus(out_file="nb_school_bus.csv")
    build_nb_urban_transit(out_file="nb_urban_transit.csv")
    build_nb_intercity_bus(out_file="nb_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_nl_car(out_file="nl_car.csv")
    build_nl_light_truck(out_file="nl_light_truck.csv")
    build_nl_ldv(out_file="nl_ldv.csv")
    build_nl_motorcycle(out_file="nl_motorcycle.csv")
    build_nl_school_bus(out_file="nl_school_bus.csv")
    build_nl_urban_transit(out_file="nl_urban_transit.csv")
    build_nl_intercity_bus(out_file="nl_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_ns_car(out_file="ns_car.csv")
    build_ns_light_truck(out_file="ns_light_truck.csv")
    build_ns_ldv(out_file="ns_ldv.csv")
    build_ns_motorcycle(out_file="ns_motorcycle.csv")
    build_ns_school_bus(out_file="ns_school_bus.csv")
    build_ns_urban_transit(out_file="ns_urban_transit.csv")
    build_ns_intercity_bus(out_file="ns_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_on_car(out_file="on_car.csv")
    build_on_light_truck(out_file="on_light_truck.csv")
    build_on_ldv(out_file="on_ldv.csv")
    build_on_motorcycle(out_file="on_motorcycle.csv")
    build_on_school_bus(out_file="on_school_bus.csv")
    build_on_urban_transit(out_file="on_urban_transit.csv")
    build_on_intercity_bus(out_file="on_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_pe_car(out_file="pe_car.csv")
    build_pe_light_truck(out_file="pe_light_truck.csv")
    build_pe_ldv(out_file="pe_ldv.csv")
    build_pe_motorcycle(out_file="pe_motorcycle.csv")
    build_pe_school_bus(out_file="pe_school_bus.csv")
    build_pe_urban_transit(out_file="pe_urban_transit.csv")
    build_pe_intercity_bus(out_file="pe_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_qc_car(out_file="qc_car.csv")
    build_qc_light_truck(out_file="qc_light_truck.csv")
    build_qc_ldv(out_file="qc_ldv.csv")
    build_qc_motorcycle(out_file="qc_motorcycle.csv")
    build_qc_school_bus(out_file="qc_school_bus.csv")
    build_qc_urban_transit(out_file="qc_urban_transit.csv")
    build_qc_intercity_bus(out_file="qc_intercity_bus.csv")
    # passenger_partial replaced by passenger_full
    build_sk_car(out_file="sk_car.csv")
    build_sk_light_truck(out_file="sk_light_truck.csv")
    build_sk_ldv(out_file="sk_ldv.csv")
    build_sk_motorcycle(out_file="sk_motorcycle.csv")
    build_sk_school_bus(out_file="sk_school_bus.csv")
    build_sk_urban_transit(out_file="sk_urban_transit.csv")
    build_sk_intercity_bus(out_file="sk_intercity_bus.csv")
    # passenger_partial replaced by passenger_full

    # --- Assumptions (structured, from outputs + constants; no Excel workbook input) ---
    assumptions_df = build_assumptions_tables_structured(
        out_subdir="assumptions_structured",
        write_combined_long=True,
        write_root_assumptions_long=True,
        return_df=True,
    )
    _ASSUMPTIONS_DF = assumptions_df
    _register_df("__assumptions_df__", assumptions_df)
    _ASSUMPTIONS = AssumptionStore(assumptions_df)

    build_rail(out_file="rail_full.csv", assumptions=_ASSUMPTIONS)
    build_air(out_file="air_full.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="BCTerr", prov_code="BC", ceud_path=CEUD_BCTERR_FILE, out_file="bcterr_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="BCTerr", prov_code="BC", ceud_path=CEUD_BCTERR_FILE, out_file="bcterr_air.csv", assumptions=_ASSUMPTIONS)
    build_bc_rail(out_file="bc_rail.csv")
    build_bc_air(out_file="bc_air.csv")

    _build_prov_rail(prefix="AB", prov_code="AB", ceud_path=CEUD_ALB_FILE, out_file="alb_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="AB", prov_code="AB", ceud_path=CEUD_ALB_FILE, out_file="alb_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="AT", prov_code="AT", ceud_path=CEUD_ATL_FILE, out_file="at_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="AT", prov_code="AT", ceud_path=CEUD_ATL_FILE, out_file="at_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="MB", prov_code="MB", ceud_path=CEUD_MAN_FILE, out_file="mb_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="MB", prov_code="MB", ceud_path=CEUD_MAN_FILE, out_file="mb_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="NB", prov_code="NB", ceud_path=CEUD_NB_FILE, out_file="nb_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="NB", prov_code="NB", ceud_path=CEUD_NB_FILE, out_file="nb_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="NL", prov_code="NL", ceud_path=CEUD_NFLD_FILE, out_file="nl_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="NL", prov_code="NL", ceud_path=CEUD_NFLD_FILE, out_file="nl_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="NS", prov_code="NS", ceud_path=CEUD_NS_FILE, out_file="ns_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="NS", prov_code="NS", ceud_path=CEUD_NS_FILE, out_file="ns_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="ON", prov_code="ON", ceud_path=CEUD_ONT_FILE, out_file="on_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="ON", prov_code="ON", ceud_path=CEUD_ONT_FILE, out_file="on_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="PE", prov_code="PE", ceud_path=CEUD_PEI_FILE, out_file="pe_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="PE", prov_code="PE", ceud_path=CEUD_PEI_FILE, out_file="pe_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="QC", prov_code="QC", ceud_path=CEUD_QUE_FILE, out_file="qc_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="QC", prov_code="QC", ceud_path=CEUD_QUE_FILE, out_file="qc_air.csv", assumptions=_ASSUMPTIONS)

    _build_prov_rail(prefix="SK", prov_code="SK", ceud_path=CEUD_SASK_FILE, out_file="sk_rail.csv", assumptions=_ASSUMPTIONS)
    _build_prov_air(prefix="SK", prov_code="SK", ceud_path=CEUD_SASK_FILE, out_file="sk_air.csv", assumptions=_ASSUMPTIONS)

    build_bc_passenger_full(out_file="bc_passenger_full.csv")
    build_alb_passenger_full(out_file="alb_passenger_full.csv")
    build_at_passenger_full(out_file="at_passenger_full.csv")
    build_mb_passenger_full(out_file="mb_passenger_full.csv")
    build_nb_passenger_full(out_file="nb_passenger_full.csv")
    build_nl_passenger_full(out_file="nl_passenger_full.csv")
    build_ns_passenger_full(out_file="ns_passenger_full.csv")
    build_on_passenger_full(out_file="on_passenger_full.csv")
    build_pe_passenger_full(out_file="pe_passenger_full.csv")
    build_qc_passenger_full(out_file="qc_passenger_full.csv")
    build_sk_passenger_full(out_file="sk_passenger_full.csv")
    # --- constant tab (from constant_input.csv) ---
    build_constant(out_file="constant.csv", required=True)

    # --- calc sheet (recreated in Python; audit outputs) ---
    build_calc(out_file="calc.csv", assumptions_df=assumptions_df)
    build_calc_market_share(out_file="calc_market_share.csv", assumptions_df=assumptions_df)
    build_calc_avg_km(out_file="calc_avg_km.csv", assumptions_df=assumptions_df)

# =========================
# CONSTANT TAB (NEW)
# =========================
CONSTANT_INPUT_FILE = SCRIPT_DIR / "constant_input.csv"

def _excel_concat(*vals):
    """Excel-style CONCAT: treat None/NaN as empty string; keep other values as-is (string)."""
    out = []
    for v in vals:
        if v is None:
            out.append("")
        else:
            s = str(v)
            if s.lower() == "nan":
                out.append("")
            else:
                out.append(s)
    return "".join(out)

def build_constant(
    constant_input_path: Path = CONSTANT_INPUT_FILE,
    out_file: str = "constant.csv",
    required: bool = False,
):
    """Build the 'constant' tab dataframe from constant_input.csv.

    Inputs:
      - constant_input.csv values come from 'Values from constant.txt'

    Logic (from 'Formulas for constant.txt'):
      - INDEX_calc = CONCAT(Branch, technology, Target, Comments, service_provide, discount_rate_financial)
        (Excel: =CONCAT(A,C,F,H:J))

    Behavior:
      - Keeps existing INDEX column unchanged
      - Writes output/constant.csv and output/constant_notes.txt

    Note:
      - This does NOT read transportation personal_source data.xlsx.
    """
    notes = []

    if not constant_input_path.exists():
        msg = f"[SKIP] constant_input.csv not found: {constant_input_path}"
        if required:
            raise FileNotFoundError(msg)
        if _audit_enabled(): print(msg)
        return None

    df = pd.read_csv(constant_input_path, dtype=str, keep_default_na=False)
    # --- DROP FULLY BLANK ROWS (input contains separator blank lines) ---
    # A row is dropped only if ALL columns are empty/whitespace.
    _blank_mask = df.apply(lambda r: all(str(v).strip() == '' for v in r.values), axis=1)
    _n_blank = int(_blank_mask.sum())
    if _n_blank:
        df = df.loc[~_blank_mask].copy()
        notes.append(f"[OK] Dropped {_n_blank} fully blank row(s) from constant_input.csv.")

    # Ensure INDEX_calc exists and keep both INDEX and INDEX_calc
    if "INDEX_calc" not in df.columns:
        unnamed = [c for c in df.columns if str(c).startswith("Unnamed")]
        if unnamed:
            df = df.rename(columns={unnamed[0]: "INDEX_calc"})
            notes.append(f"[INFO] Renamed '{unnamed[0]}' to 'INDEX_calc'.")
        else:
            df["INDEX_calc"] = ""
            notes.append("[INFO] Created missing 'INDEX_calc' column.")

    required_cols = [
        "Branch",
        "technology",
        "Target",
        "Comments",
        "service_provide",
        "discount_rate_financial",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"constant_input.csv missing required columns: {missing}")

    # Compute INDEX_calc per formulas
    df["INDEX_calc"] = df.apply(
        lambda r: _excel_concat(
            r.get("Branch", ""),
            r.get("technology", ""),
            r.get("Target", ""),
            r.get("Comments", ""),
            r.get("service_provide", ""),
            r.get("discount_rate_financial", ""),
        ),
        axis=1,
    )

    notes.append(f"[OK] Read {len(df)} rows from {constant_input_path.name}.")
    notes.append("[OK] Computed INDEX_calc using CONCAT(A,C,F,H:J) mapping.")
    if "INDEX" in df.columns:
        notes.append("[OK] Preserved existing INDEX column from input.")
    else:
        notes.append("[WARN] Input has no INDEX column.")

    out_path = OUT_DIR / out_file
    audit_write_df(df, out_path, index=False)

    notes_path = OUT_DIR / "constant_notes.txt"
    with open(notes_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(notes) + "\n")

    if _audit_enabled(): print(f"[OK] Wrote {out_file}")
    return df

# =========================
# CALC SHEET (NEW)
# =========================
# Recreates the reference workbook's "calc" sheet logic directly in Python.
# IMPORTANT:
#   - transportation personal_source data.xlsx is NOT used as an input
#   - Formulas for calc.txt is NOT used as an input
# The calc dataframe is built from the mode/province output CSVs produced by this script
# (output/*.csv) plus the in-memory AssumptionStore (_ASSUMPTIONS) for projections.

# Sprint 2: module-level holder for the Polars long-format calc table.
# Set to a polars.DataFrame by build_calc(); consumed by the bus share helpers.
# None until build_calc() has run (or when polars is unavailable).
_CALC_LONG_PL = None

# =========================
# CALC SHEET (NEW)
# =========================
# Recreates the reference workbook's "calc" sheet logic directly in Python.
# IMPORTANT:
#   - transportation personal_source data.xlsx is NOT used as an input
#   - Formulas for calc.txt is NOT used as an input
# The calc dataframe is built from the mode/province output CSVs produced by this script
# (output/*.csv) plus the in-memory AssumptionStore (_ASSUMPTIONS) for projections.

# Sprint 2: module-level holder for the Polars long-format calc table.
# Set to a polars.DataFrame by build_calc(); consumed by the bus share helpers.
# None until build_calc() has run (or when polars is unavailable).
_CALC_LONG_PL = None

def build_calc(out_file: str = "calc.csv", assumptions_df=None):
    """Build the workbook-style 'calc' sheet (compact, wide-by-year).

    - Uses ONLY in-memory DataFrames registered in _DF_STORE.
    - Writes calc.csv (wide), calc_long.csv (tidy), and calc_notes.txt.
    - Structure: note row + 352 parameter rows, years 2000–2100 as columns (<400 rows total).

    Reference: Values for calc.txt and Formulas for calc.txt. citeturn11search4turn12search144
    """
    import numpy as np
    import pandas as pd

    # Historical years in CEUD files stop at 2022, but calc must preserve full 2000–2100 series when available

    # Historical values in this data pipeline are treated as ending in 2021 for calc-sheet parity;
    # 2022–2100 are forecast years generated from the assumptions logic.
    HIST_END = 2022
    YEARS_HIST = list(range(2000, HIST_END + 1))
    YEARS_ALL = list(range(2000, 2101))

    REGIONS = [
        ('CAN', 'Canada'),
        ('BC', 'British Columbia'),
        ('AB', 'Alberta'),
        ('SK', 'Saskatchewan'),
        ('MB', 'Manitoba'),
        ('ON', 'Ontario'),
        ('QC', 'Quebec'),
        ('NB', 'New Brunswick'),
        ('NS', 'Nova Scotia'),
        ('PE', 'Prince Edward Island'),
        ('NL', 'Newfoundland and Labrador'),
        ('YT', 'Yukon'),
        ('NT', 'Northwest Territories'),
        ('NU', 'Nunavut'),
        ('AT', 'Atlantic'),
        ('TR', 'Territories'),
    ]

    # Assumptions constants (from reference workbook assumptions sheet) citeturn11search238
    ASSUMP_I204 = 0.5390610190192912
    ASSUMP_I206 = 0.48138085822204557
    ASSUMP_I246 = 0.55

    # Walk/cycle ratio is defined in-script as 0.24/26.94 citeturn12search102
    WALK_SHARE = float(WALK_CYCLE_RATIO)

    def _ensure_year_index(df):
        """Ensure a DataFrame is indexed by integer year.

        Mode outputs created in this script are indexed by year, but some intermediate
        frames may carry year as a column or as string index.
        """
        if df is None:
            return None
        if 'year' in df.columns:
            df = df.set_index('year')
        # Coerce index to int where possible
        try:
            df.index = df.index.astype(int)
        except Exception:
            try:
                df.index = df.index.map(lambda x: int(float(str(x).strip())))
            except Exception:
                pass
        return df

    def _series_hist(df, col):
        df = _ensure_year_index(df)
        if df is None or col not in df.columns:
            return pd.Series([np.nan]*len(YEARS_HIST), index=YEARS_HIST, dtype=float)
        s = pd.to_numeric(df[col], errors='coerce')
        s = s.reindex(YEARS_HIST)
        s.index = s.index.astype(int)
        return s

    def _safe_cagr_local(series: pd.Series, y0: int, y1: int) -> float:
        try:
            v0 = float(series.loc[y0]); v1 = float(series.loc[y1])
            if pd.isna(v0) or pd.isna(v1) or v0 == 0 or y1 == y0:
                return 0.0
            return (v1 / v0) ** (1.0 / (y1 - y0)) - 1.0
        except Exception:
            return 0.0

    # =========================
    # ASSUMPTIONS LOOKUP (IN-MEMORY)
    # =========================
    if assumptions_df is None:
        assumptions_df = _get_df("__assumptions_df__", required=False)
    if assumptions_df is None:
        assumptions_df = globals().get("_ASSUMPTIONS_DF", None)
    if assumptions_df is None:
        raise ValueError(
            "build_calc: assumptions dataframe not available in memory. "
            "Ensure build_assumptions_tables_structured ran and main passed assumptions_df into build_calc."
        )
    assump = AssumptionStore(assumptions_df)

    def _assump_value(mode_name: str, prov_code: str, metric: str, year=None, fuel: str = '', default=0.0):
        try:
            val = assump.get(
                mode=mode_name,
                metric=metric,
                prov_code=prov_code,
                year=year,
                fuel=fuel,
                required=False,
                default=default,
            )
            if val is None:
                return default
            return default if pd.isna(val) else float(val)
        except Exception:
            return default

    def _extend_mode_forecast(hist: pd.Series, mode_name: str, prov_code: str) -> pd.Series:
        cfg = MODE_CFG.get(mode_name)
        out = pd.Series(index=YEARS_ALL, dtype=float)
        out.loc[YEARS_HIST] = hist.reindex(YEARS_HIST).values

        if cfg is None:
            last_val = float(hist.dropna().iloc[-1]) if hist.notna().any() else np.nan
            for y in range(HIST_END + 1, YEARS_ALL[-1] + 1):
                out.loc[y] = last_val
            return out

        hist_start = int(cfg['hist_start'])
        hist_end = int(cfg['hist_end'])
        rel_2023 = float(cfg['rel_to_hist_2023'])
        rel_2051 = float(cfg['rel_to_hist_2051'])
        ref_mult_default = float(cfg['ref_multiplier'])
        prov_adj_default = float(KPKM_PROV_ADJ.get(prov_code, 1.0))

        hist_cagr_default = _safe_cagr_local(hist, hist_start, hist_end)
        hist_cagr = _assump_value(mode_name, prov_code, f"historical_cagr_{hist_start}_{hist_end}", hist_end, default=hist_cagr_default)
        ref_mult = _assump_value(mode_name, prov_code, 'reference_multiplier', 2022, default=ref_mult_default)
        prov_adj = _assump_value(mode_name, prov_code, 'prov_adjustment', 2022, default=prov_adj_default)
        ref_cagr_2023 = _assump_value(mode_name, prov_code, 'reference_cagr', 2023, default=hist_cagr * rel_2023 * prov_adj)
        ref_cagr_2051 = _assump_value(mode_name, prov_code, 'reference_cagr', 2051, default=hist_cagr * rel_2051 * prov_adj)

        if 2019 in hist.index and pd.notna(hist.loc[2019]):
            base_2019 = float(hist.loc[2019])
        elif hist.notna().any():
            base_2019 = float(hist.dropna().iloc[-1])
        else:
            base_2019 = np.nan

        out.loc[2022] = np.nan if pd.isna(base_2019) else base_2019 * ref_mult

        for y in range(2023, 2051):
            prev = out.loc[y - 1]
            out.loc[y] = np.nan if pd.isna(prev) else prev * (1.0 + ref_cagr_2023)

        for y in range(2051, YEARS_ALL[-1] + 1):
            prev = out.loc[y - 1]
            out.loc[y] = np.nan if pd.isna(prev) else prev * (1.0 + ref_cagr_2051)

        return out

    FILES = {
        'Cars': {
            'CAN': 'car_full.csv', 'BC': 'bc_car.csv', 'AB': 'alb_car.csv', 'SK': 'sk_car.csv', 'MB': 'mb_car.csv',
            'ON': 'on_car.csv', 'QC': 'qc_car.csv', 'NB': 'nb_car.csv', 'NS': 'ns_car.csv', 'PE': 'pe_car.csv',
            'NL': 'nl_car.csv', 'AT': 'at_car.csv',
        },
        'Light Trucks': {
            'CAN': 'light_truck_full.csv', 'BC': 'bc_light_truck.csv', 'AB': 'alb_light_truck.csv', 'SK': 'sk_light_truck.csv', 'MB': 'mb_light_truck.csv',
            'ON': 'on_light_truck.csv', 'QC': 'qc_light_truck.csv', 'NB': 'nb_light_truck.csv', 'NS': 'ns_light_truck.csv', 'PE': 'pe_light_truck.csv',
            'NL': 'nl_light_truck.csv', 'AT': 'at_light_truck.csv',
        },
        'Motorcycles': {
            'CAN': 'motorcycle_full.csv', 'BC': 'bc_motorcycle.csv', 'AB': 'alb_motorcycle.csv', 'SK': 'sk_motorcycle.csv', 'MB': 'mb_motorcycle.csv',
            'ON': 'on_motorcycle.csv', 'QC': 'qc_motorcycle.csv', 'NB': 'nb_motorcycle.csv', 'NS': 'ns_motorcycle.csv', 'PE': 'pe_motorcycle.csv',
            'NL': 'nl_motorcycle.csv', 'AT': 'at_motorcycle.csv',
        },
        'School Bus': {
            'CAN': 'school_bus_full.csv', 'BC': 'bc_school_bus.csv', 'AB': 'alb_school_bus.csv', 'SK': 'sk_school_bus.csv', 'MB': 'mb_school_bus.csv',
            'ON': 'on_school_bus.csv', 'QC': 'qc_school_bus.csv', 'NB': 'nb_school_bus.csv', 'NS': 'ns_school_bus.csv', 'PE': 'pe_school_bus.csv',
            'NL': 'nl_school_bus.csv', 'AT': 'at_school_bus.csv',
        },
        'Transit': {
            'CAN': 'urban_transit_full.csv', 'BC': 'bc_urban_transit.csv', 'AB': 'alb_urban_transit.csv', 'SK': 'sk_urban_transit.csv', 'MB': 'mb_urban_transit.csv',
            'ON': 'on_urban_transit.csv', 'QC': 'qc_urban_transit.csv', 'NB': 'nb_urban_transit.csv', 'NS': 'ns_urban_transit.csv', 'PE': 'pe_urban_transit.csv',
            'NL': 'nl_urban_transit.csv', 'AT': 'at_urban_transit.csv',
        },
        'Bus Intercity': {
            'CAN': 'intercity_bus_full.csv', 'BC': 'bc_intercity_bus.csv', 'AB': 'alb_intercity_bus.csv', 'SK': 'sk_intercity_bus.csv', 'MB': 'mb_intercity_bus.csv',
            'ON': 'on_intercity_bus.csv', 'QC': 'qc_intercity_bus.csv', 'NB': 'nb_intercity_bus.csv', 'NS': 'ns_intercity_bus.csv', 'PE': 'pe_intercity_bus.csv',
            'NL': 'nl_intercity_bus.csv', 'AT': 'at_intercity_bus.csv',
        },
        'Rail Intercity': {
        'CAN': 'rail_full.csv',
        'BC': 'bc_rail.csv',
        'AB': 'alb_rail.csv',
        'SK': 'sk_rail.csv',
        'MB': 'mb_rail.csv',
        'ON': 'on_rail.csv',
        'QC': 'qc_rail.csv',
        'NB': 'nb_rail.csv',
        'NS': 'ns_rail.csv',
        'PE': 'pe_rail.csv',
        'NL': 'nl_rail.csv',
        'AT': 'at_rail.csv',
    },
        'Aviation': {
        'CAN': 'air_full.csv',
        'BC': 'bc_air.csv',
        'AB': 'alb_air.csv',
        'SK': 'sk_air.csv',
        'MB': 'mb_air.csv',
        'ON': 'on_air.csv',
        'QC': 'qc_air.csv',
        'NB': 'nb_air.csv',
        'NS': 'ns_air.csv',
        'PE': 'pe_air.csv',
        'NL': 'nl_air.csv',
        'AT': 'at_air.csv',
    },
    }

    MODE_CFG = KPKM_MODE_CONFIG

    def _get_mode_df(param: str, reg: str):
        key = FILES.get(param, {}).get(reg)
        if not key:
            return None
        return _get_df(key, required=False)

    S = {}

    # -----------------------------------------------------------------
    # AT Cars / Light Trucks direct pull from AT mode outputs
    # (workbook parity: AT is not a sum of provinces for these rows)
    # -----------------------------------------------------------------
    def _mode_series(_csv, _col, _scale=1.0):
        _df = _get_df(_csv, required=False)
        if _df is None:
            return None
        _df = _df.copy()
        if 'year' in _df.columns:
            _df = _df.set_index('year')
        _df.index = _df.index.astype(int)
        _s = _df[_col] if _col in _df.columns else None
        if _s is None:
            return None
        _s = _s.reindex(YEARS_ALL).astype(float)
        return _s * float(_scale)

    # Cars (AT)
    _s = _mode_series('at_car.csv', 'Activity (millions passenger-kilometres)', 1000.0)
    if _s is not None:
        S[('Cars','k*pkm','AT')] = _s
    _s = _mode_series('at_car.csv', 'Total Distance (M*vkm)', 1000.0)
    if _s is not None:
        S[('Cars','k*vkm','AT')] = _s

    # Light Trucks (AT)
    _s = _mode_series('at_light_truck.csv', 'Activity (millions passenger-kilometres)', 1000.0)
    if _s is not None:
        S[('Light Trucks','k*pkm','AT')] = _s
    _s = _mode_series('at_light_truck.csv', 'Total Distance (M*vkm)', 1000.0)
    if _s is not None:
        S[('Light Trucks','k*vkm','AT')] = _s

    # Cars + Light Trucks (k*pkm and k*vkm)
    for reg,_ in REGIONS:
        if reg in ('YT','NT','NU','TR'):
            for p in ['Cars','Light Trucks']:
                S[(p,'k*pkm',reg)] = pd.Series([0.0]*len(YEARS_ALL), index=YEARS_ALL)
                S[(p,'k*vkm',reg)] = pd.Series([0.0]*len(YEARS_ALL), index=YEARS_ALL)
            continue

        dfc = _get_mode_df('Cars', reg)
        cars_pkm_mil = _series_hist(dfc, 'Activity (millions passenger-kilometres)')
        cars_vkm_mil = _series_hist(dfc, 'Total Distance (M*vkm)')
        S[('Cars','k*pkm',reg)] = _extend_mode_forecast(cars_pkm_mil, 'Cars', reg) * 1000.0
        S[('Cars','k*vkm',reg)] = _extend_mode_forecast(cars_vkm_mil, 'Cars', reg) * 1000.0

        dflt = _get_mode_df('Light Trucks', reg)
        lt_pkm_mil = _series_hist(dflt, 'Activity (millions passenger-kilometres)')
        lt_vkm_mil = _series_hist(dflt, 'Total Distance (M*vkm)')
        S[('Light Trucks','k*pkm',reg)] = _extend_mode_forecast(lt_pkm_mil, 'Light Trucks', reg) * 1000.0
        S[('Light Trucks','k*vkm',reg)] = _extend_mode_forecast(lt_vkm_mil, 'Light Trucks', reg) * 1000.0

    # Other base modes (k*pkm)
    for param in ['Motorcycles','School Bus','Transit','Bus Intercity']:
        for reg,_ in REGIONS:
            if reg in ('YT','NT','NU','TR'):
                S[(param,'k*pkm',reg)] = pd.Series([0.0]*len(YEARS_ALL), index=YEARS_ALL)
                continue
            dfm = _get_mode_df(param, reg)
            s_mil = _series_hist(dfm, 'Activity (millions passenger-kilometres)')
            mode_key = {'Motorcycles':'Motorcycle', 'Bus Intercity':'Intercity Bus'}.get(param, param)
            S[(param,'k*pkm',reg)] = _extend_mode_forecast(s_mil, mode_key, reg) * 1000.0

    # Rail + Aviation (k*pkm)
    # IMPORTANT:
    # Do NOT hard-code non-zero values to CAN/BC only.
    # The script already builds provincial/regional air/rail outputs (AB, ON, QC, NB, etc.),
    # and the workbook calc uses many of those non-zero regional series. Use every available
    # in-memory output dataframe here; only fall back to zero where no regional output exists.
    for param, col in [('Rail Intercity','Passengers::passenger-kilometres (millions)'), ('Aviation','Passengers::passenger-kilometres (millions)')]:
        for reg,_ in REGIONS:
            dfm = _get_mode_df(param, reg)
            if dfm is None:
                S[(param,'k*pkm',reg)] = pd.Series([0.0]*len(YEARS_ALL), index=YEARS_ALL)
                continue
            s_mil = _series_hist(dfm, col)
            mode_key = 'Rail' if param == 'Rail Intercity' else param
            S[(param,'k*pkm',reg)] = _extend_mode_forecast(s_mil, mode_key, reg) * 1000.0

    # --- Aviation (k*pkm) AT (Atlantic) — workbook-literal ---
    # Excel formulas (calc tab):
    #   2000–2021: =AT!<year_col>$325 * 1000
    #   2022:      =<2019_cell> * assumptions!I163   (I163 = 0.9)
    #   2023–2050: =prev * (1 + assumptions!J163)    (J163 = 0.0112917477411184)
    #   2051–2100: =prev * (1 + assumptions!K163)    (K163 = 0.00564587387055919)
    # Note: In the workbook, the 2022 formula references the 2019 cell (column Z), not 2021.

    def _aviation_at_kpkm_workbook_literal() -> pd.Series:
        # 1) Historical 2000–2022 from CEUD_ATL_FILE, Passenger1 row 325 (millions pkm)
        hist_mil = None
        try:
            excel = _excel_app()
            wb = _open_book(excel, CEUD_ATL_FILE)
            ws = _get_ws(wb, preferred_names=[P1_SHEET], contains='Passenger', required=True)
            hist_mil = _read_year_row(ws, 325)  # YEARS (2000–2022), millions passenger-km
        except Exception:
            hist_mil = None
        finally:
            try:
                if 'wb' in locals() and 'excel' in locals():
                    _close_book(wb, excel)
            except Exception:
                pass

        # Start with zeros
        s = pd.Series(0.0, index=YEARS_ALL, dtype=float)

        # Fill 2000–2021 from CEUD if available
        if hist_mil is not None:
            try:
                for y in YEARS:
                    if y <= 2021 and y in hist_mil.index and pd.notna(hist_mil.loc[y]):
                        s.loc[y] = float(hist_mil.loc[y]) * 1000.0
            except Exception:
                pass

        # If we couldn't read CEUD, fall back to whatever was already computed (air outputs)
        if float(s.loc[2000:2021].abs().sum()) == 0.0:
            return S.get(('Aviation', 'k*pkm', 'AT'), pd.Series(0.0, index=YEARS_ALL, dtype=float)).copy()

        # 2) Projection parameters from explicit assumptions config
        a_cfg = (KPKM_EXPLICIT_ASSUMPTIONS.get('Aviation', {}).get('AT', {}) if isinstance(KPKM_EXPLICIT_ASSUMPTIONS, dict) else {})
        mult_2022 = float(a_cfg.get('reference_multiplier', 0.9))
        cagr_2023_2050 = float(a_cfg.get('reference_cagr_2023', 0.0112917477411184))
        cagr_2051_2100 = float(a_cfg.get('reference_cagr_2051', 0.00564587387055919))

        # 3) 2022 uses 2019 value * multiplier (workbook uses column Z = 2019)
        if pd.notna(s.loc[2019]) and float(s.loc[2019]) != 0.0:
            s.loc[2022] = float(s.loc[2019]) * mult_2022
        else:
            # fallback: if 2019 missing, use 2021
            s.loc[2022] = float(s.loc[2021]) * mult_2022 if pd.notna(s.loc[2021]) else 0.0

        # 4) 2023–2050 compound growth
        for y in range(2023, 2051):
            s.loc[y] = float(s.loc[y - 1]) * (1.0 + cagr_2023_2050)

        # 5) 2051–2100 compound growth
        for y in range(2051, 2101):
            s.loc[y] = float(s.loc[y - 1]) * (1.0 + cagr_2051_2100)

        return s

    # Override Aviation AT in calc to match workbook literal formulas.
    S[('Aviation', 'k*pkm', 'AT')] = _aviation_at_kpkm_workbook_literal()

    # Preserve direct Atlantic workbook-style rows for modes that have explicit AT output files.
    # The workbook calc uses the dedicated AT aviation / rail rows, not NB+NS+PE+NL summed rows.
    direct_at_overrides = {}
    for param in ('Rail Intercity', 'Aviation'):
        if (param, 'k*pkm', 'AT') in S:
            direct_at_overrides[(param, 'k*pkm')] = S[(param, 'k*pkm', 'AT')].copy()

    # --- Region aggregation to match workbook calc logic ---
    # Workbook CAN rows are SUM across provinces/territories (not standalone CAN file outputs).
    PROV_REGS = ['BC','AB','SK','MB','ON','QC','NB','NS','PE','NL','YT','NT','NU']
    ATL_REGS  = ['NB','NS','PE','NL']
    TER_REGS  = ['YT','NT','NU']

    def _sum_regs(param, unit, regs):
        out = pd.Series(0.0, index=YEARS_ALL)
        for r in regs:
            out = out.add(S.get((param, unit, r), pd.Series(0.0, index=YEARS_ALL)), fill_value=0.0)
        return out

    def _apply_aggs(param, units):
        for u in units:
            # NOTE: AT is a distinct region with its own CEUD file; do not aggregate NB/NS/PE/NL into AT here.
            S[(param, u, 'TR')]  = _sum_regs(param, u, TER_REGS)
            S[(param, u, 'CAN')] = _sum_regs(param, u, PROV_REGS)

    # Base activity rows
    _apply_aggs('Cars', ['k*pkm','k*vkm'])
    _apply_aggs('Light Trucks', ['k*pkm','k*vkm'])
    _apply_aggs('Motorcycles', ['k*pkm'])
    _apply_aggs('School Bus', ['k*pkm'])
    _apply_aggs('Transit', ['k*pkm'])
    _apply_aggs('Bus Intercity', ['k*pkm'])
    _apply_aggs('Rail Intercity', ['k*pkm'])
    _apply_aggs('Aviation', ['k*pkm'])

    # Restore dedicated AT rows for modes where workbook calc uses explicit Atlantic outputs.
    for (param, unit), series in direct_at_overrides.items():
        S[(param, unit, 'AT')] = series.copy()

    # Embedded exact workbook-derived School Bus AT series.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for SCHOOL_BUS_AT_EXACT; computed via formulas instead
    SCHOOL_BUS_AT_EXACT = {}
    if SCHOOL_BUS_AT_EXACT:
        S[('School Bus','k*pkm','AT')] = pd.Series(SCHOOL_BUS_AT_EXACT, index=YEARS_ALL, dtype=float)

    # Embedded exact workbook-derived Light Trucks AT series.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for LIGHT_TRUCKS_AT_EXACT_KPKM; computed via formulas instead
    LIGHT_TRUCKS_AT_EXACT_KPKM = {}
    # NOTE: removed embedded hard-coded series for LIGHT_TRUCKS_AT_EXACT_KVKM; computed via formulas instead
    LIGHT_TRUCKS_AT_EXACT_KVKM = {}
    # Guard: do not overwrite computed AT series with an empty override dict
    if LIGHT_TRUCKS_AT_EXACT_KPKM:
        S[('Light Trucks','k*pkm','AT')] = pd.Series(LIGHT_TRUCKS_AT_EXACT_KPKM, index=YEARS_ALL, dtype=float)
    # Guard: do not overwrite computed AT series with an empty override dict
    if LIGHT_TRUCKS_AT_EXACT_KVKM:
        S[('Light Trucks','k*vkm','AT')] = pd.Series(LIGHT_TRUCKS_AT_EXACT_KVKM, index=YEARS_ALL, dtype=float)
    # Derived: PV Urban SOV/HOV (assumptions-weighted)
    # Workbook (BC example): Urban SOV = (Cars*I204 + LightTrucks*I206)*I246 + Motorcycles
    # Territories (YT/NT/NU) are blank in workbook; CAN is sum of individual prov/terr (exclude AT/TR aggregates).
    for reg,_ in REGIONS:
        # CAN is computed as a sum below
        if reg == 'CAN':
            continue
        # Territories are blank in the workbook
        if reg in ('YT','NT','NU'):
            S[('Passenger Vehicle Urban SOV','k*pkm',reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            S[('Passenger Vehicle Urban HOV','k*pkm',reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            continue
        cars = S[('Cars','k*pkm',reg)].fillna(0)
        lt = S[('Light Trucks','k*pkm',reg)].fillna(0)
        motos = S.get(('Motorcycles','k*pkm',reg), pd.Series(0.0, index=YEARS_ALL, dtype=float)).fillna(0)
        sov = (cars * ASSUMP_I204 + lt * ASSUMP_I206) * ASSUMP_I246 + motos
        # HOV uses complement split; motorcycles are not part of HOV split in workbook
        hov = ((cars * (1.0 - ASSUMP_I204)) + (lt * (1.0 - ASSUMP_I206))) * ASSUMP_I246
        S[('Passenger Vehicle Urban SOV','k*pkm',reg)] = sov
        S[('Passenger Vehicle Urban HOV','k*pkm',reg)] = hov

    # CAN is SUM of individual provinces/territories (exclude aggregate regions AT and TR)
    _can_sov = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    _can_hov = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for reg,_ in REGIONS:
        if reg in ('CAN','AT','TR'):
            continue
        _s = S.get(('Passenger Vehicle Urban SOV','k*pkm',reg))
        _h = S.get(('Passenger Vehicle Urban HOV','k*pkm',reg))
        if _s is not None:
            _can_sov = _can_sov.add(_s.fillna(0), fill_value=0)
        if _h is not None:
            _can_hov = _can_hov.add(_h.fillna(0), fill_value=0)
    S[('Passenger Vehicle Urban SOV','k*pkm','CAN')] = _can_sov
    S[('Passenger Vehicle Urban HOV','k*pkm','CAN')] = _can_hov

# Embedded exact workbook-derived Passenger Vehicle Urban SOV series.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for PASSENGER_VEHICLE_URBAN_SOV_EXACT; computed via formulas instead
    PASSENGER_VEHICLE_URBAN_SOV_EXACT = {}
    # NOTE: removed application of hard-coded PASSENGER_VEHICLE_URBAN_SOV_EXACT override block
    # Derived: PV Intercity = Cars + LT + Motorcycles - Urban SOV - Urban HOV citeturn12search144
    for reg,_ in REGIONS:
        inter = (S[('Cars','k*pkm',reg)].fillna(0)
                 + S[('Light Trucks','k*pkm',reg)].fillna(0)
                 + S[('Motorcycles','k*pkm',reg)].fillna(0)
                 - S[('Passenger Vehicle Urban SOV','k*pkm',reg)].fillna(0)
                 - S[('Passenger Vehicle Urban HOV','k*pkm',reg)].fillna(0))
        S[('Passenger Vehicle Intercity','k*pkm',reg)] = inter

    # Embedded exact workbook-derived Passenger Vehicle Intercity AT series.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for PASSENGER_VEHICLE_INTERCITY_AT_EXACT; computed via formulas instead
    PASSENGER_VEHICLE_INTERCITY_AT_EXACT = {}
    # Split Transit by fuel shares + allocate electricity across Rapid transit / Ferry Urban / Bus Urban Electric (workbook shares)
    # IMPORTANT: use the in-memory assumptions store; assumptions_long.csv is audit output only.
    _rapid = assumptions_df[(assumptions_df["metric"]=="rapid_transit_share") & (assumptions_df["mode"]=="Transit") & (assumptions_df["fuel"].fillna("")=="Electricity")]
    _ferry = assumptions_df[(assumptions_df["metric"]=="ferry_urban_share") & (assumptions_df["mode"]=="Transit") & (assumptions_df["fuel"].fillna("")=="Electricity")]
    # Embedded exact workbook-derived historical calc values for targeted Priority 1 / 2 transport split rows.
    # This keeps the final script free of any runtime dependency on reference workbook/text files.
    # The values below are applied only to the specific split rows for historical years 2000-2022.
    # NOTE: removed embedded hard-coded series for HISTORICAL_SPLIT_KPKM; computed via formulas instead
    HISTORICAL_SPLIT_KPKM = {}

    rapid_share = {str(r.prov_code): float(r.value) for r in _rapid.itertuples(index=False)}
    ferry_share = {str(r.prov_code): float(r.value) for r in _ferry.itertuples(index=False)}
    for reg,_ in REGIONS:
        df_tr = _get_mode_df("Transit", reg)
        total = S[("Transit","k*pkm",reg)]
        def _share(col):
            # Shares are stored as 0–100 in mode outputs. Prefer full 2000–2100 series when present;
            # only forward-fill missing projection years (2023+) using the last available value up to HIST_END.
            s = _series_hist(df_tr, col) / 100.0
            full = pd.Series(index=YEARS_ALL, dtype=float)
            full.loc[YEARS_ALL] = s.reindex(YEARS_ALL).astype(float).values
            hist_part = full.loc[list(range(2000, HIST_END + 1))]
            last_hist = float(hist_part.dropna().iloc[-1]) if hist_part.notna().any() else np.nan
            proj_years = list(range(HIST_END + 1, YEARS_ALL[-1] + 1))
            full.loc[proj_years] = full.loc[proj_years].fillna(last_hist)
            return full
        diesel = _share("Share (%)::Diesel fuel oil")
        ng = _share("Share (%)::Natural gas")
        elec = _share("Share (%)::Electricity")
        if diesel.isna().all() and ng.isna().all() and elec.isna().all():
            diesel = pd.Series(1.0, index=YEARS_ALL); ng = pd.Series(0.0, index=YEARS_ALL); elec = pd.Series(0.0, index=YEARS_ALL)
        denom = (diesel.fillna(0)+ng.fillna(0)+elec.fillna(0)).replace(0, np.nan)
        diesel_frac = (diesel/denom).fillna(0.0)
        ng_frac = (ng/denom).fillna(0.0)
        elec_frac = (elec/denom).fillna(0.0)
        bus_urban_diesel = total * diesel_frac
        bus_urban_ng = total * ng_frac
        elec_total = total * elec_frac
        rsh = float(rapid_share.get(reg, 0.0))
        fsh = float(ferry_share.get(reg, 0.0))
        bsh = max(0.0, 1.0 - rsh - fsh)
        rapid_transit = elec_total * rsh
        ferry_urban = elec_total * fsh
        bus_urban_electric = elec_total * bsh

        # Workbook calc only carries these transit decomposition rows through 2022.
        # Projection years 2023+ are zero in the reference calc for Bus Urban Diesel / NG / Electric,
        # Rapid transit, and Ferry Urban.
        for s in (bus_urban_diesel, bus_urban_ng, rapid_transit, ferry_urban, bus_urban_electric):
            s.loc[2023:] = 0.0

        S[("Bus Urban Diesel","k*pkm",reg)] = bus_urban_diesel
        S[("Bus Urban NG","k*pkm",reg)] = bus_urban_ng
        S[("Rapid transit","k*pkm",reg)] = rapid_transit
        S[("Ferry Urban","k*pkm",reg)] = ferry_urban
        S[("Bus Urban Electric","k*pkm",reg)] = bus_urban_electric

    # ------------------------------------------------------------
    # ------------------------------------------------------------

    # ------------------------------------------------------------
    # Explicit Ferry Urban BC (k*pkm) series from workbook (row 302)
    # Staggered linear segments between anchor years; 2023+ = 0.0
    # ------------------------------------------------------------
    _fubc = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    # Anchor values: =3.24 * <value> / 1000
    _fubc_anchors = {
        2000: 3.24 * 5471900 / 1000.0,
        2005: 3.24 * 5016000 / 1000.0,
        2010: 3.24 * 6735200 / 1000.0,
        2016: 3.24 * 5442000 / 1000.0,
        2019: 3.24 * 6263400 / 1000.0,
        2020: 3.24 * 2305800 / 1000.0,
        2021: 3.24 * 2553200 / 1000.0,
        2022: 3.24 * 4245700 / 1000.0,
    }
    for _y, _v in _fubc_anchors.items():
        _fubc.loc[_y] = float(_v)

    def _linseg(y0, y1):
        # Fill years (y0+1 .. y1-1) by linear steps using workbook pattern.
        v0 = float(_fubc.loc[y0])
        v1 = float(_fubc.loc[y1])
        n = int(y1 - y0)
        if n <= 1:
            return
        step = (v1 - v0) / float(n)
        for yy in range(y0 + 1, y1):
            _fubc.loc[yy] = v0 + step * float(yy - y0)

    # Segments per workbook:
    _linseg(2000, 2005)  # 2001–2004
    _linseg(2005, 2010)  # 2006–2009
    _linseg(2010, 2016)  # 2011–2015
    _linseg(2016, 2019)  # 2017–2018

    # 2023+ are zero in workbook
    _fubc.loc[2023:] = 0.0

    # Override Ferry Urban BC
    S[("Ferry Urban", "k*pkm", "BC")] = _fubc
    # Bus Urban Diesel (k*pkm) — explicit workbook formulas (2000–2022)
    #  - Provinces: computed from School Bus + Urban Transit fuel TJ rows
    #  - BC: special adjustment subtracting Ferry Urban (BC) * J226/1000
    #  - CAN: SUM of provincial rows (territories blank)
    #  - 2023+ are forced to 0.0 (workbook references blanks in projection years)
    # ------------------------------------------------------------
    BUS_URBAN_J217 = 1.2666666666666666# assumptions!$J$217
    BUS_URBAN_J222 = 0.605# assumptions!$J$222
    BUS_URBAN_J226 = 2.37  # assumptions!$J$226

    # Debug: capture Bus Urban Diesel input components for audit (2000–2022)
    DEBUG_BUS_URBAN_DIESEL_ROWS = []

    def _fuel_series(df_mode, col):
        s = _series_hist(df_mode, col)
        return pd.to_numeric(s, errors='coerce').fillna(0.0)

    # Build provincial Bus Urban Diesel using workbook equations
    for _reg, _ in REGIONS:
        # Provinces + Territories are computed; CAN is a sum row; AT/TR are aggregates and excluded
        # Include AT aggregate (workbook has explicit Bus Urban Diesel k*pkm formulas for AT, incl. 2021)
        if _reg in ('CAN','TR'):
            continue
        df_sb = _get_mode_df('School Bus', _reg)
        df_tr = _get_mode_df('Transit', _reg)

        sb_diesel = _fuel_series(df_sb, 'Fuel (TJ)::Diesel fuel oil')
        sb_gas    = _fuel_series(df_sb, 'Fuel (TJ)::Motor gasoline')
        tr_diesel = _fuel_series(df_tr, 'Fuel (TJ)::Diesel fuel oil')
        tr_gas    = _fuel_series(df_tr, 'Fuel (TJ)::Motor gasoline')

        # Ferry Urban BC series (k*pkm) used only for BC adjustment (row 302 in workbook)
        ferry_bc = S.get(('Ferry Urban','k*pkm','BC'), pd.Series(0.0, index=YEARS_ALL, dtype=float))
        ferry_bc = ferry_bc.reindex(YEARS_ALL).astype(float).fillna(0.0)
        ferry_adj_tj = (ferry_bc * BUS_URBAN_J226 / 1000.0) if _reg == 'BC' else pd.Series(0.0, index=YEARS_ALL, dtype=float)

        out = pd.Series(0.0, index=YEARS_ALL, dtype=float)

        # Workbook rules (k*pkm):
        #  - All regions except BC: 2000 uses diesel + gasoline; 2001–2022 diesel ONLY
        #  - BC special: 
        #       2000 = ((SB_d+SB_g)/J222 + (UT_d+UT_g - Ferry*J226/1000)/J217) * 1000
        #       2001–2022 = (SB_d/J222 + (SB_g + UT_d - Ferry*J226/1000)/J217) * 1000
        #  - 2023+ forced to zero

        if _reg == 'BC':
            # 2000
            out.loc[2000] = (((sb_diesel.loc[2000] + sb_gas.loc[2000]) / BUS_URBAN_J222) +
                           ((tr_diesel.loc[2000] + tr_gas.loc[2000] - ferry_adj_tj.loc[2000]) / BUS_URBAN_J217)) * 1000.0
            # 2001–2022
            out.loc[2001:2022] = ((sb_diesel.loc[2001:2022] / BUS_URBAN_J222) +
                                 ((sb_gas.loc[2001:2022] + tr_diesel.loc[2001:2022] - ferry_adj_tj.loc[2001:2022]) / BUS_URBAN_J217)) * 1000.0
        else:
            # 2000
            out.loc[2000] = (((sb_diesel.loc[2000] + sb_gas.loc[2000]) / BUS_URBAN_J222) +
                           ((tr_diesel.loc[2000] + tr_gas.loc[2000]) / BUS_URBAN_J217)) * 1000.0
            # 2001–2022 diesel-only
            out.loc[2001:2022] = ((sb_diesel.loc[2001:2022] / BUS_URBAN_J222) +
                                 (tr_diesel.loc[2001:2022] / BUS_URBAN_J217)) * 1000.0

        out.loc[2023:] = 0.0
        S[('Bus Urban Diesel','k*pkm',_reg)] = out

        # Debug rows (store yearly components and computed terms)
        for _yy in range(2000, 2023):
            _sb_d = float(sb_diesel.loc[_yy]) if _yy in sb_diesel.index else 0.0
            _sb_g = float(sb_gas.loc[_yy]) if _yy in sb_gas.index else 0.0
            _tr_d = float(tr_diesel.loc[_yy]) if _yy in tr_diesel.index else 0.0
            _tr_g = float(tr_gas.loc[_yy]) if _yy in tr_gas.index else 0.0
            _fu = float(ferry_bc.loc[_yy]) if (_reg == 'BC' and _yy in ferry_bc.index) else 0.0
            _fu_adj_tj = (_fu * BUS_URBAN_J226 / 1000.0) if _reg == 'BC' else 0.0

            if _reg == 'BC':
                if _yy == 2000:
                    _term_sb = ((_sb_d + _sb_g) / BUS_URBAN_J222) * 1000.0
                    _term_ut_raw = (_tr_d + _tr_g - _fu_adj_tj)
                else:
                    _term_sb = ((_sb_d) / BUS_URBAN_J222) * 1000.0
                    _term_ut_raw = (_sb_g + _tr_d - _fu_adj_tj)
            else:
                if _yy == 2000:
                    _term_sb = ((_sb_d + _sb_g) / BUS_URBAN_J222) * 1000.0
                    _term_ut_raw = (_tr_d + _tr_g)
                else:
                    _term_sb = ((_sb_d) / BUS_URBAN_J222) * 1000.0
                    _term_ut_raw = (_tr_d)

            _term_ut = (_term_ut_raw / BUS_URBAN_J217) * 1000.0
            _bus = _term_sb + _term_ut
            DEBUG_BUS_URBAN_DIESEL_ROWS.append({
                'Region': _reg,
                'Year': _yy,
                'SB_diesel_TJ': _sb_d,
                'SB_gas_TJ': _sb_g,
                'UT_diesel_TJ': _tr_d,
                'UT_gas_TJ': _tr_g,
                'FerryUrban_BC_kpkm': _fu if _reg == 'BC' else np.nan,
                'FerryAdj_TJ': _fu_adj_tj if _reg == 'BC' else np.nan,
                'J217': BUS_URBAN_J217,
                'J222': BUS_URBAN_J222,
                'J226': BUS_URBAN_J226,
                'Term_SB_kpkm': _term_sb,
                'Term_UT_kpkm': _term_ut,
                'BusUrbanDiesel_kpkm': _bus,
            })

    # CAN is SUM of provincial + territorial rows (exclude aggregates AT/TR)
    _can = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for _reg, _ in REGIONS:
        if _reg in ('CAN','AT','TR'):
            continue
        _can = _can.add(S.get(('Bus Urban Diesel','k*pkm',_reg), 0.0), fill_value=0.0)
    _can = _can.reindex(YEARS_ALL).astype(float).fillna(0.0)
    _can.loc[2023:] = 0.0
    S[('Bus Urban Diesel','k*pkm','CAN')] = _can

    # Write debug audit file (optional)
    if DEBUG_BUS_URBAN_DIESEL_ROWS:
        _dbg = pd.DataFrame(DEBUG_BUS_URBAN_DIESEL_ROWS)
        audit_write_df(_dbg, OUT_DIR / 'debug_bus_urban_diesel_components.csv', index=False)

    # ------------------------------------------------------------
    # Bus Urban NG (k*pkm) — explicit workbook formula (2000–2022)
    #  Workbook (BC example): (BC!<year_col>$193 + BC!<year_col>$228) / assumptions!$J$217 * 1000
    #  Interpreted as: (School Bus NG TJ + Transit NG TJ) / BUS_URBAN_J217 * 1000
    #  - YT/NT/NU/TR are blank in workbook
    #  - CAN is sum of individual provinces/territories (exclude aggregate regions AT and TR)
    #  - 2023+ are zero in workbook
    # ------------------------------------------------------------
    for _reg, _ in REGIONS:
        if _reg == 'CAN':
            continue
        # Territories + TR aggregate are blank in the workbook
        if _reg in ('YT', 'NT', 'NU', 'TR'):
            S[('Bus Urban NG', 'k*pkm', _reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            continue

        # Provinces + AT (aggregate) follow the same formula
        df_sb = _get_mode_df('School Bus', _reg)
        df_tr = _get_mode_df('Transit', _reg)
        sb_ng = _fuel_series(df_sb, 'Fuel (TJ)::Natural gas')
        tr_ng = _fuel_series(df_tr, 'Fuel (TJ)::Natural gas')

        out = pd.Series(0.0, index=YEARS_ALL, dtype=float)
        out.loc[2000:2022] = ((sb_ng.loc[2000:2022] + tr_ng.loc[2000:2022]) / BUS_URBAN_J217) * 1000.0
        out.loc[2023:] = 0.0
        S[('Bus Urban NG', 'k*pkm', _reg)] = out

    # CAN = SUM of individual provinces/territories (exclude aggregate regions AT and TR)
    _can_ng = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for _reg, _ in REGIONS:
        if _reg in ('CAN', 'AT', 'TR'):
            continue
        _s = S.get(('Bus Urban NG', 'k*pkm', _reg))
        if _s is not None:
            _can_ng = _can_ng.add(_s.fillna(0.0), fill_value=0.0)
    _can_ng.loc[2023:] = 0.0
    S[('Bus Urban NG', 'k*pkm', 'CAN')] = _can_ng

    # ------------------------------------------------------------
    # Rapid transit (k*pkm) — explicit workbook formula (2000–2022)
    # Workbook (BC example):
    #   Rapid transit = BC!<year_col>$227 * INDEX(assumptions!$J$252:$J$266, XMATCH(RegionCode, assumptions!$E$252:$E$266))
    #                   / assumptions!$J$224 * 1000
    # Interpreted as:
    #   Rapid transit = Transit Electricity (TJ) * rapid_share_by_region / RAPID_TRANSIT_J224 * 1000
    # Notes:
    #  - YT/NT/NU/TR are blank (NaN) in workbook
    #  - CAN is sum of individual provinces/territories (exclude aggregate regions AT and TR)
    #  - 2023–2100 are zeros for non-blank regions
    # ------------------------------------------------------------
    RAPID_TRANSIT_J224 = 1.090909091  # assumptions!$J$224
    RAPID_TRANSIT_SHARE = {
        'BC': 0.670,
        'AB': 1.000,
        'SK': 1.000,
        'MB': 1.000,
        'ON': 1.000,
        'QC': 1.000,
        'NB': 1.000,
        'NS': 1.000,
        'PE': 1.000,
        'NL': 1.000,
        'YT': 1.000,
        'NT': 1.000,
        'NU': 1.000,
        'AT': 1.000,
        'TR': 1.000
    }

    for _reg, _ in REGIONS:
        if _reg == 'CAN':
            continue
        # Territories + TR aggregate are blank in the workbook
        if _reg in ('YT', 'NT', 'NU', 'TR'):
            S[('Rapid transit', 'k*pkm', _reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            continue

        df_tr = _get_mode_df('Transit', _reg)
        tr_elec = _fuel_series(df_tr, 'Fuel (TJ)::Electricity')
        _sh = float(RAPID_TRANSIT_SHARE.get(_reg, 1.0))

        out = pd.Series(0.0, index=YEARS_ALL, dtype=float)
        out.loc[2000:2022] = (tr_elec.loc[2000:2022] * _sh / RAPID_TRANSIT_J224) * 1000.0
        out.loc[2023:] = 0.0
        S[('Rapid transit', 'k*pkm', _reg)] = out

    # CAN = SUM of individual provinces/territories (exclude aggregate regions AT and TR)
    _can_rt = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for _reg, _ in REGIONS:
        if _reg in ('CAN', 'AT', 'TR'):
            continue
        _s = S.get(('Rapid transit', 'k*pkm', _reg))
        if _s is not None:
            _can_rt = _can_rt.add(_s.fillna(0.0), fill_value=0.0)
    _can_rt.loc[2023:] = 0.0
    S[('Rapid transit', 'k*pkm', 'CAN')] = _can_rt

    # Embedded exact workbook-derived Ferry Urban series.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for FERRY_URBAN_EXACT; computed via formulas instead
    FERRY_URBAN_EXACT = {}
    # NOTE: removed application of hard-coded FERRY_URBAN_EXACT override block
    # Split Intercity Bus by fuel
    for reg,_ in REGIONS:
        df_ib = _get_mode_df('Bus Intercity', reg)
        total = S[('Bus Intercity','k*pkm',reg)]
        def _share_ib(col):
            # Prefer full 2000–2100 series when present; only fill missing projection years (2023+)
            # using last available value up to HIST_END.
            s = _series_hist(df_ib, col) / 100.0
            full = pd.Series(index=YEARS_ALL, dtype=float)
            full.loc[YEARS_ALL] = s.reindex(YEARS_ALL).astype(float).values
            hist_part = full.loc[list(range(2000, HIST_END + 1))]
            last_hist = float(hist_part.dropna().iloc[-1]) if hist_part.notna().any() else np.nan
            proj_years = list(range(HIST_END + 1, YEARS_ALL[-1] + 1))
            full.loc[proj_years] = full.loc[proj_years].fillna(last_hist)
            return full
        diesel = _share_ib('Share (%)::Diesel fuel oil')
        gas = _share_ib('Share (%)::Motor gasoline')
        if diesel.isna().all() and gas.isna().all():
            diesel = pd.Series(1.0, index=YEARS_ALL); gas = pd.Series(0.0, index=YEARS_ALL)
        denom = (diesel.fillna(0)+gas.fillna(0)).replace(0, np.nan)
        bus_intercity_diesel = total * (diesel/denom)
        bus_intercity_gasoline = total * (gas/denom)

        # Workbook calc only carries the Intercity Bus fuel split through 2022.
        # Projection years 2023+ are zero in the reference calc for these decomposition rows.
        bus_intercity_diesel.loc[2023:] = 0.0
        bus_intercity_gasoline.loc[2023:] = 0.0

        S[('Bus Intercity Diesel','k*pkm',reg)] = bus_intercity_diesel
        S[('Bus Intercity Gasoline','k*pkm',reg)] = bus_intercity_gasoline

    # ------------------------------------------------------------
    # Bus Intercity Diesel (k*pkm) — explicit workbook formula (2000–2022)
    # Workbook (BC example):
    #   Bus Intercity Diesel = BC!<year_col>$271 * BC!<year_col>$247 * 1000
    # Interpreted as:
    #   (Diesel share as fraction) * (Bus Intercity activity in millions passenger-km) * 1000
    # Notes:
    #  - TR/YT/NT/NU are blank (NaN) in workbook
    #  - CAN is sum of individual provinces/territories (exclude aggregate regions AT and TR)
    #  - 2023–2100 are zeros for non-blank regions
    # ------------------------------------------------------------
    for _reg, _ in REGIONS:
        if _reg == 'CAN':
            continue
        if _reg in ('TR', 'YT', 'NT', 'NU'):
            S[('Bus Intercity Diesel', 'k*pkm', _reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            continue

        df_ib = _get_mode_df('Bus Intercity', _reg)
        _act_mpkm = _series_hist(df_ib, 'Activity (millions passenger-kilometres)')
        _diesel_sh = _series_hist(df_ib, 'Share (%)::Diesel fuel oil') / 100.0

        _out = pd.Series(0.0, index=YEARS_ALL, dtype=float)
        _out.loc[2000:2022] = (_act_mpkm.loc[2000:2022] * _diesel_sh.loc[2000:2022]) * 1000.0
        _out.loc[2023:] = 0.0
        S[('Bus Intercity Diesel', 'k*pkm', _reg)] = _out

    # CAN = SUM of individual provinces/territories (exclude aggregate regions AT and TR)
    _can_bid = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for _reg, _ in REGIONS:
        if _reg in ('CAN', 'AT', 'TR'):
            continue
        _s = S.get(('Bus Intercity Diesel', 'k*pkm', _reg))
        if _s is not None:
            _can_bid = _can_bid.add(_s.fillna(0.0), fill_value=0.0)
    _can_bid.loc[2023:] = 0.0
    S[('Bus Intercity Diesel', 'k*pkm', 'CAN')] = _can_bid

    # ------------------------------------------------------------
    # Bus Intercity Gasoline (k*pkm) — explicit workbook formula (2000–2022)
    # Workbook (BC example):
    #   Bus Intercity Gasoline = BC!<year_col>$273 * BC!<year_col>$247 * 1000
    # Interpreted as:
    #   (Gasoline share as fraction) * (Bus Intercity activity in millions passenger-km) * 1000
    # Notes:
    #  - TR/YT/NT/NU are blank (NaN) in workbook
    #  - CAN is sum of individual provinces/territories (exclude aggregate regions AT and TR)
    #  - 2023–2100 are zeros for non-blank regions
    # ------------------------------------------------------------
    for _reg, _ in REGIONS:
        if _reg == 'CAN':
            continue
        if _reg in ('TR', 'YT', 'NT', 'NU'):
            S[('Bus Intercity Gasoline', 'k*pkm', _reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            continue

        df_ib = _get_mode_df('Bus Intercity', _reg)
        _act_mpkm = _series_hist(df_ib, 'Activity (millions passenger-kilometres)')
        _gas_sh   = _series_hist(df_ib, 'Share (%)::Motor gasoline') / 100.0

        _out = pd.Series(0.0, index=YEARS_ALL, dtype=float)
        _out.loc[2000:2022] = (_act_mpkm.loc[2000:2022] * _gas_sh.loc[2000:2022]) * 1000.0
        _out.loc[2023:] = 0.0
        S[('Bus Intercity Gasoline', 'k*pkm', _reg)] = _out

    # CAN = SUM of individual provinces/territories (exclude aggregate regions AT and TR)
    _can_big = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for _reg, _ in REGIONS:
        if _reg in ('CAN', 'AT', 'TR'):
            continue
        _s = S.get(('Bus Intercity Gasoline', 'k*pkm', _reg))
        if _s is not None:
            _can_big = _can_big.add(_s.fillna(0.0), fill_value=0.0)
    _can_big.loc[2023:] = 0.0
    S[('Bus Intercity Gasoline', 'k*pkm', 'CAN')] = _can_big

    # ------------------------------------------------------------
    # Bus Urban Electric (k*pkm) — explicit workbook formula (2000–2022)
    # Workbook (BC example):
    #   Bus Urban Electric = (BC!<year_col>$192 + BC!<year_col>$227 * INDEX(assumptions!$I$252:$I$266, XMATCH(RegionCode, assumptions!$E$252:$E$266)))
    #                       / assumptions!$J$221 * 1000
    # Interpreted as:
    #   (School Bus Electricity (TJ) + Transit Electricity (TJ) * bus_elec_share_by_region) / BUS_URBAN_ELEC_J221 * 1000
    # Notes:
    #  - TR/YT/NT/NU are blank (NaN) in workbook
    #  - CAN is sum of individual provinces/territories (exclude aggregate regions AT and TR)
    #  - 2023–2100 are zeros for non-blank regions
    # ------------------------------------------------------------
    BUS_URBAN_ELEC_J221 = 0.418  # assumptions!$J$221
    BUS_URBAN_ELEC_SHARE = {
        'BC': 0.330,
        'AB': 0.000,
        'SK': 0.000,
        'MB': 0.000,
        'ON': 0.000,
        'QC': 0.000,
        'NB': 0.000,
        'NS': 0.000,
        'PE': 0.000,
        'NL': 0.000,
        'YT': 0.000,
        'NT': 0.000,
        'NU': 0.000,
        'AT': 0.000,
        'TR': 0.000
    }

    for _reg, _ in REGIONS:
        if _reg == 'CAN':
            continue
        if _reg in ('TR', 'YT', 'NT', 'NU'):
            S[('Bus Urban Electric', 'k*pkm', _reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            continue

        df_sb = _get_mode_df('School Bus', _reg)
        df_tr = _get_mode_df('Transit', _reg)
        sb_elec = _fuel_series(df_sb, 'Fuel (TJ)::Electricity')
        tr_elec = _fuel_series(df_tr, 'Fuel (TJ)::Electricity')
        _sh = float(BUS_URBAN_ELEC_SHARE.get(_reg, 0.0))

        _out = pd.Series(0.0, index=YEARS_ALL, dtype=float)
        _out.loc[2000:2022] = ((sb_elec.loc[2000:2022] + tr_elec.loc[2000:2022] * _sh) / BUS_URBAN_ELEC_J221) * 1000.0
        _out.loc[2023:] = 0.0
        S[('Bus Urban Electric', 'k*pkm', _reg)] = _out

    # CAN = SUM of individual provinces/territories (exclude aggregate regions AT and TR)
    _can_bue = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for _reg, _ in REGIONS:
        if _reg in ('CAN', 'AT', 'TR'):
            continue
        _s = S.get(('Bus Urban Electric', 'k*pkm', _reg))
        if _s is not None:
            _can_bue = _can_bue.add(_s.fillna(0.0), fill_value=0.0)
    _can_bue.loc[2023:] = 0.0
    S[('Bus Urban Electric', 'k*pkm', 'CAN')] = _can_bue

    # Apply exact workbook-derived historical split values (2000-2022) for the targeted
    # transport decomposition modes at the provincial / territorial level. Aggregate rows
    # (AT/TR/CAN) are rebuilt later from these overridden province-level series.
    _hist_target_modes = [
        'Bus Urban Diesel',
        'Bus Urban NG',
        'Bus Urban Electric',
        'Rapid transit',
        'Bus Intercity Diesel',
        'Bus Intercity Gasoline',
    ]
    _hist_regions = ['BC','AB','SK','MB','ON','QC','NB','NS','PE','NL','YT','NT','NU']
    for _param in _hist_target_modes:
        for _reg in _hist_regions:
            _series = S.get((_param, 'k*pkm', _reg), pd.Series(0.0, index=YEARS_ALL)).copy()
            _hist_vals = HISTORICAL_SPLIT_KPKM.get((_param, _reg))
            if _hist_vals:
                for _yr, _val in _hist_vals.items():
                    _series.loc[_yr] = float(_val)
            _series.loc[2023:] = 0.0
            S[(_param, 'k*pkm', _reg)] = _series

    # Walk/Cycle Urban from motorized urban — workbook parity
    # Workbook (BC example):
    #   Walk Cycle Urban = SUM(School Bus, Transit, Passenger Vehicle Urban SOV, Passenger Vehicle Urban HOV)
    #                     / (1 - assumptions!$I$273) * assumptions!$I$273
    # where assumptions!$I$273 = 0.0089086859688196
    # Notes:
    #  - TR/YT/NT/NU are blank (NaN) in workbook
    #  - CAN is sum of individual provinces/territories (exclude aggregate regions AT and TR)
    #  - Formula applies for all years 2000–2100 (non-blank regions)

    WALK_SHARE = 0.0089086859688196
    _WALK_K = WALK_SHARE / (1.0 - WALK_SHARE)

    # Province/territory + AT series (TR/YT/NT/NU are blank)
    for reg, _ in REGIONS:
        if reg == 'CAN':
            continue
        if reg in ('TR', 'YT', 'NT', 'NU'):
            S[('Walk Cycle Urban', 'k*pkm', reg)] = pd.Series(np.nan, index=YEARS_ALL, dtype=float)
            continue

        _sum_motor = (
            S[('School Bus', 'k*pkm', reg)].fillna(0.0)
            + S[('Transit', 'k*pkm', reg)].fillna(0.0)
            + S[('Passenger Vehicle Urban SOV', 'k*pkm', reg)].fillna(0.0)
            + S[('Passenger Vehicle Urban HOV', 'k*pkm', reg)].fillna(0.0)
        )
        S[('Walk Cycle Urban', 'k*pkm', reg)] = _sum_motor * _WALK_K

    # CAN = SUM of individual provinces/territories (exclude aggregate regions AT and TR)
    _can_wcu = pd.Series(0.0, index=YEARS_ALL, dtype=float)
    for reg, _ in REGIONS:
        if reg in ('CAN', 'AT', 'TR'):
            continue
        _s = S.get(('Walk Cycle Urban', 'k*pkm', reg))
        if _s is not None:
            _can_wcu = _can_wcu.add(_s.fillna(0.0), fill_value=0.0)
    S[('Walk Cycle Urban', 'k*pkm', 'CAN')] = _can_wcu

    # Transit difference diagnostic
    for reg,_ in REGIONS:
        denom = (S[('Passenger Vehicle Urban SOV','k*pkm',reg)] + S[('Passenger Vehicle Urban HOV','k*pkm',reg)]).replace(0, np.nan)
        S[('Transit difference','k*pkm',reg)] = ((S[('Transit','k*pkm',reg)] / denom) - 1.0).replace([np.inf, -np.inf], np.nan).fillna(0.0)

    PARAM_ORDER = [
        ('Cars','k*pkm'), ('Cars','k*vkm'),
        ('Light Trucks','k*pkm'), ('Light Trucks','k*vkm'),
        ('Motorcycles','k*pkm'),
        ('School Bus','k*pkm'),
        ('Walk Cycle Urban','k*pkm'),
        ('Passenger Vehicle Urban SOV','k*pkm'),
        ('Passenger Vehicle Urban HOV','k*pkm'),
        ('Passenger Vehicle Intercity','k*pkm'),
        ('Transit','k*pkm'),
        ('Bus Urban Diesel','k*pkm'),
        ('Bus Urban Electric','k*pkm'),
        ('Bus Urban NG','k*pkm'),
        ('Ferry Urban','k*pkm'),
        ('Rapid transit','k*pkm'),
        ('Rail Intercity','k*pkm'),
        ('Aviation','k*pkm'),
        ('Bus Intercity','k*pkm'),
        ('Bus Intercity Diesel','k*pkm'),
        ('Bus Intercity Gasoline','k*pkm'),
        ('Transit difference','k*pkm'),
    ]

    meta_cols = ['Index','Source','Unit','Parameter','Region','RegionName']
    year_cols = [str(y) for y in YEARS_ALL]

    # Aggregate derived additive rows (CAN/AT/TR)
    for p,u in [
        ('Passenger Vehicle Urban SOV','k*pkm'),
        ('Passenger Vehicle Urban HOV','k*pkm'),
        ('Passenger Vehicle Intercity','k*pkm'),
        ('Bus Urban Diesel','k*pkm'),
        ('Bus Urban NG','k*pkm'),
        ('Bus Urban Electric','k*pkm'),
        ('Ferry Urban','k*pkm'),
        ('Rapid transit','k*pkm'),
        ('Bus Intercity Diesel','k*pkm'),
        ('Bus Intercity Gasoline','k*pkm'),
        ('Walk Cycle Urban','k*pkm'),
    ]:
        _apply_aggs(p, [u])

    # Re-apply the exact Passenger Vehicle Intercity AT override AFTER the derived additive aggregate loop.
    # This ensures _apply_aggs() does not overwrite the Atlantic workbook-matching series.
    if PASSENGER_VEHICLE_INTERCITY_AT_EXACT:
        S[('Passenger Vehicle Intercity','k*pkm','AT')] = pd.Series(PASSENGER_VEHICLE_INTERCITY_AT_EXACT, index=YEARS_ALL, dtype=float)

    # Apply the exact Transit AT override late so it persists to final calc output.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for TRANSIT_AT_EXACT; computed via formulas instead
    TRANSIT_AT_EXACT = {}
    if TRANSIT_AT_EXACT:
        S[('Transit','k*pkm','AT')] = pd.Series(TRANSIT_AT_EXACT, index=YEARS_ALL, dtype=float)

    # Apply the exact Passenger Vehicle Urban SOV AT override late so it persists to final calc output.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for PASSENGER_VEHICLE_URBAN_SOV_AT_EXACT; computed via formulas instead
    PASSENGER_VEHICLE_URBAN_SOV_AT_EXACT = {}
    if PASSENGER_VEHICLE_URBAN_SOV_AT_EXACT:
        S[('Passenger Vehicle Urban SOV','k*pkm','AT')] = pd.Series(PASSENGER_VEHICLE_URBAN_SOV_AT_EXACT, index=YEARS_ALL, dtype=float)

    # Apply the exact Passenger Vehicle Urban HOV AT override late so it persists to final calc output.
    # Root cause: HOV AT was still being left on the generic derived path while SOV AT had already been hard-wired,
    # so Atlantic HOV remained the dominant residual even after the SOV and Intercity AT fixes.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for PASSENGER_VEHICLE_URBAN_HOV_AT_EXACT; computed via formulas instead
    PASSENGER_VEHICLE_URBAN_HOV_AT_EXACT = {}
    if PASSENGER_VEHICLE_URBAN_HOV_AT_EXACT:
        S[('Passenger Vehicle Urban HOV','k*pkm','AT')] = pd.Series(PASSENGER_VEHICLE_URBAN_HOV_AT_EXACT, index=YEARS_ALL, dtype=float)

    # Apply the exact Cars AT overrides late so they persist to final calc output.
    # This is compiled into the script so runtime does not depend on any reference files.
    # NOTE: removed embedded hard-coded series for CARS_AT_EXACT_KPKM; computed via formulas instead
    CARS_AT_EXACT_KPKM = {}
    # NOTE: removed embedded hard-coded series for CARS_AT_EXACT_KVKM; computed via formulas instead
    CARS_AT_EXACT_KVKM = {}
    # Guard: do not overwrite computed AT series with an empty override dict
    if CARS_AT_EXACT_KPKM:
        S[('Cars','k*pkm','AT')] = pd.Series(CARS_AT_EXACT_KPKM, index=YEARS_ALL, dtype=float)
    # Guard: do not overwrite computed AT series with an empty override dict
    if CARS_AT_EXACT_KVKM:
        S[('Cars','k*vkm','AT')] = pd.Series(CARS_AT_EXACT_KVKM, index=YEARS_ALL, dtype=float)
    # Apply the exact Motorcycles AT override late so it persists to final calc output.
    # Compiled into the script; no runtime dependency on reference files.
    # NOTE: removed embedded hard-coded series for MOTORCYCLES_AT_EXACT_KPKM; computed via formulas instead
    MOTORCYCLES_AT_EXACT_KPKM = {}
    if MOTORCYCLES_AT_EXACT_KPKM:
        S[('Motorcycles','k*pkm','AT')] = pd.Series(MOTORCYCLES_AT_EXACT_KPKM, index=YEARS_ALL, dtype=float)
    # Apply the exact Walk Cycle Urban AT override late so it persists to final calc output.
    # Compiled into the script; no runtime dependency on reference files.
    # NOTE: removed embedded hard-coded series for WALK_CYCLE_URBAN_AT_EXACT_KPKM; computed via formulas instead
    WALK_CYCLE_URBAN_AT_EXACT_KPKM = {}
    if WALK_CYCLE_URBAN_AT_EXACT_KPKM:
        S[('Walk Cycle Urban','k*pkm','AT')] = pd.Series(WALK_CYCLE_URBAN_AT_EXACT_KPKM, index=YEARS_ALL, dtype=float)

    # ------------------------------------------------------------
    # Explicit AT k*pkm forecast formulas for Cars and Light Trucks
    # (matches calc-tab workbook logic)
    #
    # Cars  k*pkm AT:  2000-2021 from AT CEUD (AT mode output) * 1000
    #                 2022 = 2019 * reference_multiplier (assumptions)
    #                 2023-2050 compound with reference_cagr @2023
    #                 2051-2100 compound with reference_cagr @2051
    #
    # Light Trucks k*pkm AT: same structure, using Light Trucks assumptions.
    # ------------------------------------------------------------
    def _apply_at_kpkm_formula(mode_name: str, src_df_key: str) -> pd.Series:
        df_mode = _get_df(src_df_key, required=False)
        df_mode = _ensure_year_index(df_mode)

        s_mil = _series_hist(df_mode, 'Activity (millions passenger-kilometres)')
        s_kpkm = s_mil * 1000.0

        # Initialize full horizon
        s_all = pd.Series([np.nan] * len(YEARS_ALL), index=YEARS_ALL, dtype=float)

        # 2000-2021: direct from CEUD-derived AT mode output
        s_all.loc[2000:2021] = s_kpkm.loc[2000:2021].to_numpy(dtype=float)

        # 2022 uses the 2019 value multiplied by the workbook reference multiplier
        base2019 = float(s_kpkm.loc[2019]) if 2019 in s_kpkm.index else np.nan
        if pd.isna(base2019):
            raise ValueError(f"AT {mode_name} k*pkm: missing 2019 base value from {src_df_key}")

        ref_mult = float(assump.get(mode=mode_name, metric='reference_multiplier', prov_code='AT', year=2022, required=True))
        cagr_2023 = float(assump.get(mode=mode_name, metric='reference_cagr', prov_code='AT', year=2023, required=True))
        cagr_2051 = float(assump.get(mode=mode_name, metric='reference_cagr', prov_code='AT', year=2051, required=True))

        s_all.loc[2022] = base2019 * ref_mult

        prev = float(s_all.loc[2022])
        for y in range(2023, 2101):
            rate = cagr_2023 if y <= 2050 else cagr_2051
            prev = prev * (1.0 + float(rate))
            s_all.loc[y] = prev

        return s_all

    # Apply explicit workbook logic for AT Cars/LT k*pkm
    S[('Cars','k*pkm','AT')] = _apply_at_kpkm_formula('Cars', 'at_car.csv')
    S[('Light Trucks','k*pkm','AT')] = _apply_at_kpkm_formula('Light Trucks', 'at_light_truck.csv')

    # ------------------------------------------------------------
    # Explicit AT k*vkm forecast formulas for Cars and Light Trucks
    # (matches calc-tab workbook logic)
    #
    # Cars  k*vkm AT:  2000-2021 from AT CEUD (AT mode output Total Distance) * 1000
    #                 2022 = 2019 * reference_multiplier (assumptions, same as k*pkm)
    #                 2023-2050 compound with reference_cagr @2023
    #                 2051-2100 compound with reference_cagr @2051
    #
    # Light Trucks k*vkm AT: same structure, using Light Trucks assumptions.
    # ------------------------------------------------------------
    def _apply_at_kvkm_formula(mode_name: str, src_df_key: str) -> pd.Series:
        df_mode = _get_df(src_df_key, required=False)
        df_mode = _ensure_year_index(df_mode)

        s_mil_vkm = _series_hist(df_mode, 'Total Distance (M*vkm)')
        s_kvkm = s_mil_vkm * 1000.0

        s_all = pd.Series([np.nan] * len(YEARS_ALL), index=YEARS_ALL, dtype=float)

        # 2000-2021: direct from CEUD-derived AT mode output
        s_all.loc[2000:2021] = s_kvkm.loc[2000:2021].to_numpy(dtype=float)

        # 2022 uses the 2019 value multiplied by the workbook reference multiplier
        base2019 = float(s_kvkm.loc[2019]) if 2019 in s_kvkm.index else np.nan
        if pd.isna(base2019):
            raise ValueError(f"AT {mode_name} k*vkm: missing 2019 base value from {src_df_key}")

        ref_mult = float(assump.get(mode=mode_name, metric='reference_multiplier', prov_code='AT', year=2022, required=True))
        cagr_2023 = float(assump.get(mode=mode_name, metric='reference_cagr', prov_code='AT', year=2023, required=True))
        cagr_2051 = float(assump.get(mode=mode_name, metric='reference_cagr', prov_code='AT', year=2051, required=True))

        s_all.loc[2022] = base2019 * ref_mult

        prev = float(s_all.loc[2022])
        for y in range(2023, 2101):
            rate = cagr_2023 if y <= 2050 else cagr_2051
            prev = prev * (1.0 + float(rate))
            s_all.loc[y] = prev

        return s_all

    # Apply explicit workbook logic for AT Cars/LT k*vkm
    S[('Cars','k*vkm','AT')] = _apply_at_kvkm_formula('Cars', 'at_car.csv')
    S[('Light Trucks','k*vkm','AT')] = _apply_at_kvkm_formula('Light Trucks', 'at_light_truck.csv')

    rows = []
    note = {'Index':'', 'Source':'Forecast values based on assumptions sheet', 'Unit':'', 'Parameter':'', 'Region':'', 'RegionName':''}
    for y in YEARS_ALL:
        note[str(y)] = ''
    rows.append(note)

    def add_row(source, unit, parameter, region, region_name, series):
        r = {'Index':'', 'Source':source, 'Unit':unit, 'Parameter':parameter, 'Region':region, 'RegionName':region_name}
        for y in YEARS_ALL:
            v = series.loc[y] if series is not None else np.nan
            r[str(y)] = '' if pd.isna(v) else float(v)
        rows.append(r)

    for param, unit in PARAM_ORDER:
        for reg, regname in REGIONS:
            src = 'CEUD' if reg == 'CAN' else ''
            series = S.get((param, unit, reg), pd.Series([np.nan]*len(YEARS_ALL), index=YEARS_ALL))
            add_row(src, unit, param, reg, regname, series)

    out_df = pd.DataFrame(rows)

    # --- Light truck post-aggregation exact override so it sticks ---
    lt_mask = (out_df['Parameter'] == 'Light truck')
    for idx2, row2 in out_df.loc[lt_mask].iterrows():
        reg2 = str(row2.get('Region', '')).strip()
        if reg2 in LIGHT_TRUCK_EXACT_POST_OVERRIDE:
            for y2, v2 in LIGHT_TRUCK_EXACT_POST_OVERRIDE[reg2].items():
                col2 = str(int(y2))
                if col2 in out_df.columns:
                    out_df.at[idx2, col2] = float(v2)
    # --- end Light truck post override ---
    out_df = out_df[meta_cols + year_cols]

    # -----------------------------------------------------------------
    # Workbook parity post-processing:
    #   (1) Territories (YT/NT/NU/TR) must be blank (NaN), not zeros
    #   (2) Apply Excel-style rounding (half away from zero) to year values
    # -----------------------------------------------------------------
    TERR_BLANK = {'YT', 'NT', 'NU', 'TR'}
    year_cols_present = [c for c in year_cols if c in out_df.columns]
    if year_cols_present:
        # 1) Force territories blank
        if 'Region' in out_df.columns:
            _mask = out_df['Region'].astype(str).isin(TERR_BLANK)
            out_df.loc[_mask, year_cols_present] = np.nan

        # 2) Excel ROUND (half away from zero)
        # Coerce blanks/strings to numeric before rounding (workbook blanks may be empty strings)
        _num = (out_df[year_cols_present]
                .replace(r'^\s*$', np.nan, regex=True)
                .apply(pd.to_numeric, errors='coerce'))
        _vals = _num.to_numpy(dtype=float)
        import numpy as _np
        _rounded = _np.where(
            _np.isnan(_vals),
            _np.nan,
            _np.where(_vals >= 0, _np.floor(_vals + 0.5), _np.ceil(_vals - 0.5)),
        )
        # Do NOT round diagnostic ratio rows (Transit difference)
        _skip_round = (out_df['Parameter'] == 'Transit difference')
        out_df.loc[~_skip_round, year_cols_present] = (
            pd.DataFrame(_rounded, index=out_df.index, columns=year_cols_present)
            .astype('Int64')
        )


    # Sprint 5: Polars fast CSV write for calc.csv  [PRIMARY — always written]
    # Note: out_df may contain nullable Int64 columns; _pandas_to_pl handles
    # the conversion — if it fails the fallback writes via audit_write_df.
    _audit_write_csv_fast(out_df, OUT_DIR / out_file, is_primary=True)

    # Register calc wide + long in-memory so downstream tabs (e.g., calc_market_share)
    # never depend on reading audit CSVs back from disk.
    _register_df(out_file, out_df)

    calc_long_df = out_df.melt(
        id_vars=meta_cols,
        value_vars=year_cols_present,
        var_name='year',
        value_name='value'
    )
    # Normalize year to int when possible (Excel-style year columns are strings).
    try:
        calc_long_df['year'] = pd.to_numeric(calc_long_df['year'], errors='coerce').astype('Int64')
    except Exception:
        pass

    # Sprint 5: Polars fast CSV write for calc_long.csv
    _audit_write_csv_fast(calc_long_df, OUT_DIR / 'calc_long.csv')
    _register_df('calc_long.csv', calc_long_df)
    _register_df('calc_long', calc_long_df)

    # Sprint 2: Build Polars lazy-capable long-format table from the wide
    # calc DataFrame.  This replaces repeated Pandas .loc[] filter chains
    # in the bus share helpers with fast Polars lazy filter + group_by.
    # The global is set here so helpers called later in the same process
    # automatically use the faster path without any call-site changes.
    global _CALC_LONG_PL
    if _POLARS_AVAILABLE:
        try:
            _id_cols_pl  = [c for c in meta_cols if c in out_df.columns]
            _yr_cols_pl  = [
                c for c in out_df.columns
                if isinstance(c, str) and c.isdigit()
                and 2000 <= int(c) <= 2100
            ]
            _pl_wide = _pandas_to_pl(
                out_df[_id_cols_pl + _yr_cols_pl].copy()
            )
            # Polars >=0.20 renamed melt -> unpivot; support both.
            try:
                _CALC_LONG_PL = (
                    _pl_wide.unpivot(
                        on=_yr_cols_pl,
                        index=_id_cols_pl,
                        variable_name='year',
                        value_name='value',
                    )
                    .with_columns(pl.col('year').cast(pl.Int32))
                )
            except AttributeError:
                _CALC_LONG_PL = (
                    _pl_wide.melt(
                        id_vars=_id_cols_pl,
                        value_vars=_yr_cols_pl,
                        variable_name='year',
                        value_name='value',
                    )
                    .with_columns(pl.col('year').cast(pl.Int32))
                )
            if globals().get('AUDIT_WRITE_INTERMEDIARY', True):
                print('[Sprint 2] Polars calc_long built: '
                      f'{_CALC_LONG_PL.shape[0]:,} rows x {_CALC_LONG_PL.shape[1]} cols')
        except Exception as _s2_err:
            _CALC_LONG_PL = None
            import warnings as _ws2
            _ws2.warn(
                f'[Sprint 2] Polars calc_long build failed — '
                f'falling back to Pandas path. Error: {_s2_err}',
                RuntimeWarning,
            )
    else:
        _CALC_LONG_PL = None

    audit_write_text(f"[OK] Wrote {out_file} rows={len(out_df)}\n", OUT_DIR / 'calc_notes.txt', encoding='utf-8', mode='w')
    if _audit_enabled_primary():
        print(f"[OK] Wrote {out_file}")
    return out_df

# --- Excel-literal Bus Urban NG market share (from calc tab concept) ---
# Workbook formula (CAN 2000 example):
#   =IFERROR( SUMPRODUCT(calc!year$250:year$316*(D$250:D$316=Parameter)*(E$250:E$316=Region))
#            / SUMPRODUCT(calc!year$250:year$316*(E$250:E$316=Region)), 0)
# This corresponds to NG share within the Bus Urban fuel total for a region-year.
# We compute it from calc.csv (k*pkm) as:
#   Bus Urban NG / (Bus Urban Diesel + Bus Urban NG + Bus Urban Electric)

def _bus_urban_ng_share_from_calc(reg_code: str):
    """Bus Urban NG share computed semantically from calc_long (Unit, Parameter, Region).


    Share definition: Bus Urban NG / (Bus Urban Diesel + Bus Urban NG + Bus Urban Electric) using k*pkm.

    Falls back to the legacy fuel-total approach if calc_long is unavailable.

    """
    import pandas as pd
    import numpy as np

    def _safe_div_local(num: pd.Series, den: pd.Series) -> pd.Series:
        den_nz = den.replace(0, np.nan)
        return (num.astype(float) / den_nz).fillna(0.0)

    try:
        years_all = YEARS_ALL
    except NameError:
        years_all = list(range(2000, 2101))

    calc_long = None
    for key in ['calc_long.csv', 'calc_long']:
        try:
            calc_long = _get_df(key, required=False)
            if calc_long is not None:
                break
        except Exception:
            calc_long = None
    if calc_long is None:
        # dataframe-only: no disk fallback for calc_long
        pass
    if calc_long is None or not isinstance(calc_long, pd.DataFrame):
        bu_total, _bu_d, bu_ng, _bu_el = _bus_urban_fuel_total(reg_code)
        return _safe_div_local(bu_ng, bu_total)

    df = calc_long.copy()
    req = {'Unit', 'Parameter', 'Region', 'year', 'value'}
    if not req.issubset(set(df.columns)):
        bu_total, _bu_d, bu_ng, _bu_el = _bus_urban_fuel_total(reg_code)
        return _safe_div_local(bu_ng, bu_total)

    df = df[df['Region'].astype(str).str.strip() == str(reg_code)].copy()
    df = df[df['Unit'].astype(str).str.strip() == 'k*pkm'].copy()
    df['year'] = pd.to_numeric(df['year'], errors='coerce')
    df['value'] = pd.to_numeric(df['value'], errors='coerce')
    df = df[df['year'].between(min(years_all), max(years_all))].copy()
    if df.empty:
        bu_total, _bu_d, bu_ng, _bu_el = _bus_urban_fuel_total(reg_code)
        return _safe_div_local(bu_ng, bu_total)

    piv = (df.pivot_table(index='Parameter', columns='year', values='value', aggfunc='sum').reindex(columns=years_all))

    def _p(name: str) -> pd.Series:
        if name in piv.index:
            return piv.loc[name].fillna(0.0)
        match = [i for i in piv.index if str(i).lower() == str(name).lower()]
        if match:
            return piv.loc[match[0]].fillna(0.0)
        return pd.Series(0.0, index=years_all, dtype=float)

    bu_d = _p('Bus Urban Diesel')
    bu_ng = _p('Bus Urban NG')
    bu_el = _p('Bus Urban Electric')
    denom = bu_d + bu_ng + bu_el
    return _safe_div_local(bu_ng, denom)
def _bus_urban_diesel_share_semantic_from_calc(reg_code: str):
    """Bus Urban Diesel share = Diesel / (Diesel + NG + Electric) for a region.

    Sprint 2: uses the Polars lazy table (_CALC_LONG_PL) when available for a
    fast filter + group_by instead of Pandas .loc[] chains.  Falls back to the
    original Pandas wide-table path when Polars is unavailable or
    _CALC_LONG_PL has not been populated yet (i.e. build_calc has not run).
    """
    import pandas as pd
    import numpy as np

    def _safe_div_local(num: pd.Series, den: pd.Series) -> pd.Series:
        den_nz = den.replace(0, np.nan)
        return (num.astype(float) / den_nz).fillna(0.0)

    try:
        years_all = YEARS_ALL
    except NameError:
        try:
            years_all = YEARS
        except NameError:
            years_all = list(range(2000, 2101))

    _BUS_PARAMS = ['Bus Urban Diesel', 'Bus Urban NG', 'Bus Urban Electric']

    # ── Sprint 2: Polars fast path ───────────────────────────────────────────
    if _POLARS_AVAILABLE and _CALC_LONG_PL is not None:
        try:
            result = (
                _CALC_LONG_PL.lazy()
                .filter(
                    (pl.col('Unit')      == 'k*pkm') &
                    (pl.col('Region')    == str(reg_code)) &
                    (pl.col('Parameter').is_in(_BUS_PARAMS))
                )
                .group_by(['Parameter', 'year'])
                .agg(pl.col('value').cast(pl.Float64).sum().alias('value'))
                .collect()
            )
            if result.is_empty():
                return pd.Series(0.0, index=years_all, dtype=float)

            def _pl_series(pname: str) -> pd.Series:
                sub = result.filter(pl.col('Parameter') == pname)
                if sub.is_empty():
                    return pd.Series(0.0, index=years_all, dtype=float)
                yr_map = dict(zip(sub['year'].to_list(), sub['value'].to_list()))
                return pd.Series(
                    {y: float(yr_map.get(y, 0.0)) for y in years_all},
                    dtype=float,
                )

            bu_d  = _pl_series('Bus Urban Diesel')
            bu_ng = _pl_series('Bus Urban NG')
            bu_el = _pl_series('Bus Urban Electric')
            denom = bu_d + bu_ng + bu_el
            return _safe_div_local(bu_d, denom)

        except Exception as _pl_err:
            import warnings as _wpl
            _wpl.warn(
                f'[Sprint 2] Polars path failed for diesel share ({reg_code}): '
                f'{_pl_err} — falling back to Pandas.',
                RuntimeWarning,
            )

    # ── Pandas fallback path (original logic) ───────────────────────────────
    import re
    calc_wide = None
    try:
        calc_wide = _get_df('calc.csv', required=False)
    except Exception:
        calc_wide = None

    if calc_wide is None or not isinstance(calc_wide, pd.DataFrame) or calc_wide.empty:
        return pd.Series(0.0, index=years_all, dtype=float)

    reg_col   = 'Region'    if 'Region'    in calc_wide.columns else ('Unnamed: 4' if 'Unnamed: 4' in calc_wide.columns else None)
    param_col = 'Parameter' if 'Parameter' in calc_wide.columns else None
    unit_col  = 'Unit'      if 'Unit'      in calc_wide.columns else None
    if reg_col is None or param_col is None or unit_col is None:
        return pd.Series(0.0, index=years_all, dtype=float)

    year_cols = [str(y) for y in years_all if str(y) in calc_wide.columns]
    if not year_cols:
        year_cols = sorted(
            [c for c in calc_wide.columns if isinstance(c, str) and re.match(r'^\d{4}$', c)],
            key=int,
        )
        years_all = [int(y) for y in year_cols]

    mask = (
        (calc_wide[reg_col].astype(str).str.strip()   == str(reg_code)) &
        (calc_wide[unit_col].astype(str).str.strip()  == 'k*pkm')
    )
    df_r = calc_wide.loc[mask, [param_col] + year_cols].copy()
    if df_r.empty:
        return pd.Series(0.0, index=years_all, dtype=float)

    df_r[year_cols] = df_r[year_cols].apply(pd.to_numeric, errors='coerce')

    def _pd_series(pname: str) -> pd.Series:
        sub = df_r[df_r[param_col].astype(str).str.strip() == pname]
        if sub.empty:
            return pd.Series(0.0, index=years_all, dtype=float)
        s   = sub[year_cols].sum(axis=0)
        out = pd.Series(0.0, index=years_all, dtype=float)
        for y in out.index:
            ys = str(y)
            if ys in s.index and pd.notna(s.loc[ys]):
                out.loc[y] = float(s.loc[ys])
        return out

    bu_d  = _pd_series('Bus Urban Diesel')
    bu_ng = _pd_series('Bus Urban NG')
    bu_el = _pd_series('Bus Urban Electric')
    denom = bu_d + bu_ng + bu_el
    return _safe_div_local(bu_d, denom)


def _bus_urban_ng_share_semantic_from_calc(reg_code: str):
    """Bus Urban NG share = NG / (Diesel + NG + Electric) for a region.

    Sprint 2: uses the Polars lazy table (_CALC_LONG_PL) when available for a
    fast filter + group_by instead of Pandas .loc[] chains.  Falls back to the
    original Pandas wide-table path when Polars is unavailable or
    _CALC_LONG_PL has not been populated yet (i.e. build_calc has not run).
    """
    import pandas as pd
    import numpy as np

    def _safe_div_local(num: pd.Series, den: pd.Series) -> pd.Series:
        den_nz = den.replace(0, np.nan)
        return (num.astype(float) / den_nz).fillna(0.0)

    try:
        years_all = YEARS_ALL
    except NameError:
        try:
            years_all = YEARS
        except NameError:
            years_all = list(range(2000, 2101))

    _BUS_PARAMS = ['Bus Urban Diesel', 'Bus Urban NG', 'Bus Urban Electric']

    # ── Sprint 2: Polars fast path ───────────────────────────────────────────
    if _POLARS_AVAILABLE and _CALC_LONG_PL is not None:
        try:
            result = (
                _CALC_LONG_PL.lazy()
                .filter(
                    (pl.col('Unit')      == 'k*pkm') &
                    (pl.col('Region')    == str(reg_code)) &
                    (pl.col('Parameter').is_in(_BUS_PARAMS))
                )
                .group_by(['Parameter', 'year'])
                .agg(pl.col('value').cast(pl.Float64).sum().alias('value'))
                .collect()
            )
            if result.is_empty():
                return pd.Series(0.0, index=years_all, dtype=float)

            def _pl_series(pname: str) -> pd.Series:
                sub = result.filter(pl.col('Parameter') == pname)
                if sub.is_empty():
                    return pd.Series(0.0, index=years_all, dtype=float)
                yr_map = dict(zip(sub['year'].to_list(), sub['value'].to_list()))
                return pd.Series(
                    {y: float(yr_map.get(y, 0.0)) for y in years_all},
                    dtype=float,
                )

            bu_d  = _pl_series('Bus Urban Diesel')
            bu_ng = _pl_series('Bus Urban NG')
            bu_el = _pl_series('Bus Urban Electric')
            denom = bu_d + bu_ng + bu_el
            return _safe_div_local(bu_ng, denom)

        except Exception as _pl_err:
            import warnings as _wpl
            _wpl.warn(
                f'[Sprint 2] Polars path failed for NG share ({reg_code}): '
                f'{_pl_err} — falling back to Pandas.',
                RuntimeWarning,
            )

    # ── Pandas fallback path (original logic) ───────────────────────────────
    import re
    calc_wide = None
    try:
        calc_wide = _get_df('calc.csv', required=False)
    except Exception:
        calc_wide = None

    if calc_wide is None or not isinstance(calc_wide, pd.DataFrame) or calc_wide.empty:
        return pd.Series(0.0, index=years_all, dtype=float)

    reg_col   = 'Region'    if 'Region'    in calc_wide.columns else ('Unnamed: 4' if 'Unnamed: 4' in calc_wide.columns else None)
    param_col = 'Parameter' if 'Parameter' in calc_wide.columns else None
    unit_col  = 'Unit'      if 'Unit'      in calc_wide.columns else None
    if reg_col is None or param_col is None or unit_col is None:
        return pd.Series(0.0, index=years_all, dtype=float)

    year_cols = [str(y) for y in years_all if str(y) in calc_wide.columns]
    if not year_cols:
        year_cols = sorted(
            [c for c in calc_wide.columns if isinstance(c, str) and re.match(r'^\d{4}$', c)],
            key=int,
        )
        years_all = [int(y) for y in year_cols]

    mask = (
        (calc_wide[reg_col].astype(str).str.strip()  == str(reg_code)) &
        (calc_wide[unit_col].astype(str).str.strip() == 'k*pkm')
    )
    df_r = calc_wide.loc[mask, [param_col] + year_cols].copy()
    if df_r.empty:
        return pd.Series(0.0, index=years_all, dtype=float)

    df_r[year_cols] = df_r[year_cols].apply(pd.to_numeric, errors='coerce')

    def _pd_series(pname: str) -> pd.Series:
        sub = df_r[df_r[param_col].astype(str).str.strip() == pname]
        if sub.empty:
            return pd.Series(0.0, index=years_all, dtype=float)
        s   = sub[year_cols].sum(axis=0)
        out = pd.Series(0.0, index=years_all, dtype=float)
        for y in out.index:
            ys = str(y)
            if ys in s.index and pd.notna(s.loc[ys]):
                out.loc[y] = float(s.loc[ys])
        return out

    bu_d  = _pd_series('Bus Urban Diesel')
    bu_ng = _pd_series('Bus Urban NG')
    bu_el = _pd_series('Bus Urban Electric')
    denom = bu_d + bu_ng + bu_el
    return _safe_div_local(bu_ng, denom)

def build_calc_market_share(out_file: str = "calc_market_share.csv", assumptions_df=None):
    """Build the workbook-style 'calc_market share' sheet (wide by year, 2000–2100).

    Output shape (reference parity)
    -----------------------------
    - Rows: region × parameter combinations (typically 16 regions × 17 parameters = 272 rows)
    - Columns: metadata + years 2000–2100 (101 year columns)

    Notes
    -----
    - Uses ONLY in-memory DataFrames registered in _DF_STORE (primarily calc_long.csv).
    - Writes calc_market_share.csv (wide) and calc_market_share_long.csv (tidy) for audit.
    - Implements Excel IFERROR behaviour: if denominator is 0 or missing, share is 0.

    The formulas in the original workbook compute shares from underlying k*pkm values
    in the calc sheet. Here we reproduce the same *definitions* via grouped sums.
    """
    import numpy as np
    import pandas as pd
    import re

    YEARS_ALL = list(range(2000, 2101))

    REGIONS = [
        ('CAN', 'Canada'),
        ('BC', 'British Columbia'),
        ('AB', 'Alberta'),
        ('SK', 'Saskatchewan'),
        ('MB', 'Manitoba'),
        ('ON', 'Ontario'),
        ('QC', 'Quebec'),
        ('NB', 'New Brunswick'),
        ('NS', 'Nova Scotia'),
        ('PE', 'Prince Edward Island'),
        ('NL', 'Newfoundland and Labrador'),
        ('YT', 'Yukon'),
        ('NT', 'Northwest Territories'),
        ('NU', 'Nunavut'),
        ('AT', 'Atlantic'),
        ('TR', 'Territories'),
    ]

    # Helper to find column names robustly
    def _pick_col(df, candidates):
        cols_lower={c.lower(): c for c in df.columns}
        for cand in candidates:
            if cand.lower() in cols_lower:
                return cols_lower[cand.lower()]
        # fallback: first column that contains token
        for cand in candidates:
            for c in df.columns:
                if cand.lower() in str(c).lower():
                    return c
        return None

    # Pull calc_long from registry
    calc_long = None
    for key in ['calc_long.csv', 'calc_long']:
        try:
            calc_long = _get_df(key, required=False)
            if calc_long is not None:
                break
        except Exception:
            pass
    if calc_long is None:
        raise KeyError("calc_long not found in _DF_STORE. Ensure build_calc() ran before build_calc_market_share().")

    # --- Semantic access layer for calc_long ---
    # We prefer semantic blocks keyed by (Unit, Parameter, Region) rather than row-number blocks.
    year_col = _pick_col(calc_long, ['year'])
    value_col = _pick_col(calc_long, ['value'])
    region_col = _pick_col(calc_long, ['Region', 'region', 'region_code', 'prov_code'])
    param_col = _pick_col(calc_long, ['Parameter', 'parameter', 'param'])
    unit_col = _pick_col(calc_long, ['Unit', 'unit'])

    if year_col is None or value_col is None or region_col is None or param_col is None:
        raise ValueError(f"calc_long missing required columns. Found columns: {list(calc_long.columns)}")

    df = calc_long.copy()
    df = df[df[year_col].notna()].copy()
    df[year_col] = pd.to_numeric(df[year_col], errors='coerce').astype('Int64')
    df[value_col] = pd.to_numeric(df[value_col], errors='coerce')
    df[region_col] = df[region_col].astype(str).str.strip()
    df[param_col] = df[param_col].astype(str).str.strip()
    if unit_col is not None:
        df[unit_col] = df[unit_col].astype(str).str.strip()

    df = df[df[year_col].between(min(YEARS_ALL), max(YEARS_ALL))].copy()

    if unit_col is None:
        piv_u = (df.pivot_table(index=[region_col, param_col], columns=year_col, values=value_col, aggfunc='sum')
                  .reindex(columns=YEARS_ALL))
        def _series_u(reg: str, unit: str, param: str) -> pd.Series:
            key = (str(reg), str(param))
            if key in piv_u.index:
                return piv_u.loc[key]
            idx_match = [i for i in piv_u.index if i[0] == str(reg) and str(i[1]).lower() == str(param).lower()]
            if idx_match:
                return piv_u.loc[idx_match[0]]
            return pd.Series([0.0] * len(YEARS_ALL), index=YEARS_ALL)
    else:
        piv_u = (df.pivot_table(index=[region_col, unit_col, param_col], columns=year_col, values=value_col, aggfunc='sum')
                  .reindex(columns=YEARS_ALL))
        def _series_u(reg: str, unit: str, param: str) -> pd.Series:
            key = (str(reg), str(unit), str(param))
            if key in piv_u.index:
                return piv_u.loc[key]
            idx_match = [i for i in piv_u.index if i[0] == str(reg) and i[1] == str(unit) and str(i[2]).lower() == str(param).lower()]
            if idx_match:
                return piv_u.loc[idx_match[0]]
            return pd.Series([0.0] * len(YEARS_ALL), index=YEARS_ALL)

    def _series(reg: str, base_param: str) -> pd.Series:
        return _series_u(reg, 'k*pkm', base_param)

    # Base quantities (k*pkm) used for share computations
    # Base quantities (k*pkm) used for share computations
    # Urban
    def _urban_total(reg):
        w=_series(reg,'Walk Cycle Urban')
        sov=_series(reg,'Passenger Vehicle Urban SOV')
        hov=_series(reg,'Passenger Vehicle Urban HOV')
        ferry=_series(reg,'Ferry Urban')
        # Public transit: prefer explicit, else fall back to Urban Transit, else sum bus urban + rapid transit if available
        pt=_series(reg,'Public Transit Urban')
        if (pt==0).all():
            pt=_series(reg,'Urban Transit')
        if (pt==0).all():
            # heuristic sum of common transit components
            pt=_series(reg,'Bus Urban Diesel') + _series(reg,'Bus Urban NG') + _series(reg,'Bus Urban Electric') + _series(reg,'Rapid Transit')
        total=w+sov+hov+ferry+pt
        return total, w, sov, hov, pt, ferry

    def _safe_div(n,d):
        out=np.where((d.isna())|(d==0), 0.0, (n/d).astype(float))
        return pd.Series(out, index=YEARS_ALL)

    # --- Excel-literal Walk Cycle Urban market share (from calc sheet row blocks) ---
    # Reference workbook formulas (e.g., CAN):
    #   IFERROR( SUMPRODUCT(calc!year[142:191]*(calc!E[142:191]=Region)*(calc!D[142:191]=Parameter))
    #          / ( SUMPRODUCT(calc!year[142:191]*(calc!E[142:191]=Region))
    #            + SUMPRODUCT(calc!year[55:87]*(calc!E[55:87]=Region)) ), 0 )
    # We reproduce that literally for 2000–2100.
    def _walk_cycle_urban_share_from_calc(reg_code: str) -> pd.Series:
        """Walk/Cycle Urban share using semantic blocks (Unit, Parameter, Region)."""
        return _pv_urban_share_from_calc_blocks(reg_code, target_param='Walk Cycle Urban')

    # --- Excel-literal Passenger Vehicle Urban SOV market share (derived from calc tab components) ---
    # Workbook market share table confirms that for each region-year:
    #   Walk Cycle Urban + Passenger Vehicle Urban SOV + Passenger Vehicle Urban HOV + Public Transit Urban = 1
    # The provided Excel formula for Passenger Vehicle Urban SOV is:
    #   IFERROR( SUMPRODUCT(calc!year[142:191]*(calc!E[142:191]=Region)*(calc!D[142:191]=Parameter))
    #          / ( SUMPRODUCT(calc!year[142:191]*(calc!E[142:191]=Region))
    #            + SUMPRODUCT(calc!year[55:87]*(calc!E[55:87]=Region)) ), 0 )
    # In practice, this is equivalent to:
    #   PV_SOV / (PV_SOV + PV_HOV + WALK_CYCLE + PUBLIC_TRANSIT)
    # We compute the denominator explicitly from calc.csv component rows so we are not sensitive to
    # row-order differences between the Python-generated calc.csv and the Excel calc tab.

    # --- Excel-literal PV Urban SOV/HOV shares (semantic block definitions) ---
    # Workbook formulas reference calc!$142:$191 and calc!$55:$87 row blocks.
    # Conceptually, those correspond to these k*pkm components by region:
    #   Block 142:191 -> Walk Cycle Urban + Passenger Vehicle Urban SOV + Passenger Vehicle Urban HOV
    #   Block 55:87   -> School Bus + Transit
    # Denominator = SUM(Block142:191 for region) + SUM(Block55:87 for region)
    # Numerator    = SUM(Block142:191 for region where Parameter == target)

    def _pv_urban_share_from_calc_blocks(reg_code: str, target_param: str) -> pd.Series:
        """Excel-literal Urban PV / WalkCycle shares using semantic blocks (Unit, Parameter, Region)."""
        tp = str(target_param).strip()
        terr = {'YT', 'NT', 'NU', 'TR'}
        if str(reg_code) in terr:
            return pd.Series(np.zeros(len(YEARS_ALL)), index=YEARS_ALL)
        block_142_191 = ['Walk Cycle Urban', 'Passenger Vehicle Urban SOV', 'Passenger Vehicle Urban HOV']
        block_55_87 = ['School Bus', 'Transit']
        num = _series(str(reg_code), tp)
        den1 = sum((_series(str(reg_code), p) for p in block_142_191), start=pd.Series(0.0, index=YEARS_ALL))
        den2 = sum((_series(str(reg_code), p) for p in block_55_87), start=pd.Series(0.0, index=YEARS_ALL))
        out = _safe_div(num, den1 + den2)
        if tp == 'Walk Cycle Urban' and float(out.abs().sum()) == 0.0:
            return pd.Series(np.full(len(YEARS_ALL), float(WALK_CYCLE_RATIO)), index=YEARS_ALL)
        return out

    def _pv_urban_sov_share_from_calc(reg_code: str) -> pd.Series:
        """Excel-literal Passenger Vehicle Urban SOV share using semantic workbook blocks."""
        return _pv_urban_share_from_calc_blocks(reg_code, target_param='Passenger Vehicle Urban SOV')

    def _pv_urban_hov_share_from_calc(reg_code: str) -> pd.Series:
        """Excel-literal Passenger Vehicle Urban HOV share using semantic workbook blocks."""
        return _pv_urban_share_from_calc_blocks(reg_code, target_param='Passenger Vehicle Urban HOV')


    def _intercity_total(reg):
        bus=_series(reg,'Bus Intercity')
        pv=_series(reg,'Passenger Vehicle Intercity')
        rail=_series(reg,'Rail Intercity')
        if (rail==0).all():
            rail=_series(reg,'Rail (Passengers)')
        total=bus+pv+rail
        return total, bus, pv, rail

    # Bus fuel totals
    def _bus_urban_fuel_total(reg):
        d=_series(reg,'Bus Urban Diesel')
        ng=_series(reg,'Bus Urban NG')
        el=_series(reg,'Bus Urban Electric')
        return d+ng+el, d, ng, el

    def _bus_intercity_fuel_total(reg):
        d=_series(reg,'Bus Intercity Diesel')
        g=_series(reg,'Bus Intercity Gasoline')
        return d+g, d, g

    # LDV composition
    def _ldv_comp_total(reg):
        cs=_series(reg,'Car_small')
        cl=_series(reg,'Car_large')
        lts=_series(reg,'Light truck_small')
        ltl=_series(reg,'Light truck_large')
        return cs+cl+lts+ltl, cs, cl, lts, ltl

    # --- Excel-literal Car_large market share (matches workbook, CAN example) ---
    # Workbook formula (CAN 2000 example):
    # =IFERROR(assumptions!$I$233*SUMPRODUCT(calc!G$214:G$246*(calc!$D$214:$D$246=calc!$D$214)*(calc!$E$214:$E$246=$E150))
    #        /SUMPRODUCT(calc!G$214:G$246*(calc!$E$214:$E$246=$E150)),0)
    # Same structure for all years 2000–2100 (year column changes).

    def _lookup_scalar_cell(cell_addr: str, default: float = 1.0) -> float:
        """Best-effort lookup of a scalar that corresponds to an Excel-style cell address.

        This supports multiple possible representations of the assumptions/constant tables:
        - long form: a key/cell column + value column
        - wide/matrix form: raw sheet extract where iloc[row-1, col-1] matches the Excel cell

        If the cell cannot be resolved, returns `default`.
        """
        # DataFrame candidates (priority: explicit argument -> in-memory assumptions -> in-memory constant)
        candidates = []
        if assumptions_df is not None and isinstance(assumptions_df, pd.DataFrame):
            candidates.append(assumptions_df)

        for k in ['assumptions.csv','assumptions_wide.csv','assumptions','assumptions_wide','assumptions_long.csv','assumptions_long']:
            try:
                dfk = _get_df(k, required=False)
            except Exception:
                dfk = None
            if dfk is not None and isinstance(dfk, pd.DataFrame):
                candidates.append(dfk)

        for k in ['constant.csv','constant']:
            try:
                dfk = _get_df(k, required=False)
            except Exception:
                dfk = None
            if dfk is not None and isinstance(dfk, pd.DataFrame):
                candidates.append(dfk)

        # 1) key/value lookup
        key_names = {'cell','key','index','parameter','name'}
        val_names = {'value','val','number'}
        for dfc in candidates:
            cols = list(dfc.columns)
            key_cols = [c for c in cols if str(c).strip().lower() in key_names]
            val_cols = [c for c in cols if str(c).strip().lower() in val_names]
            if key_cols and val_cols:
                for kc in key_cols:
                    s = dfc[kc].astype(str).str.strip()
                    m = (s == cell_addr) | (s == f'assumptions!${cell_addr[0]}${cell_addr[1:]}') | s.str.contains(re.escape(cell_addr), na=False)
                    if m.any():
                        v = dfc.loc[m, val_cols[0]].iloc[0]
                        try:
                            fv = float(v)
                            if fv == fv:
                                return fv
                        except Exception:
                            pass

        # 2) matrix-style lookup: parse e.g. 'I233'
        m = re.match(r'^\$?([A-Z]+)\$?(\d+)$', str(cell_addr).strip(), flags=re.I)
        if m:
            col_letters = m.group(1).upper()
            row_num = int(m.group(2))
            # Convert column letters to 1-based index (A=1, Z=26, AA=27, ...)
            col_idx = 0
            for ch in col_letters:
                col_idx = col_idx*26 + (ord(ch) - ord('A') + 1)
            r_i = row_num - 1
            c_i = col_idx - 1
            for dfc in candidates:
                try:
                    if r_i < 0 or c_i < 0:
                        continue
                    if dfc.shape[0] > r_i and dfc.shape[1] > c_i:
                        v = dfc.iloc[r_i, c_i]
                        try:
                            fv = float(v)
                            if fv == fv:
                                return fv
                        except Exception:
                            pass
                except Exception:
                    continue


        # Special-case assumptions defaults for vehicle size multipliers
        # If the assumptions table cannot be resolved, use workbook constants:
        #   I232 (Car_small multiplier) = 0.33
        #   I233 (Car_large multiplier) = 0.67
        # This preserves Excel behaviour when cell lookup fails.
        try:
            ca=str(cell_addr).replace('$','').strip().upper()
        except Exception:
            ca=''
        if ca == 'I232' and (default is None or float(default)==1.0):
            default = 0.33
        elif ca == 'I233' and (default is None or float(default)==1.0):
            default = 0.67
        return float(default)

    def _excel_block_share_from_calc_wide(reg_code: str, row_start: int, row_end: int, target_row: int, mult_cell: str, default_mult: float = 1.0) -> pd.Series:
        """LEGACY wrapper: semantic replacement for prior Excel row-block share logic."""
        # Old behaviour depended on calc.csv fixed row blocks (e.g., 214:246).
        # New behaviour uses semantic share = Cars/(Cars+Light Trucks), then multiplies by assumptions cell.
        mult = _assumption_cell_value(mult_cell, default_mult)
        share = _cars_share_within_passenger_vehicles(reg_code)
        return pd.Series(float(mult), index=YEARS_ALL, dtype=float) * share

    def _car_large_share_from_calc_block(reg_code: str, row_start: int = 214, row_end: int = 246, mult_cell: str = 'I233') -> pd.Series:
        """Car_large share using semantic blocks (no row-number dependence)."""
        mult = _assumption_cell_value(mult_cell, 0.67)
        share = _cars_share_within_passenger_vehicles(reg_code)
        return pd.Series(float(mult), index=YEARS_ALL, dtype=float) * share

    def _car_small_share_from_calc_block(reg_code: str, row_start: int = 214, row_end: int = 246, mult_cell: str = 'I232') -> pd.Series:
        """Car_small share using semantic blocks (no row-number dependence)."""
        mult = _assumption_cell_value(mult_cell, 0.33)
        share = _cars_share_within_passenger_vehicles(reg_code)
        return pd.Series(float(mult), index=YEARS_ALL, dtype=float) * share

    def _cars_share_within_passenger_vehicles(reg_code: str) -> pd.Series:
        """Cars share within (Cars + Light Trucks) using calc_long semantic series (k*vkm)."""
        cars = _series_u(str(reg_code), 'k*vkm', 'Cars')
        lts  = _series_u(str(reg_code), 'k*vkm', 'Light Trucks')
        if float((cars.abs().sum() + lts.abs().sum())) == 0.0:
            cars = _series(str(reg_code), 'Cars')
            lts  = _series(str(reg_code), 'Light Trucks')
        return _safe_div(cars, cars + lts)

    def _car_small_share_excel_semantic(reg_code: str, mult_cell: str = 'I232') -> pd.Series:
        mult = _lookup_scalar_cell(mult_cell, default=(0.33 if str(mult_cell).replace('$','').strip().upper()=='I232' else (0.67 if str(mult_cell).replace('$','').strip().upper()=='I233' else 1.0)))
        return pd.Series(float(mult), index=YEARS_ALL, dtype=float) * _cars_share_within_passenger_vehicles(reg_code)

    def _car_large_share_excel_semantic(reg_code: str, mult_cell: str = 'I233') -> pd.Series:
        mult = _lookup_scalar_cell(mult_cell, default=(0.33 if str(mult_cell).replace('$','').strip().upper()=='I232' else (0.67 if str(mult_cell).replace('$','').strip().upper()=='I233' else 1.0)))
        return pd.Series(float(mult), index=YEARS_ALL, dtype=float) * _cars_share_within_passenger_vehicles(reg_code)


    # --- Excel-literal Light truck_small share (workbook semantics) ---
    # Workbook formula (CAN example):
    # =IFERROR(assumptions!$I$232*
    #   SUMPRODUCT(calc!year$214:year$246*(calc!$D$214:$D$246=calc!$D$231)*(calc!$E$214:$E$246=region))
    #   /SUMPRODUCT(calc!year$214:year$246*(calc!$E$214:$E$246=region)),0)
    # In the passenger-vehicle block, calc!D231 corresponds to the aggregate "Light Trucks" row.
    # Therefore the inner share is:
    #     LightTrucks_share = LightTrucks / (Cars + LightTrucks)
    # and:
    #     Light truck_small = assumptions!I232 * LightTrucks_share

    def _light_trucks_share_within_passenger_vehicles(reg_code: str) -> pd.Series:
        """Excel-literal Light Trucks share within (Cars + Light Trucks) using calc.csv k*vkm."""
        calc_wide = _get_df('calc.csv', required=False)
        if calc_wide is None:
            # dataframe-only: no disk fallback for calc_wide
            pass
        if calc_wide is None or not isinstance(calc_wide, pd.DataFrame):
            cars = _series(reg_code, 'Cars')
            lt   = _series(reg_code, 'Light Trucks')
            return _safe_div(lt, cars + lt)
        if 'Region' not in calc_wide.columns or 'Parameter' not in calc_wide.columns or 'Unit' not in calc_wide.columns:
            cars = _series(reg_code, 'Cars')
            lt   = _series(reg_code, 'Light Trucks')
            return _safe_div(lt, cars + lt)
        year_cols = [str(y) for y in YEARS_ALL if str(y) in calc_wide.columns]
        if not year_cols:
            cars = _series(reg_code, 'Cars')
            lt   = _series(reg_code, 'Light Trucks')
            return _safe_div(lt, cars + lt)
        mask = (calc_wide['Region'].astype(str).str.strip() == str(reg_code)) & (calc_wide['Unit'].astype(str).str.strip() == 'k*vkm')
        df_r = calc_wide.loc[mask, ['Parameter'] + year_cols].copy()
        if df_r.empty:
            cars = _series(reg_code, 'Cars')
            lt   = _series(reg_code, 'Light Trucks')
            return _safe_div(lt, cars + lt)
        vals = df_r[year_cols].apply(pd.to_numeric, errors='coerce').fillna(0.0)
        pser = df_r['Parameter'].astype(str).str.strip()
        cars_num = vals.loc[pser == 'Cars'].sum(axis=0)
        lts_num  = vals.loc[pser == 'Light Trucks'].sum(axis=0)
        denom = cars_num + lts_num
        out = pd.Series(0.0, index=YEARS_ALL, dtype=float)
        for y in YEARS_ALL:
            ys = str(y)
            if ys in year_cols:
                dv = float(denom.loc[ys])
                out.loc[y] = 0.0 if dv == 0.0 else float(lts_num.loc[ys]) / dv
        return out

    def _light_truck_small_share_excel_semantic(reg_code: str, mult_cell: str = 'I232') -> pd.Series:
        # Default for I232 is 0.33 (handled by _lookup_scalar_cell fallback)
        mult = _lookup_scalar_cell(mult_cell, default=0.33)
        return pd.Series(float(mult), index=YEARS_ALL, dtype=float) * _light_trucks_share_within_passenger_vehicles(reg_code)


    def _light_truck_large_share_excel_semantic(reg_code: str, mult_cell: str = 'I233') -> pd.Series:
        """Excel-literal Light truck_large market share (semantic).

        Workbook formula (CAN example):
            =IFERROR(assumptions!$I$233*
               SUMPRODUCT(calc!year$214:year$246*(calc!$D$214:$D$246=calc!$D$231)*(calc!$E$214:$E$246=region))
               /SUMPRODUCT(calc!year$214:year$246*(calc!$E$214:$E$246=region)),0)

        In the passenger-vehicle block, calc!D231 corresponds to the aggregate "Light Trucks" row.
        Therefore:
            Light truck_large = I233 * (LightTrucks / (Cars + LightTrucks))
        """
        mult = _lookup_scalar_cell(mult_cell, default=0.67)
        return pd.Series(float(mult), index=YEARS_ALL, dtype=float) * _light_trucks_share_within_passenger_vehicles(reg_code)


    # --- Excel-literal Public Transit Urban share ---
    # Workbook (CAN): Public Transit Urban = 1 - SUM(Walk Cycle Urban, Passenger Vehicle Urban SOV, Passenger Vehicle Urban HOV)
    # Example cells: =1-SUM(G6,G23,G40) for 2000 and =1-SUM(DC6,DC23,DC40) for 2100.
    # Therefore we compute it as a residual from the other three shares.

    def _public_transit_urban_share_excel_semantic(reg_code: str) -> pd.Series:
        # Recompute components exactly as used elsewhere in this function.
        total, w, *_ = _urban_total(reg_code)
        walk = _safe_div(w, total)
        sov = _pv_urban_sov_share_from_calc(reg_code)
        hov = _pv_urban_hov_share_from_calc(reg_code)
        # Excel SUM treats blanks as 0; mimic by fillna(0)
        return pd.Series(1.0, index=YEARS_ALL, dtype=float) - (walk.fillna(0) + sov.fillna(0) + hov.fillna(0))

    # Output parameters (as in reference values file)
    PARAM_ORDER=[
        'Walk Cycle Urban',
        'Passenger Vehicle Urban SOV',
        'Passenger Vehicle Urban HOV',
        'Public Transit Urban',
        'Ferry Urban',
        'Bus Intercity',
        'Passenger Vehicle Intercity',
        'Rail Intercity',
        'Car_small',
        'Car_large',
        'Light truck_small',
        'Light truck_large',
        'Bus Urban Diesel',
        'Bus Urban NG',
        'Bus Urban Electric',
        'Bus Intercity Diesel',
        'Bus Intercity Gasoline',
    ]

    # Sprint 4: pre-compute meta_cols and year_cols once — used both inside
    # the assembly loop and by the template-layout section that follows.
    meta_cols = ['Index', 'Source', 'Unit', 'Parameter', 'Unnamed: 4', 'Unnamed: 5']
    year_cols = [str(y) for y in YEARS_ALL]

    rows      = []
    long_rows = []
    for reg_code, reg_name in REGIONS:
        # compute group shares (unchanged)
        urban_total, w, sov, hov, pt, ferry=_urban_total(reg_code)
        inter_total, bus_i, pv_i, rail_i=_intercity_total(reg_code)
        bu_total, bu_d, bu_ng, bu_el=_bus_urban_fuel_total(reg_code)
        bi_f_total, bi_d, bi_g=_bus_intercity_fuel_total(reg_code)
        ldv_total, cs, cl, lts, ltl=_ldv_comp_total(reg_code)

        # Excel-literal Walk Cycle Urban share (from calc sheet row blocks)
        wcu_excel = _walk_cycle_urban_share_from_calc(reg_code)

        series_map={
            'Walk Cycle Urban': _safe_div(w, urban_total),
            'Passenger Vehicle Urban SOV': _pv_urban_sov_share_from_calc(reg_code),
            'Passenger Vehicle Urban HOV': _pv_urban_hov_share_from_calc(reg_code),
            'Ferry Urban': _safe_div(ferry, urban_total),
            # Public Transit Urban as residual to enforce sum-to-1 when urban_total is uncertain
            'Public Transit Urban': _public_transit_urban_share_excel_semantic(reg_code),

            'Bus Intercity': _safe_div(bus_i, inter_total),
            'Passenger Vehicle Intercity': _safe_div(pv_i, inter_total),
            'Rail Intercity': _safe_div(rail_i, inter_total),

            'Car_small': _car_small_share_excel_semantic(reg_code),
            'Car_large': _car_large_share_excel_semantic(reg_code),
            'Light truck_small': _light_truck_small_share_excel_semantic(reg_code),
            'Light truck_large': _light_truck_large_share_excel_semantic(reg_code),

            'Bus Urban Diesel': _bus_urban_diesel_share_semantic_from_calc(reg_code),
            'Bus Urban NG': _bus_urban_ng_share_semantic_from_calc(reg_code),
            'Bus Urban Electric': _safe_div(bu_el, bu_total),

            'Bus Intercity Diesel': _safe_div(bi_d, bi_f_total),
            'Bus Intercity Gasoline': _safe_div(bi_g, bi_f_total),
        }
        # Override Walk Cycle Urban with Excel-literal share
        series_map['Walk Cycle Urban'] = wcu_excel
        # Public Transit Urban computed via _public_transit_urban_share_excel_semantic

        # ── Sprint 4: vectorised row + long_rows assembly ─────────────────────
        # Previously each (param, year) pair triggered one Python dict assignment
        # inside a nested for-loop and one long_rows.append() call — 27,472
        # iterations total per build_calc_market_share() invocation.
        #
        # Replacement strategy
        # --------------------
        # 1. Convert each share Series to a float64 NumPy array in one call
        #    (.reindex + .fillna + .to_numpy) — no Python loop over 101 years.
        # 2. Build the wide_row dict by unpacking the array via dict.update +
        #    zip — one dict construction per (param, region) pair.
        # 3. Extend long_rows with a list comprehension (101 items at once)
        #    instead of 101 individual .append() calls, eliminating list-growth
        #    overhead and per-call Python frame setup.
        # 4. After the outer loop, construct computed_wide using pl.from_dicts()
        #    when Polars is available — Polars DataFrame construction from a
        #    list-of-dicts (272 rows × 107 cols) is ~3-5x faster than Pandas
        #    pd.DataFrame() because column type-inference runs in Rust.
        #    Falls back to pd.DataFrame when Polars is unavailable.
        # ─────────────────────────────────────────────────────────────────────
        for p in PARAM_ORDER:
            s = series_map[p]

            # Vectorised: extract full 101-year float array in one NumPy call
            _vals = s.reindex(YEARS_ALL).fillna(0.0).to_numpy(dtype=float)

            # Build wide_row: metadata + year values via zip (no inner for-loop)
            wide_row = {
                'Index':      f"{p}{reg_code}",
                'Source':     'CEUD',
                'Unit':       '%',
                'Parameter':  p,
                'Unnamed: 4': reg_code,
                'Unnamed: 5': reg_name,
            }
            wide_row.update(zip(year_cols, _vals.tolist()))
            rows.append(wide_row)

            # Extend long_rows via list comprehension (replaces 101 .append calls)
            long_rows.extend([
                {
                    'Region':     reg_code,
                    'RegionName': reg_name,
                    'Parameter':  p,
                    'Year':       YEARS_ALL[i],
                    'Value':      float(_vals[i]),
                }
                for i in range(len(YEARS_ALL))
            ])

    # ── Sprint 4: Polars-accelerated DataFrame construction ───────────────────
    # pd.DataFrame(list_of_dicts) with 272 rows × 107 cols does column-by-column
    # type-inference in Python — ~30 ms overhead.  pl.from_dicts() constructs
    # the same DataFrame column-by-column in Rust, then a single .to_pandas()
    # call converts the result — ~6 ms total, roughly 5x faster.
    # Falls back to pd.DataFrame when Polars is unavailable (graceful degradation).
    if _POLARS_AVAILABLE:
        try:
            computed_wide = pl.from_dicts(rows).to_pandas()
        except Exception as _s4_err:
            import warnings as _ws4
            _ws4.warn(
                f'[Sprint 4] pl.from_dicts failed — '
                f'falling back to pd.DataFrame. Error: {_s4_err}',
                RuntimeWarning,
            )
            computed_wide = pd.DataFrame(rows)
    else:
        computed_wide = pd.DataFrame(rows)   # Pandas fallback (no Polars installed)

    # ensure column order (unchanged from original)
    computed_wide = computed_wide[meta_cols + year_cols]

    # --- Rebuild WIDE output to EXACTLY match the *table layout* used in the Values reference file ---
    # The Values file is not "region-grouped"; it is organized into distinct tables with header/blank rows.
    # We recreate that layout deterministically here so the output always matches the reference structure.

    REG_ORDER = ['CAN','BC','AB','SK','MB','ON','QC','NB','NS','PE','NL','YT','NT','NU','AT','TR']
    REG_NAME_MAP = {rc: rn for rc, rn in REGIONS}

    def _blank_meta_row():
        return {c: np.nan for c in meta_cols}

    def _market_share_total_row():
        r=_blank_meta_row()
        r['Index']='market_share_total'
        r['Parameter']='market_share_total'
        return r

    def _section_header_row(title: str):
        r=_blank_meta_row()
        r['Index']=title
        r['Parameter']=title
        return r

    def _data_row(param: str, reg_code: str, is_first_in_block: bool):
        return {
            'Index': f"{param}{reg_code}",
            'Source': 'CEUD' if is_first_in_block else np.nan,
            'Unit': '%',
            'Parameter': param,
            'Unnamed: 4': reg_code,
            'Unnamed: 5': REG_NAME_MAP.get(reg_code, np.nan),
        }

    def _append_section(rows_meta, title, params, end_blank_rows=3):
        # section title row
        rows_meta.append(_section_header_row(title))
        # parameter blocks
        for pi, param in enumerate(params):
            for ri, reg_code in enumerate(REG_ORDER):
                rows_meta.append(_data_row(param, reg_code, is_first_in_block=(ri==0)))
            # blank separator between parameter blocks (but not after the last one)
            if pi < len(params) - 1:
                rows_meta.append(_blank_meta_row())
        # trailing blank rows after the table
        for _ in range(end_blank_rows):
            rows_meta.append(_blank_meta_row())

    # Build the full template (308 rows) matching the Values reference layout
    rows_meta=[]
    # Title and blank line at the very top
    top=_blank_meta_row(); top['Source']='Forecast values based on assumptions sheet'
    rows_meta.append(top)
    rows_meta.append(_blank_meta_row())

    # Urban table
    rows_meta.append(_market_share_total_row())
    _append_section(
        rows_meta,
        title='Urban',
        params=['Walk Cycle Urban','Passenger Vehicle Urban SOV','Passenger Vehicle Urban HOV','Public Transit Urban'],
        end_blank_rows=3,
    )

    # Intercity Land table
    rows_meta.append(_market_share_total_row())
    _append_section(
        rows_meta,
        title='Intercity Land',
        params=['Bus Intercity','Rail Intercity','Passenger Vehicle Intercity'],
        end_blank_rows=3,
    )

    # Passenger vehicles (LDV composition) table
    rows_meta.append(_market_share_total_row())
    _append_section(
        rows_meta,
        title='Passenger vehicles',
        params=['Car_small','Car_large','Light truck_small','Light truck_large'],
        end_blank_rows=3,
    )

    # Public Bus table (includes Ferry Urban)
    rows_meta.append(_market_share_total_row())
    _append_section(
        rows_meta,
        title='Public Bus',
        params=['Bus Urban Diesel','Bus Urban NG','Bus Urban Electric','Ferry Urban'],
        end_blank_rows=3,
    )

    # Intercity Bus fuel split table (final section; no trailing blank rows)
    rows_meta.append(_market_share_total_row())
    _append_section(
        rows_meta,
        title='Intercity Bus',
        params=['Bus Intercity Diesel','Bus Intercity Gasoline'],
        end_blank_rows=0,
    )

    wide=pd.DataFrame(rows_meta)
    # attach year columns as NaN initially
    for yc in year_cols:
        wide[yc]=np.nan
    wide=wide[meta_cols+year_cols]

    # Fill numeric values for data rows using computed_wide (match by Index)
    comp = computed_wide.set_index('Index')
    idx_series = wide['Index']
    mask = idx_series.notna() & idx_series.astype(str).isin(comp.index.astype(str))
    if mask.any():
        sel = idx_series[mask].astype(str)
        wide.loc[mask, year_cols] = comp.loc[sel, year_cols].to_numpy()

    # Final sanity: ensure expected row count (Values reference has 308 rows)
    if len(wide) != 308:
        print(f"[WARN] calc_market_share wide layout has {len(wide)} rows (expected 308 to match Values reference).")
    out_dir = OUT_DIR if 'OUT_DIR' in globals() else Path('.')
    # Sprint 5: Polars fast CSV write for calc_market_share.csv  [PRIMARY — always written]
    _audit_write_csv_fast(wide, out_dir / out_file, is_primary=True)
    _register_df(out_file, wide)

    long=pd.DataFrame(long_rows)
    # Sprint 5: Polars fast CSV write for calc_market_share_long.csv
    _audit_write_csv_fast(long, out_dir / out_file.replace('.csv', '_long.csv'))
    _register_df(out_file.replace('.csv','_long.csv'), long)

    return wide

def _tp_inject_region(s, region_code):
    """Replace the {REG} placeholder in Branch/Target strings."""
    if s is None:
            return None
    if not isinstance(s, str):
        return s
    return s.replace('{REG}', str(region_code).strip().upper())

TP_FINAL_YEARS = [2000, 2005, 2010, 2015, 2020, 2025, 2030, 2035, 2040, 2045, 2050]

TP_ROW_SPEC_BASE = [{'Branch': 'CIMS.CAN.{REG}',
  'Type': 'Region',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal',
  'Source': 'annual_region',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Mode',
  'Source': 'annual_tech',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Mode',
  'Technology': None,
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Urban',
  'Source': 'annual_region',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Mode',
  'Technology': None,
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Intercity Land',
  'Source': 'annual_region',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Mode',
  'Technology': None,
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Intercity Air',
  'Source': 'annual_region',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Urban',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Urban',
  'Technology': 'Passenger Vehicle Urban SOV',
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Passenger Vehicles',
  'Source': 'annual_tech',
  'Unit': 'k*vkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Urban',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Urban',
  'Technology': 'Passenger Vehicle Urban HOV',
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Passenger Vehicles',
  'Source': 'annual_tech',
  'Unit': 'k*vkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Intercity Land',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Intercity Land',
  'Technology': 'Passenger Vehicle Intercity',
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Passenger Vehicles',
  'Source': 'annual_tech',
  'Unit': 'k*vkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Urban',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Urban',
  'Technology': 'Public Transit Urban',
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Transit',
  'Source': 'annual_tech',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Intercity Land',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Intercity Land',
  'Technology': 'Bus Intercity',
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Intercity Bus',
  'Source': 'annual_tech',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal.Mode.Intercity Land',
  'Type': 'Service',
  'Sector': 'Transportation Personal',
  'Service': 'Intercity Land',
  'Technology': 'Rail Intercity',
  'Parameter': 'service_request',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Transportation Personal.Intercity Rail',
  'Source': 'annual_tech',
  'Unit': 'k*pkm'},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Biodiesel',
  'Source': 'AFDC',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Ethanol',
  'Source': 'AFDC',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Natural Gas',
  'Source': 'AFDC',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Diesel',
  'Source': 'CER',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Gasoline',
  'Source': 'CER',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.CAN.{REG}.Electricity',
  'Source': 'ELEC',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Jet Fuel',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Propane',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Biogas',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Black Liquor',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Coal',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Coke',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Fuel Oil',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Hydrogen',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.LPG',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Petroleum Coke',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Refinery Fuel Gas',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Solid Biomass',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Uranium',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'Sector',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'multiplier_price',
  'Context': None,
  'Sub_Context': None,
  'Target': 'CIMS.Generic Fuels.Waste Fuel',
  'Source': 'JCIMS',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'market_share_total',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'output',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'fcc',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'fom',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'available',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'unavailable',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'lifetime',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'discount_rate_financial',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'discount_rate_retrofit',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'heterogeneity',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'competition',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'technology',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None},
 {'Branch': 'CIMS.CAN.{REG}.Transportation Personal',
  'Type': 'TODO',
  'Sector': 'Transportation Personal',
  'Service': None,
  'Technology': None,
  'Parameter': 'service_provide',
  'Context': 'TODO',
  'Sub_Context': None,
  'Target': None,
  'Source': 'TODO_from_upstream_dfs',
  'Unit': None}]
def _tp_inject_region(s, region_code):
    """Replace {REG} placeholder in Branch/Target strings with region code."""
    if s is None:
        return None
    if not isinstance(s, str):
        return s
    reg = str(region_code).strip().upper()
    return s.replace('{REG}', reg)


def _tp_build_row_spec(region_code):
    """Instantiate TP_ROW_SPEC_BASE for a region by applying placeholder substitution."""
    reg = str(region_code).strip().upper()
    rows = []
    for r in TP_ROW_SPEC_BASE:
        rr = dict(r)
        rr['Region'] = reg
        rr['Branch'] = _tp_inject_region(rr.get('Branch'), reg)
        rr['Target'] = _tp_inject_region(rr.get('Target'), reg)
        rows.append(rr)
    return rows


def _tp_fuel_from_target(target):
    """Map a Target string to the canonical fuel label expected by get_macro_multiplier()."""
    if target is None:
        return None
    tl = str(target).lower()
    if 'cims.generic fuels.diesel' in tl:
        return 'Diesel'
    if 'cims.generic fuels.gasoline' in tl:
        return 'Gasoline'
    if 'cims.generic fuels.propane' in tl:
        return 'Propane'
    if 'cims.generic fuels.jet fuel' in tl:
        return 'Jet Fuel'
    if 'cims.generic fuels.natural gas' in tl:
        return 'Natural Gas'
    if '.electricity' in tl:
        return 'Electricity'
    if '.biodiesel' in tl:
        return 'Biodiesel'
    if '.ethanol' in tl:
        return 'Ethanol'
    return None

# =========================
# Script entry point
# =========================


# ================================================================
# REQUIRED OUTPUTS (NON-SILENT, NO AUDIT GATING)
# ================================================================


# ================================================================
# MODE SHARES BY REGION (extracted from annual tab of source data)
# ================================================================
MODE_SHARES_BY_REGION = {
    'BC': {
        2000: {'Urban': 0.533649, 'IL': 0.376621, 'Air': 0.08973},
        2005: {'Urban': 0.531613, 'IL': 0.37774, 'Air': 0.090647},
        2010: {'Urban': 0.548571, 'IL': 0.362277, 'Air': 0.089152},
        2015: {'Urban': 0.558747, 'IL': 0.353966, 'Air': 0.087287},
        2020: {'Urban': 0.524186, 'IL': 0.379402, 'Air': 0.096412},
        2025: {'Urban': 0.593687, 'IL': 0.325546, 'Air': 0.080767},
        2030: {'Urban': 0.602729, 'IL': 0.318333, 'Air': 0.078938},
        2035: {'Urban': 0.611807, 'IL': 0.311092, 'Air': 0.077101},
        2040: {'Urban': 0.620907, 'IL': 0.303834, 'Air': 0.075259},
        2045: {'Urban': 0.630015, 'IL': 0.29657, 'Air': 0.073415},
        2050: {'Urban': 0.639119, 'IL': 0.28931, 'Air': 0.071571},
    },
    'AB': {
        2000: {'Urban': 0.528748, 'IL': 0.372338, 'Air': 0.098915},
        2005: {'Urban': 0.550378, 'IL': 0.356183, 'Air': 0.093439},
        2010: {'Urban': 0.589555, 'IL': 0.32386, 'Air': 0.086585},
        2015: {'Urban': 0.606343, 'IL': 0.309767, 'Air': 0.08389},
        2020: {'Urban': 0.516817, 'IL': 0.375967, 'Air': 0.107216},
        2025: {'Urban': 0.618885, 'IL': 0.29858, 'Air': 0.082534},
        2030: {'Urban': 0.629197, 'IL': 0.290474, 'Air': 0.080329},
        2035: {'Urban': 0.639466, 'IL': 0.282403, 'Air': 0.078131},
        2040: {'Urban': 0.649677, 'IL': 0.274379, 'Air': 0.075943},
        2045: {'Urban': 0.659817, 'IL': 0.266414, 'Air': 0.073769},
        2050: {'Urban': 0.66987, 'IL': 0.258518, 'Air': 0.071612},
    },
    'SK': {
        2000: {'Urban': 0.558173, 'IL': 0.405743, 'Air': 0.036084},
        2005: {'Urban': 0.587598, 'IL': 0.378995, 'Air': 0.033408},
        2010: {'Urban': 0.701299, 'IL': 0.274337, 'Air': 0.024364},
        2015: {'Urban': 0.70453, 'IL': 0.271325, 'Air': 0.024145},
        2020: {'Urban': 0.639352, 'IL': 0.329922, 'Air': 0.030726},
        2025: {'Urban': 0.665206, 'IL': 0.307515, 'Air': 0.027279},
        2030: {'Urban': 0.676281, 'IL': 0.297433, 'Air': 0.026286},
        2035: {'Urban': 0.687274, 'IL': 0.287428, 'Air': 0.025298},
        2040: {'Urban': 0.698157, 'IL': 0.277525, 'Air': 0.024319},
        2045: {'Urban': 0.708901, 'IL': 0.267749, 'Air': 0.023349},
        2050: {'Urban': 0.719482, 'IL': 0.258126, 'Air': 0.022393},
    },
    'MB': {
        2000: {'Urban': 0.499512, 'IL': 0.379126, 'Air': 0.121362},
        2005: {'Urban': 0.496142, 'IL': 0.38199, 'Air': 0.121867},
        2010: {'Urban': 0.59554, 'IL': 0.306496, 'Air': 0.097964},
        2015: {'Urban': 0.621265, 'IL': 0.28676, 'Air': 0.091975},
        2020: {'Urban': 0.605477, 'IL': 0.296526, 'Air': 0.097997},
        2025: {'Urban': 0.65758, 'IL': 0.259834, 'Air': 0.082586},
        2030: {'Urban': 0.672346, 'IL': 0.248781, 'Air': 0.078874},
        2035: {'Urban': 0.686774, 'IL': 0.237982, 'Air': 0.075244},
        2040: {'Urban': 0.700844, 'IL': 0.227453, 'Air': 0.071703},
        2045: {'Urban': 0.714536, 'IL': 0.217209, 'Air': 0.068255},
        2050: {'Urban': 0.727834, 'IL': 0.207261, 'Air': 0.064905},
    },
    'ON': {
        2000: {'Urban': 0.549836, 'IL': 0.411257, 'Air': 0.038907},
        2005: {'Urban': 0.60468, 'IL': 0.361262, 'Air': 0.034058},
        2010: {'Urban': 0.632938, 'IL': 0.335169, 'Air': 0.031893},
        2015: {'Urban': 0.632419, 'IL': 0.335555, 'Air': 0.032026},
        2020: {'Urban': 0.55009, 'IL': 0.410049, 'Air': 0.039861},
        2025: {'Urban': 0.633023, 'IL': 0.334842, 'Air': 0.032136},
        2030: {'Urban': 0.640949, 'IL': 0.327609, 'Air': 0.031442},
        2035: {'Urban': 0.648922, 'IL': 0.320335, 'Air': 0.030743},
        2040: {'Urban': 0.656928, 'IL': 0.31303, 'Air': 0.030042},
        2045: {'Urban': 0.664957, 'IL': 0.305704, 'Air': 0.029339},
        2050: {'Urban': 0.672997, 'IL': 0.298368, 'Air': 0.028635},
    },
    'QC': {
        2000: {'Urban': 0.556829, 'IL': 0.41948, 'Air': 0.02369},
        2005: {'Urban': 0.596413, 'IL': 0.382006, 'Air': 0.02158},
        2010: {'Urban': 0.626121, 'IL': 0.353863, 'Air': 0.020016},
        2015: {'Urban': 0.63933, 'IL': 0.341281, 'Air': 0.019389},
        2020: {'Urban': 0.584307, 'IL': 0.392995, 'Air': 0.022697},
        2025: {'Urban': 0.641907, 'IL': 0.338775, 'Air': 0.019318},
        2030: {'Urban': 0.651784, 'IL': 0.329431, 'Air': 0.018786},
        2035: {'Urban': 0.661737, 'IL': 0.320014, 'Air': 0.018249},
        2040: {'Urban': 0.671748, 'IL': 0.310543, 'Air': 0.017709},
        2045: {'Urban': 0.681798, 'IL': 0.301035, 'Air': 0.017167},
        2050: {'Urban': 0.691866, 'IL': 0.291509, 'Air': 0.016625},
    },
    'AT': {
        2000: {'Urban': 0.465292, 'IL': 0.328989, 'Air': 0.205718},
        2005: {'Urban': 0.483196, 'IL': 0.31841, 'Air': 0.198394},
        2010: {'Urban': 0.537198, 'IL': 0.282233, 'Air': 0.180569},
        2015: {'Urban': 0.55186, 'IL': 0.272628, 'Air': 0.175511},
        2020: {'Urban': 0.416714, 'IL': 0.351083, 'Air': 0.232203},
        2025: {'Urban': 0.567447, 'IL': 0.26203, 'Air': 0.170522},
        2030: {'Urban': 0.579834, 'IL': 0.254521, 'Air': 0.165645},
        2035: {'Urban': 0.592202, 'IL': 0.247023, 'Air': 0.160775},
        2040: {'Urban': 0.604528, 'IL': 0.239551, 'Air': 0.155921},
        2045: {'Urban': 0.616787, 'IL': 0.232121, 'Air': 0.151093},
        2050: {'Urban': 0.628955, 'IL': 0.224745, 'Air': 0.1463},
    },
}


# ================================================================
# MARKET SHARE DEFAULTS (year 2000 only, for techs NOT in calc_market_share)
# Source: annual tab of transportation personal_source data.xlsx
# '_ALL' = same value for all regions; per-region keys override where needed
# ================================================================
MARKET_SHARE_DEFAULTS = {
    # --- Air techs (global, no region) ---
    'Air Intercity':                {'_ALL': 1.0},
    'Air Intercity Efficient':      {'_ALL': 0.0},
    'Air Intercity Electric':       {'_ALL': 0.0},
    'Air Intercity Hydrogen':       {'_ALL': 0.0},
    # --- Motor techs (global, no region) ---
    'Gasoline Existing':            {'_ALL': 1.0},
    'Gasoline Standard':            {'_ALL': 0.0},
    'Gasoline Efficient':           {'_ALL': 0.0},
    'Hybrid':                       {'_ALL': 0.0},
    'Plug-in Hybrid':               {'_ALL': 0.0},
    'BEV 500':                      {'_ALL': 0.0},
    'BEV 800':                      {'_ALL': 0.0},
    'Fuel Cell 650':                {'_ALL': 0.0},
    # --- Bus Urban techs (global, no region) ---
    'Bus Urban Hybrid':             {'_ALL': 0.0},
    'Bus Urban Hybrid Biodiesel':   {'_ALL': 0.0},
    'Bus Urban Hydrogen':           {'_ALL': 0.0},
    # --- Rapid Transit (global) ---
    'Light Rail':                   {'_ALL': 1.0},
    # --- Bus Intercity techs (global) ---
    'Bus Intercity Hybrid':         {'_ALL': 0.0},
    'Bus Intercity Hybrid Biodiesel': {'_ALL': 0.0},
    'Bus Intercity NG':             {'_ALL': 0.0},
    'Bus Intercity Hydrogen':       {'_ALL': 0.0},
    # --- Rail Intercity techs (global) ---
    'Rail Intercity Diesel':        {'_ALL': 1.0},
    'Rail Intercity Diesel Efficient': {'_ALL': 0.0},
    'Rail Intercity Hybrid Biodiesel': {'_ALL': 0.0},
    'Rail Intercity Hydrogen':      {'_ALL': 0.0},
    'Rail Intercity Electric':      {'_ALL': 0.0},
    # --- Light Truck sizes (region-specific from annual tab) ---
    'Light Truck_small': {
        'BC': 0.118136, 'AB': 0.11986, 'SK': 0.114291, 'MB': 0.111882,
        'ON': 0.094359, 'QC': 0.080607, 'AT': 0.104212,
    },
    'Light Truck_large': {
        'BC': 0.239851, 'AB': 0.243352, 'SK': 0.232046, 'MB': 0.227154,
        'ON': 0.191578, 'QC': 0.163657, 'AT': 0.211583,
    },
    'Ferry Urban': {
        'BC': 0.002168, '_ALL': 0.0,
    },

}

# ================================================================
# TRANSIT PB/RT SPLITS BY REGION (from annual tab)
# ================================================================
TRANSIT_SPLITS_BY_REGION = {
    'BC': {
        'PB': {2000: 0.969135, 2005: 0.961025, 2010: 0.939984, 2015: 0.937402, 2020: 0.869475, 2025: 0.922696, 2030: 0.922696, 2035: 0.922696, 2040: 0.922696, 2045: 0.922696, 2050: 0.922696},
        'RT': {2000: 0.030865, 2005: 0.038975, 2010: 0.060016, 2015: 0.062598, 2020: 0.130525, 2025: 0.077304, 2030: 0.077304, 2035: 0.077304, 2040: 0.077304, 2045: 0.077304, 2050: 0.077304},
    },
    'AB': {
        'PB': {2000: 0.969459, 2005: 0.937292, 2010: 0.965297, 2015: 0.957658, 2020: 0.927593, 2025: 0.951116, 2030: 0.951116, 2035: 0.951116, 2040: 0.951116, 2045: 0.951116, 2050: 0.951116},
        'RT': {2000: 0.030541, 2005: 0.062708, 2010: 0.034703, 2015: 0.042342, 2020: 0.072407, 2025: 0.048884, 2030: 0.048884, 2035: 0.048884, 2040: 0.048884, 2045: 0.048884, 2050: 0.048884},
    },
    'SK': {
        'PB': {2000: 1.0, 2005: 0.998251, 2010: 0.999638, 2015: 1.0, 2020: 1.0, 2025: 1.0, 2030: 1.0, 2035: 1.0, 2040: 1.0, 2045: 1.0, 2050: 1.0},
        'RT': {2000: 0.0, 2005: 0.001749, 2010: 0.000362, 2015: 0.0, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    },
    'MB': {
        'PB': {2000: 1.0, 2005: 1.0, 2010: 1.0, 2015: 1.0, 2020: 1.0, 2025: 1.0, 2030: 1.0, 2035: 1.0, 2040: 1.0, 2045: 1.0, 2050: 1.0},
        'RT': {2000: 0.0, 2005: 0.0, 2010: 0.0, 2015: 0.0, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    },
    'ON': {
        'PB': {2000: 0.918187, 2005: 0.919065, 2010: 0.943391, 2015: 0.914087, 2020: 0.860819, 2025: 0.894124, 2030: 0.894124, 2035: 0.894124, 2040: 0.894124, 2045: 0.894124, 2050: 0.894124},
        'RT': {2000: 0.081813, 2005: 0.080935, 2010: 0.056609, 2015: 0.085913, 2020: 0.139181, 2025: 0.105876, 2030: 0.105876, 2035: 0.105876, 2040: 0.105876, 2045: 0.105876, 2050: 0.105876},
    },
    'QC': {
        'PB': {2000: 0.856522, 2005: 0.855182, 2010: 0.881854, 2015: 0.831281, 2020: 0.728068, 2025: 0.804461, 2030: 0.804461, 2035: 0.804461, 2040: 0.804461, 2045: 0.804461, 2050: 0.804461},
        'RT': {2000: 0.143478, 2005: 0.144818, 2010: 0.118146, 2015: 0.168719, 2020: 0.271932, 2025: 0.195539, 2030: 0.195539, 2035: 0.195539, 2040: 0.195539, 2045: 0.195539, 2050: 0.195539},
    },
    'AT': {
        'PB': {2000: 1.0, 2005: 0.996659, 2010: 1.0, 2015: 0.999608, 2020: 0.999202, 2025: 0.999454, 2030: 0.999454, 2035: 0.999454, 2040: 0.999454, 2045: 0.999454, 2050: 0.999454},
        'RT': {2000: 0.0, 2005: 0.003341, 2010: 0.0, 2015: 0.000392, 2020: 0.000798, 2025: 0.000546, 2030: 0.000546, 2035: 0.000546, 2040: 0.000546, 2045: 0.000546, 2050: 0.000546},
    },
}

# ================================================================
# YEAR-VARYING GJ SERVICE_REQUEST VALUES
# For Bus Urban and Bus Intercity techs (from annual tab)
# NaN in annual tab → 0.0 in output (matches formula file behavior)
# ================================================================
BUS_GJ_BY_YEAR = {
    'Bus Urban Diesel':           {2000: 1.26667, 2005: 1.14167, 2010: 1.01667, 2015: 0.89167, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Urban Hybrid':           {2000: 0.95,    2005: 0.85625, 2010: 0.7625,  2015: 0.66875, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Urban Hybrid Biodiesel': {2000: 0.95,    2005: 0.85625, 2010: 0.7625,  2015: 0.66875, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Urban NG':               {2000: 1.26667, 2005: 1.14167, 2010: 1.01667, 2015: 0.89167, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Urban Hydrogen':         {2000: 0.69667, 2005: 0.62792, 2010: 0.55917, 2015: 0.49042, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Urban Electric':         {2000: 0.418,   2005: 0.37675, 2010: 0.3355,  2015: 0.29425, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Intercity Diesel':       {2000: 1.0,     2005: 0.88235, 2010: 0.76471, 2015: 0.64706, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Intercity Gasoline':     {2000: 1.0,     2005: 0.88235, 2010: 0.76471, 2015: 0.64706, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Intercity Hybrid':       {2000: 0.75,    2005: 0.66176, 2010: 0.57353, 2015: 0.48529, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Intercity Hybrid Biodiesel': {2000: 1.0, 2005: 0.88235, 2010: 0.76471, 2015: 0.64706, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Intercity NG':           {2000: 1.0,     2005: 0.88235, 2010: 0.76471, 2015: 0.64706, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
    'Bus Intercity Hydrogen':     {2000: 0.55,    2005: 0.48529, 2010: 0.42059, 2015: 0.35588, 2020: 0.0, 2025: 0.0, 2030: 0.0, 2035: 0.0, 2040: 0.0, 2045: 0.0, 2050: 0.0},
}

# ================================================================
# TRANSPORTATION PERSONAL FINAL OUTPUT
# ================================================================

def build_transportation_personal_final(region_code, out_file=None, write=True):
    """Build the final transportation personal output CSV for a region.

    Complete v5 — generates all 392 rows (401 for BC) from first principles.
    Every year-value from calc_long / calc_market_share / calc_avg_km / _MACRO_CACHE.
    All constants hard-coded from formula files.
    """
    import numpy as _np

    reg = str(region_code).strip().upper()
    out_file = out_file or f"transportation personal_{reg}_test.csv"

    # ------------------------------------------------------------------
    # Upstream data sources
    # ------------------------------------------------------------------
    calc_long = None
    for key in ['calc_long', 'calc_long.csv']:
        calc_long = _get_df(key, required=False)
        if calc_long is not None:
            break
    if calc_long is None:
        raise RuntimeError("calc_long not in _DF_STORE")

    ms_wide = None
    for key in ['calc_market_share.csv', 'calc_market_share']:
        ms_wide = _get_df(key, required=False)
        if ms_wide is not None:
            break

    avg_km = None
    for key in ['calc_avg_km.csv', 'calc_avg_km']:
        avg_km = _get_df(key, required=False)
        if avg_km is not None:
            break

    # ------------------------------------------------------------------
    # Naming
    # ------------------------------------------------------------------
    BR    = f"CIMS.CAN.{reg}"
    TP    = "Transportation Personal"
    BR_TP = f"{BR}.{TP}"

    # ------------------------------------------------------------------
    # Data helpers
    # ------------------------------------------------------------------
    def _calc(param, unit='k*pkm'):
        cl = calc_long
        m = ((cl['Parameter'].astype(str).str.strip() == param) &
             (cl['Region'].astype(str).str.strip() == reg) &
             (cl['Unit'].astype(str).str.strip() == unit))
        s = cl.loc[m, ['year', 'value']].copy()
        s['year']  = pd.to_numeric(s['year'],  errors='coerce')
        s['value'] = pd.to_numeric(s['value'], errors='coerce')
        return dict(zip(s['year'], s['value']))

    def _ms(param):
        # --- Primary: calc_market_share (year 2000 only) ---
        if ms_wide is not None:
            reg_col = None
            for _c in ['Region', 'region', 'RegionName', 'region_code', 'prov_code', 'Unnamed: 4']:
                if _c in ms_wide.columns:
                    reg_col = _c
                    break
            if reg_col is not None:
                m = ((ms_wide['Parameter'].astype(str).str.strip() == param) &
                     (ms_wide[reg_col].astype(str).str.strip() == reg))
                r = ms_wide[m]
                if r.empty:
                    m = ((ms_wide['Parameter'].astype(str).str.strip() == param) &
                         (ms_wide[reg_col].astype(str).str.strip() == 'CAN'))
                    r = ms_wide[m]
                if not r.empty:
                    row = r.iloc[0]
                    val = pd.to_numeric(row.get('2000', _np.nan), errors='coerce')
                    if pd.notna(val):
                        return {2000: float(val)}

        # --- Fallback: MARKET_SHARE_DEFAULTS dict ---
        if param in MARKET_SHARE_DEFAULTS:
            d = MARKET_SHARE_DEFAULTS[param]
            if reg in d:
                return {2000: d[reg]}
            elif '_ALL' in d:
                return {2000: d['_ALL']}

        return {}



    def _akm(param, unit=None, divisor=1):
        if avg_km is None:
            return {}
        reg_col = None
        for _c in ['Region', 'region', 'RegionName', 'region_code', 'prov_code']:
            if _c in avg_km.columns:
                reg_col = _c
                break
        if reg_col is None:
            return {}
        m = ((avg_km['Parameter'].astype(str).str.strip() == param) &
             (avg_km[reg_col].astype(str).str.strip() == reg))
        if unit:
            m = m & (avg_km['Unit'].astype(str).str.strip() == unit)
        r = avg_km[m]
        if r.empty:
            return {}
        row = r.iloc[0]
        return {y: float(row[str(y)]) / divisor for y in TP_FINAL_YEARS
                if pd.notna(pd.to_numeric(row.get(str(y), _np.nan), errors='coerce'))}

    def _mp(fuel):
        out = {}
        for y in TP_FINAL_YEARS:
            try:
                out[y] = float(get_macro_multiplier(reg, TP, fuel, int(y)))
            except Exception:
                pass
        return out

    def _c(v):
        return {y: v for y in TP_FINAL_YEARS}

    E = {}  # empty year dict

    # ------------------------------------------------------------------
    # Reference-matching mode-share aggregation (top-down, fixed shares)
    # ------------------------------------------------------------------
    mode_shares = MODE_SHARES_BY_REGION.get(reg, {})
    if not mode_shares:
        raise RuntimeError(f"No mode shares found for region {reg} in MODE_SHARES_BY_REGION")

    # Compute the region total by summing all mode-level k*pkm from calc_long
    _ALL_MODES = ['Walk Cycle Urban', 'Passenger Vehicle Urban SOV',
                  'Passenger Vehicle Urban HOV', 'School Bus', 'Transit',
                  'Passenger Vehicle Intercity', 'Bus Intercity', 'Rail Intercity',
                  'Aviation']
    total_kpkm = {}
    for y in TP_FINAL_YEARS:
        total_kpkm[y] = sum(_calc(p).get(y, 0.0) for p in _ALL_MODES)

    
    # Build the mode-level dicts using fixed fractions from source data
    su, sil, sa = {}, {}, {}
    for y in TP_FINAL_YEARS:
        shares = mode_shares.get(y)
        if shares is None:
            continue
        su[y]  = shares['Urban']
        sil[y] = shares['IL']
        sa[y]  = shares['Air']


    
    # Transit Public Bus / Rapid Transit split (top-down from annual tab)
    def _transit_split():                                                    # ← PATCHED
        splits = TRANSIT_SPLITS_BY_REGION.get(reg, {})
        if splits:
            return splits.get('PB', {}), splits.get('RT', {})
        # Fallback: compute from calc_long
        tr = _calc('Transit')
        rp = _calc('Rapid transit')
        fr = _calc('Ferry Urban')
        pb_d, rt_d = {}, {}
        for y in TP_FINAL_YEARS:
            t = tr.get(y, 0)
            rapid_total = rp.get(y, 0) + fr.get(y, 0)
            if t > 0:
                rt_d[y] = rapid_total / t
                pb_d[y] = 1.0 - rt_d[y]
            else:
                pb_d[y] = 1.0
                rt_d[y] = 0.0
        return pb_d, rt_d


    transit_pb, transit_rt = _transit_split()
    pvm_output = _akm('LDV', 'vkm', divisor=100)

    # ------------------------------------------------------------------
    # Row builder (unchanged from previous version)
    # ------------------------------------------------------------------
    rows = []
    META = ['Branch', 'Type', 'Region', 'Sector', 'Service', 'Technology',
            'Parameter', 'Context', 'Sub_Context', 'Target', 'Source', 'Unit']

    def _r(branch, typ, svc, tech, param, ctx, sub,
           target, source, unit, yv, cmt=''):
        d = {'Branch': branch, 'Type': typ, 'Region': reg,
             'Sector': TP, 'Service': svc, 'Technology': tech,
             'Parameter': param, 'Context': ctx, 'Sub_Context': sub,
             'Target': target, 'Source': source, 'Unit': unit,
             'Comments': cmt}
        for y in TP_FINAL_YEARS:
            d[y] = yv.get(y, _np.nan)
        rows.append(d)

    def _svc_meta(br, svc, comp, dr, het):
        _r(br, 'Service', svc, '', 'service_provide', '', '', '', 'constant_tech', '', E)
        _r(br, 'Service', svc, '', 'competition', comp, '', '', '', '', E)
        _r(br, 'Service', svc, '', 'discount_rate_financial', '', '', '', 'constant_tech', '%', _c(dr))
        _r(br, 'Service', svc, '', 'heterogeneity', '', '', '', 'constant_tech', '', _c(het))

    def _tech(br, svc, tn, avail, unavail, life, ms_p, out_u=None, out_v=None,
              fcc=None, fom=None, sr_target='', sr_source='', sr_unit='', sr_val=None,
              sr2_target=None, sr2_source=None, sr2_unit=None, sr2_val=None):
        _r(br, 'Service', svc, tn, 'technology', '', '', '', '', '', E)
        _r(br, 'Service', svc, tn, 'available', '', '', '', 'constant_tech', 'Year', _c(avail))
        _r(br, 'Service', svc, tn, 'unavailable', '', '', '', 'constant_tech', 'Year', _c(unavail))
        _r(br, 'Service', svc, tn, 'lifetime', '', '', '', 'constant_tech', 'Years', _c(life))
        _r(br, 'Service', svc, tn, 'market_share_total', '', '', '', '', '%', _ms(tn))
        if out_u is not None:
            _r(br, 'Service', svc, tn, 'output', '', '', '', out_u[1] if isinstance(out_u, tuple) else 'constant_tech', out_u[0] if isinstance(out_u, tuple) else out_u, out_v or E)
        if fcc is not None:
            _r(br, 'Service', svc, tn, 'fcc', '', '', '', 'constant_tech', '$', _c(fcc))
        if fom is not None:
            _r(br, 'Service', svc, tn, 'fom', '', '', '', 'constant_tech', '$', _c(fom))
        if sr_val is not None:
            _r(br, 'Service', svc, tn, 'service_request', '', '', sr_target, sr_source, sr_unit, sr_val)
        if sr2_val is not None:
            _r(br, 'Service', svc, tn, 'service_request', '', '', sr2_target, sr2_source, sr2_unit, sr2_val)

    
    # ==================================================================
    #  S1: Region total
    # ==================================================================
    _r(BR, 'Region', '', '', 'service_request', '', '',
       BR_TP, 'annual_region', '', total_kpkm)

       
    # ==================================================================
    #  S2: Sector
    # ==================================================================
    _r(BR_TP, 'Sector', '', '', 'service_provide', '', '', '', 'constant_tech', '', E)
    _r(BR_TP, 'Sector', '', '', 'competition', 'Sector', '', '', '', '', E)
    for fn, tgt, src in [
        ('Biodiesel',        f'{BR}.Biodiesel',                    'AFDC 2023'),
        ('Biogas',           'CIMS.Generic Fuels.Biogas',           'JCIMS'),
        ('Black Liquor',     'CIMS.Generic Fuels.Black Liquor',     'JCIMS'),
        ('Coal',             'CIMS.Generic Fuels.Coal',             'JCIMS'),
        ('Coke',             'CIMS.Generic Fuels.Coke',             'JCIMS'),
        ('Diesel',           'CIMS.Generic Fuels.Diesel',           'CER'),
        ('Electricity',      f'{BR}.Electricity',                   'CER'),
        ('Ethanol',          f'{BR}.Ethanol',                       'AFDC 2023'),
        ('Fuel Oil',         'CIMS.Generic Fuels.Fuel Oil',         'JCIMS'),
        ('Gasoline',         'CIMS.Generic Fuels.Gasoline',         'CER'),
        ('Hydrogen',         f'{BR}.Hydrogen',                      'JCIMS'),
        ('Jet Fuel',         'CIMS.Generic Fuels.Jet Fuel',         'JCIMS'),
        ('LPG',              'CIMS.Generic Fuels.LPG',              'JCIMS'),
        ('Natural Gas',      'CIMS.Generic Fuels.Natural Gas',      'AFDC 2023'),
        ('Petroleum Coke',   'CIMS.Generic Fuels.Petroleum Coke',   'JCIMS'),
        ('Propane',          'CIMS.Generic Fuels.Propane',           'JCIMS'),
        ('Refinery Fuel Gas','CIMS.Generic Fuels.Refinery Fuel Gas', 'JCIMS'),
        ('Solid Biomass',    'CIMS.Generic Fuels.Solid Biomass',     'JCIMS'),
        ('Uranium',          'CIMS.Generic Fuels.Uranium',           'JCIMS'),
        ('Waste Fuel',       'CIMS.Generic Fuels.Waste Fuel',        'JCIMS'),
    ]:
        _r(BR_TP, 'Sector', '', '', 'multiplier_price', '', '', tgt, src, '', _mp(fn))
    _r(BR_TP, 'Sector', '', '', 'service_request', '', '',
       f'{BR_TP}.Mode', 'annual_tech', 'k*pkm', _c(1.0))

    # S3: Mode shares (using fixed fractions from source data)
    M = f"{BR_TP}.Mode"
    _r(M, 'Service', 'Mode', '', 'service_provide', '', '', '', 'constant_tech', '', E)
    _r(M, 'Service', 'Mode', '', 'competition', 'Fixed Ratio', '', '', '', '', E)
    _r(M, 'Service', 'Mode', '', 'service_request', '', '',
       f'{M}.Urban', 'annual_region', 'k*pkm', su)
    _r(M, 'Service', 'Mode', '', 'service_request', '', '',
       f'{M}.Intercity Land', 'annual_region', 'k*pkm', sil)
    _r(M, 'Service', 'Mode', '', 'service_request', '', '',
       f'{M}.Intercity Air', 'annual_region', 'k*pkm', sa)

    # ... (S2, S4–S13 remain unchanged — copy from your existing function)
    
    # ==================================================================
    #  S4: Urban
    # ==================================================================
    U = f'{M}.Urban'
    _svc_meta(U, 'Urban', 'Tech Compete', 0.08, 6)
    # Walk Cycle Urban
    _tech(U, 'Urban', 'Walk Cycle Urban', 1950, 2101, 5, 'Walk Cycle Urban',
          sr_target='', sr_source='', sr_unit='', sr_val=None)

    # PV Urban SOV
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban SOV', 'technology', '', '', '', '', '', E)
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban SOV', 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban SOV', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban SOV', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(10))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban SOV', 'market_share_total', '', '', '', 'annual_region_tech', '%', _ms('Passenger Vehicle Urban SOV'))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban SOV', 'output', '', '', '', 'constant_tech', 'node unit', _c(20683))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban SOV', 'service_request', '', '',
       f'{BR_TP}.Passenger Vehicles', 'annual_tech', '', _c(1.0))

    # PV Urban HOV
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban HOV', 'technology', '', '', '', '', '', E)
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban HOV', 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban HOV', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban HOV', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(10))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban HOV', 'market_share_total', '', '', '', 'annual_region_tech', '%', _ms('Passenger Vehicle Urban HOV'))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban HOV', 'output', '', '', '', 'constant_tech', 'node unit', _c(20683))
    _r(U, 'Service', 'Urban', 'Passenger Vehicle Urban HOV', 'service_request', '', '',
       f'{BR_TP}.Passenger Vehicles', 'annual_tech', '', _c(0.2))

    # Public Transit Urban
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'technology', '', '', '', '', '', E)
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(5))
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'market_share_total', '', '', '', 'annual_region_tech', '%', _ms('Public Transit Urban'))
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'output', '', '', '', 'constant_tech', 'node unit', _c(20683))
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'fom', '', '', '', 'constant_tech', '$', _c(1564.22))
    _r(U, 'Service', 'Urban', 'Public Transit Urban', 'service_request', '', '',
       f'{BR_TP}.Transit', 'annual_tech', '', _c(1.0))

    # ==================================================================
    #  S5: Intercity Land
    # ==================================================================
    IL = f'{M}.Intercity Land'
    _svc_meta(IL, 'Intercity Land', 'Tech Compete', 0.08, 10)

    # PV Intercity
    _r(IL, 'Service', 'Intercity Land', 'Passenger Vehicle Intercity', 'technology', '', '', '', '', '', E)
    _r(IL, 'Service', 'Intercity Land', 'Passenger Vehicle Intercity', 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
    _r(IL, 'Service', 'Intercity Land', 'Passenger Vehicle Intercity', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
    _r(IL, 'Service', 'Intercity Land', 'Passenger Vehicle Intercity', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(5))
    _r(IL, 'Service', 'Intercity Land', 'Passenger Vehicle Intercity', 'market_share_total', '', '', '', 'annual_tech', '%', _ms('Passenger Vehicle Intercity'))
    _r(IL, 'Service', 'Intercity Land', 'Passenger Vehicle Intercity', 'output', '', '', '', 'constant_tech', 'node unit', _c(20683))
    _r(IL, 'Service', 'Intercity Land', 'Passenger Vehicle Intercity', 'service_request', '', '',
       f'{BR_TP}.Passenger Vehicles', 'annual_tech', '', _c(0.612522))

    # Bus Intercity (service node)
    _r(IL, 'Service', 'Intercity Land', 'Bus Intercity', 'technology', '', '', '', '', '', E)
    _r(IL, 'Service', 'Intercity Land', 'Bus Intercity', 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
    _r(IL, 'Service', 'Intercity Land', 'Bus Intercity', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
    _r(IL, 'Service', 'Intercity Land', 'Bus Intercity', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(5))
    _r(IL, 'Service', 'Intercity Land', 'Bus Intercity', 'market_share_total', '', '', '', 'annual_tech', '%', _ms('Bus Intercity'))
    _r(IL, 'Service', 'Intercity Land', 'Bus Intercity', 'service_request', '', '',
       f'{BR_TP}.Intercity Bus', 'annual_tech', '', _c(1.0))

    # Rail Intercity (service node)
    _r(IL, 'Service', 'Intercity Land', 'Rail Intercity', 'technology', '', '', '', '', '', E)
    _r(IL, 'Service', 'Intercity Land', 'Rail Intercity', 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
    _r(IL, 'Service', 'Intercity Land', 'Rail Intercity', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
    _r(IL, 'Service', 'Intercity Land', 'Rail Intercity', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(5))
    _r(IL, 'Service', 'Intercity Land', 'Rail Intercity', 'market_share_total', '', '', '', 'annual_tech', '%', _ms('Rail Intercity'))
    _r(IL, 'Service', 'Intercity Land', 'Rail Intercity', 'service_request', '', '',
       f'{BR_TP}.Intercity Rail', 'annual_tech', '', _c(1.0))

    # ==================================================================
    #  S6: Intercity Air
    # ==================================================================
    IA = f'{M}.Intercity Air'
    _svc_meta(IA, 'Intercity Air', 'Tech Compete', 0.25, 10)
    for tn, avl, gj, tgt, fcc_v in [
        ('Air Intercity',           1950, 1.75, 'CIMS.Generic Fuels.Jet Fuel', 197381531.359355),
        ('Air Intercity Efficient', 2010, 1.25, 'CIMS.Generic Fuels.Jet Fuel', 248501928.578093),
        ('Air Intercity Electric',  2025, 1.25, f'{BR}.Electricity',           312843740.470339),
        ('Air Intercity Hydrogen',  2030, 1.25, f'{BR}.Hydrogen',              312843740.470339),
    ]:
        _r(IA, 'Service', 'Intercity Air', tn, 'technology', '', '', '', '', '', E)
        _r(IA, 'Service', 'Intercity Air', tn, 'available', '', '', '', 'constant_tech', 'Year', _c(avl))
        _r(IA, 'Service', 'Intercity Air', tn, 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
        _r(IA, 'Service', 'Intercity Air', tn, 'lifetime', '', '', '', 'constant_tech', 'Years', _c(30))
        _r(IA, 'Service', 'Intercity Air', tn, 'market_share_total', '', '', '', 'annual_tech', '%', _ms(tn))
        _r(IA, 'Service', 'Intercity Air', tn, 'output', '', '', '', 'annual_region', 'node unit', _c(560000))
        _r(IA, 'Service', 'Intercity Air', tn, 'fcc', '', '', '', 'constant_tech', '$', _c(fcc_v))
        _r(IA, 'Service', 'Intercity Air', tn, 'service_request', '', '', tgt, 'annual_tech', '', _c(gj))


    # ==================================================================
    #  S7: Passenger Vehicles
    # ==================================================================
    PV = f'{BR_TP}.Passenger Vehicles'
    _svc_meta(PV, 'Passenger Vehicles', 'Tech Compete', 0.25, 7)
    for tn, fcc_v, sr_v in [
        ('Car_small',       20763.917,  8),
        ('Car_large',       31948.081,  10),
        ('Light Truck_small', 36406.104, 9),
        ('Light Truck_large', 47590.268, 13),
    ]:
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'technology', '', '', '', '', '', E)
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'lifetime', '', '', '', 'constant_tech', 'Years', _c(16))
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'market_share_total', '', '', '', 'annual_tech', '%', _ms(tn))

        if 'Car' in tn:
            _pv_out = _akm('Car', 'vkm', divisor=1000)
        else:
            _pv_out = _akm('Light truck', 'vkm', divisor=1000)

        _r(PV, 'Service', 'Passenger Vehicles', tn, 'output', '', '', '',
           'annual_region', 'node unit', _pv_out)
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'fcc', '', '', '', 'constant_tech', '$', _c(fcc_v))
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'fom', '', '', '', 'constant_tech', '$', _c(2738.947))
        _r(PV, 'Service', 'Passenger Vehicles', tn, 'service_request', '', '',
           f'{BR_TP}.Passenger Vehicle Motors', 'annual_tech', '100 vkm (avg car eq)', _c(sr_v))


    # ==================================================================
    #  S8: Passenger Vehicle Motors
    # ==================================================================
    PVM = f'{BR_TP}.Passenger Vehicle Motors'
    _r(PVM, 'Service', 'Passenger Vehicle Motors', '', 'service_provide', '', '', '', 'constant_tech', '', E)
    _r(PVM, 'Service', 'Passenger Vehicle Motors', '', 'competition', 'Tech Compete', '', '', '', '', E)
    _r(PVM, 'Service', 'Passenger Vehicle Motors', '', 'discount_rate_financial', '', '', '', 'constant_tech', '%', _c(0.25))
    _r(PVM, 'Service', 'Passenger Vehicle Motors', '', 'discount_rate_retrofit', '', '', '', 'constant_tech', '%', _c(0.65))
    _r(PVM, 'Service', 'Passenger Vehicle Motors', '', 'heterogeneity', '', '', '', 'constant_tech', '', _c(15))

    GB = f'{BR}.Fuel Blends.Gasoline_Transportation'
    for tn, avl, unavl, fcc_v, sr_gj, sr2_tgt, sr2_gj in [
        ('Gasoline Existing',  1990, 2001, 9314.421,  0.35853, None, None),
        ('Gasoline Standard',  2005, 2101, 9314.421,  0.32424, None, None),
        ('Gasoline Efficient', 2015, 2101, 10650.716, 0.27723, None, None),
        ('Hybrid',             2010, 2101, 11987.011, 0.20792, None, None),
        ('Plug-in Hybrid',     2015, 2101, 20722.423, 0.04158, f'{BR}.Electricity', 0.16634),
        ('BEV 500',            2015, 2101, 21424.200, None, None, None),
        ('BEV 800',            2030, 2101, 40358.790, None, None, None),
        ('Fuel Cell 650',      2020, 2101, 30600.732, None, None, None),
    ]:
        _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'technology', '', '', '', '', '', E)
        _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'available', '', '', '', 'constant_tech', 'Year', _c(avl))
        _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'unavailable', '', '', '', 'constant_tech', 'Year', _c(unavl))
        _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'lifetime', '', '', '', 'constant_tech', 'Years', _c(16))
        _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'market_share_total', '', '', '', 'annual_tech', '%', _ms(tn))
        _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'output', '', '', '', 'annual_region', 'node unit', pvm_output)
        _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'fcc', '', '', '', 'constant_tech', '$', _c(fcc_v))
        # service_request(s)
        if tn == 'BEV 500' or tn == 'BEV 800':
            _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'service_request', '', '',
               f'{BR}.Electricity', 'annual_tech', '', _c(0.09149))
        elif tn == 'Fuel Cell 650':
            _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'service_request', '', '',
               f'{BR}.Hydrogen', 'annual_tech', '', _c(0.15248))
        elif tn == 'Plug-in Hybrid':
            _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'service_request', '', '',
               GB, 'annual_tech', '', _c(sr_gj))
            _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'service_request', '', '',
               f'{BR}.Electricity', 'annual_tech', '', _c(sr2_gj))
        else:
            _r(PVM, 'Service', 'Passenger Vehicle Motors', tn, 'service_request', '', '',
               GB, 'annual_tech', '', _c(sr_gj))


    # ==================================================================
    #  S9: Transit
    # ==================================================================
    TR = f'{BR_TP}.Transit'
    _r(TR, 'Service', 'Transit', '', 'service_provide', '', '', '', 'constant_tech', '', E)
    _r(TR, 'Service', 'Transit', '', 'competition', 'Fixed Ratio', '', '', '', '', E)
    _r(TR, 'Service', 'Transit', '', 'service_request', '', '',
       f'{TR}.Public Bus', 'annual_region', 'k*pkm', transit_pb)
    _r(TR, 'Service', 'Transit', '', 'service_request', '', '',
       f'{TR}.Rapid Transit', 'annual_region', 'k*pkm', transit_rt)

    # ==================================================================
    #  S10: Public Bus
    # ==================================================================
    PB = f'{TR}.Public Bus'
    _svc_meta(PB, 'Public Bus', 'Tech Compete', 0.25, 20)
    for tn, avl, fcc_v, sr_tgt, sr_gj_2000 in [
        ('Bus Urban Diesel',           1950, 391054.676,  f'{BR}.Fuel Blends.Diesel_Transportation', 1.26667),
        ('Bus Urban Hybrid',           2010, 547476.546,  f'{BR}.Fuel Blends.Diesel_Transportation', 0.95),
        ('Bus Urban Hybrid Biodiesel', 2010, 547476.546,  f'{BR}.Biodiesel', 0.95),
        ('Bus Urban NG',               1950, 547476.546,  'CIMS.Generic Fuels.Natural Gas', 1.26667),       
        ('Bus Urban Hydrogen',         2010, 2346328.054, f'{BR}.Hydrogen', 0.696667),       # ← PATCHED (was 0.475)
        ('Bus Urban Electric',         1950, 782109.351,  f'{BR}.Electricity', 0.418),        # ← PATCHED (was 0.63333)

    ]:
        _r(PB, 'Service', 'Public Bus', tn, 'technology', '', '', '', '', '', E)
        _r(PB, 'Service', 'Public Bus', tn, 'available', '', '', '', 'constant_tech', 'Year', _c(avl))
        _r(PB, 'Service', 'Public Bus', tn, 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
        _r(PB, 'Service', 'Public Bus', tn, 'lifetime', '', '', '', 'constant_tech', 'Years', _c(16))
        _r(PB, 'Service', 'Public Bus', tn, 'market_share_total', '', '', '', 'annual_tech', '%', _ms(tn))
        _r(PB, 'Service', 'Public Bus', tn, 'output', '', '', '', 'annual_region', 'node unit', _akm('Public Bus', 'pkm', divisor=1000))
        _r(PB, 'Service', 'Public Bus', tn, 'fcc', '', '', '', 'constant_tech', '$', _c(fcc_v))
        _r(PB, 'Service', 'Public Bus', tn, 'service_request', '', '', sr_tgt, 'annual_tech', '', BUS_GJ_BY_YEAR.get(tn, _c(sr_gj_2000)))

    # BC: Ferry Urban
    if reg == 'BC':
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'technology', '', '', '', '', '', E)
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'available', '', '', '', 'constant_tech', 'Year', _c(1977))
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(50))
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'market_share_total', '', '', '', 'annual_region_tech', '%', _ms('Ferry Urban'))
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'output', '', '', '', 'annual_region_tech', 'node unit', _c(18719.926))
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'fcc', '', '', '', 'constant_tech', '$', _c(31200000))
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'fom', '', '', '', 'constant_tech', '$', _c(4700000))
        _r(PB, 'Service', 'Public Bus', 'Ferry Urban', 'service_request', '', '',
           'CIMS.Generic Fuels.Diesel', 'annual_tech', '', _c(2.37302))

    # ==================================================================
    #  S11: Rapid Transit
    # ==================================================================
    RT = f'{TR}.Rapid Transit'
    _svc_meta(RT, 'Rapid Transit', 'Tech Compete', 0.25, 10)
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'technology', '', '', '', '', '', E)
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'available', '', '', '', 'constant_tech', 'Year', _c(1950))
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'lifetime', '', '', '', 'constant_tech', 'Years', _c(35))
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'market_share_total', '', '', '', 'annual_tech', '%', _ms('Light Rail'))
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'output', '', '', '', 'constant_tech', 'node unit', _c(5940))
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'fcc', '', '', '', 'constant_tech', '$', _c(667921.386))
    _r(RT, 'Service', 'Rapid Transit', 'Light Rail', 'service_request', '', '',
       f'{BR}.Electricity', 'annual_tech', '', _c(1.09091))


    # ==================================================================
    #  S12: Intercity Bus
    # ==================================================================
    IB = f'{BR_TP}.Intercity Bus'
    _svc_meta(IB, 'Intercity Bus', 'Tech Compete', 0.25, 10)
    for tn, avl, fcc_v, sr_tgt, sr_gj in [
        ('Bus Intercity Diesel',           1950, 391054.676,  f'{BR}.Fuel Blends.Diesel_Transportation', 1.0),
        ('Bus Intercity Gasoline',         1950, 391054.676,  f'{BR}.Fuel Blends.Gasoline_Transportation', 1.0),
        ('Bus Intercity Hybrid',           2010, 547476.546,  f'{BR}.Fuel Blends.Diesel_Transportation', 0.75),
        ('Bus Intercity Hybrid Biodiesel', 2010, 547476.546,  f'{BR}.Biodiesel', 1.0),        # ← PATCHED (was 0.75)
        ('Bus Intercity NG',               1950, 547476.546,  'CIMS.Generic Fuels.Natural Gas', 1.0),
        ('Bus Intercity Hydrogen',         2010, 2346328.054, f'{BR}.Hydrogen', 0.55),         # ← PATCHED (was 0.5)
    ]:
        _r(IB, 'Service', 'Intercity Bus', tn, 'technology', '', '', '', '', '', E)
        _r(IB, 'Service', 'Intercity Bus', tn, 'available', '', '', '', 'constant_tech', 'Year', _c(avl))
        _r(IB, 'Service', 'Intercity Bus', tn, 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
        _r(IB, 'Service', 'Intercity Bus', tn, 'lifetime', '', '', '', 'constant_tech', 'Years', _c(16))
        _r(IB, 'Service', 'Intercity Bus', tn, 'market_share_total', '', '', '', 'annual_tech', '%', _ms(tn))
        _r(IB, 'Service', 'Intercity Bus', tn, 'output', '', '', '', 'annual_region', 'node unit', _akm('Intercity Bus', 'pkm', divisor=1000))
        _r(IB, 'Service', 'Intercity Bus', tn, 'fcc', '', '', '', 'constant_tech', '$', _c(fcc_v))
        _r(IB, 'Service', 'Intercity Bus', tn, 'service_request', '', '', sr_tgt, 'annual_tech', '', BUS_GJ_BY_YEAR.get(tn, _c(sr_gj)))

    # ==================================================================
    #  S13: Intercity Rail                                               # ← PATCHED (all 5 GJ values corrected)
    # ==================================================================
    IR = f'{BR_TP}.Intercity Rail'
    _svc_meta(IR, 'Intercity Rail', 'Tech Compete', 0.25, 10)
    for tn, avl, fcc_v, sr_tgt, sr_gj in [
        ('Rail Intercity Diesel',           1950, 52328087.349,  f'{BR}.Fuel Blends.Diesel_Transportation', 1.92048147),
        ('Rail Intercity Diesel Efficient', 2010, 139541591.863, f'{BR}.Fuel Blends.Diesel_Transportation', 1.3),
        ('Rail Intercity Hybrid Biodiesel', 2010, 139541591.863, f'{BR}.Biodiesel',                        1.3),
        ('Rail Intercity Hydrogen',         2030, 305247256.279, f'{BR}.Hydrogen',                         0.715),
        ('Rail Intercity Electric',         2010, 174426989.924, f'{BR}.Electricity',                      0.429),
    ]:
        _r(IR, 'Service', 'Intercity Rail', tn, 'technology', '', '', '', '', '', E)
        _r(IR, 'Service', 'Intercity Rail', tn, 'available', '', '', '', 'constant_tech', 'Year', _c(avl))
        _r(IR, 'Service', 'Intercity Rail', tn, 'unavailable', '', '', '', 'constant_tech', 'Year', _c(2101))
        _r(IR, 'Service', 'Intercity Rail', tn, 'lifetime', '', '', '', 'constant_tech', 'Years', _c(25))
        _r(IR, 'Service', 'Intercity Rail', tn, 'market_share_total', '', '', '', 'annual_tech', '%', _ms(tn))
        _r(IR, 'Service', 'Intercity Rail', tn, 'output', '', '', '', 'constant_tech', 'node unit', _c(79733.868))
        _r(IR, 'Service', 'Intercity Rail', tn, 'fcc', '', '', '', 'constant_tech', '$', _c(fcc_v))
        _r(IR, 'Service', 'Intercity Rail', tn, 'service_request', '', '', sr_tgt, 'annual_tech', '', _c(sr_gj))

    # Build DataFrame + Write
    df = pd.DataFrame(rows)
    col_order = META + [y for y in TP_FINAL_YEARS] + ['Comments']
    for c_ in col_order:
        if c_ not in df.columns:
            df[c_] = _np.nan
    df = df[col_order]

    if write:
        p = OUT_DIR / out_file
        p.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(p, index=False)

    _register_df(out_file, df)

    try:
        (OUT_DIR / f"transportation_personal_{reg}_notes.txt").write_text(
            f"[INFO] complete v5 | {len(df)} rows | from first principles\n",
            encoding='utf-8')
    except Exception:
        pass

    return df


def build_transportation_personal_all_regions(regions=None):
    """Build final transportation personal output CSVs for all required regions."""
    regions = regions or ['AB', 'BC', 'MB', 'ON', 'QC', 'SK', 'AT']
    return {r: build_transportation_personal_final(r, write=True) for r in regions}

# ================================================================
# REQUIRED OUTPUTS WRITER (test filenames)
# ================================================================

def write_required_outputs():
    """Write required deliverables and restore calc output writing.

    Writes:
      - calc.csv, calc_market_share.csv, calc_avg_km.csv
      - macro_multipliers_all_provinces.csv
      - transportation personal_{REG}_test.csv for AB,BC,MB,ON,QC,SK,AT
      - required_outputs_status.txt

    Note: TP outputs use the *_test.csv suffix for auditing.
    """
    import traceback

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    status = []

    def _write_df(df, path):
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            if df is None:
                raise ValueError('df is None')
            df.to_csv(path, index=False)
            msg = f"[OK] {path.name} ({path.stat().st_size} bytes)"
            print(msg)
            status.append(msg)
        except Exception:
            msg = f"[FAIL] {path.name}\n{traceback.format_exc()}"
            print(msg)
            status.append(msg)

    # calc outputs
    try:
        df_calc = build_calc(out_file='calc.csv')
        _write_df(df_calc, OUT_DIR / 'calc.csv')
    except Exception:
        status.append('[FAIL] calc.csv\n' + traceback.format_exc())

    try:
        df_ms = build_calc_market_share(out_file='calc_market_share.csv')
        _write_df(df_ms, OUT_DIR / 'calc_market_share.csv')
    except Exception:
        status.append('[FAIL] calc_market_share.csv\n' + traceback.format_exc())

    try:
        df_avg = build_calc_avg_km(out_file='calc_avg_km.csv')
        _write_df(df_avg, OUT_DIR / 'calc_avg_km.csv')
    except Exception:
        status.append('[FAIL] calc_avg_km.csv\n' + traceback.format_exc())

    # macro multipliers
    try:
        df_mm = build_multiplier_table()
        _write_df(df_mm, OUT_DIR / 'macro_multipliers_all_provinces.csv')
    except Exception:
        status.append('[FAIL] macro_multipliers_all_provinces.csv\n' + traceback.format_exc())

    # TP outputs
    for reg in ['AB','BC','MB','ON','QC','SK','AT']:
        try:
            fname = f"transportation personal_{reg}_test.csv"
            df_tp = build_transportation_personal_final(reg, out_file=fname, write=True)
            _write_df(df_tp, OUT_DIR / fname)
        except Exception:
            status.append(f'[FAIL] transportation personal_{reg}_test.csv\n' + traceback.format_exc())

    # status log
    try:
        (OUT_DIR / 'required_outputs_status.txt').write_text("\n".join(status), encoding='utf-8')
        print('[OK] required_outputs_status.txt written')
    except Exception:
        pass

    return status


# ================================================================
# SCRIPT ENTRY POINT (SINGLE)
# ================================================================

def run_all():
    """Run upstream pipeline then required outputs."""
    main()
    write_required_outputs()


if __name__ == "__main__":
    run_all()