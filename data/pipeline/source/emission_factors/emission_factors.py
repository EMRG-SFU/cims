"""
Emission Factors Extraction Script
===================================
Primary source : Canada NIR 2025, Annex 6 Emission Factors Tables
                 (EN_Annex6_Emission_Factors_Tables.xlsx)
Secondary sources: see per-fuel comments below
Output columns : year, fuel, emissions_type, emission_source, source, value, units

All output values are in tonnes of emissions per GJ (e.g. tCO2/GJ, tCH4/GJ, tN2O/GJ).

─────────────────────────────────────────────────────────────────────────────
FUELS CONFIGURED
─────────────────────────────────────────────────────────────────────────────
Fossil combustion fuels (CO2/CH4/N2O from NIR Annex 6):
  Natural Gas         CO2: Table A6.1-1 (annual, Canada)
                      CH4/N2O: Table A6.1-3 (Residential/Commercial/Agriculture)
  Heavy Fuel Oil      CO2/CH4/N2O: Table A6.1-6 (Industrial)
  Light Fuel Oil      CO2/CH4/N2O: Table A6.1-6 (Residential)
  Gasoline            CO2/CH4/N2O: Table A6.1-6 (Motor Gasoline)
  Diesel              CO2/CH4/N2O: Table A6.1-15 (HDDV Advanced Control)
  Kerosene            CO2/CH4/N2O: Table A6.1-6 (Residential)
  Jet Fuel            CO2/CH4/N2O: Table A6.1-15 (Aviation Turbo Fuel)
  Refinery Fuel Gas   CO2: Table A6.1-7 annual (Refineries and Others)
                      CH4: Table A6.1-9 annual with ranges
                      N2O: Table A6.1-6 (Still Gas - Refineries and Others)
  Propane             CO2: Table A6.1-5 (Residential)
                      CH4/N2O: Table A6.1-5 (Residential)
  Butane              CO2/CH4/N2O: Table A6.1-5
  Ethane              CO2/CH4/N2O: Table A6.1-5
  LPG                 Not in NIR -- CO2: IPCC 2006 Vol.2 Table 1.4
                                   CH4/N2O: IPCC 2006 Vol.2 Table 2.3 (Manufacturing)
  Petroleum Coke      CO2: Table A6.1-7 annual (Refineries and Others)
                      CH4: Table A6.1-6
                      N2O: Table A6.1-8 annual (Refineries and Others)
  Waste               CO2: Table A6.1-14 annual with ranges
                      CH4/N2O: Table A6.1-14

Coal:
  Thermal Coal        CO2/CH4/N2O: CEEDC_coal_electricity_energy_emission_factors_260313.xlsx
                      (Canada weighted-average t/GJ, rows 227/237/246;
                       derived from NIR Tables A6.1-10 and A6.1-12)
  Metallurgical Coal  All emission factors = 0; emissions accounted for elsewhere in the model

Process / non-energy fuels (CO2 from NIR; CH4/N2O = 0 -- not reported for non-energy use):
  Natural Gas Feedstock   CO2: 38 g CO2/m3 hardcoded (NIR A6.2.4 prose, Cheminfo 2005). 
  Petrochemical Feedstock CO2: Table A6.2-10
  Naphtha Specialties     CO2: Table A6.2-10
  Lubricants              CO2: Table A6.2-10
  Other Non-Energy Products CO2: Table A6.2-10

Biogenic fuels (CO2 biogenic -- labelled "emissions_biomass" in CIMS output, not counted in
  Energy sector totals per NIR Table A6.6-1 note a):
  Ethanol             CO2: Table A6.1-15 (Renewable Fuels / Ethanol, col_index 2: 1508.04 g/L)
                      CH4/N2O: Table A6.1-6 (Motor Gasoline -- NIR footnote **)
  Biodiesel           CO2: Table A6.1-15 (Renewable Fuels / Biodiesel, col_index 2: 2472.2 g/L)
                      CH4/N2O: Table A6.1-15 (HDDV Advanced Control -- NIR footnote ***)
  Black Liquor        CO2/CH4/N2O: Table A6.6-1 (Spent Pulping Liquor, col_index 3/4/5)
                      Energy content: 14.0 GJ/t (Statistics Canada Cat. no. 57-601-X)
  Solid Biomass       CO2/CH4/N2O: Table A6.6-1 (Wood Fuel / Wood Waste, col_index 3/4/5)
                      Energy content: 18.0 GJ/t (Statistics Canada Cat. no. 57-601-X)

Renewable drop-in fuels (CO2 biogenic -- labelled "emissions_biomass" in CIMS output;
  Emission factors assumed same as fossil fuel counterpart:
  Renewable Natural Gas      CO2: Table A6.1-1 (Canada, same as Natural Gas)
                             CH4/N2O: Table A6.1-3 (same as Natural Gas)
  SAF                        CO2/CH4/N2O: Table A6.1-15 (Aviation Turbo Fuel)
  Renewable Diesel           CO2/CH4/N2O: Table A6.1-15 (HDDV Advanced Control)
  Renewable Gasoline         CO2/CH4/N2O: Table A6.1-6 (Motor Gasoline)

Zero-EF fuels (all gases = 0; see inline comments for rationale):
  Uranium             Nuclear fission -- no carbon combustion
  Hydrogen            H2 combustion produces water only 
  Electricity         Energy carrier -- emissions attributed at generation
  Asphalt             Non-combusted material -- no point-of-use GHG emissions

─────────────────────────────────────────────────────────────────────────────
"""

import re
import sys
import warnings
from math import floor, log10
from pathlib import Path
import openpyxl
import pandas as pd
import polars as pl

_current_file = Path(__file__)
_project_root = _current_file.parent.parent.parent
_utils_path   = _project_root / "data" / "pipeline" / "utils"
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from utils.extensions.data_extensions import extend_series_constant
from utils.controls_conversions import load_energy_conversions


# ==============================================================================
# CONFIGURATION
# ==============================================================================
BASE_PATH        = Path(r"C:\cims\data")
MAPPINGS_PATH    = BASE_PATH / "mappings_conversions"
RAW_PATH         = BASE_PATH / "raw_data"

NIR_XLSX         = RAW_PATH / "eccc" / "nir" / "EN_Annex6_Emission_Factors_Tables.xlsx"
NIR_PDF          = RAW_PATH / "eccc" / "nir" / "EN_Annex6_Emission_Factors.pdf"
COAL_XLSX        = RAW_PATH / "CEEDC" / "CEEDC_coal_electricity_energy_emission_factors_260313.xlsx"
IPCC_PDF         = RAW_PATH / "IPCC" / "V2_2_Ch2_Stationary_Combustion.pdf"
CONVERSIONS_FILE = MAPPINGS_PATH / "energy_conversions.csv"
ENERGY_MAP_FILE  = MAPPINGS_PATH / "energy_map.csv"
OUTPUT_DIR       = BASE_PATH / "processed_data" / "emission_factors"

YEAR_START = 2000
YEAR_END   = 2100

def _detect_nir_year(xlsx_path: str) -> int:
    """
    Detect the most recent data year in the NIR xlsx by reading the year
    column of Table A6.1-1 (Natural Gas CO2 -- the most complete annual series).
    Falls back to the current calendar year if detection fails.
    """
    import datetime
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb["Table A6.1–1"]
        years = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)) and 1990 <= cell <= 2100:
                    years.append(int(cell))
        wb.close()
        return max(years) if years else datetime.date.today().year
    except Exception:
        return datetime.date.today().year

NIR_YEAR   = _detect_nir_year(NIR_XLSX)
NIR_SOURCE = f"Canada NIR Annex 6 (ECCC {NIR_YEAR})"
IPCC_SOURCE     = "IPCC 2006 Guidelines for National GHG Inventories, Vol. 2"
COAL_SOURCE_CO2 = f"{COAL_XLSX.name} (derived from Canada NIR Annex 6, Table A6.1-10)"
COAL_SOURCE_CH4 = f"{COAL_XLSX.name} (derived from Canada NIR Annex 6, Table A6.1-12)"
COAL_SOURCE_N2O = f"{COAL_XLSX.name} (derived from Canada NIR Annex 6, Table A6.1-12)"
ASSUMPTION      = "Assumption"

DEFAULT_EMISSION_SOURCE = "combustion"



def validate_fuel_map(fuels: list, path: Path = ENERGY_MAP_FILE) -> None:
    """
    Cross-check the FUELS list against the CIMS column of energy_map.csv.
    Prints a warning for any fuel present in one but not the other.
    """
    if not path.exists():
        print(f"  ⚠️  Energy map not found at {path} — skipping fuel validation")
        return

    map_df     = pd.read_csv(path, dtype=str)
    map_names  = set(map_df["CIMS"].dropna().str.strip())
    fuel_names = {f["fuel_name"] for f in fuels}

    missing_from_script = map_names - fuel_names
    missing_from_map    = fuel_names - map_names

    if missing_from_script:
        for name in sorted(missing_from_script):
            print(f"  ⚠️  In energy_map.csv but not in FUELS: '{name}'")
    if missing_from_map:
        for name in sorted(missing_from_map):
            print(f"  ⚠️  In FUELS but not in energy_map.csv: '{name}'")
    if not missing_from_script and not missing_from_map:
        print("  ✅ Fuel map validation passed — all fuels accounted for")


# Load energy conversions once at module level so FUELS can reference values
_EC = load_energy_conversions()


def _gj_per_L(csv_name: str) -> float:
    """Return GJ/L from a GJ/m3 CSV entry (divide by 1000)."""
    return _EC[csv_name.lower()] / 1000


def _gj_per_m3(csv_name: str) -> float:
    """Return GJ/m3 directly from the CSV."""
    return _EC[csv_name.lower()]


def _gj_per_kg(csv_name: str) -> float:
    """Return GJ/kg directly from the CSV (mass-based fuels)."""
    return _EC[csv_name.lower()]


def _sigfig(x: float, sig: int = 6) -> float:
    """Round x to sig significant figures."""
    if x == 0:
        return 0.0
    d = sig - 1 - int(floor(log10(abs(x))))
    return round(x, d)
    """Round x to sig significant figures."""
    if x == 0:
        return 0.0
    d = sig - 1 - int(floor(log10(abs(x))))
    return round(x, d)


# ==============================================================================
# FUEL GAS CONFIG HELPERS
# Build repetitive gas config dicts so the FUELS list stays concise.
# Produce identical output to writing the dicts longhand — no logic change.
# ==============================================================================

def static_gases(sheet: str, sector: str, subsector: str = None) -> list:
    """
    Return CO2/CH4/N2O gas configs that all read from the same static_sector
    table (same sheet, sector, and optional subsector).
    Covers the majority of liquid and gaseous fossil fuels.
    """
    return [
        {"gas": g, "table_type": "static_sector", "sheet": sheet,
         "sector": sector, "gas_col": g, "subsector": subsector}
        for g in ("CO2", "CH4", "N2O")
    ]


def zero_gases(reason: str) -> list:
    """
    Return CO2/CH4/N2O gas configs that are all zero by assumption.
    Covers fuels where all GHG emissions are zero at point of use
    (e.g. Uranium, Hydrogen, Electricity, Metallurgical Coal, Asphalt).
    """
    return [
        {"gas": g, "table_type": "zero_assumption", "reason": reason}
        for g in ("CO2", "CH4", "N2O")
    ]


def process_gases(sheet: str, sector: str, col_index: int = None) -> list:
    """
    Return CO2/CH4/N2O gas configs for non-energy process fuels.
    CO2 is read from a static_sector table; CH4 and N2O are zero because
    the NIR only reports CO2 for non-energy use (NIR A6.2.4).
    """
    co2_cfg = {"gas": "CO2", "table_type": "static_sector", "sheet": sheet,
               "sector": sector, "gas_col": "CO2"}
    if col_index is not None:
        co2_cfg["col_index"] = col_index
    zero_reason = "Non-energy process use -- NIR A6 reports CO2 only"
    return [
        co2_cfg,
        {"gas": "CH4", "table_type": "zero_assumption", "reason": zero_reason},
        {"gas": "N2O", "table_type": "zero_assumption", "reason": zero_reason},
    ]


def biogenic_gases(sheet: str, sector: str, subsector: str = None) -> list:
    """
    Return CH4 and N2O gas configs for biogenic / renewable drop-in fuels.
    These use the same static_sector table as the fossil counterpart.

    CO2 is specified inline per fuel (not here) because:
      - A6.1-x tables only report CH4/N2O for biogenic fuels -- no CO2 column.
      - CO2 must be read from the fossil-counterpart table (static_sector or
        annual_regional depending on the fuel), with "biogenic": True to signal
        that build_cims_table should label it "emissions_biomass".
    """
    return [
        {"gas": "CH4", "table_type": "static_sector", "sheet": sheet,
         "sector": sector, "gas_col": "CH4", "subsector": subsector},
        {"gas": "N2O", "table_type": "static_sector", "sheet": sheet,
         "sector": sector, "gas_col": "N2O", "subsector": subsector},
    ]


# ==============================================================================
# FUELS
# ==============================================================================
FUELS = [

    # -- Gaseous fossil fuels --------------------------------------------------

    {
        "fuel_name":      "Natural Gas",
        "energy_content": _gj_per_m3("Natural Gas"),   # GJ/m3
        "source_unit":    "g/m3",
        "gases": [
            {"gas": "CO2", "table_type": "annual_regional",
             "sheet": "Table A6.1–1", "region_col": "Canada"},
            {"gas": "CH4", "table_type": "static_sector",
             "sheet": "Table A6.1–3",
             "sector": "Residential, Construction, Commercial/Institutional, Agriculture",
             "gas_col": "CH4"},
            {"gas": "N2O", "table_type": "static_sector",
             "sheet": "Table A6.1–3",
             "sector": "Residential, Construction, Commercial/Institutional, Agriculture",
             "gas_col": "N2O"},
        ],
    },
    {
        # Still gas (refinery fuel gas) has mixed source units across gases:
        #   CO2: Table A6.1-7 in g/m3 (gas volume basis)
        #   CH4: Table A6.1-9 in g/m3 (gas volume basis)
        #   N2O: Table A6.1-6 in g/L  (liquid volume basis)
        #
        # For CO2 and CH4, energy_content of 0.0373 GJ/m3 is used as a proxy for still gas in gas form (same as natural gas). Still gas
        # is primarily methane and light hydrocarbons (i.e. ethane). Using natural gas energy content is a defensible approximation that 
        # avoids an uncertain density conversion. Still gas is likely slightly richer than pipeline gas, so 0.0373 GJ/m3 may modestly 
        # understate energy content, meaning t/GJ factors could be slightly overstated. Source: CER Conversions (natural gas proxy).
        # 
        # For N2O, the NIR value of 0.00002 g/L (Table A6.1-6) is on a liquid volume basis and is pre-converted using the CER still gas
        # liquid energy content (41.73 GJ/m3 = 0.04173 GJ/L): 
        #   0.00002 g/L / 0.04173 GJ/L = 4.794e-7 tN2O/GJ
        "fuel_name":      "Refinery Fuel Gas",
        "energy_content": _gj_per_m3("Natural Gas"),   # GJ/m3 (natural gas proxy for gas-phase still gas)
        "source_unit":    "g/m3",
        "gases": [
            {"gas": "CO2", "table_type": "annual_regional",
             "sheet": "Table A6.1–7", "region_col": "Refineries and Othersc"},  # footnote letter needed to disambiguate from petcoke column
            {"gas": "CH4", "table_type": "annual_with_ranges",
             "sheet": "Table A6.1–9", "value_col_idx": 2, "year_col_idx": 1},
            {"gas": "N2O", "table_type": "direct_value",
             "value": 4.794e-10,  # tN2O/GJ -- NIR Table A6.1-6 (0.00002 g/L) / CER 0.04173 GJ/L / 1,000,000
             "source_override": NIR_SOURCE + ", Table A6.1-6 (Still Gas, liquid basis, converted using CER 0.04173 GJ/L)"},
        ],
    },
    {
        "fuel_name":      "Propane",
        "energy_content": _gj_per_L("Propane"),        # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–5", "Propane", subsector="Residential"),
    },
    {
        "fuel_name":      "Butane",
        "energy_content": _gj_per_L("Butane"),         # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–5", "Butane"),
    },
    {
        "fuel_name":      "Ethane",
        "energy_content": _gj_per_L("Ethane"),         # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–5", "Ethane"),
    },
    {
        # LPG is not present in NIR Annex 6. CO2 from IPCC 2006 Vol.2 Table 1.4: 
        #   63,100 kg CO2/TJ = 0.0631 tCO2/GJ.
        # CH4/N2O from IPCC 2006 Vol.2 Table 2.3 (Manufacturing Industries sector):
        #   CH4 = 5 kg/TJ = 0.000005 tCH4/GJ; N2O = 0.1 kg/TJ = 0.0000001 tN2O/GJ.
        "fuel_name":      "LPG",
        "energy_content": _gj_per_L("LPG"),            # GJ/L
        "source_unit":    "g/L",
        "gases": [
            {"gas": "CO2", "table_type": "direct_value",
             "value": 0.0631,     # tCO2/GJ -- IPCC 2006 Vol.2 Table 1.4
             "source_override": IPCC_SOURCE + ", Table 1.4"},
            {"gas": "CH4", "table_type": "direct_value",
             "value": 0.000005,   # tCH4/GJ -- IPCC 2006 Vol.2 Table 2.3 (Manufacturing)
             "source_override": IPCC_SOURCE + ", Table 2.3"},
            {"gas": "N2O", "table_type": "direct_value",
             "value": 0.0000001,  # tN2O/GJ -- IPCC 2006 Vol.2 Table 2.3 (Manufacturing)
             "source_override": IPCC_SOURCE + ", Table 2.3"},
        ],
    },

    # -- Liquid fossil fuels ---------------------------------------------------

    {
        "fuel_name":      "Heavy Fuel Oil",
        "energy_content": _gj_per_L("Heavy fuel oil"), # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–6", "Heavy Fuel Oil", subsector="Industrial"),
    },
    {
        "fuel_name":      "Light Fuel Oil",
        "energy_content": _gj_per_L("Heating Oil"),    # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–6", "Light Fuel Oil", subsector="Residential"),
    },
    {
        "fuel_name":      "Gasoline",
        "energy_content": _gj_per_L("Motor gasoline"), # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–6", "Motor Gasolinee"),
    },
    {
        "fuel_name":      "Diesel",
        "energy_content": _gj_per_L("Diesel"),         # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–15", "Heavy-duty Diesel Vehicles (HDDVs)",
                              subsector="Advanced Control"),
    },
    {
        "fuel_name":      "Kerosene",
        "energy_content": _gj_per_L("Kerosene"),       # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–6", "Kerosene", subsector="Residential"),
    },
    {
        "fuel_name":      "Jet Fuel",
        "energy_content": _gj_per_L("Jet Fuel (Jet A-1)"),  # GJ/L
        "source_unit":    "g/L",
        "gases": static_gases("Table A6.1–15", "Aviation Turbo Fuel"),
    },
    {
        # CO2: Table A6.1-7 in g/L (liquid volume, refinery accounting convention)
        # CH4: Table A6.1-6 in g/L (liquid volume, refinery accounting convention)
        # N2O: Table A6.1-8 in g/m3 (liquid volume, same convention, just unscaled)
        # Petcoke is physically a solid but tracked volumetrically alongside other refined petroleum products in refinery accounting. 
        # All three gases are on a liquid-equivalent volume basis. CO2 and CH4 use 0.04238 GJ/L (CER).N2O uses 42.38 GJ/m3 via 
        # energy_content_override -- identical to 0.04238 GJ/L (x1000 L/m3) -- to match the g/m3 source units in Table A6.1-8.
        "fuel_name":      "Petroleum Coke",
        "energy_content": _gj_per_L("Petroleum coke"), # GJ/L
        "source_unit":    "g/L",
        "gases": [
            {"gas": "CO2", "table_type": "annual_regional",
             "sheet": "Table A6.1–7", "region_col": "Refineries and Othersb"},  # footnote letter needed to disambiguate from still gas column
            {"gas": "CH4", "table_type": "static_sector",
             "sheet": "Table A6.1–6", "sector": "Petroleum Coke", "gas_col": "CH4"},
            {"gas": "N2O", "table_type": "annual_regional",
             "sheet": "Table A6.1–8", "region_col": "Refineries and Others",
             "energy_content_override": _EC["petroleum coke"]},  # GJ/m3 liquid = GJ/L x1000, matches g/m3 source units
        ],
    },

    # -- Coal ------------------------------------------------------------------

    {
        # Source:CEEDC_coal_electricity_energy_emission_factors_260313.xlsx. Province-level weighted averages derived from NIR Tables A6.1-10 (CO2) and A6.1-12 (CH4, N2O). 
        # Values already in t/GJ -- no unit conversion needed. Canada national average: rows 227 (CO2), 237 (CH4), 246 (N2O), years 1990-2030.
        "fuel_name":      "Thermal Coal",
        "energy_content": 1.0,      # Placeholder -- values already in t/GJ
        "source_unit":    "t/GJ",
        "gases": [
            {"gas": "CO2", "table_type": "row_index",
             "xlsx_path": COAL_XLSX, "sheet": "NIR_Coal_Elec",
             "row_index": 226, "year_row_index": 0, "data_col_start": 3,
             "source_override": COAL_SOURCE_CO2},
            {"gas": "CH4", "table_type": "row_index",
             "xlsx_path": COAL_XLSX, "sheet": "NIR_Coal_Elec",
             "row_index": 236, "year_row_index": 0, "data_col_start": 3,
             "source_override": COAL_SOURCE_CH4},
            {"gas": "N2O", "table_type": "row_index",
             "xlsx_path": COAL_XLSX, "sheet": "NIR_Coal_Elec",
             "row_index": 245, "year_row_index": 0, "data_col_start": 3,
             "source_override": COAL_SOURCE_N2O},
        ],
    },

    # -- Waste fuel ------------------------------------------------------------

    {
        # Source units are kg/GJ -- divide by 1000 to get t/GJ (no energy_content needed).
        "fuel_name":           "Waste",
        "energy_content":      None,
        "units_are_kg_per_GJ": True,
        "source_unit":         "kg/GJ",
        "gases": [
            {"gas": "CO2", "table_type": "annual_with_ranges",
             "sheet": "Table A6.1–14", "value_col_idx": 2, "year_col_idx": 1},
            # CH4/N2O use annual_with_ranges so that if ECCC begins reporting annual values in a future NIR update they are captured
            # automatically. Currently these values are static across all years in Table A6.1-14(CH4 col index 3, N2O col index 4) 
            # so the output will be constant year-over-year until the NIR reports variation.
            {"gas": "CH4", "table_type": "annual_with_ranges",
             "sheet": "Table A6.1–14", "value_col_idx": 3, "year_col_idx": 1},
            {"gas": "N2O", "table_type": "annual_with_ranges",
             "sheet": "Table A6.1–14", "value_col_idx": 4, "year_col_idx": 1},
        ],
    },

    # -- Process / non-energy fuels --------------------------------------------

    {
        # CO2 from NIR Section A6.2.4 found in text, not table (Cheminfo Services 2005). CH4/N2O not reported for non-energy process use.
        "fuel_name":               "Natural Gas Feedstock",
        "energy_content":          _gj_per_m3("Natural Gas"),  # GJ/m3  (same as Natural Gas)
        "source_unit":             "g/m3",
        "emission_source_override": "process",
        "gases": [
            {"gas": "CO2", "table_type": "hardcoded",
             "value": 38.0},  # g CO2/m3  (NIR Annex 6, Section A6.2.4, Cheminfo Services 2005)
            {"gas": "CH4", "table_type": "zero_assumption",
             "reason": "Non-energy process use -- NIR A6 reports CO2 only"},
            {"gas": "N2O", "table_type": "zero_assumption",
             "reason": "Non-energy process use -- NIR A6 reports CO2 only"},
        ],
    },
    {
        "fuel_name":               "Petrochemical Feedstock",
        "energy_content":          _gj_per_L("Petrochemical feedstock"),  # GJ/L
        "source_unit":             "g/L",
        "emission_source_override": "process",
        "gases": process_gases("Table A6.2–10", "Petrochemical Feedstocksc", col_index=3),
    },
    {
        "fuel_name":               "Naphtha Specialties",
        "energy_content":          _gj_per_L("Naphtha specialties"),      # GJ/L
        "source_unit":             "g/L",
        "emission_source_override": "process",
        "gases": process_gases("Table A6.2–10", "Naphthase", col_index=3),
    },
    {
        "fuel_name":               "Lubricants",
        "energy_content":          _gj_per_L("Lubes and greases"),        # GJ/L
        "source_unit":             "g/L",
        "emission_source_override": "process",
        "gases": process_gases("Table A6.2–10", "Lubricating Oils and Greasesf", col_index=3),
    },
    {
        "fuel_name":               "Other Non-Energy Products",
        "energy_content":          _gj_per_L("Other products"),           # GJ/L
        "source_unit":             "g/L",
        "emission_source_override": "process",
        "gases": process_gases("Table A6.2–10", "Petroleum Used for Other Productsh", col_index=3),
    },

    # -- Biogenic fuels --------------------------------------------------------

    {
        # CO2 from NIR Table A6.1-15 (Renewable Fuels / Ethanol), col_index 2: 1508.04 g/L.
        # CH4/N2O: NIR footnote ** directs use of gasoline factors (Table A6.1-15, HDDV not applicable
        # for transport; use A6.1-6 Motor Gasoline per NIR note in the script header).
        # CO2 labelled "emissions_biomass" in CIMS output.
        "fuel_name":      "Ethanol",
        "energy_content": _gj_per_L("Ethanol"),        # GJ/L
        "source_unit":    "g/L",
        "gases": [
            {"gas": "CO2", "table_type": "static_sector",
             "sheet": "Table A6.1\u201315", "sector": "Ethanol",
             "gas_col": "CO2", "col_index": 2, "biogenic": True},
        ] + biogenic_gases("Table A6.1\u20136", "Motor Gasolinee"),
    },
    {
        # CO2 from NIR Table A6.1-15 (Renewable Fuels / Biodiesel), col_index 2: 2472.2 g/L.
        # CH4/N2O: NIR footnote *** directs use of diesel factors (Table A6.1-15 HDDV Advanced Control).
        # CO2 labelled "emissions_biomass" in CIMS output.
        "fuel_name":      "Biodiesel",
        "energy_content": _gj_per_L("Biodiesel"),      # GJ/L
        "source_unit":    "g/L",
        "gases": [
            {"gas": "CO2", "table_type": "static_sector",
             "sheet": "Table A6.1\u201315", "sector": "Biodiesel",
             "gas_col": "CO2", "col_index": 2, "biogenic": True},
        ] + biogenic_gases("Table A6.1\u201315", "Heavy-duty Diesel Vehicles (HDDVs)",
                           subsector="Advanced Control"),
    },
    {
        # CO2/CH4/N2O from NIR Table A6.6-1 (Spent Pulping Liquor / Industrial Combustion).
        # Source: NIR table based on NCASI (2011/2012). Units: g/kg fuel.
        # CO2 labelled "emissions_biomass" in CIMS output.
        # Energy content: Statistics Canada, Energy Statistics Handbook, Cat. no. 57-601-X, Appendix A.
        "fuel_name":               "Black Liquor",
        "energy_content_kg":       _gj_per_kg("Black liquor"),  # GJ/kg
        "energy_content":          None,    # Not used -- energy_content_kg takes precedence
        "source_unit":             "g/kg",
        "emission_source_override": "combustion",
        "gases": [
            {"gas": "CO2", "table_type": "static_sector",
             "sheet": "Table A6.6\u20131", "sector": "Spent Pulping Liquor",
             "gas_col": "CO2", "col_index": 3, "biogenic": True},
            {"gas": "CH4", "table_type": "static_sector",
             "sheet": "Table A6.6\u20131", "sector": "Spent Pulping Liquor",
             "gas_col": "CH4", "col_index": 4},
            {"gas": "N2O", "table_type": "static_sector",
             "sheet": "Table A6.6\u20131", "sector": "Spent Pulping Liquor",
             "gas_col": "N2O", "col_index": 5},
        ],
    },
    {
        # CO2/CH4/N2O from NIR Table A6.6-1 (Wood Fuel / Wood Waste / Industrial Combustion).
        # Source: NIR table based on US EPA (2003) and NCASI (2012). Units: g/kg fuel.
        # CO2 labelled "emissions_biomass" in CIMS output.
        # Energy content: Statistics Canada, Energy Statistics Handbook, Cat. no. 57-601-X, Appendix A.
        "fuel_name":               "Solid Biomass",
        "energy_content_kg":       _gj_per_kg("Solid Biomass"),  # GJ/kg
        "energy_content":          None,
        "source_unit":             "g/kg",
        "emission_source_override": "combustion",
        "gases": [
            {"gas": "CO2", "table_type": "static_sector",
             "sheet": "Table A6.6\u20131", "sector": "Wood Fuel / Wood Waste",
             "gas_col": "CO2", "col_index": 3, "biogenic": True},
            {"gas": "CH4", "table_type": "static_sector",
             "sheet": "Table A6.6\u20131", "sector": "Wood Fuel / Wood Waste",
             "gas_col": "CH4", "col_index": 4},
            {"gas": "N2O", "table_type": "static_sector",
             "sheet": "Table A6.6\u20131", "sector": "Wood Fuel / Wood Waste",
             "gas_col": "N2O", "col_index": 5},
        ],
    },

    # -- Renewable drop-in fuels -----------------------------------------------
    # "Renewable" fuels are chemically similar to their fossil fuel counterparts and combust
    # under similar conditions. Emission factors assumed same as fossil fuel counterpart.
    # CO2 labelled "emissions_biomass" in CIMS output.
    {
        # RNG assumed to be upgraded to pipeline quality -- chemically the same as fossil NG
        # once upgraded. Burns in the same appliances; CH4/N2O combustion behaviour
        # indistinguishable. Note: this assumption may not hold for non-upgraded biogas.
        # CO2: annual_regional from Table A6.1-1 (same as Natural Gas), biogenic label only.
        "fuel_name":      "Renewable Natural Gas",
        "energy_content": _gj_per_m3("Natural Gas"),   # GJ/m3  (same as Natural Gas)
        "source_unit":    "g/m3",
        "gases": [
            {"gas": "CO2", "table_type": "annual_regional",
             "sheet": "Table A6.1\u20131", "region_col": "Canada", "biogenic": True},
        ] + biogenic_gases("Table A6.1\u20133",
                           "Residential, Construction, Commercial/Institutional, Agriculture"),
    },
    {
        # CO2 EF: same as Jet Fuel (A6.1-15, Aviation Turbo Fuel); biogenic label only.
        "fuel_name":      "SAF",
        "energy_content": _gj_per_L("Jet Fuel (Jet A-1)"),  # GJ/L  (same as Jet Fuel)
        "source_unit":    "g/L",
        "gases": [
            {"gas": "CO2", "table_type": "static_sector",
             "sheet": "Table A6.1–15", "sector": "Aviation Turbo Fuel",
             "gas_col": "CO2", "biogenic": True},
        ] + biogenic_gases("Table A6.1–15", "Aviation Turbo Fuel"),
    },
    {
        # CO2 EF: same as Diesel (A6.1-15, HDDV Advanced Control); biogenic label only.
        "fuel_name":      "Renewable Diesel",
        "energy_content": _gj_per_L("Diesel"),              # GJ/L  (same as Diesel -- drop-in fuel)
        "source_unit":    "g/L",
        "gases": [
            {"gas": "CO2", "table_type": "static_sector",
             "sheet": "Table A6.1–15", "sector": "Heavy-duty Diesel Vehicles (HDDVs)",
             "gas_col": "CO2", "subsector": "Advanced Control", "biogenic": True},
        ] + biogenic_gases("Table A6.1–15", "Heavy-duty Diesel Vehicles (HDDVs)",
                           subsector="Advanced Control"),
    },
    {
        # CO2 EF: same as Gasoline (A6.1-6, Motor Gasoline); biogenic label only.
        "fuel_name":      "Renewable Gasoline",
        "energy_content": _gj_per_L("Motor gasoline"),      # GJ/L  (same as Gasoline -- drop-in fuel)
        "source_unit":    "g/L",
        "gases": [
            {"gas": "CO2", "table_type": "static_sector",
             "sheet": "Table A6.1–6", "sector": "Motor Gasolinee",
             "gas_col": "CO2", "biogenic": True},
        ] + biogenic_gases("Table A6.1–6", "Motor Gasolinee"),
    },

    # -- Zero-EF fuels ---------------------------------------------------------
    # These fuels are retained in FUELS so energy_map.csv validation passes and
    # their zero-EF treatment is documented, but "exclude_from_output": True
    # prevents them from appearing in either output CSV.

    {
        # Metallurgical coal is a feedstock input -- emissions accounted for elsewhere in the model.
        "fuel_name":            "Metallurgical Coal",
        "exclude_from_output":  True,
        "emission_source_override": "Process",
        "energy_content":       1.0,
        "source_unit":          "N/A",
        "gases": zero_gases("Emissions accounted for elsewhere in the model"),
    },
    {
        # Nuclear fission -- no carbon combustion; zero GHG at point of use.
        "fuel_name":            "Uranium",
        "exclude_from_output":  True,
        "emission_source_override": "combustion",
        "energy_content":       1.0,
        "source_unit":          "N/A",
        "gases": zero_gases("Nuclear fission -- no carbon combustion; zero GHG at point of use"),
    },
    {
        # H2 combustion produces water only; no carbon emissions at point of use.
        # Upstream production emissions (i.e. grey/blue/green H2) excluded from combustion EFs.
        "fuel_name":            "Hydrogen",
        "exclude_from_output":  True,
        "energy_content":       1.0,
        "source_unit":          "N/A",
        "gases": zero_gases("H2 combustion produces water only; no carbon emissions at point of use"),
    },
    {
        # Energy carrier -- emissions attributed at the point of generation.
        # Applying an EF at the consumption side would double-count generation emissions.
        "fuel_name":            "Electricity",
        "exclude_from_output":  True,
        "energy_content":       1.0,
        "source_unit":          "N/A",
        "gases": zero_gases("Emissions attributed at generation; applying EF at consumption would double-count"),
    },
    {
        # Non-combusted paving material. Slow bitumen oxidation over time is negligible and not reported in the NIR.
        # NIR Annex 6 contains no asphalt emission factor.
        "fuel_name":            "Asphalt",
        "exclude_from_output":  True,
        "emission_source_override": "Process",
        "energy_content":       1.0,
        "source_unit":          "N/A",
        "gases": zero_gases("Non-combusted material -- no point-of-use GHG emissions; NIR Annex 6 contains no asphalt EF"),
    },
]


# ==============================================================================
# HELPERS
# ==============================================================================

def _read_sheet_raw(xlsx_path: str, sheet: str) -> list:
    """
    Read an xlsx sheet into a plain list-of-lists.
    Raises a clear error if the file or sheet is not found.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(
            f"Source file not found: '{xlsx_path}'\n"
            f"  Expected location: {path.resolve()}\n"
            f"  Copy the xlsx into the same folder as this script and re-run."
        )
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet}' not found in '{xlsx_path}'.\n"
            f"  Available sheets: {wb.sheetnames}\n"
            f"  The NIR table layout may have changed — check the config."
        )
    ws = wb[sheet]
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    wb.close()
    return rows


def _clean(v) -> str:
    """
    Normalise a raw cell value to a plain string.
    Strips: non-breaking spaces, thousands-separator spaces, unit suffixes
    (e.g. 'g/kg'), and trailing footnote letters from numeric strings.
    """
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    s = re.sub(r"\s*g/(kg|m3|L|m\u00b3)\s*$", "", s).strip()
    if re.match(r"^\d[\d ]*[\d.]+[a-z]?$", s) and " " in s:
        s = s.replace(" ", "")
    s = re.sub(r"^(\d+\.?\d*)[a-z,\s]+$", r"\1", s)
    return s


def g_to_t_per_GJ(ef_g_per_unit: float, energy_content_GJ: float) -> float:
    """
    Convert g/[unit] to t/GJ.
      g/unit / GJ/unit = g/GJ  -->  / 1,000,000  -->  t/GJ
    """
    return (ef_g_per_unit / energy_content_GJ) / 1_000_000


def expand_years(year_ef: dict, year_start: int, year_end: int,
                  label: str = "") -> dict:
    """
    Expand a sparse {year: value} dict to cover every year in [year_start, year_end].
    - Back-fills years before the first known year.
    - Linearly interpolates gaps between known years.
    - Forward-fills years after the last known year.
    Warns if year_end exceeds the last known data year (forward-fill applied).
    """
    known_years = sorted(year_ef)
    first_yr, last_yr = known_years[0], known_years[-1]
    if last_yr < year_end:
        warnings.warn(
            f"{label or 'Unknown'}: last data year is {last_yr}, "
            f"but YEAR_END={year_end}. "
            f"Values for {last_yr + 1}–{year_end} will be forward-filled from {last_yr}.",
            stacklevel=3,
        )
    result = {}
    for yr in range(year_start, year_end + 1):
        if yr in year_ef:
            result[yr] = year_ef[yr]
        elif yr < first_yr:
            result[yr] = year_ef[first_yr]
        elif yr > last_yr:
            result[yr] = year_ef[last_yr]
        else:
            lo = max(y for y in known_years if y < yr)
            hi = min(y for y in known_years if y > yr)
            slope = (year_ef[hi] - year_ef[lo]) / (hi - lo)
            result[yr] = year_ef[lo] + slope * (yr - lo)
    return result


# ==============================================================================
# PARSERS
# ==============================================================================

def parse_annual_regional(xlsx_path: str, sheet: str, region_col: str) -> dict:
    """
    Parse a NIR table that has year rows and region columns.
    Returns {year: EF_value} for the chosen region column.

    region_col is matched using startswith() so that trailing footnote letters
    (e.g. "Refineries and Othersb") do not need to be included. If multiple
    columns start with the same prefix the first match is used and a warning
    is printed so the ambiguity can be investigated.
    """
    rows = _read_sheet_raw(xlsx_path, sheet)

    header_row_idx = next(
        (i for i, row in enumerate(rows) if any(_clean(v) == "Year" for v in row)),
        None,
    )
    if header_row_idx is None:
        raise ValueError(f"Could not find 'Year' header in {sheet}")

    year_row     = rows[header_row_idx]
    year_col_idx = next(i for i, v in enumerate(year_row) if _clean(v) == "Year")

    def _find_region(row, region_col):
        """Return column index using startswith matching; warn if ambiguous."""
        matches = [i for i, v in enumerate(row) if _clean(v).startswith(region_col)]
        if len(matches) > 1:
            print(f"  ⚠  WARNING: Multiple columns start with '{region_col}' "
                  f"in {sheet} -- using first match. Check for ambiguity.")
        return matches[0] if matches else None

    # Region label may be on the same row as Year, or on the next row (two-row headers)
    region_col_idx = _find_region(year_row, region_col)
    if region_col_idx is not None:
        data_start_idx = header_row_idx + 1
    else:
        region_row     = rows[header_row_idx + 1]
        region_col_idx = _find_region(region_row, region_col)
        if region_col_idx is None:
            available = [_clean(v) for v in region_row if _clean(v)]
            raise ValueError(f"Region '{region_col}' not found. Available: {available}")
        data_start_idx = header_row_idx + 2

    result = {}
    for row in rows[data_start_idx:]:
        try:
            yr  = int(float(row[year_col_idx]))
            val = float(_clean(row[region_col_idx]))
            result[yr] = val
        except (TypeError, ValueError, IndexError):
            continue
    return result


def parse_annual_with_ranges(xlsx_path: str, sheet: str, value_col_idx: int,
                              year_col_idx: int = 1) -> dict:
    """
    Parse a NIR table where the year column may contain single years ('2005')
    or year ranges ('1990-96', '2018-2023').
    Returns {year: value}.
    """
    rows = _read_sheet_raw(xlsx_path, sheet)
    result = {}
    for row in rows:
        yr_raw  = _clean(row[year_col_idx])  if len(row) > year_col_idx  else ""
        val_raw = _clean(row[value_col_idx]) if len(row) > value_col_idx else ""
        if not yr_raw or not val_raw:
            continue
        try:
            val = float(val_raw)
        except ValueError:
            continue
        if re.match(r"^\d{4}$", yr_raw):
            result[int(yr_raw)] = val
        else:
            m = re.match(r"^(\d{4})[\u2013-](\d{2,4})$", yr_raw)
            if m:
                start   = int(m.group(1))
                end_raw = m.group(2)
                end     = int(str(start)[:2] + end_raw) if len(end_raw) == 2 else int(end_raw)
                for y in range(start, end + 1):
                    result[y] = val
    return result


def parse_static_sector(xlsx_path: str, sheet: str, sector: str, gas_col: str,
                         subsector: str = None, col_index: int = None) -> float:
    """
    Parse a NIR table with sector/subsector rows and gas columns.
    Returns a single EF float value.

    subsector=None  : value taken directly from the sector row.
    subsector=<str> : finds the sector heading first, then the subsector row beneath it.
    col_index=<int> : use this column index directly (for non-standard table headers).
    """
    rows = _read_sheet_raw(xlsx_path, sheet)

    if col_index is not None:
        gas_col_idx = col_index
        data_start  = 1
    else:
        gas_row_idx = next(
            (i for i, row in enumerate(rows)
             if any(_clean(v) in ("CH4", "CO2", "N2O", "CO2a", "CH4b", "N2Oc") for v in row)),
            None,
        )
        if gas_row_idx is None:
            raise ValueError(f"Could not locate gas header row in {sheet}")
        gas_row = rows[gas_row_idx]
        gas_col_idx = next(
            (i for i, v in enumerate(gas_row) if _clean(v).startswith(gas_col)), None
        )
        if gas_col_idx is None:
            available = [_clean(v) for v in gas_row if _clean(v)]
            raise ValueError(f"Gas column '{gas_col}' not found. Available: {available}")
        data_start = gas_row_idx + 1

    if subsector is None:
        for row in rows[data_start:]:
            if _clean(row[1]) == sector:
                try:
                    return float(_clean(row[gas_col_idx]))
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Could not parse value for sector '{sector}' / gas '{gas_col}'"
                    )
        raise ValueError(f"Sector '{sector}' not found in {sheet}")

    in_sector = False
    for row in rows[data_start:]:
        label = _clean(row[1])
        if label == sector:
            in_sector = True
            continue
        if in_sector:
            if label == subsector:
                try:
                    return float(_clean(row[gas_col_idx]))
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Could not parse value for sector '{sector}' / "
                        f"subsector '{subsector}' / gas '{gas_col}'"
                    )
            if label and not any(_clean(v) for v in row[2:] if v):
                break
    raise ValueError(
        f"Subsector '{subsector}' under sector '{sector}' not found in {sheet}"
    )


def parse_row_by_index(xlsx_path: str, sheet: str, row_index: int,
                       year_row_index: int = 0, data_col_start: int = 3) -> dict:
    """
    Parse a row where values are already in final units (no conversion needed).
    Year headers are read from year_row_index (0-based row number).
    Data columns start at data_col_start (0-based column index).
    Returns {year: value}.
    """
    rows     = _read_sheet_raw(xlsx_path, sheet)
    year_row = rows[year_row_index]
    data_row = rows[row_index]
    result   = {}
    for col_idx in range(data_col_start, len(year_row)):
        year_val = year_row[col_idx]
        data_val = data_row[col_idx] if col_idx < len(data_row) else None
        if year_val is None or data_val is None:
            continue
        try:
            result[int(year_val)] = float(data_val)
        except (ValueError, TypeError):
            continue
    return result


# ==============================================================================
# BUILD RECORDS
# ==============================================================================

def _row(year, fuel, gas, emission_source, source, units, value) -> dict:
    """Return a single output record as a dict."""
    return {
        "year":            year,
        "fuel":            fuel,
        "emissions_type":  gas,
        "emission_source": emission_source,
        "source":          source,
        "value":           value,
        "units":           units,
    }


def build_records(year_start: int = YEAR_START, year_end: int = YEAR_END) -> pl.DataFrame:
    """
    Iterate over FUELS and produce one record per (fuel, gas, year).
    Returns a polars DataFrame sorted by fuel, emissions_type, year.
    """
    records = []

    for fuel in FUELS:
        fuel_name      = fuel["fuel_name"]
        energy_content = fuel.get("energy_content_kg") or fuel["energy_content"]
        emission_src   = fuel.get("emission_source_override", DEFAULT_EMISSION_SOURCE)
        kg_per_gj      = fuel.get("units_are_kg_per_GJ", False)

        for gas_cfg in fuel["gases"]:
            gas        = gas_cfg["gas"]
            table_type = gas_cfg["table_type"]
            sheet      = gas_cfg.get("sheet")
            units      = f"t{gas}/GJ"

            if table_type == "annual_regional":
                ec       = gas_cfg.get("energy_content_override", energy_content)
                raw      = parse_annual_regional(NIR_XLSX, sheet, gas_cfg["region_col"])
                expanded = expand_years(raw, year_start, NIR_YEAR, f"{fuel_name} {gas}")
                expanded = extend_series_constant(pd.Series(expanded), base_year=NIR_YEAR, end_year=year_end + 1)
                for yr, val in expanded.items():
                    records.append(_row(yr, fuel_name, gas, emission_src, NIR_SOURCE, units,
                                        _sigfig(g_to_t_per_GJ(val, ec))))

            elif table_type == "annual_with_ranges":
                raw      = parse_annual_with_ranges(NIR_XLSX, sheet,
                                                     gas_cfg["value_col_idx"],
                                                     gas_cfg.get("year_col_idx", 1))
                expanded = expand_years(raw, year_start, NIR_YEAR, f"{fuel_name} {gas}")
                expanded = extend_series_constant(pd.Series(expanded), base_year=NIR_YEAR, end_year=year_end + 1)
                for yr, val in expanded.items():
                    converted = val / 1000 if kg_per_gj else g_to_t_per_GJ(val, energy_content)
                    records.append(_row(yr, fuel_name, gas, emission_src, NIR_SOURCE, units, converted))

            elif table_type == "static_sector":
                val = parse_static_sector(NIR_XLSX, sheet, gas_cfg["sector"],
                                           gas_cfg["gas_col"],
                                           gas_cfg.get("subsector"),
                                           gas_cfg.get("col_index"))
                converted = val / 1000 if kg_per_gj else g_to_t_per_GJ(val, energy_content)
                for yr in range(year_start, year_end + 1):
                    records.append(_row(yr, fuel_name, gas, emission_src, NIR_SOURCE, units, converted))

            elif table_type == "hardcoded":
                converted = _sigfig(g_to_t_per_GJ(gas_cfg["value"], energy_content))
                for yr in range(year_start, year_end + 1):
                    records.append(_row(yr, fuel_name, gas, emission_src, NIR_SOURCE, units, converted))

            elif table_type == "direct_value":
                src = gas_cfg.get("source_override", NIR_SOURCE)
                val = _sigfig(gas_cfg["value"])
                for yr in range(year_start, year_end + 1):
                    records.append(_row(yr, fuel_name, gas, emission_src, src, units, val))

            elif table_type == "row_index":
                raw      = parse_row_by_index(gas_cfg["xlsx_path"], gas_cfg["sheet"],
                                               gas_cfg["row_index"],
                                               gas_cfg.get("year_row_index", 0),
                                               gas_cfg.get("data_col_start", 3))
                expanded = expand_years(raw, year_start, NIR_YEAR, f"{fuel_name} {gas}")
                expanded = extend_series_constant(pd.Series(expanded), base_year=NIR_YEAR, end_year=year_end + 1)
                src      = gas_cfg.get("source_override", NIR_SOURCE)
                for yr, val in expanded.items():
                    records.append(_row(yr, fuel_name, gas, emission_src, src, units, _sigfig(val)))

            elif table_type == "zero_assumption":
                src = gas_cfg.get("source_override", ASSUMPTION)
                for yr in range(year_start, year_end + 1):
                    records.append(_row(yr, fuel_name, gas, emission_src, src, units, 0.0))

            else:
                raise ValueError(f"Unknown table_type '{table_type}' for {fuel_name} / {gas}")

    return (
        pl.DataFrame(records)
        .with_columns(pl.col("year").cast(pl.Int32))
        .sort(["fuel", "emissions_type", "year"])
    )


# ==============================================================================
# MAIN
# ==============================================================================

PROCESS_FUELS = {
    "Natural Gas Feedstock", "Petrochemical Feedstock",
    "Naphtha Specialties", "Lubricants", "Other Non-Energy Products",
}

# Derived from FUELS -- any fuel with exclude_from_output=True is also expected
# to be all-zero, so no separate maintenance needed. If those two properties ever
# diverge, replace this with a hardcoded set.
EXPECTED_ZERO_FUELS = {f["fuel_name"] for f in FUELS if f.get("exclude_from_output")}
# Fuel/gas combos where zero CO2 is expected (biogenic fuels now carry actual EFs
# labelled as emissions_biomass; this set is intentionally empty but retained for
# any future fuels where CO2 is genuinely zero rather than biogenic)
EXPECTED_ZERO_CO2: set[str] = set()


def _is_expected_zero(fuel_name: str, gas: str) -> bool:
    """
    Return True if an all-zero value is expected and should not trigger a warning.
    Centralises the zero-expectation logic used by both validate() and the summary.
    Biogenic fuels are no longer in EXPECTED_ZERO_CO2 -- they now carry actual EFs
    labelled as emissions_biomass in the CIMS output.
    """
    if fuel_name in EXPECTED_ZERO_FUELS:
        return True
    if gas == "CO2" and fuel_name in EXPECTED_ZERO_CO2:
        return True
    return False


def validate(df: pl.DataFrame, year_start: int, year_end: int) -> list[str]:
    """
    Run sanity checks on the output DataFrame.
    Returns a list of warning strings (empty = all clear).
    """
    issues = []
    expected_years = year_end - year_start + 1

    for fuel in FUELS:
        name = fuel["fuel_name"]
        for gas in ("CO2", "CH4", "N2O"):
            subset = df.filter(
                (pl.col("fuel") == name) & (pl.col("emissions_type") == gas)
            )

            # Check row count
            if len(subset) != expected_years:
                issues.append(
                    f"ROW COUNT  {name} / {gas}: "
                    f"expected {expected_years} rows, got {len(subset)}"
                )
                continue

            vals = subset["value"].to_list()
            all_zero = all(v == 0.0 for v in vals)
            any_negative = any(v < 0.0 for v in vals)

            if any_negative:
                issues.append(f"NEGATIVE   {name} / {gas}: contains negative values")

            if all_zero and not _is_expected_zero(name, gas) and name not in PROCESS_FUELS:
                issues.append(
                    f"ALL ZEROS  {name} / {gas}: all values are zero "
                    f"(may indicate a failed table lookup)"
                )

    return issues


def build_cims_table(df: pl.DataFrame) -> pl.DataFrame:
    """
    Transform the primary emission-factors DataFrame into the CIMS parameter
    table format.  All logic is derived directly from the emission_factors.csv columns;
    no new data is introduced.

    Mapping
    -------
    Branch      : "CIMS.Generic Fuels.<fuel>"   (fuel = Service column)
    Type        : "Service"
    Region      : "CIMS"
    Sector      : "" (blank)
    Service     : fuel column from emission_factors.csv
    Technology  : "" (blank)
    Parameter   : "emissions_biomass" for biogenic-CO2 rows (biogenic fuels, CO2 only);
                  "emissions" for all other rows
    Context     : emissions_type  (CO2 / CH4 / N2O)
    Sub_Context : "Process" if emission_source == "process" else "Combustion" from emission_factors.csv
    Target      : "" (blank)
    Source      : source column from emission_factors.csv
    Unit        : units column from emission_factors.csv
    Year        : year column from emission_factors.csv
    Value       : value column from emission_factors.csv
    """
    # Fuels whose CO2 emissions are biogenic and must be labelled "emissions_biomass"
    _BIOGENIC_FUELS = pl.Series([
        "Ethanol", "Biodiesel", "Black Liquor", "Solid Biomass",
        "Renewable Natural Gas", "SAF", "Renewable Diesel", "Renewable Gasoline",
    ])

    return (
        df.with_columns([
            (pl.lit("CIMS.Generic Fuels.") + pl.col("fuel")).alias("Branch"),
            pl.lit("Service").alias("Type"),
            pl.lit("CIMS").alias("Region"),
            pl.lit("").alias("Sector"),
            pl.col("fuel").alias("Service"),
            pl.lit("").alias("Technology"),
            pl.when(
                pl.col("fuel").is_in(_BIOGENIC_FUELS) & (pl.col("emissions_type") == "CO2")
            )
              .then(pl.lit("emissions_biomass"))
              .otherwise(pl.lit("emissions"))
              .alias("Parameter"),
            pl.col("emissions_type").alias("Context"),
            pl.when(pl.col("emission_source") == "process")
              .then(pl.lit("Process"))
              .otherwise(pl.lit("Combustion"))
              .alias("Sub_Context"),
            pl.lit("").alias("Target"),
        ])
        .select([
            "Branch", "Type", "Region", "Sector", "Service",
            "Technology", "Parameter", "Context", "Sub_Context",
            "Target", "source", "units", "year", "value",
        ])
        .rename({"source": "Source", "units": "Unit", "year": "Year", "value": "Value"})
        .sort(["Service", "Context", "Year"])
    )


def main() -> None:
    """Run the full emission factors pipeline and write output CSVs."""
    print("=" * 80)
    print("EMISSION FACTORS EXTRACTION")
    print("=" * 80)
    print(f"  NIR source:  {NIR_XLSX.name}  (detected year: {NIR_YEAR})")
    print(f"  Coal source: {COAL_XLSX.name}")
    print(f"  Year range:  {YEAR_START}–{YEAR_END}")
    print("=" * 80)

    # -- Fuel map validation ---------------------------------------------------
    print("\nValidating fuels against energy_map.csv...")
    validate_fuel_map(FUELS)

    df = build_records()

    # -- Validation ------------------------------------------------------------
    issues = validate(df, YEAR_START, YEAR_END)
    if issues:
        print("\n⚠️  VALIDATION WARNINGS")
        for issue in issues:
            print(f"   ⚠️  {issue}")
    else:
        print("\n✅ Validation passed — no issues found")

    # -- Write outputs ---------------------------------------------------------
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Fuels with "exclude_from_output": True are retained in FUELS for documentation
    # and energy_map validation, but are excluded from both output CSVs.
    _excluded = {f["fuel_name"] for f in FUELS if f.get("exclude_from_output")}
    df_out = df.filter(~pl.col("fuel").is_in(pl.Series(list(_excluded))))

    output_path = OUTPUT_DIR / "emission_factors.csv"
    df_out.write_csv(output_path)
    print(f"\n✅ Emission factors complete")
    print(f"   Total rows:          {len(df_out):,}")
    print(f"   Fuels processed:     {df_out['fuel'].n_unique()}")
    print(f"   Years covered:       {df_out['year'].min()} – {df_out['year'].max()}")
    print(f"   Excluded (zero-EF):  {sorted(EXPECTED_ZERO_FUELS)}")
    print(f"   Saved to:            {output_path}")

    cims_path = OUTPUT_DIR / "emission_factors_CIMS.csv"
    cims_df = build_cims_table(df_out)
    cims_df.write_csv(cims_path)
    print(f"\n✅ CIMS parameter table complete")
    print(f"   Total rows:          {len(cims_df):,}")
    print(f"   Saved to:            {cims_path}")

    # -- Summary ---------------------------------------------------------------
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    for fuel in FUELS:
        for gas in ("CO2", "CH4", "N2O"):
            subset = df.filter(
                (pl.col("fuel") == fuel["fuel_name"]) & (pl.col("emissions_type") == gas)
            )
            if len(subset) > 0:
                mn   = subset["value"].min()
                mx   = subset["value"].max()
                flag = " ← all zero" if mn == 0.0 and mx == 0.0 and not _is_expected_zero(fuel["fuel_name"], gas) else ""
                print(f"  {fuel['fuel_name']:35s} {gas}: "
                      f"min={mn:.6e}  max={mx:.6e}  (n={len(subset)}){flag}")
    print("=" * 80)

    if issues:
        print("\n❌ Exiting with errors — review warnings above.")
        sys.exit(1)


if __name__ == "__main__":
    main()
