"""
Canada Coal Mining Production by Region, 2000–2024
===================================================

OUTPUT
------
A long CSV with columns:
    Region | Variable | Unit | Year | Value

Variables:
    Coal Mining                              — total production (kt)
    Coal Mining.Coal.Metallurgical Finishing — met share as fraction of total (% of kt, held flat to 2100)

USAGE
-----
Place the four input files in the same folder as this script, then run:

    python canada_coal_mining_by_region.py

The script writes canada_coal_mining_by_region_2000_2024.csv to the same folder.

DEPENDENCIES
------------
    pip install polars openpyxl

INPUT FILES
-----------
1.  25100048.csv                                — StatCan table 25-10-0048
2.  25100046.csv                                — StatCan table 25-10-0046
3.  bcannualcoalproduction.csv                  — BC Ministry annual production
4.  st98-2025-coal-production-demand-data.xlsx  — AER ST98 2025

================================================================================
DATA SOURCES AND METHODOLOGY
================================================================================

PRIMARY SOURCES
---------------

[SC48]  Statistics Canada, Table 25-10-0048-01
        "Coal statistics, monthly, by geography and coal type"
        Coverage: January 1946 – December 2007 (monthly)
        Variables used: "Total coal production, all types" by province
        URL: https://www150.statcan.gc.ca/t1/tbl1/en/dtbl/25100048

[SC46]  Statistics Canada, Table 25-10-0046-01
        "Coal production and coal use, monthly, by geography and coal type"
        Coverage: January 2008 – present (monthly)
        Variables used: "Total all coal types and uses" and "Bituminous, metallurgical"
        filtered to Coal volume == "Production", by province
        URL: https://www150.statcan.gc.ca/t1/tbl1/en/dtbl/25100046

[BC]    BC Ministry of Energy, Mines and Low Carbon Innovation
        "British Columbia Annual Coal Production and Value"
        Coverage: 1835–2020 (annual)
        URL: https://open.canada.ca/data/en/dataset/
             387692b1-ff47-4a68-ab36-e88591474e7a

[AER]   Alberta Energy Regulator, ST98: Alberta Energy Outlook 2025
        "Figure S7.1 — Alberta Marketable Coal Production"
        Read from the "Figures" sheet of st98-2025-coal-production-demand-data.xlsx
        Columns used: Year | Subbituminous (Mt) | Thermal Bituminous (Mt)
                            | Metallurgical Bituminous (Mt)
        Coverage: 1999–2024 actuals (rows where year > 2024 are forecasts, excluded)
        URL: https://www.aer.ca/data-and-performance-reports/statistical-reports/
             alberta-energy-outlook-st98/coal/coal-production

REGIONAL METHODOLOGY — TOTAL COAL PRODUCTION
---------------------------------------------

StatCan [SC48] and [SC46] are monthly series. Annual values are the sum of
12 monthly observations. Where StatCan suppresses monthly values (STATUS == 'x'),
those months contribute 0 to the sum; the number of suppressed months is counted
explicitly and used for scaling where needed.

BRITISH COLUMBIA
  2000–2020 → [BC] annual production CSV (tonnes converted to kt).
              The BC Ministry series is used in preference to [SC48]/[SC46]
              because it covers all suppressed years continuously. For 2000–2013
              it agrees closely with StatCan monthly sums; for 2014–2020 StatCan
              suppresses all BC monthly data (STATUS='x'), making the BC Ministry
              the only available source.
  2021–2024 → All 12 months of BC data remain suppressed in [SC46] and the BC
              Ministry CSV ends at 2020. BC production is estimated as:
                  BC = Canada_total (SC46) − Alberta − Saskatchewan
                       − New Brunswick − Nova Scotia
              This residual approach works because Alberta is complete in [SC46]
              for these years, and Saskatchewan/NB/NS are zero.

ALBERTA
  2000–2007 → [SC48]: all 12 months present, no suppression.
  2008–2013 → [SC46]: all 12 months present, no suppression.
  2014–2019 → [SC46] suppresses all Alberta monthly data. Total estimated as:
                  AER_metallurgical + AER_thermal_bituminous + AER_subbituminous
              from [AER] Figure S7.1, converted from Mt to kt.
  2020–2024 → [SC46]: all 12 months present, no suppression.

SASKATCHEWAN
  2000–2001 → [SC48]: complete.
  2002      → [SC48]: 6 of 12 months suppressed. Available months scaled to annual:
                  estimate = observed_sum × 12 / (12 − n_suppressed_months)
  2003      → [SC48]: all 12 months suppressed. The Canada total implies ~10,700 kt
              unaccounted. Treated as zero, consistent with reported mine curtailments
              at Bienfait and Boundary Dam (the gap is flagged in console output).
  2004–2007 → [SC48]: complete.
  2008–2013 → [SC46]: complete.
  2014–2018 → [SC46]: partially suppressed. Scaled using same method as 2002.
  2019+     → Zero (Boundary Dam and Bienfait operations ceased ~2019).

NEW BRUNSWICK
  2000–2003 → [SC48]: complete. Small production from the Minto colliery.
  2004–2009 → StatCan marks NB as suppressed ('x'), but Canada-level totals show
              small real Minto production through 2009. Allocated as the residual
              of the Canada total after subtracting the other four regions.
  2010+     → Zero (Minto closed ~2010; later 'x' codes are a survey artefact).

NOVA SCOTIA
  2000–2001 → [SC48]: complete.
  2002      → [SC48]: 6 of 12 months suppressed; scaled as above.
  2003+     → Zero (Phalen/Princess collieries closed ~2001).

REGIONAL METHODOLOGY — METALLURGICAL COAL PRODUCTION
------------------------------------------------------

Canada produces metallurgical (coking) coal almost exclusively in BC and Alberta.
Saskatchewan, New Brunswick, and Nova Scotia produce only thermal, lignite, or
subbituminous coal and are assigned zero metallurgical production for all years.

BRITISH COLUMBIA
  2000–2007 → [SC48] does not distinguish met from thermal bituminous. BC's met
              share is estimated using the first year available from [SC46]:
                  met_share_2008 = (BC_total − BC_thermal_bituminous) / BC_total
                                 ≈ 0.963  (Quinsam thermal mine still active)
              Applied to BC total production for 2000–2007.
  2008–2013 → [SC46] reports both "Bituminous, metallurgical" and "Bituminous,
              thermal" for BC with no suppression:
                  met = total − thermal
  2014–2024 → StatCan suppresses both met and thermal for BC from 2014 onward, so
              the split cannot be derived from [SC46]. A fixed share of 95% is applied,
              consistent with the BC Geological Survey statement that "over 95% of coal
              currently produced in BC is metallurgical coal."
              Source: https://www2.gov.bc.ca/gov/content/industry/mineral-exploration-mining/
                      british-columbia-geological-survey/geology/coal-overview

ALBERTA
  2000–2024 → [AER] Figure S7.1 "Metallurgical Bituminous" column throughout.
              [SC46] reports Alberta met coal but suppresses it from 2014 onward.
              AER provides a consistent series from 1999. For 2008–2013 the AER
              values differ from [SC46] by ≤2% (0.1 Mt rounding in AER).

CROSS-CHECK
-----------
StatCan [SC46] publishes a Canada-level "Bituminous, metallurgical" production
total for 2008–2013 (suppressed from 2014 onward). Our regional sum for those
years differs from this total by −408 to +317 kt (<1.5%), within the rounding
inherent in AER's 0.1 Mt resolution.

KNOWN GAPS AND CAVEATS
-----------------------
• 2003 Saskatchewan: ~10,700 kt appears in the StatCan Canada total but cannot
  be regionally allocated. Treated as zero; gap is flagged in console output.
• New Brunswick 2004–2009 is allocated from the Canada residual to reflect
  small suppressed Minto mine production. Later NB 'x' codes are treated as zero.
• 2019–2020: The sum of BC (Ministry) + Alberta (StatCan) falls ~3,700–4,200 kt
  short of the StatCan Canada total, likely reflecting rounding differences
  between BC Ministry and StatCan monthly data, and possibly small suppressed
  amounts in other provinces.
• BC 2021–2024 met coal assumes zero thermal output (consistent with post-2013
  trend); all BC production is classified as metallurgical.
• AER 2024 values are from the ST98 2025 edition (most recent published actuals).

================================================================================
"""

import sys
import polars as pl
from pathlib import Path
from openpyxl import load_workbook

_project_root = Path(__file__).parent.parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from utils.data_extensions import extend_cagr_periods, compute_cagr, load_cagr_assumptions, extend_constant
from utils.controls_conversions import load_control_config, DATA_START, PROJECTION_END, BASE_PATH

# ── File paths ───────────────────────────────────────────────────────────────────
MAPPINGS_PATH    = BASE_PATH / 'mappings_conversions'
REGION_MAP_FILE  = MAPPINGS_PATH / 'region_map.csv'
PATH_48          = BASE_PATH / 'raw_data/stats_can/activity/coal_production/25100048.csv'
PATH_46          = BASE_PATH / 'raw_data/stats_can/activity/coal_production/25100046.csv'
PATH_BC          = BASE_PATH / 'raw_data/bc_gov/bcannualcoalproduction.csv'
PATH_AER         = BASE_PATH / 'raw_data/ab_gov/st98-2025-coal-production-demand-data.xlsx'
OUTPUT           = BASE_PATH / 'processed_data/activity/coal_mining.csv'

_config        = load_control_config()
LAST_DATA_YEAR = _config["last_data_year"]
_archived      = _config["archived_data"]
EXTEND_TO      = PROJECTION_END
YEARS          = list(range(DATA_START, LAST_DATA_YEAR["stat_can_coal"] + 1))
REGIONS    = ['British Columbia', 'Alberta', 'Saskatchewan', 'New Brunswick', 'Nova Scotia']

ASSUMPTIONS_FILE = BASE_PATH / 'raw_data/assumptions/activity_cagr_projections.csv'
CAGR_START, CAGR_END, CAGR_PERIODS = load_cagr_assumptions('Coal Mining', ASSUMPTIONS_FILE)

# Per-region overrides — one explicit annual rate per period.
CAGR_OVERRIDES: dict[str, tuple[float, ...]] = {
}


def scale_partial(observed_sum: float, n_suppressed: int, n_total: int = 12) -> float:
    """Scale a partial-year sum to annual using available months."""
    available = n_total - n_suppressed
    if available > 0 and observed_sum > 0:
        return observed_sum * n_total / available
    return observed_sum


def residual_to_nonnegative(*parts: float) -> float:
    """Return a residual clipped at zero to avoid tiny negative rounding artefacts."""
    residual = parts[0] - sum(parts[1:])
    return residual if residual > 0 else 0.0


def main() -> pl.DataFrame:
    """Build coal mining DataFrame and write to CSV."""

    print("Loading data...")

    # StatCan 25-10-0048 — Polars lazy read; extract year from REF_DATE at load time.
    # VALUE is stored as String in the StatCan CSV (suppressed rows contain nothing,
    # not a numeric null), so cast to Float64 — invalid entries become null.
    sc48 = (
        pl.scan_csv(PATH_48)
        .with_columns([
            pl.col('REF_DATE').str.slice(0, 4).cast(pl.Int32).alias('year'),
            pl.col('VALUE').cast(pl.Float64, strict=False),
        ])
        .collect()
    )

    # StatCan 25-10-0046
    sc46 = (
        pl.scan_csv(PATH_46)
        .with_columns([
            pl.col('REF_DATE').str.slice(0, 4).cast(pl.Int32).alias('year'),
            pl.col('VALUE').cast(pl.Float64, strict=False),
        ])
        .collect()
    )

    # BC Ministry annual production (tonnes → kt)
    # The CSV has metadata footer rows; read everything as String so Polars doesn't
    # choke on non-numeric Year values, then cast with strict=False to drop them.
    bc_annual_kt: dict[int, float] = {
        int(r['year']): r['kt']
        for r in (
            pl.scan_csv(PATH_BC, infer_schema_length=0)   # all columns as String
            .select([
                pl.col('Year').cast(pl.Int32, strict=False).alias('year'),
                pl.col('Coal_tonnes')
                  .str.replace_all(',', '')
                  .str.strip_chars()
                  .cast(pl.Float64, strict=False)
                  .truediv(1000)                           # tonnes → kt
                  .alias('kt'),
            ])
            .drop_nulls()
            .collect()
            .to_dicts()
        )
    }

    # AER ST98 2025 — "Figures" sheet has an irregular layout that openpyxl handles
    # better than any DataFrame reader; keep openpyxl here.
    # Layout: col[1]=Year, col[2]=Subbituminous (Mt), col[3]=Thermal Bituminous (Mt),
    #         col[4]=Metallurgical Bituminous (Mt). Rows where col[1] > 2024 are forecasts.
    wb_aer = load_workbook(PATH_AER, read_only=True)
    ab_subbit_kt: dict[int, float] = {}
    ab_thermal_kt: dict[int, float] = {}
    ab_met_kt:    dict[int, float] = {}
    for row in wb_aer['Figures'].iter_rows(values_only=True):
        yr = row[1]
        if not isinstance(yr, (int, float)):
            continue
        yr = int(yr)
        if yr > 2024:
            continue
        if row[2] is not None: ab_subbit_kt[yr] = float(row[2]) * 1000   # Mt → kt
        if row[3] is not None: ab_thermal_kt[yr] = float(row[3]) * 1000
        if row[4] is not None: ab_met_kt[yr]    = float(row[4]) * 1000


    # ── Canada total (StatCan anchor, always complete) ───────────────────────────────
    # Stored as a plain dict {year: kt} — only ever used as a scalar lookup.

    _can48 = (
        sc48
        .filter(
            (pl.col('Coal and coke, components') == 'Total coal production, all types') &
            (pl.col('GEO') == 'Canada') &
            pl.col('year').is_between(2000, 2007)
        )
        .group_by('year')
        .agg(pl.col('VALUE').sum().alias('kt'))
    )

    _can46 = (
        sc46
        .filter(
            (pl.col('Coal types and uses') == 'Total all coal types and uses') &
            (pl.col('Coal volume') == 'Production') &
            (pl.col('GEO') == 'Canada') &
            pl.col('year').is_between(2008, LAST_DATA_YEAR["stat_can_coal"])
        )
        .group_by('year')
        .agg(pl.col('VALUE').sum().alias('kt'))
    )

    canada_kt: dict[int, float] = {
        r['year']: r['kt']
        for r in pl.concat([_can48, _can46]).to_dicts()
    }


    # ── Pre-aggregate StatCan lookups ────────────────────────────────────────────────
    # Materialise annual sums and missing-month counts for every (geo, year) combination
    # up front, then look up from dicts in the loop. This avoids re-filtering the full
    # DataFrames on every iteration.

    def _build_sc48_lookups() -> tuple[dict, dict]:
        """Returns (annual_sum, missing_count) dicts keyed by (geo, year)."""
        filt = sc48.filter(
            pl.col('Coal and coke, components') == 'Total coal production, all types'
        )
        sums = (
            filt.group_by(['GEO', 'year'])
            .agg(pl.col('VALUE').sum().alias('kt'))
            .to_dicts()
        )
        missing = (
            filt.filter(pl.col('VALUE').is_null())
            .group_by(['GEO', 'year'])
            .agg(pl.len().alias('n'))
            .to_dicts()
        )
        s = {(r['GEO'], r['year']): r['kt']  for r in sums}
        m = {(r['GEO'], r['year']): r['n']   for r in missing}
        return s, m

    def _build_sc46_lookups() -> tuple[dict, dict, dict]:
        """Returns (total_sum, thermal_sum, missing_count) dicts keyed by (geo, year)."""
        prod = sc46.filter(pl.col('Coal volume') == 'Production')

        total_sums = (
            prod.filter(pl.col('Coal types and uses') == 'Total all coal types and uses')
            .group_by(['GEO', 'year'])
            .agg(pl.col('VALUE').sum().alias('kt'))
            .to_dicts()
        )
        thermal_sums = (
            prod.filter(pl.col('Coal types and uses') == 'Bituminous, thermal')
            .group_by(['GEO', 'year'])
            .agg(pl.col('VALUE').sum().alias('kt'))
            .to_dicts()
        )
        missing = (
            prod.filter(
                (pl.col('Coal types and uses') == 'Total all coal types and uses') &
                pl.col('VALUE').is_null()
            )
            .group_by(['GEO', 'year'])
            .agg(pl.len().alias('n'))
            .to_dicts()
        )
        ts = {(r['GEO'], r['year']): r['kt'] for r in total_sums}
        th = {(r['GEO'], r['year']): r['kt'] for r in thermal_sums}
        m  = {(r['GEO'], r['year']): r['n']  for r in missing}
        return ts, th, m

    sc48_sums, sc48_miss = _build_sc48_lookups()
    sc46_sums, sc46_thermal, sc46_miss = _build_sc46_lookups()


    # ── Scalar lookup helpers ────────────────────────────────────────────────────────

    def sc48_sum(geo: str, year: int) -> float:
        """Annual production sum from 25-10-0048 (suppressed months count as 0)."""
        return sc48_sums.get((geo, year), 0.0)

    def sc48_missing(geo: str, year: int) -> int:
        """Count of suppressed (NaN) months in 25-10-0048 for a region-year."""
        return sc48_miss.get((geo, year), 0)

    def sc46_sum(geo: str, year: int) -> float:
        """Annual total production sum from 25-10-0046 (suppressed months count as 0)."""
        return sc46_sums.get((geo, year), 0.0)

    def sc46_thermal_sum(geo: str, year: int) -> float:
        """Annual thermal bituminous sum from 25-10-0046."""
        return sc46_thermal.get((geo, year), 0.0)

    def sc46_missing(geo: str, year: int) -> int:
        """Count of suppressed months in 25-10-0046 for a region-year."""
        return sc46_miss.get((geo, year), 0)


    # ── BC met share (2008) — anchor for 2000–2007 met coal estimate ─────────────────
    _bc_total_2008    = sc46_sum('British Columbia', 2008)
    _bc_thermal_2008  = sc46_thermal_sum('British Columbia', 2008)
    BC_MET_SHARE_2008 = (_bc_total_2008 - _bc_thermal_2008) / _bc_total_2008
    print(f"BC met share (2000–2007 estimate, from StatCan 25-10-0046 2008): {BC_MET_SHARE_2008:.4f}")

    # 2014–2024: StatCan suppresses both met and thermal for BC; split is unknown from
    # primary data. A fixed 95% share is used, consistent with the BC Geological Survey:
    # "Over 95% of coal currently produced in BC is metallurgical coal."
    # Source: https://www2.gov.bc.ca/gov/content/industry/mineral-exploration-mining/
    #         british-columbia-geological-survey/geology/coal-overview
    BC_MET_SHARE_2014_PLUS = 0.95


    # ── Build production tables ───────────────────────────────────────────────────────
    # The year-by-year conditional logic is inherently imperative; plain dicts are the
    # right structure here. The results are assembled into a Polars DataFrame at the end.

    print("Building production tables...")

    total: dict[tuple[int, str], float] = {}
    met:   dict[tuple[int, str], float] = {}

    for y in YEARS:

        # ── BRITISH COLUMBIA ──────────────────────────────────────────────────────

        if y in bc_annual_kt:
            total[y, 'British Columbia'] = bc_annual_kt[y]          # [BC] 2000–2020
        elif y <= 2007:
            total[y, 'British Columbia'] = sc48_sum('British Columbia', y)   # [SC48]
        # 2021–2024: filled via residual after all other regions are set

        if y <= 2007:
            # No met/thermal split in [SC48]; apply 2008 met share
            met[y, 'British Columbia'] = total[y, 'British Columbia'] * BC_MET_SHARE_2008
        elif y <= 2013:
            # [SC46] has separate met and thermal for BC, both complete
            bc_thm = sc46_thermal_sum('British Columbia', y)
            met[y, 'British Columbia'] = total[y, 'British Columbia'] - bc_thm
        else:
            # 2014+: StatCan suppresses the met/thermal split for BC.
            # Apply the 95% fixed share from the BC Geological Survey.
            # For 2021–2024 the total is not yet known (filled below); set to 0.0
            # as a placeholder — these entries are overwritten in the residual block.
            met[y, 'British Columbia'] = total.get((y, 'British Columbia'), 0.0) * BC_MET_SHARE_2014_PLUS

        # ── ALBERTA ───────────────────────────────────────────────────────────────

        if y <= 2007:
            total[y, 'Alberta'] = sc48_sum('Alberta', y)                    # [SC48]
        elif y <= 2013:
            total[y, 'Alberta'] = sc46_sum('Alberta', y)                    # [SC46]
        elif y <= 2019:
            # [SC46] suppressed 2014–2019; use [AER] sum of all types
            total[y, 'Alberta'] = (
                ab_met_kt.get(y, 0) + ab_thermal_kt.get(y, 0) + ab_subbit_kt.get(y, 0)
            )
        else:
            total[y, 'Alberta'] = sc46_sum('Alberta', y)                    # [SC46] 2020+

        met[y, 'Alberta'] = ab_met_kt.get(y, 0.0)                           # [AER] all years

        # ── SASKATCHEWAN ──────────────────────────────────────────────────────────

        if y <= 2001:
            total[y, 'Saskatchewan'] = sc48_sum('Saskatchewan', y)
        elif y == 2002:
            total[y, 'Saskatchewan'] = scale_partial(
                sc48_sum('Saskatchewan', y), sc48_missing('Saskatchewan', y)
            )
        elif y == 2003:
            total[y, 'Saskatchewan'] = 0.0   # all months suppressed; assumed zero
        elif y <= 2007:
            total[y, 'Saskatchewan'] = sc48_sum('Saskatchewan', y)
        elif y <= 2013:
            total[y, 'Saskatchewan'] = sc46_sum('Saskatchewan', y)
        elif y <= 2018:
            total[y, 'Saskatchewan'] = scale_partial(
                sc46_sum('Saskatchewan', y), sc46_missing('Saskatchewan', y)
            )
        else:
            total[y, 'Saskatchewan'] = 0.0   # mines closed ~2019

        met[y, 'Saskatchewan'] = 0.0   # no metallurgical coal

        # ── NEW BRUNSWICK ─────────────────────────────────────────────────────────

        if y <= 2003:
            val = sc48_sum('New Brunswick', y)
            total[y, 'New Brunswick'] = val if val > 0 else 0.0
        elif 2004 <= y <= 2009:
            # StatCan marks NB as suppressed ('x'), but Canada-level totals show
            # real Minto production through 2009. Allocate NB as the Canada residual
            # after subtracting the other four regions (NS is zero here, so 0.0).
            total[y, 'New Brunswick'] = residual_to_nonnegative(
                canada_kt[y],
                total[y, 'British Columbia'],
                total[y, 'Alberta'],
                total[y, 'Saskatchewan'],
                0.0,   # Nova Scotia
            )
        else:
            total[y, 'New Brunswick'] = 0.0   # Minto closed ~2010; later 'x' is artefactual

        met[y, 'New Brunswick'] = 0.0   # no metallurgical coal

        # ── NOVA SCOTIA ───────────────────────────────────────────────────────────

        if y <= 2001:
            total[y, 'Nova Scotia'] = sc48_sum('Nova Scotia', y)
        elif y == 2002:
            total[y, 'Nova Scotia'] = scale_partial(
                sc48_sum('Nova Scotia', y), sc48_missing('Nova Scotia', y)
            )
        else:
            total[y, 'Nova Scotia'] = 0.0   # collieries closed ~2001

        met[y, 'Nova Scotia'] = 0.0   # no metallurgical coal


    # ── BC residual from StatCan Canada total ────────────────────────────────────────

    for y in range(_archived["bc_coal"] + 1, LAST_DATA_YEAR["stat_can_coal"] + 1):
        bc_total = residual_to_nonnegative(
            canada_kt[y],
            total[y, 'Alberta'],
            total[y, 'Saskatchewan'],
            total[y, 'New Brunswick'],
            total[y, 'Nova Scotia'],
        )
        total[y, 'British Columbia'] = bc_total
        met[y,   'British Columbia'] = bc_total * BC_MET_SHARE_2014_PLUS   # 95% met share (BC Geological Survey)


    # ── Assemble raw data ─────────────────────────────────────────────────────────────

    df_raw = pl.DataFrame([
        {'Year': y, 'Region': r,
         'total': round(total[y, r], 1),
         'met':   round(met[y, r],   1)}
        for y in YEARS for r in REGIONS
    ])

    violations = df_raw.filter(pl.col('met') > pl.col('total') + 0.1)
    if violations.height > 0:
        for row in violations.to_dicts():
            print(f"  WARNING: met > total for {row['Region']} {row['Year']} "
                  f"(met={row['met']:.1f}, total={row['total']:.1f})")


    # ── Map NIR region names to CIMS codes ───────────────────────────────────────────

    region_map = pl.read_csv(REGION_MAP_FILE, columns=["CIMS", "NIR"])

    df_raw = (
        df_raw
        .join(region_map, left_on="Region", right_on="NIR", how="left")
        .with_columns(pl.col("CIMS").fill_null(pl.col("Region")).alias("Region"))
        .drop("CIMS")
    )


    # ── Build long-format combined DataFrame ─────────────────────────────────────────

    df_raw = df_raw.with_columns(
        pl.when(pl.col("total") > 0)
        .then(pl.col("met") / pl.col("total"))
        .otherwise(0.0)
        .alias("met_pct")
    )

    coal_total = (
        df_raw
        .select(["Region", "Year", pl.col("total").alias("Value")])
        .with_columns([
            pl.lit("Coal Mining").alias("Variable"),
            pl.lit("kt").alias("Unit"),
            pl.lit("Stats Can/Prov Gvts").alias("Source"),
        ])
        .select(["Region", "Variable", "Unit", "Source", "Year", "Value"])
    )

    coal_met_pct = (
        df_raw
        .select(["Region", "Year", pl.col("met_pct").alias("Value")])
        .with_columns([
            pl.lit("Coal Mining.Coal.Metallurgical Finishing").alias("Variable"),
            pl.lit("% of kt").alias("Unit"),
            pl.lit("Stats Can/Prov Gvts").alias("Source"),
        ])
        .select(["Region", "Variable", "Unit", "Source", "Year", "Value"])
    )

    combined = pl.concat([coal_total, coal_met_pct]).sort(["Region", "Variable", "Year"])


    # ── Console cross-check (historical) ─────────────────────────────────────────────

    regional_sums = (
        combined
        .filter((pl.col("Variable") == "Coal Mining") & (pl.col("Year") <= LAST_DATA_YEAR["stat_can_coal"]))
        .group_by("Year")
        .agg(pl.col("Value").sum().alias("our_sum"))
        .sort("Year")
    )

    print("\nCross-check: regional sum vs. StatCan Canada total (kt)")
    print(f"{'Year':>6}  {'Our Sum':>10}  {'StatCan':>10}  {'Gap':>8}")
    for row in regional_sums.to_dicts():
        y    = row['Year']
        our  = row['our_sum']
        sc   = canada_kt[y]
        gap  = our - sc
        flag = "  ← see methodology notes" if abs(gap) > 1000 else ""
        print(f"{y:>6}  {our:>10,.1f}  {sc:>10,.1f}  {gap:>+8,.1f}{flag}")


    # ── Extend to 2100 using CAGR with dampeners ─────────────────────────────────────
    #
    #   Coal Mining total: raw CAGR computed over CAGR_START–CAGR_END (2014–2024),
    #   then applied with near-zero dampeners across three projection periods so that
    #   growth effectively stops. Metallurgical Finishing share (% of kt): held flat
    #   via extend_constant.

    last_year = combined["Year"].max()

    future_rows = []
    for keys, grp in combined.group_by(["Region", "Variable", "Unit"], maintain_order=True):
        region, variable, unit = keys
        s = grp.sort("Year").to_pandas().set_index("Year")["Value"]

        base_val = s.loc[CAGR_END] if CAGR_END in s.index else s.get(int(last_year), 0.0)
        if unit == "% of kt" or base_val <= 0:
            extended = extend_constant(s, end_year=EXTEND_TO)
            for year, value in extended.loc[last_year + 1:].items():
                future_rows.append({
                    "Region": region, "Variable": variable, "Unit": unit,
                    "Source": "Assumptions",
                    "Year": int(year), "Value": float(value),
                })
        else:
            raw_cagr = compute_cagr(s, CAGR_START, CAGR_END)
            projected = extend_cagr_periods(
                base_val,
                raw_cagr,
                CAGR_PERIODS,
                CAGR_OVERRIDES.get(region),
            )
            for year, value in projected.items():
                if int(year) <= int(last_year):
                    continue
                future_rows.append({
                    "Region": region, "Variable": variable, "Unit": unit,
                    "Source": "Assumptions",
                    "Year": int(year), "Value": float(value),
                })

    combined = pl.concat([combined, pl.DataFrame(future_rows)]).sort(["Region", "Variable", "Year"])


    # ── Save ──────────────────────────────────────────────────────────────────────────

    print("\nWriting output CSV...")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    combined.write_csv(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"Rows:          {combined.height:,}")
    print(f"Regions:       {combined['Region'].n_unique()}")
    print(f"Variables:     {sorted(combined['Variable'].unique().to_list())}")
    print(f"Years covered: {combined['Year'].min()} - {combined['Year'].max()}")

    return combined


if __name__ == '__main__':
    main()
